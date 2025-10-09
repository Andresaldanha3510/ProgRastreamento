import eventlet
eventlet.monkey_patch(psycopg=False) 


import uuid
import xml.etree.ElementTree as ET
import json
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, make_response, send_from_directory, Response
from datetime import datetime, timedelta, date
import requests
import logging
import os
import math
import re
from dotenv import load_dotenv
import boto3
from werkzeug.utils import secure_filename
import io
from openpyxl import Workbook
from sqlalchemy.exc import IntegrityError
from sqlalchemy import or_, and_, func, extract, UniqueConstraint, LargeBinary
from flask_login import UserMixin, login_user, logout_user, login_required, current_user
from num2words import num2words
from collections import defaultdict
import pytesseract
from PIL import Image
import openrouteservice
import click
from pathlib import Path
from functools import wraps
from cryptography.fernet import Fernet
from werkzeug.security import generate_password_hash, check_password_hash
from cryptography.hazmat.primitives.serialization.pkcs12 import load_key_and_certificates
from cryptography.x509.oid import NameOID
import zipfile
from flask import send_file
from flask_socketio import emit, join_room, leave_room
from xhtml2pdf import pisa
from extensions import db, CTeEmitido
from flask_mail import Message
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle
from openpyxl.utils import get_column_letter




# 1. Importa as extensões E os modelos do novo arquivo 'extensions.py'
from extensions import db, migrate, socketio, login_manager, mail, CertificadoDigital, NFeImportada

# Seus decoradores (sem alterações)
# ... (cole seus 3 decoradores aqui: admin_required, owner_required, master_required) ...

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin:
            flash('Acesso restrito ao administrador.', 'error')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

def owner_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role != 'Owner':
            flash('Acesso restrito ao proprietário do sistema.', 'error')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

def master_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or (current_user.role not in ['Admin', 'Master']):
            flash('Acesso restrito a administradores ou masters.', 'error')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function
    
# Lógica do .env (sem alterações)
env_path = Path(__file__).resolve().with_name('.env')
if not env_path.exists(): raise FileNotFoundError(f'Arquivo .env não encontrado em {env_path}')
load_dotenv(dotenv_path=env_path)

# ... (seu código de validação do .env continua aqui) ...
required_r2 = [
    'CLOUDFLARE_R2_ENDPOINT', 'CLOUDFLARE_R2_ACCESS_KEY', 'CLOUDFLARE_R2_SECRET_KEY',
    'CLOUDFLARE_R2_BUCKET', 'CLOUDFLARE_R2_PUBLIC_URL',
]
missing = [k for k in required_r2 if not os.getenv(k)]
if missing:
    raise ValueError('Variáveis faltando no .env: ' + ', '.join(missing))

print('R2 config carregada:')
for k in required_r2:
    print(f'  {k}: {os.getenv(k)}')

# Criação e configuração da aplicação
app = Flask(__name__)
app.config.update(
    # ... (seu bloco de config continua aqui) ...
    MAIL_SERVER='smtp.gmail.com', MAIL_PORT=587, MAIL_USE_TLS=True,
    MAIL_USERNAME='trackgo789@gmail.com', MAIL_PASSWORD='mmoa moxc juli sfbe',
    MAIL_DEFAULT_SENDER='trackgo789@gmail.com',
    SQLALCHEMY_DATABASE_URI=os.getenv('DATABASE_URL', 'sqlite:///transport.db'),
    SQLALCHEMY_TRACK_MODIFICATIONS=False,
    SECRET_KEY=os.getenv('SECRET_KEY', 'w9z$kL2mNpQvR7tYxJ3hF8gWcPqB5vM2nZ4rT6yU'),
    CLOUDFLARE_R2_ENDPOINT=os.getenv('CLOUDFLARE_R2_ENDPOINT'),
    CLOUDFLARE_R2_ACCESS_KEY=os.getenv('CLOUDFLARE_R2_ACCESS_KEY'),
    CLOUDFLARE_R2_SECRET_KEY=os.getenv('CLOUDFLARE_R2_SECRET_KEY'),
    CLOUDFLARE_R2_BUCKET=os.getenv('CLOUDFLARE_R2_BUCKET'),
    CLOUDFLARE_R2_PUBLIC_URL=(lambda url: url.split('=', 1)[-1])(os.getenv('CLOUDFLARE_R2_PUBLIC_URL', '')),
    SEFAZ_AMBIENTE=os.getenv('SEFAZ_AMBIENTE', 'PRODUCAO'),
    NFE_API_URL=os.getenv('NFE_API_URL'),
)
GEOAPIFY_API_KEY = os.getenv('GEOAPIFY_API_KEY', '7cd423ef184f48f0b770682cbebe11d0')
OPENROUTESERVICE_API_KEY = os.getenv('OPENROUTESERVICE_API_KEY')
Maps_API_KEY = os.getenv('Maps_API_KEY')

# Lógica da chave de criptografia
ENCRYPTION_KEY = os.getenv('ENCRYPTION_KEY')
if not ENCRYPTION_KEY: raise ValueError("ENCRYPTION_KEY não definida no .env!")
cipher_suite = Fernet(ENCRYPTION_KEY.encode())
app.cipher_suite = cipher_suite # Associa o cipher ao app para uso posterior



# 2. Seção correta de inicialização das extensões
db.init_app(app)
migrate.init_app(app, db)
socketio.init_app(app, async_mode='eventlet')
mail.init_app(app)
login_manager.init_app(app)
login_manager.login_view = 'login'

# Configurações de logging e variáveis globais
last_geocode_time = {}
GEOCODE_INTERVAL_SECONDS = 600
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# 3. Import do sefaz_service após a inicialização
from sefaz_service import consultar_notas_sefaz

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(Usuario, int(user_id))

class Motorista(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    empresa_id = db.Column(db.Integer, db.ForeignKey('empresa.id'), nullable=False)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    # Aba: Dados Pessoais
    nome = db.Column(db.String(100), nullable=False)
    telefone = db.Column(db.String(20), nullable=True)
    cpf = db.Column(db.String(14), unique=True, nullable=False, index=True)
    data_nascimento = db.Column(db.Date, nullable=True)  # Mantido como opcional para flexibilidade
    nacionalidade = db.Column(db.String(50), default='Nacional')
    naturalidade = db.Column(db.String(100), nullable=True)
    estado_civil = db.Column(db.String(50), nullable=True)
    sexo = db.Column(db.String(20), nullable=True)
    pai = db.Column(db.String(100), nullable=True)
    mae = db.Column(db.String(100), nullable=True)
    data_admissao = db.Column(db.Date, nullable=True)
    situacao = db.Column(db.String(50), default='NORMAL / LIBERADO')
    data_desativacao = db.Column(db.Date, nullable=True)
    classificacao = db.Column(db.String(50), nullable=True)
    cod_departamento = db.Column(db.String(50), nullable=True)
    numero_ficha = db.Column(db.String(50), nullable=True)
    foto = db.Column(db.String(500), nullable=True)
    anexos = db.Column(db.String(2048), nullable=True)

    # Aba: Endereço
    cep = db.Column(db.String(9), nullable=True)
    tipo_logradouro = db.Column(db.String(50), nullable=True)
    logradouro = db.Column(db.String(255), nullable=True)
    numero = db.Column(db.String(20), nullable=True)
    complemento = db.Column(db.String(100), nullable=True)
    bairro = db.Column(db.String(100), nullable=True)
    cidade = db.Column(db.String(100), nullable=True)
    uf = db.Column(db.String(2), nullable=True)
    email = db.Column(db.String(120), nullable=True)
    tipo_imovel = db.Column(db.String(50), nullable=True)
    tempo_residencia = db.Column(db.String(50), nullable=True)

    # Aba: Documentação
    cnh_numero = db.Column(db.String(30), unique=True, nullable=False, index=True)
    cnh_data_primeira = db.Column(db.Date, nullable=True)
    
    
    cnh_data_vencimento = db.Column(db.Date, nullable=True)
    cnh_categoria = db.Column(db.String(10), nullable=True)
    # --- FIM DA CORREÇÃO ---
    
    cnh_cod_seguranca = db.Column(db.String(20), nullable=True)
    rg = db.Column(db.String(20), nullable=True)
    rg_uf = db.Column(db.String(2), nullable=True)
    pis = db.Column(db.String(20), nullable=True)
    inss = db.Column(db.String(20), nullable=True)
    titulo_eleitor = db.Column(db.String(20), nullable=True)
    ctps = db.Column(db.String(20), nullable=True)
    funcao = db.Column(db.String(100), nullable=True)
    mopp_numero = db.Column(db.String(20), nullable=True)
    mopp_vencimento = db.Column(db.Date, nullable=True)
    salario_base = db.Column(db.Float, nullable=True, default=0.0)

    # Aba: Contatos de Referência
    contato_nome = db.Column(db.String(100), nullable=True)
    contato_tipo_ref = db.Column(db.String(50), nullable=True)
    contato_tipo_fone = db.Column(db.String(50), nullable=True)
    contato_telefone = db.Column(db.String(20), nullable=True)
    contato_operadora = db.Column(db.String(50), nullable=True)
    contato_obs = db.Column(db.String(255), nullable=True)


    
    
    usuario = db.relationship('Usuario', backref='motorista', uselist=False)
    viagens = db.relationship('Viagem', foreign_keys='Viagem.motorista_id', backref='motorista_formal')

    # Propriedades para compatibilidade
    @property
    def cpf_cnpj(self):
        return self.cpf

    @property
    def cnh(self):
        return self.cnh_numero

    @property
    def validade_cnh(self):
        return self.cnh_data_vencimento
    
    @property
    def validade_mopp(self):
        return self.mopp_vencimento

# Em app.py

# Em app.py

class UnidadeNegocio(db.Model):
    """Representa uma empresa, filial ou centro de custo para lançamentos financeiros."""
    __tablename__ = 'unidade_negocio'
    id = db.Column(db.Integer, primary_key=True)
    empresa_id = db.Column(db.Integer, db.ForeignKey('empresa.id'), nullable=False)
    
    nome = db.Column(db.String(150), nullable=False)
    razao_social = db.Column(db.String(255), nullable=True)
    cnpj = db.Column(db.String(14), unique=True, nullable=True)
    is_matriz = db.Column(db.Boolean, default=False, nullable=False)
    
    lancamentos = db.relationship('LancamentoFluxoCaixa', back_populates='unidade_negocio', lazy='dynamic')
    
    __table_args__ = (db.UniqueConstraint('empresa_id', 'nome', name='_empresa_nome_uc'),)

    def __repr__(self):
        return f'<UnidadeNegocio {self.nome}>'

    # ESTE MÉTODO É A SOLUÇÃO
    def to_dict(self):
        """Converte o objeto UnidadeNegocio para um dicionário serializável."""
        return {
            'id': self.id,
            'empresa_id': self.empresa_id,
            'nome': self.nome,
            'razao_social': self.razao_social,
            'cnpj': self.cnpj,
            'is_matriz': self.is_matriz
        }

class ConfiguracaoBorracharia(db.Model):
    __tablename__ = 'configuracao_borracharia'
    id = db.Column(db.Integer, primary_key=True)
    empresa_id = db.Column(db.Integer, db.ForeignKey('empresa.id'), nullable=False, unique=True)
    
    # Configurações de vida útil e alertas
    vida_util_dot_anos = db.Column(db.Integer, default=7)
    alerta_dot_dias = db.Column(db.Integer, default=90)
    
    # Configurações de sulco
    sulco_minimo_recapagem_mm = db.Column(db.Float, default=3.0)
    sulco_minimo_descarte_mm = db.Column(db.Float, default=1.6)
    
    # Configurações de recapagem
    max_recapagens = db.Column(db.Integer, default=2)
    km_alerta_recapagem = db.Column(db.Integer, default=5000)
    
    # Relacionamento
    empresa = db.relationship('Empresa', backref=db.backref('config_borracharia', uselist=False))
    
    def __repr__(self):
        return f'<ConfiguracaoBorracharia {self.empresa_id}>'

class Pneu(db.Model):
    __tablename__ = 'pneu'
    id = db.Column(db.Integer, primary_key=True)
    empresa_id = db.Column(db.Integer, db.ForeignKey('empresa.id'), nullable=False)
    veiculo_id = db.Column(db.Integer, db.ForeignKey('veiculo.id'), nullable=True)
    
    # Identificação do pneu
    numero_fogo = db.Column(db.String(50), nullable=False)
    marca = db.Column(db.String(100), nullable=False)
    modelo = db.Column(db.String(100), nullable=False)
    dimensao = db.Column(db.String(50), nullable=False)
    dot = db.Column(db.String(20), nullable=False)
    
    # Dados de compra
    data_compra = db.Column(db.Date, nullable=False)
    valor_compra = db.Column(db.Float, nullable=False)
    fornecedor = db.Column(db.String(200), nullable=True)
    
    # Status e localização
    status = db.Column(db.String(50), default='Estoque', nullable=False, index=True)
    # Status: Estoque, Em Uso, Recapando, Descartado
    posicao = db.Column(db.String(100), nullable=True)
    # Ex: "Eixo 1 - Dianteiro Esquerdo"
    
    # Dados de uso
    km_instalacao = db.Column(db.Float, nullable=True)
    data_instalacao = db.Column(db.Date, nullable=True)
    km_remocao = db.Column(db.Float, nullable=True)
    data_remocao = db.Column(db.Date, nullable=True)
    
    # Recapagem
    numero_recapagens = db.Column(db.Integer, default=0)
    data_ultima_recapagem = db.Column(db.Date, nullable=True)
    custo_total_recapagens = db.Column(db.Float, default=0.0)
    
    # Medições de sulco
    sulco_atual_mm = db.Column(db.Float, nullable=True)
    data_ultima_medicao = db.Column(db.Date, nullable=True)
    
    # Motivo de descarte
    motivo_descarte = db.Column(db.Text, nullable=True)
    data_descarte = db.Column(db.Date, nullable=True)
    
    # Timestamps
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    # Relacionamentos
    empresa = db.relationship('Empresa', backref='pneus')
    veiculo = db.relationship('Veiculo', backref='pneus')
    medicoes = db.relationship('MedicaoSulco', backref='pneu', cascade="all, delete-orphan")
    movimentacoes = db.relationship('MovimentacaoPneu', backref='pneu', cascade="all, delete-orphan")
    
    # Constraint para evitar duplicatas
    __table_args__ = (
        db.UniqueConstraint('empresa_id', 'numero_fogo', name='_empresa_numero_fogo_uc'),
    )
    
    @property
    def idade_anos(self):
        """Calcula a idade do pneu baseada no DOT"""
        try:
            # Extrai os últimos 4 dígitos do DOT (semana + ano)
            if len(self.dot) >= 4:
                ano_fabricacao = int(self.dot[-2:])
                # Se for menor que 30, assume 20xx, senão 19xx
                if ano_fabricacao < 30:
                    ano_fabricacao += 2000
                else:
                    ano_fabricacao += 1900
                
                ano_atual = datetime.now().year
                return ano_atual - ano_fabricacao
        except:
            pass
        return 0
    
    @property
    def precisa_alerta_dot(self):
        """Verifica se precisa alerta por idade"""
        config = ConfiguracaoBorracharia.query.filter_by(empresa_id=self.empresa_id).first()
        if not config:
            return False
        return self.idade_anos >= config.vida_util_dot_anos
    
    @property
    def precisa_recapagem(self):
        """Verifica se precisa ir para recapagem"""
        config = ConfiguracaoBorracharia.query.filter_by(empresa_id=self.empresa_id).first()
        if not config or not self.sulco_atual_mm:
            return False
        return self.sulco_atual_mm <= config.sulco_minimo_recapagem_mm
    
    @property
    def precisa_descarte(self):
        """Verifica se precisa ser descartado"""
        config = ConfiguracaoBorracharia.query.filter_by(empresa_id=self.empresa_id).first()
        if not config:
            return False
        
        # Por sulco
        if self.sulco_atual_mm and self.sulco_atual_mm <= config.sulco_minimo_descarte_mm:
            return True
        
        # Por número máximo de recapagens
        if self.numero_recapagens >= config.max_recapagens:
            return True
        
        return False
    
    def __repr__(self):
        return f'<Pneu {self.numero_fogo}>'

class MedicaoSulco(db.Model):
    __tablename__ = 'medicao_sulco'
    id = db.Column(db.Integer, primary_key=True)
    pneu_id = db.Column(db.Integer, db.ForeignKey('pneu.id'), nullable=False)
    
    sulco_mm = db.Column(db.Float, nullable=False)
    data_medicao = db.Column(db.Date, default=date.today)
    observacoes = db.Column(db.Text, nullable=True)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=True)
    
    usuario = db.relationship('Usuario')
    
    def __repr__(self):
        return f'<MedicaoSulco {self.pneu_id}: {self.sulco_mm}mm>'

class MovimentacaoPneu(db.Model):
    __tablename__ = 'movimentacao_pneu'
    id = db.Column(db.Integer, primary_key=True)
    pneu_id = db.Column(db.Integer, db.ForeignKey('pneu.id'), nullable=False)
    
    tipo_movimentacao = db.Column(db.String(50), nullable=False)
    # Tipos: INSTALACAO, REMOCAO, RECAPAGEM_ENVIO, RECAPAGEM_RETORNO, DESCARTE
    
    data_movimentacao = db.Column(db.DateTime, default=datetime.utcnow)
    veiculo_id = db.Column(db.Integer, db.ForeignKey('veiculo.id'), nullable=True)
    posicao_origem = db.Column(db.String(100), nullable=True)
    posicao_destino = db.Column(db.String(100), nullable=True)
    km_veiculo = db.Column(db.Float, nullable=True)
    
    # Dados específicos para recapagem
    custo_recapagem = db.Column(db.Float, nullable=True)
    fornecedor_recapagem = db.Column(db.String(200), nullable=True)
    
    observacoes = db.Column(db.Text, nullable=True)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=True)
    
    veiculo = db.relationship('Veiculo')
    usuario = db.relationship('Usuario')
    
    def __repr__(self):
        return f'<MovimentacaoPneu {self.tipo_movimentacao}>'
    

    
class FolhaPagamento(db.Model):
    __tablename__ = 'folha_pagamento'
    id = db.Column(db.Integer, primary_key=True)
    motorista_id = db.Column(db.Integer, db.ForeignKey('motorista.id'), nullable=False)
    empresa_id = db.Column(db.Integer, db.ForeignKey('empresa.id'), nullable=False)
    
    mes_referencia = db.Column(db.Integer, nullable=False, index=True)
    ano_referencia = db.Column(db.Integer, nullable=False, index=True)
    
    salario_base_registro = db.Column(db.Float, default=0.0) # Salva o salário base no momento da criação
    total_proventos = db.Column(db.Float, default=0.0)
    total_descontos = db.Column(db.Float, default=0.0)
    salario_liquido = db.Column(db.Float, default=0.0)
    
    status = db.Column(db.String(20), default='Em Aberto', index=True) # Em Aberto, Fechada, Paga
    data_criacao = db.Column(db.DateTime, default=datetime.utcnow)
    data_fechamento = db.Column(db.DateTime, nullable=True)
    data_pagamento = db.Column(db.DateTime, nullable=True)
    meio_pagamento = db.Column(db.String(50), nullable=True)
    comprovante_url = db.Column(db.String(500), nullable=True)
    observacoes = db.Column(db.Text, nullable=True)
    viagem_id = db.Column(db.Integer, db.ForeignKey('viagem.id'), nullable=True)
    viagem = db.relationship('Viagem') # Facilita o acesso aos dados da viagem

    motorista = db.relationship('Motorista', backref='folhas_pagamento')
    itens = db.relationship('ItemFolhaPagamento', backref='folha', lazy='dynamic', cascade="all, delete-orphan")

    __table_args__ = (db.UniqueConstraint('motorista_id', 'mes_referencia', 'ano_referencia', name='_motorista_mes_ano_uc'),)

    def __repr__(self):
        return f'<FolhaPagamento {self.id} para Motorista {self.motorista_id} - {self.mes_referencia}/{self.ano_referencia}>'

    def recalcular_totais(self):
        """Recalcula os totais com base nos itens."""
        proventos = db.session.query(func.sum(ItemFolhaPagamento.valor)).filter(
            ItemFolhaPagamento.folha_pagamento_id == self.id,
            ItemFolhaPagamento.tipo == 'Provento'
        ).scalar() or 0.0

        descontos = db.session.query(func.sum(ItemFolhaPagamento.valor)).filter(
            ItemFolhaPagamento.folha_pagamento_id == self.id,
            ItemFolhaPagamento.tipo == 'Desconto'
        ).scalar() or 0.0
        
        self.total_proventos = self.salario_base_registro + proventos
        self.total_descontos = descontos
        self.salario_liquido = self.total_proventos - self.total_descontos


class ItemFolhaPagamento(db.Model):
    __tablename__ = 'item_folha_pagamento'
    id = db.Column(db.Integer, primary_key=True)
    folha_pagamento_id = db.Column(db.Integer, db.ForeignKey('folha_pagamento.id'), nullable=False)
    
    # --- LINHA ADICIONADA AQUI ---
    viagem_id = db.Column(db.Integer, db.ForeignKey('viagem.id'), nullable=True, index=True)
    
    tipo = db.Column(db.String(10), nullable=False, index=True) # 'Provento' ou 'Desconto'
    descricao = db.Column(db.String(255), nullable=False)
    valor = db.Column(db.Float, nullable=False)

    def __repr__(self):
        return f'<ItemFolhaPagamento {self.id} - {self.tipo}: {self.descricao}>'

class Manutencao(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    veiculo_id = db.Column(db.Integer, db.ForeignKey('veiculo.id'), nullable=False)
    empresa_id = db.Column(db.Integer, db.ForeignKey('empresa.id'), nullable=False)
    itens = db.relationship('ManutencaoItem', backref='manutencao', lazy=True, cascade="all, delete-orphan")

    data_entrada = db.Column(db.DateTime, default=datetime.utcnow)
    data_saida = db.Column(db.DateTime, nullable=True)
    odometro = db.Column(db.BigInteger, nullable=False)
    custo_total = db.Column(db.Float, nullable=True)
    servicos_executados = db.Column(db.Text, nullable=True)
    status = db.Column(db.String(50), nullable=False) # Ex: Agendada, Em Andamento, Concluída
    tipo_manutencao = db.Column(db.String(50), nullable=False, default='Corretiva') # Ex: Preventiva, Corretiva
    descricao_problema = db.Column(db.Text, nullable=True)

    veiculo_plano_veiculo_id = db.Column(db.Integer)
    veiculo_plano_plano_id = db.Column(db.Integer)

    __table_args__ = (db.ForeignKeyConstraint(
        ['veiculo_plano_veiculo_id', 'veiculo_plano_plano_id'],
        ['veiculo_plano.veiculo_id', 'veiculo_plano.plano_id'],
    ),)

    veiculo = db.relationship('Veiculo', back_populates='manutencoes')
    veiculo_plano_associado = db.relationship('VeiculoPlano', back_populates='manutencoes')

    def __repr__(self):
        return f'<Manutencao {self.id} para Veiculo {self.veiculo_id}>'


class Licenca(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    empresa_id = db.Column(db.Integer, db.ForeignKey('empresa.id'), nullable=False, unique=True)
    plano = db.Column(db.String(50), nullable=False, default='Básico')
    max_usuarios = db.Column(db.Integer, nullable=False, default=5)
    max_veiculos = db.Column(db.Integer, nullable=False, default=5)
    data_expiracao = db.Column(db.Date, nullable=True)
    ativo = db.Column(db.Boolean, default=True, nullable=False)

    def __repr__(self):
        return f'<Licenca {self.id} - Plano {self.plano} para Empresa {self.empresa_id}>'

class Convite(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(120), nullable=False, index=True)
    token = db.Column(db.String(36), unique=True, nullable=False, index=True)
    usado = db.Column(db.Boolean, default=False)
    data_criacao = db.Column(db.DateTime, default=datetime.utcnow)
    data_expiracao = db.Column(db.DateTime, nullable=False)
    criado_por = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    role = db.Column(db.String(20), nullable=False, default='Motorista')
    empresa_id = db.Column(db.Integer, db.ForeignKey('empresa.id'), nullable=True)

class Empresa(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    razao_social = db.Column(db.String(200), nullable=False)
    nome_fantasia = db.Column(db.String(200), nullable=True)
    cnpj = db.Column(db.String(14), unique=True, nullable=False, index=True)
    inscricao_estadual = db.Column(db.String(20), nullable=True)
    endereco = db.Column(db.String(255), nullable=False)
    cidade = db.Column(db.String(100), nullable=False)
    estado = db.Column(db.String(2), nullable=False)
    cep = db.Column(db.String(8), nullable=False)
    telefone = db.Column(db.String(11), nullable=True)
    email_contato = db.Column(db.String(120), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    licenca = db.relationship('Licenca', backref='empresa', uselist=False, cascade="all, delete-orphan")
    usuarios = db.relationship('Usuario', backref='empresa', lazy=True)

    def __repr__(self):
        return f'<Empresa {self.razao_social}>'

class Cobranca(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    cliente_id = db.Column(db.Integer, db.ForeignKey('cliente.id'), nullable=False)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    valor_total = db.Column(db.Float, nullable=False)
    data_emissao = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    data_vencimento = db.Column(db.Date, nullable=False)
    data_pagamento = db.Column(db.DateTime, nullable=True)
    status = db.Column(db.String(20), nullable=False, default='Pendente', index=True)
    meio_pagamento = db.Column(db.String(50), nullable=True)
    observacoes = db.Column(db.Text, nullable=True)
    
    # --- ADICIONADO ---
    empresa_id = db.Column(db.Integer, db.ForeignKey('empresa.id'), nullable=False)
    
    cliente = db.relationship('Cliente', backref='cobrancas')
    
    # ▼▼▼ LINHA CORRIGIDA PARA RESOLVER O PROBLEMA DE EXCLUSÃO ▼▼▼
    usuario = db.relationship('Usuario', backref=db.backref('cobrancas_geradas', cascade="all, delete-orphan"))
    
    viagens = db.relationship('Viagem', backref='cobranca', lazy='dynamic')

    @property
    def is_vencida(self):
        return self.data_vencimento < datetime.utcnow().date() and self.status == 'Pendente'

    def __repr__(self):
        return f'<Cobranca {self.id} - Cliente {self.cliente.nome_razao_social}>'

class CentroCusto(db.Model):
    __tablename__ = 'centro_custo'
    id = db.Column(db.Integer, primary_key=True)
    empresa_id = db.Column(db.Integer, db.ForeignKey('empresa.id'), nullable=False)
    nome = db.Column(db.String(150), nullable=False)

    __table_args__ = (db.UniqueConstraint('empresa_id', 'nome', name='_empresa_nome_cc_uc'),)

    def to_dict(self):
        return {'id': self.id, 'nome': self.nome}

    def __repr__(self):
        return f'<CentroCusto {self.nome}>'

class Romaneio(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    data_emissao = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    observacoes = db.Column(db.Text, nullable=True)
    viagem_id = db.Column(db.Integer, db.ForeignKey('viagem.id'), nullable=False, unique=True)
    
    # --- ADICIONADO ---
    empresa_id = db.Column(db.Integer, db.ForeignKey('empresa.id'), nullable=False)
    
    viagem = db.relationship('Viagem', backref=db.backref('romaneio', uselist=False))
    itens = db.relationship('ItemRomaneio', backref='romaneio', lazy=True, cascade="all, delete-orphan")

    @property
    def total_volumes(self):
        return len(self.itens) if self.itens else 0

    @property
    def peso_total(self):
        return sum(item.peso_total_item for item in self.itens) if self.itens else 0.0
    
class RateioVeiculo(db.Model):
    """Modelo para rateio de despesas entre veículos"""
    __tablename__ = 'rateio_veiculos'
    
    id = db.Column(db.Integer, primary_key=True)
    
    # CORREÇÃO APLICADA AQUI: de 'lancamentos_fluxo_caixa.id' para 'lancamento_fluxo_caixa.id'
    lancamento_id = db.Column(db.Integer, db.ForeignKey('lancamento_fluxo_caixa.id'), nullable=False)
    
    # Esta linha você já corrigiu perfeitamente!
    veiculo_id = db.Column(db.Integer, db.ForeignKey('veiculo.id'), nullable=False)
    
    valor_rateado = db.Column(db.Numeric(10, 2), nullable=False)
    percentual = db.Column(db.Numeric(5, 2), nullable=False)
    data_criacao = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relacionamentos
    lancamento = db.relationship('LancamentoFluxoCaixa', backref=db.backref('rateios', cascade="all, delete-orphan"))
    veiculo = db.relationship('Veiculo', backref='rateios')


class ItemRomaneio(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    produto_descricao = db.Column(db.String(255), nullable=False)
    quantidade = db.Column(db.Integer, nullable=False, default=1)
    embalagem = db.Column(db.String(50), nullable=True)
    peso_kg = db.Column(db.Float, nullable=True, default=0.0)
    romaneio_id = db.Column(db.Integer, db.ForeignKey('romaneio.id'), nullable=False)

    @property
    def peso_total_item(self):
        return (self.peso_kg or 0.0) * (self.quantidade or 0)

class Veiculo(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    empresa_id = db.Column(db.Integer, db.ForeignKey('empresa.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    placa = db.Column(db.String(7), unique=True, nullable=False, index=True)
    modelo = db.Column(db.String(100), nullable=False)
    categoria = db.Column(db.String(50), nullable=True)
    status = db.Column(db.String(50), nullable=True, default='Disponível')
    marca = db.Column(db.String(100), nullable=True)
    ano_fabricacao = db.Column(db.Integer, nullable=True)
    ano_modelo = db.Column(db.Integer, nullable=True)
    cor = db.Column(db.String(50), nullable=True)
    combustivel = db.Column(db.String(50), nullable=True)
    is_administrativo = db.Column(db.Boolean, default=False, nullable=False, index=True)


    renavam = db.Column(db.String(11), nullable=True, unique=True)
    chassi = db.Column(db.String(17), nullable=True, unique=True)
    numero_motor = db.Column(db.String(50), nullable=True)
    crlv_numero = db.Column(db.String(50), nullable=True)
    crlv_vencimento = db.Column(db.Date, nullable=True)
    seguro_numero = db.Column(db.String(100), nullable=True)
    seguro_seguradora = db.Column(db.String(100), nullable=True)
    seguro_vencimento = db.Column(db.Date, nullable=True)

    capacidade_carga_kg = db.Column(db.Float, nullable=True)
    peso_bruto_total_kg = db.Column(db.Float, nullable=True)
    eixos = db.Column(db.Integer, nullable=True)
    cilindrada = db.Column(db.String(20), nullable=True)
    potencia_cv = db.Column(db.Integer, nullable=True)
    tanque_combustivel_litros = db.Column(db.Integer, nullable=True)
    consumo_medio_km_l = db.Column(db.Float, nullable=True)

    km_atual = db.Column(db.Float, nullable=True)
    valor_aquisicao = db.Column(db.Float, nullable=True)
    data_aquisicao = db.Column(db.Date, nullable=True)
    ultima_manutencao = db.Column(db.Date, nullable=True)
    km_ultima_manutencao = db.Column(db.Float, nullable=True)
    proxima_manutencao = db.Column(db.Date, nullable=True)
    motorista_padrao_id = db.Column(db.Integer, db.ForeignKey('motorista.id'), nullable=True)

    observacoes = db.Column(db.Text, nullable=True)
    fotos_urls = db.Column(db.Text, nullable=True)

    viagens = db.relationship('Viagem', backref='veiculo', lazy=True, cascade="all, delete-orphan")
    manutencoes = db.relationship('Manutencao', back_populates='veiculo', lazy='dynamic', cascade="all, delete-orphan")
    planos_associados = db.relationship('VeiculoPlano', back_populates='veiculo', cascade="all, delete-orphan")

    @property
    def km_rodados(self):
        return self.km_atual
    
    @km_rodados.setter
    def km_rodados(self, value):
        self.km_atual = value

    @property
    def ano(self):
        return self.ano_modelo
        
    def to_dict(self):
        veiculo_display = f"{self.marca or ''} {self.modelo or ''}"
        ano_display = self.ano_modelo or self.ano_fabricacao or "N/A"
        km_display = f"{int(self.km_atual)} km" if self.km_atual is not None else "Não informado"
        data_cadastro_str = self.created_at.isoformat() if self.created_at else None

        return {
            'id': self.id,
            'placa': self.placa,
            'veiculo': veiculo_display.strip(),
            'modelo': self.modelo or '',
            'marca': self.marca or '',
            'ano': ano_display,
            'status': self.status or 'Disponível',
            'created_at': data_cadastro_str,
            'quilometragem': self.km_atual
        }
    
    def __repr__(self):
        return f'<Veiculo {self.marca} {self.modelo} {self.placa}>'
    
class PlanoDeManutencao(db.Model):
    __tablename__ = 'plano_de_manutencao'
    id = db.Column(db.Integer, primary_key=True)
    descricao = db.Column(db.String(150), unique=True, nullable=False)
    intervalo_km_padrao = db.Column(db.Integer)
    intervalo_meses_padrao = db.Column(db.Integer)
    veiculos = db.relationship('VeiculoPlano', back_populates='plano', cascade="all, delete-orphan")
    insumos_associados = db.relationship('PlanoInsumo', back_populates='plano', cascade="all, delete-orphan")

    def __repr__(self):
        return f'<PlanoDeManutencao {self.descricao}>'

class VeiculoPlano(db.Model):
    __tablename__ = 'veiculo_plano'
    veiculo_id = db.Column(db.Integer, db.ForeignKey('veiculo.id'), primary_key=True)
    plano_id = db.Column(db.Integer, db.ForeignKey('plano_de_manutencao.id'), primary_key=True)
    
    intervalo_km = db.Column(db.Integer)
    intervalo_meses = db.Column(db.Integer)
    
    km_ultima_manutencao = db.Column(db.Integer, nullable=True)
    data_ultima_manutencao = db.Column(db.Date, nullable=True)

    veiculo = db.relationship('Veiculo', back_populates='planos_associados')
    plano = db.relationship('PlanoDeManutencao', back_populates='veiculos')
    manutencoes = db.relationship('Manutencao', back_populates='veiculo_plano_associado')

class ManutencaoItem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    manutencao_id = db.Column(db.Integer, db.ForeignKey('manutencao.id'), nullable=False)
    data = db.Column(db.Date, nullable=False, default=date.today)
    descricao = db.Column(db.String(255), nullable=False)
    quantidade = db.Column(db.Float, nullable=False, default=1)
    custo_unitario = db.Column(db.Float, nullable=False, default=0)

    @property
    def custo_total_item(self):
        return self.quantidade * self.custo_unitario

class Insumo(db.Model):
    __tablename__ = 'insumo'
    id = db.Column(db.Integer, primary_key=True)
    empresa_id = db.Column(db.Integer, db.ForeignKey('empresa.id'), nullable=False)
    
    descricao = db.Column(db.String(200), nullable=False)
    codigo_peca = db.Column(db.String(50), nullable=True)
    custo_unitario_medio = db.Column(db.Float, nullable=True, default=0.0)
    
    quantidade_em_estoque = db.Column(db.Float, nullable=False, server_default='0')
    ponto_ressuprimento = db.Column(db.Float, nullable=True)

    __table_args__ = (db.UniqueConstraint('descricao', 'empresa_id', name='uq_insumo_descricao_empresa'),)

    def __repr__(self):
        return f'<Insumo {self.descricao}>'

class PlanoInsumo(db.Model):
    __tablename__ = 'plano_insumo'
    plano_id = db.Column(db.Integer, db.ForeignKey('plano_de_manutencao.id'), primary_key=True)
    insumo_id = db.Column(db.Integer, db.ForeignKey('insumo.id'), primary_key=True)
    
    quantidade = db.Column(db.Float, nullable=False, default=1)
    
    plano = db.relationship('PlanoDeManutencao', back_populates='insumos_associados')
    insumo = db.relationship('Insumo')

from werkzeug.security import generate_password_hash, check_password_hash
from flask_login import UserMixin

class Usuario(db.Model, UserMixin):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(100), nullable=False)
    sobrenome = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False, index=True)
    senha_hash = db.Column(db.String(256), nullable=False)
    telefone = db.Column(db.String(11), nullable=True)
    idioma = db.Column(db.String(20), default='Português')
    two_factor_enabled = db.Column(db.Boolean, default=False)
    two_factor_phone = db.Column(db.String(11), nullable=True)
    is_admin = db.Column(db.Boolean, default=False)
    role = db.Column(db.String(20), nullable=False, default='Motorista')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    cpf_cnpj = db.Column(db.String(14), unique=True, nullable=True, index=True)
    empresa_id = db.Column(db.Integer, db.ForeignKey('empresa.id'), nullable=True)

    def set_password(self, password):
     self.senha_hash = generate_password_hash(password, method='pbkdf2:sha256')

    def check_password(self, password):
        return check_password_hash(self.senha_hash, password)

class Destino(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    viagem_id = db.Column(db.Integer, db.ForeignKey('viagem.id'), nullable=False)
    endereco = db.Column(db.String(200), nullable=False)
    ordem = db.Column(db.Integer, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Abastecimento(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    veiculo_id = db.Column(db.Integer, db.ForeignKey('veiculo.id'), nullable=False)
    motorista_id = db.Column(db.Integer, db.ForeignKey('motorista.id'), nullable=False)
    data_abastecimento = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    litros = db.Column(db.Float, nullable=False)
    preco_por_litro = db.Column(db.Float, nullable=False)
    custo_total = db.Column(db.Float, nullable=False)
    odometro = db.Column(db.Float, nullable=False)
    anexo_comprovante = db.Column(db.String(500), nullable=True)
    viagem_id = db.Column(db.Integer, db.ForeignKey('viagem.id'), nullable=True, index=True) # Alterado para nullable=True

    veiculo = db.relationship('Veiculo', backref='abastecimentos')
    motorista = db.relationship('Motorista', backref='abastecimentos_registrados')

    # MÉTODO ADICIONADO AQUI:
    def to_dict(self):
        """Converte o objeto para um dicionário."""
        return {
            'id': self.id,
            'data_abastecimento': self.data_abastecimento.strftime('%d/%m/%Y'),
            'litros': self.litros,
            'preco_por_litro': self.preco_por_litro,
            'custo_total': self.custo_total,
            'odometro': self.odometro
        }

class CustoViagem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    viagem_id = db.Column(db.Integer, db.ForeignKey('viagem.id'), nullable=False, unique=True)
    combustivel = db.Column(db.Float, nullable=True)
    pedagios = db.Column(db.Float, nullable=True)
    alimentacao = db.Column(db.Float, nullable=True)
    hospedagem = db.Column(db.Float, nullable=True)
    outros = db.Column(db.Float, nullable=True)
    descricao_outros = db.Column(db.String(300), nullable=True)
    anexos = db.Column(db.String(500), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    viagem = db.relationship('Viagem', backref=db.backref('custo_viagem', uselist=False))
    

class Cliente(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    pessoa_tipo = db.Column(db.String(10), nullable=False)
    nome_razao_social = db.Column(db.String(200), nullable=False)
    nome_fantasia = db.Column(db.String(200), nullable=True)
    cpf_cnpj = db.Column(db.String(14), unique=True, nullable=False, index=True)
    inscricao_estadual = db.Column(db.String(20), nullable=True)
    cep = db.Column(db.String(8), nullable=False)
    logradouro = db.Column(db.String(255), nullable=False)
    numero = db.Column(db.String(20), nullable=False)
    complemento = db.Column(db.String(100), nullable=True)
    bairro = db.Column(db.String(100), nullable=False)
    cidade = db.Column(db.String(100), nullable=False)
    estado = db.Column(db.String(2), nullable=False)
    email = db.Column(db.String(120), nullable=False)
    telefone = db.Column(db.String(11), nullable=False)
    contato_principal = db.Column(db.String(100), nullable=True)
    anexos = db.Column(db.String(1000), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    cadastrado_por_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    

    empresa_id = db.Column(db.Integer, db.ForeignKey('empresa.id'), nullable=False)

    cadastrado_por = db.relationship('Usuario', backref='clientes_cadastrados')

    def __repr__(self):
        return f'<Cliente {self.id}: {self.nome_razao_social}>'
    

class LancamentoFluxoCaixa(db.Model):
    """Lançamentos manuais no fluxo de caixa (receitas e despesas não vinculadas a NFe)"""
    __tablename__ = 'lancamento_fluxo_caixa'
    id = db.Column(db.Integer, primary_key=True)
    empresa_id = db.Column(db.Integer, db.ForeignKey('empresa.id'), nullable=False)
    
    # --- CAMPO DE LIGAÇÃO COM O BANCO ---
    unidade_negocio_id = db.Column(db.Integer, db.ForeignKey('unidade_negocio.id'), nullable=False, index=True)
    
    # Dados básicos do lançamento
    tipo = db.Column(db.String(10), nullable=False, index=True)  # 'RECEITA' ou 'DESPESA'
    descricao = db.Column(db.String(255), nullable=False)
    categoria = db.Column(db.String(100), nullable=True)  # Ex: Combustível, Manutenção, Frete, etc.
    valor_total = db.Column(db.Float, nullable=False)
    centro_custo_id = db.Column(db.Integer, db.ForeignKey('centro_custo.id'), nullable=True, index=True)
    centro_custo = db.relationship('CentroCusto', backref='lancamentos')
    
    # Controle de datas
    data_lancamento = db.Column(db.DateTime, default=datetime.utcnow)
    data_vencimento = db.Column(db.Date, nullable=False, index=True)
    data_pagamento = db.Column(db.Date, nullable=True)
    tem_rateio = db.Column(db.Boolean, nullable=False, server_default='false')
    
    # Status e controle
    status_pagamento = db.Column(db.String(20), default='PENDENTE', nullable=False, index=True)
    
    # Dados opcionais
    fornecedor_cliente = db.Column(db.String(255), nullable=True)  # Nome do fornecedor/cliente
    documento_numero = db.Column(db.String(50), nullable=True)  # Número da nota, boleto, etc.
    observacoes = db.Column(db.Text, nullable=True)
    meio_pagamento = db.Column(db.String(50), nullable=True)
    
    # Controle de parcelas
    parcela_numero = db.Column(db.Integer, default=1)
    parcela_total = db.Column(db.Integer, default=1)
    lancamento_pai_id = db.Column(db.Integer, db.ForeignKey('lancamento_fluxo_caixa.id'), nullable=True)
    
    veiculo_id = db.Column(db.Integer, db.ForeignKey('veiculo.id'), nullable=True, index=True)
    
    # --- RELACIONAMENTOS (AQUI ESTÁ A CORREÇÃO) ---
    unidade_negocio = db.relationship('UnidadeNegocio', back_populates='lancamentos')
    parcelas = db.relationship('LancamentoFluxoCaixa', remote_side=[id], backref='lancamento_pai')
    veiculo = db.relationship('Veiculo', backref='lancamentos_financeiros')
    
    # Anexos de comprovantes
    anexos_urls = db.Column(db.String(2048), nullable=True)  # URLs separadas por vírgula
    
    def __repr__(self):
        return f'<LancamentoFluxoCaixa {self.id} - {self.tipo}: {self.descricao}>'

    @property
    def is_vencido(self):
        """Verifica se o lançamento está vencido"""
        return (self.data_vencimento < date.today() and 
                self.status_pagamento == 'PENDENTE')
    
class LancamentoNotaFiscal(db.Model):
    __tablename__ = 'lancamento_nota_fiscal'
    id = db.Column(db.Integer, primary_key=True)
    empresa_id = db.Column(db.Integer, db.ForeignKey('empresa.id'), nullable=False)
    
    # --- CORREÇÃO APLICADA AQUI ---
    # O tipo da coluna agora é String(44) para ser igual ao da tabela NFeImportada.
    nfe_importada_chave_acesso = db.Column(db.String(44), db.ForeignKey('nfe_importada.chave_acesso'), nullable=False)
    
    # Dados extraídos para fácil acesso
    chave_acesso = db.Column(db.String(44), nullable=False, index=True)
    emitente_cnpj = db.Column(db.String(14), nullable=False)
    emitente_nome = db.Column(db.String(255), nullable=False)
    valor_total = db.Column(db.Float, nullable=False)
    data_emissao = db.Column(db.DateTime, nullable=False)
    
    # Dados financeiros preenchidos pelo usuário
    data_vencimento = db.Column(db.Date, nullable=False, index=True)
    
    # Controle do fluxo de caixa
    status_pagamento = db.Column(db.String(20), default='A Pagar', nullable=False, index=True) # Ex: A Pagar, Pago
    data_pagamento = db.Column(db.Date, nullable=True)
    
    # Data em que o lançamento foi feito no sistema
    data_lancamento = db.Column(db.DateTime, default=datetime.utcnow)

    # O relacionamento agora usa a coluna corrigida
    nfe_original = db.relationship('NFeImportada', foreign_keys=[nfe_importada_chave_acesso])

    def __repr__(self):
        return f'<LancamentoNotaFiscal {self.id} - {self.emitente_nome} R$ {self.valor_total}>'
    parcela_numero = db.Column(db.Integer, default=1)
    parcela_total = db.Column(db.Integer, default=1)
    lancamento_pai_id = db.Column(db.Integer, db.ForeignKey('lancamento_nota_fiscal.id'))
    observacoes = db.Column(db.Text)
    

    
class Viagem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    motorista_id = db.Column(db.Integer, db.ForeignKey('motorista.id'), nullable=True)
    motorista_cpf_cnpj = db.Column(db.String(14), nullable=True, index=True)
    veiculo_id = db.Column(db.Integer, db.ForeignKey('veiculo.id'), nullable=False)
    cliente = db.Column(db.String(100), nullable=False)
    valor_recebido = db.Column(db.Float, nullable=True)
    forma_pagamento = db.Column(db.String(50), nullable=True)
    endereco_saida = db.Column(db.String(200), nullable=False)
    endereco_destino = db.Column(db.String(200), nullable=False)
    distancia_km = db.Column(db.Float, nullable=True) # Distância da API (pode ser usada como estimativa)
    data_inicio = db.Column(db.DateTime, nullable=False)
    data_fim = db.Column(db.DateTime, nullable=True)
    duracao_segundos = db.Column(db.Integer, nullable=True)
    custo = db.Column(db.Float, nullable=True)
    status = db.Column(db.String(50), nullable=False, default='pendente', index=True)
    observacoes = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    empresa_id = db.Column(db.Integer, db.ForeignKey('empresa.id'), nullable=False)
    cobranca_id = db.Column(db.Integer, db.ForeignKey('cobranca.id'), nullable=True)
    public_tracking_token = db.Column(db.String(36), default=lambda: str(uuid.uuid4()), nullable=False)
    odometro_inicial = db.Column(db.Float, nullable=True)
    odometro_final = db.Column(db.Float, nullable=True)
    route_geometry = db.Column(db.Text, nullable=True)
    
    destinos = db.relationship('Destino', backref='viagem', cascade='all, delete-orphan')
    abastecimentos = db.relationship('Abastecimento', backref='viagem', lazy=True)
    localizacoes = db.relationship('Localizacao', backref='viagem', lazy=True, cascade="all, delete-orphan")
    peso_toneladas = db.Column(db.Float, nullable=True)
    custo_motorista_variavel = db.Column(db.Float, nullable=True) 
    material_transportado = db.Column(db.String(150), nullable=True)


    

  
    @property
    def distancia_percorrida(self):
        if self.odometro_inicial is not None and self.odometro_final is not None:
            if self.odometro_final >= self.odometro_inicial:
                return self.odometro_final - self.odometro_inicial
        return 0.0 # Retorna 0 se os dados do odômetro não estiverem disponíveis

    @property
    def consumo_medio(self):
        distancia = self.distancia_percorrida # Usa a nova propriedade
        if not distancia > 0 or not self.abastecimentos:
            return 0.0
        total_litros = sum(abast.litros for abast in self.abastecimentos if abast.litros is not None)
        if total_litros > 0:
            return distancia / total_litros
        return 0.0

  

    @property
    def custo_real_completo(self):
        """
        Calcula o custo REAL e completo da viagem, incluindo custos diretos e rateados DINAMICAMENTE.
        """
        
        # --- 1. Custos Diretos da Viagem ---
        # Soma despesas como pedágio, alimentação, hospedagem, etc., registradas na viagem.
        custos_diretos_viagem = 0
        if self.custo_viagem:
            custos_diretos_viagem = (self.custo_viagem.pedagios or 0) + \
                                    (self.custo_viagem.alimentacao or 0) + \
                                    (self.custo_viagem.hospedagem or 0) + \
                                    (self.custo_viagem.outros or 0)

        # Soma o custo real de todos os abastecimentos registrados para esta viagem.
        custo_combustivel_real = sum(abast.custo_total for abast in self.abastecimentos)

        
        custo_desgaste_veiculo = 0
        if self.veiculo_id and self.distancia_percorrida > 0:
            # Calcula o custo fixo médio por KM com base no histórico do veículo
            custo_fixo_km = calcular_custo_fixo_por_km(self.veiculo_id)
            # Calcula o custo de manutenção médio por KM com base no histórico do veículo
            custo_manutencao_km = calcular_custo_manutencao_por_km(self.veiculo_id)
            
            # Multiplica a média de custo pela distância real percorrida na viagem
            custo_desgaste_veiculo = self.distancia_percorrida * (custo_fixo_km + custo_manutencao_km)

        # Custo do Motorista, com lógica flexível para pagamento variável ou fixo.
        custo_motorista = 0
        if self.custo_motorista_variavel and self.custo_motorista_variavel > 0:
            # Se houver um custo variável registrado (ex: frete por tonelada), usa-o.
            custo_motorista = self.custo_motorista_variavel
        elif self.motorista_formal and self.duracao_segundos and self.duracao_segundos > 0:
            # Senão, calcula com base no salário fixo (para viagens normais).
            salario_base = self.motorista_formal.salario_base or 0.0
            custo_hora_motorista = salario_base / 220  # Custo por hora, considerando 220h/mês
            duracao_horas = self.duracao_segundos / 3600
            custo_motorista = duracao_horas * custo_hora_motorista

        # --- 3. Soma Total ---
        # Soma todos os componentes para obter o custo real completo.
        custo_total = custos_diretos_viagem + custo_combustivel_real + custo_desgaste_veiculo + custo_motorista
        
        return custo_total

    @property
    def lucro_real(self):
        """Calcula o lucro real com base no custo completo."""
        if self.valor_recebido is None:
            # Se não houver receita, o "lucro" é negativo, igual ao custo total.
            return -self.custo_real_completo
        return self.valor_recebido - self.custo_real_completo
    @property
    def lucro_real(self):
        """Calcula o lucro real com base no custo completo."""
        if self.valor_recebido is None:
            # Se não houver receita, o "lucro" é negativo, igual ao custo total.
            return -self.custo_real_completo
        return self.valor_recebido - self.custo_real_completo
    @property
    def lucro_real(self):
        """Calcula o lucro real com base no custo completo."""
        if self.valor_recebido is None:
            # Se não houver receita, o "lucro" é negativo, igual ao custo total.
            return -self.custo_real_completo
        return self.valor_recebido - self.custo_real_completo
    
@app.route('/api/rateio/search')
@login_required
def api_search_rateio_destinos():
    """Busca unificada por veículos e centros de custo."""
    term = request.args.get('term', '').strip()
    if len(term) < 2:
        return jsonify([])

    # Busca Veículos
    veiculos = Veiculo.query.filter(
        Veiculo.empresa_id == current_user.empresa_id,
        or_(
            Veiculo.placa.ilike(f'%{term}%'),
            Veiculo.modelo.ilike(f'%{term}%')
        )
    ).limit(5).all()
    
    # Busca Centros de Custo
    centros_custo = CentroCusto.query.filter(
        CentroCusto.empresa_id == current_user.empresa_id,
        CentroCusto.nome.ilike(f'%{term}%')
    ).limit(5).all()

    resultados = []
    for v in veiculos:
        resultados.append({
            'id': v.id,
            'nome': f"{v.placa} - {v.modelo}",
            'tipo': 'veiculo',
            'icone': 'fas fa-truck'
        })
    for cc in centros_custo:
        resultados.append({
            'id': cc.id,
            'nome': cc.nome,
            'tipo': 'centro_custo',
            'icone': 'fas fa-folder'
        })
    
    return jsonify(resultados)

@app.route('/api/fluxo_caixa/marcar_pago_massa', methods=['POST'])
@login_required
@master_required
def api_marcar_pago_em_massa():
    """API para marcar múltiplos lançamentos como pagos de uma só vez."""
    data = request.get_json()
    item_ids = data.get('item_ids', [])
    data_pagamento_str = data.get('data_pagamento')
    meio_pagamento = data.get('meio_pagamento')

    if not all([item_ids, data_pagamento_str, meio_pagamento]):
        return jsonify({'success': False, 'message': 'Dados incompletos para o pagamento em massa.'}), 400

    try:
        data_pagamento = datetime.strptime(data_pagamento_str, '%Y-%m-%d').date()
        lancamentos_atualizados = 0
        
        # Otimiza a busca pelos IDs
        ids_manuais = [int(i.replace('manual_', '')) for i in item_ids if i.startswith('manual_')]
        
        if ids_manuais:
            # Atualiza todos de uma vez com o método update()
            num_atualizados = LancamentoFluxoCaixa.query.filter(
                LancamentoFluxoCaixa.id.in_(ids_manuais),
                LancamentoFluxoCaixa.empresa_id == current_user.empresa_id
            ).update({
                'status_pagamento': 'PAGO',
                'data_pagamento': data_pagamento,
                'meio_pagamento': meio_pagamento
            }, synchronize_session=False)
            lancamentos_atualizados += num_atualizados
        
        db.session.commit()
        
        return jsonify({'success': True, 'message': f'{lancamentos_atualizados} lançamento(s) foram marcados como pagos.'})

    except Exception as e:
        db.session.rollback()
        logger.error(f"Erro no pagamento em massa: {e}", exc_info=True)
        return jsonify({'success': False, 'message': f'Erro interno: {str(e)}'}), 500

@app.route('/api/centro_custo/criar', methods=['POST'])
@login_required
def api_criar_centro_custo():
    """Cria um novo centro de custo na hora."""
    data = request.json
    nome = data.get('nome', '').strip()
    if not nome:
        return jsonify({'success': False, 'message': 'Nome é obrigatório'}), 400

    try:
        # Verifica se já existe para evitar duplicatas
        existente = CentroCusto.query.filter_by(empresa_id=current_user.empresa_id, nome=nome).first()
        if existente:
            return jsonify({'success': True, 'centro_custo': existente.to_dict()})

        novo_cc = CentroCusto(empresa_id=current_user.empresa_id, nome=nome)
        db.session.add(novo_cc)
        db.session.commit()
        return jsonify({'success': True, 'centro_custo': novo_cc.to_dict()}), 201

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500
    

def get_distancia_total_periodo(veiculo_id, dias=365):
    """
    Calcula o total de KM percorridos por um veículo em viagens concluídas
    nos últimos X dias.
    """
    try:
        data_limite = datetime.utcnow() - timedelta(days=dias)
        
        # Soma a 'distancia_percorrida' (odometro_final - odometro_inicial) de todas as viagens no período.
        distancia_query = db.session.query(
            func.sum(Viagem.odometro_final - Viagem.odometro_inicial)
        ).filter(
            Viagem.veiculo_id == veiculo_id,
            Viagem.status == 'concluida',
            Viagem.data_fim >= data_limite,
            Viagem.odometro_final.isnot(None),
            Viagem.odometro_inicial.isnot(None)
        ).scalar()
        
        return distancia_query or 0.0
    except Exception as e:
        logger.error(f"Erro ao calcular distância total para veículo {veiculo_id}: {e}", exc_info=True)
        return 0.0
    
@app.route('/api/fiscal/salvar_lancamento', methods=['POST'])
@login_required
def api_salvar_lancamento():
    try:
        data = request.get_json()
        logger.info(f"Dados recebidos para lançamento: {data}")
        
        chave_acesso = data.get('chave_acesso')
        nota = NFeImportada.query.filter_by(chave_acesso=chave_acesso, empresa_id=current_user.empresa_id).first()
        if not nota:
            return jsonify({'success': False, 'message': 'Nota fiscal não encontrada.'}), 404
        if nota.status == 'PROCESSADA':
            return jsonify({'success': False, 'message': 'Esta nota já foi processada.'}), 409

        tem_rateio = data.get('tem_rateio', False)
        parcelas = data.get('parcelas', [])
        categoria = data.get('categoria', 'Fornecedores (NFe)')
        
        unidade_negocio_id = data.get('unidade_negocio_id')
        if not unidade_negocio_id:
            return jsonify({'success': False, 'message': 'A Unidade de Negócio é obrigatória.'}), 400

        if not parcelas:
            return jsonify({'success': False, 'message': 'Os dados de parcelamento são obrigatórios.'}), 400

        total_parcelas = len(parcelas)
        
        numero_nota = nota.chave_acesso[25:34]
        
        documento_pai = f"NFE-{numero_nota}-{uuid.uuid4().hex[:6]}"
        
        for i, parcela_info in enumerate(parcelas, 1):
            valor_parcela = float(parcela_info.get('valor', 0))
            data_vencimento = datetime.strptime(parcela_info.get('data_vencimento'), '%Y-%m-%d').date()
            
            descricao_base = f"NFe {numero_nota} - {nota.emitente_nome}"
            if total_parcelas > 1:
                descricao_base += f" - Parcela {i}/{total_parcelas}"

            novo_lancamento = LancamentoFluxoCaixa(
                empresa_id=current_user.empresa_id,
                unidade_negocio_id=unidade_negocio_id,
                tipo='DESPESA',
                descricao=descricao_base,
                categoria=categoria,
                valor_total=valor_parcela,
                data_vencimento=data_vencimento,
                data_lancamento=nota.data_emissao,
                fornecedor_cliente=nota.emitente_nome,
                documento_numero=documento_pai,
                observacoes=f"Chave de acesso: {chave_acesso}",
                parcela_numero=i,
                parcela_total=total_parcelas,
                tem_rateio=tem_rateio
            )
            db.session.add(novo_lancamento)
            db.session.flush()

            if tem_rateio:
                # ▼▼▼ LINHA CORRIGIDA AQUI ▼▼▼
                veiculos_rateio = data.get('rateios', [])
                if not veiculos_rateio:
                    db.session.rollback()
                    return jsonify({'success': False, 'message': 'Para rateio, os veículos são obrigatórios.'}), 400
                
                for veiculo_rateio in veiculos_rateio:
                    percentual = float(veiculo_rateio.get('percentual', 0))
                    valor_rateado_para_parcela = valor_parcela * (percentual / 100)

                    novo_rateio_db = RateioVeiculo(
                        lancamento_id=novo_lancamento.id,
                        veiculo_id=int(veiculo_rateio['veiculo_id']),
                        valor_rateado=valor_rateado_para_parcela,
                        percentual=percentual
                    )
                    db.session.add(novo_rateio_db)
        
        nota.status = 'PROCESSADA'
        db.session.commit()
        message = f'Lançamento criado com sucesso! {total_parcelas} parcela(s) registrada(s) no fluxo de caixa.'
        return jsonify({'success': True, 'message': message})

    except Exception as e:
        db.session.rollback()
        logger.error(f"Erro ao salvar lançamento NFe: {e}", exc_info=True)
        return jsonify({'success': False, 'message': f'Erro interno: {str(e)}'}), 500

@app.route('/financeiro/unidades_negocio')
@login_required
@master_required
def unidades_negocio_page():
    """Página para gerenciar as unidades de negócio."""
    unidades_objetos = UnidadeNegocio.query.filter_by(
        empresa_id=current_user.empresa_id
    ).order_by(
        UnidadeNegocio.is_matriz.desc(), UnidadeNegocio.nome
    ).all()

    # Convertendo a lista de objetos para uma lista de dicionários
    unidades_para_template = [unidade.to_dict() for unidade in unidades_objetos]
    
    # Enviando a lista de dicionários para o template
    return render_template(
        'unidades_negocio.html', 
        unidades=unidades_para_template,
        active_page='unidades_negocio'
    )

@app.route('/api/financeiro/unidade_negocio/salvar', methods=['POST'])
@login_required
@master_required
def api_salvar_unidade_negocio():
    """API para criar ou editar uma unidade de negócio."""
    data = request.json
    unidade_id = data.get('id')
    
    try:
        if unidade_id: # Edição
            unidade = UnidadeNegocio.query.get_or_404(unidade_id)
            if unidade.empresa_id != current_user.empresa_id:
                return jsonify({'success': False, 'message': 'Acesso negado.'}), 403
        else: # Criação
            unidade = UnidadeNegocio(empresa_id=current_user.empresa_id)
            db.session.add(unidade)

        unidade.nome = data.get('nome')
        unidade.razao_social = data.get('razao_social')
        unidade.cnpj = re.sub(r'\D', '', data.get('cnpj', ''))
        unidade.is_matriz = data.get('is_matriz', False)

        if unidade.is_matriz:
            UnidadeNegocio.query.filter(
                UnidadeNegocio.empresa_id == current_user.empresa_id,
                UnidadeNegocio.id != unidade.id
            ).update({'is_matriz': False})

        db.session.commit()

    
        return jsonify({
            'success': True, 
            'message': 'Unidade de negócio salva com sucesso!',
            'unidade': unidade.to_dict()
        })

    except IntegrityError:
        db.session.rollback()
        return jsonify({'success': False, 'message': 'Já existe uma unidade com este nome ou CNPJ.'}), 409
    except Exception as e:
        db.session.rollback()
        logger.error(f"Erro ao salvar unidade de negócio: {e}", exc_info=True)
        return jsonify({'success': False, 'message': str(e)}), 500
    
@app.template_filter('format_cnpj')
def format_cnpj(value):
    if not value or len(value) != 14:
        return value
    return f"{value[:2]}.{value[2:5]}.{value[5:8]}/{value[8:12]}-{value[12:]}"

@app.route('/api/financeiro/unidade_negocio/excluir/<int:id>', methods=['DELETE'])
@login_required
@master_required
def api_excluir_unidade_negocio(id):
    """API para excluir uma unidade de negócio."""
    unidade = UnidadeNegocio.query.get_or_404(id)
    if unidade.empresa_id != current_user.empresa_id:
        return jsonify({'success': False, 'message': 'Acesso negado.'}), 403
    
    if unidade.is_matriz:
        return jsonify({'success': False, 'message': 'Não é possível excluir a unidade matriz.'}), 400
        
    if LancamentoFluxoCaixa.query.filter_by(unidade_negocio_id=id).first():
        return jsonify({'success': False, 'message': 'Não é possível excluir, pois existem lançamentos associados a esta unidade.'}), 409

    try:
        db.session.delete(unidade)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Unidade excluída com sucesso.'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/fiscal/visualizar_xml/<string:chave_acesso>')
@login_required
def api_visualizar_xml(chave_acesso):
    """
    Busca e retorna o conteúdo XML de uma NF-e para visualização.
    """
    nota = NFeImportada.query.filter_by(
        chave_acesso=chave_acesso,
        empresa_id=current_user.empresa_id
    ).first()

    if not nota:
        return jsonify({'success': False, 'message': 'Nota fiscal não encontrada.'}), 404

    if not nota.xml_content:
        return jsonify({'success': False, 'message': 'Conteúdo XML não disponível para esta nota.'}), 404

    try:
        return jsonify({
            'success': True, 
            'xml_content': nota.xml_content,
            'emitente_nome': nota.emitente_nome,
            'chave_acesso': nota.chave_acesso
        })

    except Exception as e:
        logger.error(f"Erro ao visualizar XML da NFe {chave_acesso}: {e}", exc_info=True)
        return jsonify({'success': False, 'message': f'Erro ao processar XML: {e}'}), 500


@app.route('/api/fluxo_caixa/excluir_rateio', methods=['POST'])
@login_required
@master_required
def api_excluir_rateio():
    """API para excluir todos os rateios de um grupo de lançamentos, convertendo-os em despesa geral."""
    data = request.get_json()
    item_id_str = data.get('item_id', '')
    
    if not item_id_str:
        return jsonify({'success': False, 'message': 'ID do item é obrigatório'}), 400

    try:
        lancamento_id_clicado = int(item_id_str.replace('manual_', ''))
        lancamento_clicado = db.session.get(LancamentoFluxoCaixa, lancamento_id_clicado)
        
        if not lancamento_clicado or lancamento_clicado.empresa_id != current_user.empresa_id:
            return jsonify({'success': False, 'message': 'Lançamento não encontrado.'}), 404

        # Encontra todos os lançamentos do mesmo grupo
        lancamentos_do_grupo = []
        if lancamento_clicado.documento_numero and (lancamento_clicado.documento_numero.startswith('NFE-') or lancamento_clicado.documento_numero.startswith('MANUAL-')):
            lancamentos_do_grupo = LancamentoFluxoCaixa.query.filter_by(
                documento_numero=lancamento_clicado.documento_numero,
                empresa_id=current_user.empresa_id
            ).all()
        else:
            lancamentos_do_grupo.append(lancamento_clicado)

        # Itera e remove o rateio de cada um
        for lancamento in lancamentos_do_grupo:
             if lancamento.status_pagamento == 'PAGO':
                # Retorna erro se qualquer parcela do grupo já estiver paga
                return jsonify({'success': False, 'message': 'Não é possível excluir o rateio de um lançamento com parcelas já pagas.'}), 400
             
             RateioVeiculo.query.filter_by(lancamento_id=lancamento.id).delete()
             lancamento.tem_rateio = False
        
        db.session.commit()
        return jsonify({'success': True, 'message': 'Rateio excluído com sucesso! O lançamento agora é uma despesa geral.'})

    except Exception as e:
        db.session.rollback()
        logger.error(f"Erro ao excluir rateio do lançamento {item_id_str}: {e}", exc_info=True)
        return jsonify({'success': False, 'message': f'Erro interno: {str(e)}'}), 500

@app.route('/fiscal/xml/<string:chave_acesso>')
@login_required
def visualizar_xml_page(chave_acesso):
    """
    Página para visualização formatada do XML da NF-e.
    """
    nota = NFeImportada.query.filter_by(
        chave_acesso=chave_acesso,
        empresa_id=current_user.empresa_id
    ).first_or_404()

    if not nota.xml_content:
        flash('Conteúdo XML não disponível para esta nota.', 'error')
        return redirect(url_for('importar_notas_fiscais'))

    try:
        import xml.dom.minidom as minidom
        
        # Formatar o XML para exibição
        try:
            dom = minidom.parseString(nota.xml_content)
            xml_formatado = dom.toprettyxml(indent="  ")
            # Remove linhas vazias extras
            xml_formatado = '\n'.join([line for line in xml_formatado.split('\n') if line.strip()])
        except:
            xml_formatado = nota.xml_content

        return render_template('visualizar_xml.html', 
                             nota=nota, 
                             xml_content=xml_formatado)

    except Exception as e:
        logger.error(f"Erro ao formatar XML da NFe {chave_acesso}: {e}", exc_info=True)
        flash('Erro ao processar o XML da nota fiscal.', 'error')
        return redirect(url_for('importar_notas_fiscais'))

def calcular_custo_manutencao_por_km(veiculo_id, dias=365):
    """
    Calcula o custo médio de manutenção por KM para um veículo.
    Custo Total de Manutenções / KM Totais Rodados no período.
    """
    # 1. Calcula o custo total de manutenções concluídas no período.
    data_limite = datetime.utcnow() - timedelta(days=dias)
    custo_total_manutencoes = db.session.query(
        func.sum(Manutencao.custo_total)
    ).filter(
        Manutencao.veiculo_id == veiculo_id,
        Manutencao.status == 'Concluída',
        Manutencao.data_saida >= data_limite
    ).scalar() or 0.0

    # 2. Pega a distância total rodada no mesmo período.
    distancia_total = get_distancia_total_periodo(veiculo_id, dias)

    # 3. Calcula a média e retorna (evitando divisão por zero).
    if distancia_total > 0:
        return custo_total_manutencoes / distancia_total
    return 0.0

def calcular_custo_fixo_por_km(veiculo_id, dias=365):
    """
    Calcula o custo médio de despesas fixas (IPVA, Seguro) por KM.
    Custo Total de Despesas Fixas / KM Totais Rodados no período.
    """
    # 1. Calcula o custo total de despesas diversas (fixas) no período.
    data_limite = (datetime.utcnow() - timedelta(days=dias)).date()
    custo_total_fixo = db.session.query(
        func.sum(DespesaVeiculo.valor)
    ).filter(
        DespesaVeiculo.veiculo_id == veiculo_id,
        DespesaVeiculo.data >= data_limite
    ).scalar() or 0.0

    # 2. Pega a distância total rodada no mesmo período.
    distancia_total = get_distancia_total_periodo(veiculo_id, dias)
    
    # 3. Calcula a média e retorna.
    if distancia_total > 0:
        return custo_total_fixo / distancia_total
    return 0.0


@app.route('/lancar_frete_rapido', methods=['POST'])
@login_required
def lancar_frete_rapido():
    try:
        # 1. Obter todos os dados do formulário de frete
        data_str = request.form.get('data')
        veiculo_id = request.form.get('veiculo_id')
        motorista_id = request.form.get('motorista_id')
        cliente = request.form.get('cliente')
        origem = request.form.get('origem')
        material = request.form.get('material')
        peso = float(request.form.get('peso_toneladas'))
        valor_frete_total = float(request.form.get('valor_frete_total'))
        valor_por_tonelada_motorista = float(request.form.get('valor_por_tonelada_motorista'))
        
        # 2. Calcular o pagamento variável do motorista para esta viagem específica
        pagamento_motorista = peso * valor_por_tonelada_motorista

        # 3. Criar o objeto Viagem com os novos campos preenchidos
        novo_frete = Viagem(
            motorista_id=int(motorista_id),
            veiculo_id=int(veiculo_id),
            cliente=cliente,
            valor_recebido=valor_frete_total,
            endereco_saida=origem,
            endereco_destino=cliente, # Simplificando, o destino é a empresa cliente
            data_inicio=datetime.strptime(data_str, '%Y-%m-%d'),
            status='concluida', # Lançamentos de frete já entram como concluídos
            observacoes=f"Material: {material}", # Usamos observações para o material
            empresa_id=current_user.empresa_id,
            # Preenchendo os novos campos do banco de dados:
            peso_toneladas=peso,
            custo_motorista_variavel=pagamento_motorista
        )
        
        db.session.add(novo_frete)
        db.session.commit()
        
        flash('Frete lançado com sucesso!', 'success')
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Erro ao lançar frete rápido: {e}", exc_info=True)
        flash(f'Erro ao lançar frete: {e}', 'error')
        
    return redirect(url_for('iniciar_viagem_page'))

@app.route('/financeiro/folha_pagamento')
@login_required
@master_required # Apenas Admins e Masters podem ver
def folha_pagamento_dashboard():
    hoje = date.today()
    
    # Filtros
    mes_filtro = request.args.get('mes', hoje.month, type=int)
    ano_filtro = request.args.get('ano', hoje.year, type=int)
    motorista_filtro = request.args.get('search', '').strip()

    query = FolhaPagamento.query.filter(
        FolhaPagamento.empresa_id == current_user.empresa_id,
        FolhaPagamento.mes_referencia == mes_filtro,
        FolhaPagamento.ano_referencia == ano_filtro
    ).join(Motorista)

    if motorista_filtro:
        query = query.filter(Motorista.nome.ilike(f'%{motorista_filtro}%'))
    
    folhas = query.order_by(Motorista.nome).all()
    
    # Cálculos para o dashboard
    total_proventos_geral = sum(f.total_proventos for f in folhas)
    total_descontos_geral = sum(f.total_descontos for f in folhas)
    total_liquido_geral = sum(f.salario_liquido for f in folhas)

    return render_template('folha_pagamento_dashboard.html', 
                           folhas=folhas,
                           mes_filtro=mes_filtro,
                           ano_filtro=ano_filtro,
                           search_query=motorista_filtro,
                           total_proventos=total_proventos_geral,
                           total_descontos=total_descontos_geral,
                           total_liquido=total_liquido_geral,
                           ano_atual=hoje.year,
                           active_page='folha_pagamento')


def calcular_consumo_medio_real(veiculo_id, periodo_dias=90):
  
    try:
        # 1. Define o período de análise (padrão: últimos 90 dias)
        data_limite = datetime.utcnow() - timedelta(days=periodo_dias)

        # 2. Busca todos os abastecimentos do veículo no período, ordenados pelo odômetro
        abastecimentos = Abastecimento.query.filter(
            Abastecimento.veiculo_id == veiculo_id,
            Abastecimento.data_abastecimento >= data_limite
        ).order_by(Abastecimento.odometro.asc()).all()

        # 3. Se não houver pelo menos 2 registros, não é possível calcular a média
        if len(abastecimentos) < 2:
            return 0.0  # Retorna 0 se não houver dados suficientes

        total_km_rodados = 0.0
        total_litros_consumidos = 0.0

        # 4. Itera entre os abastecimentos para calcular a distância e o consumo por trecho
        for i in range(len(abastecimentos) - 1):
            abastecimento_anterior = abastecimentos[i]
            abastecimento_atual = abastecimentos[i+1]

            # Distância percorrida entre os dois abastecimentos
            distancia_trecho = abastecimento_atual.odometro - abastecimento_anterior.odometro
            
            # Consideramos que os litros do abastecimento ANTERIOR foram consumidos neste trecho
            litros_trecho = abastecimento_anterior.litros

            if distancia_trecho > 0 and litros_trecho > 0:
                total_km_rodados += distancia_trecho
                total_litros_consumidos += litros_trecho

        # 5. Calcula a média final
        if total_litros_consumidos > 0:
            media_consumo = total_km_rodados / total_litros_consumidos
            return media_consumo
        
        return 0.0

    except Exception as e:
        logger.error(f"Erro ao calcular consumo médio para veiculo {veiculo_id}: {e}", exc_info=True)
        return 0.0 # Retorna 0 em caso de erro

def obter_preco_medio_combustivel_recente(empresa_id, default=5.80, limit=20):
    """
    Busca a média de preço por litro dos últimos 'limit' abastecimentos da empresa.
    """
    try:
        # Busca o preço médio diretamente no banco de dados
        preco_medio_query = db.session.query(
            func.avg(Abastecimento.preco_por_litro)
        ).join(Veiculo).filter(
            Veiculo.empresa_id == empresa_id
        ).scalar()

        # Se houver um resultado, retorna ele. Senão, usa o valor padrão.
        return preco_medio_query if preco_medio_query else default
    except Exception as e:
        logger.error(f"Erro ao obter preço médio do combustível para empresa {empresa_id}: {e}")
        return default
    
@app.route('/api/pneus/estoque')
@login_required
def api_pneus_em_estoque():
    """API para listar pneus disponíveis no estoque para os modais."""
    try:
        pneus = Pneu.query.filter_by(
            empresa_id=current_user.empresa_id,
            status='Estoque'
        ).order_by(Pneu.numero_fogo).all()

        pneus_data = [{
            'id': p.id,
            'numero_fogo': p.numero_fogo,
            'marca': p.marca,
            'modelo': p.modelo,
            'dimensao': p.dimensao
        } for p in pneus]

        return jsonify({'success': True, 'pneus': pneus_data})
    except Exception as e:
        logger.error(f"Erro ao buscar pneus em estoque: {e}", exc_info=True)
        return jsonify({'success': False, 'message': str(e)}), 500
    

def calcular_consumo_medio_real(veiculo_id, periodo_dias=90):
    """
    Calcula o consumo médio real (km/l) de um veículo com base no histórico
    de abastecimentos nos últimos 'periodo_dias'.
    """
    try:
        data_limite = datetime.utcnow() - timedelta(days=periodo_dias)

        abastecimentos = Abastecimento.query.filter(
            Abastecimento.veiculo_id == veiculo_id,
            Abastecimento.data_abastecimento >= data_limite
        ).order_by(Abastecimento.odometro.asc()).all()

        if len(abastecimentos) < 2:
            return 0.0

        total_km_rodados = 0.0
        total_litros_consumidos = 0.0

        for i in range(len(abastecimentos) - 1):
            abastecimento_anterior = abastecimentos[i]
            abastecimento_atual = abastecimentos[i+1]

            distancia_trecho = abastecimento_atual.odometro - abastecimento_anterior.odometro
            litros_trecho = abastecimento_anterior.litros

            if distancia_trecho > 0 and litros_trecho > 0:
                total_km_rodados += distancia_trecho
                total_litros_consumidos += litros_trecho

        if total_litros_consumidos > 0:
            media_consumo = total_km_rodados / total_litros_consumidos
            return media_consumo
        
        return 0.0

    except Exception as e:
        logger.error(f"Erro ao calcular consumo médio para veiculo {veiculo_id}: {e}", exc_info=True)
        return 0.0

def obter_preco_medio_combustivel_recente(empresa_id, default=5.80):
    """
    Busca a média de preço por litro dos últimos abastecimentos da empresa.
    """
    try:
        preco_medio_query = db.session.query(
            func.avg(Abastecimento.preco_por_litro)
        ).join(Veiculo).filter(
            Veiculo.empresa_id == empresa_id
        ).scalar()

        return float(preco_medio_query) if preco_medio_query else default
    except Exception as e:
        logger.error(f"Erro ao obter preço médio do combustível para empresa {empresa_id}: {e}")
        return default

# --- Rota Principal da API de Estimativa ---

@app.route('/api/viagem/estimar_custo', methods=['POST'])
@login_required
def estimar_custo_viagem_api():
    """
    API para estimar o custo total de uma viagem (Combustível + Motorista + Desgaste do Veículo).
    Utiliza dados históricos para maior precisão.
    """
    try:
        data = request.get_json()
        veiculo_id = data.get('veiculo_id')
        motorista_id = data.get('motorista_id')
        distancia_km = data.get('distancia_km')
        duracao_segundos = data.get('duracao_segundos')

        if not all([veiculo_id, motorista_id, distancia_km is not None, duracao_segundos is not None]):
            return jsonify({'success': False, 'message': 'Dados insuficientes para estimar o custo.'}), 400

        veiculo = db.session.get(Veiculo, int(veiculo_id))
        motorista = db.session.get(Motorista, int(motorista_id))

        if not veiculo or not motorista:
            return jsonify({'success': False, 'message': 'Veículo ou motorista não encontrado.'}), 404

        # 1. Custo de Combustível (Lógica Inteligente com Fallback)
        consumo_real = calcular_consumo_medio_real(veiculo.id)
        consumo_a_ser_usado = consumo_real or veiculo.consumo_medio_km_l or 1.0
        preco_combustivel_para_calculo = obter_preco_medio_combustivel_recente(current_user.empresa_id)
        
        litros_estimados = distancia_km / consumo_a_ser_usado
        custo_combustivel = litros_estimados * preco_combustivel_para_calculo

        # 2. Custo do Motorista (Rateado por hora)
        salario_base = motorista.salario_base or 0.0
        custo_hora_motorista = salario_base / 220  # Custo/hora considerando 220h/mês
        duracao_horas = duracao_segundos / 3600
        custo_motorista = duracao_horas * custo_hora_motorista

        # 3. Custo Fixo e Manutenção do Veículo (Rateado por KM)
        # Estes valores são placeholders. No futuro, podem ser substituídos por
        # funções que calculam a média real com base nos modelos Manutencao e DespesaVeiculo.
        CUSTO_FIXO_POR_KM = 0.55       # Ex: IPVA, Seguro, etc.
        CUSTO_MANUTENCAO_POR_KM = 0.40 # Ex: Pneus, óleo, peças de desgaste.
        custo_desgaste_veiculo = (CUSTO_FIXO_POR_KM + CUSTO_MANUTENCAO_POR_KM) * distancia_km

        # 4. Totalização dos custos
        custo_total_estimado = custo_combustivel + custo_motorista + custo_desgaste_veiculo

        # 5. Retorno dos dados em formato JSON para o Frontend
        return jsonify({
            'success': True,
            'custos': {
                'combustivel': round(custo_combustivel, 2),
                'motorista': round(custo_motorista, 2),
                'veiculo': round(custo_desgaste_veiculo, 2),
                'total': round(custo_total_estimado, 2)
            }
        })

    except Exception as e:
        logger.error(f"Erro na API de estimar custo: {e}", exc_info=True)
        return jsonify({'success': False, 'message': f'Ocorreu um erro interno: {e}'}), 500




@app.route('/financeiro/folha_pagamento/gerar', methods=['POST'])
@login_required
@master_required
def gerar_folhas_mes():
    mes = request.form.get('mes', type=int)
    ano = request.form.get('ano', type=int)

    if not mes or not ano:
        flash('Mês e Ano são obrigatórios para gerar as folhas.', 'error')
        return redirect(url_for('folha_pagamento_dashboard'))

    motoristas_ativos = Motorista.query.filter_by(empresa_id=current_user.empresa_id, situacao='NORMAL / LIBERADO').all()
    criadas = 0
    atualizadas = 0
    
    for motorista in motoristas_ativos:
        folha = FolhaPagamento.query.filter_by(
            motorista_id=motorista.id,
            mes_referencia=mes,
            ano_referencia=ano
        ).first()

        if not folha:
            folha = FolhaPagamento(
                motorista_id=motorista.id,
                empresa_id=current_user.empresa_id,
                mes_referencia=mes,
                ano_referencia=ano,
                salario_base_registro=motorista.salario_base or 0.0,
                status='Em Aberto'
            )
            db.session.add(folha)
            db.session.flush() # Garante que a folha tenha um ID antes de adicionar itens
            criadas += 1
        elif folha.status != 'Em Aberto':
            continue # Pula folhas já fechadas ou pagas
        else:
            atualizadas += 1

        # --- LÓGICA DE INTEGRAÇÃO DE FRETES ---
        # 1. Busca todas as viagens concluídas do motorista no mês que são fretes (custo > 0)
        viagens_do_mes = Viagem.query.filter(
            Viagem.motorista_id == motorista.id,
            Viagem.status == 'concluida',
            Viagem.custo_motorista_variavel > 0, # Filtra apenas fretes com valor
            extract('month', Viagem.data_fim) == mes,
            extract('year', Viagem.data_fim) == ano
        ).all()

        # 2. Para cada viagem, adiciona um provento na folha se ainda não existir
        for viagem in viagens_do_mes:
            item_existente = ItemFolhaPagamento.query.filter_by(
                folha_pagamento_id=folha.id,
                viagem_id=viagem.id # Checa se o item já foi lançado para essa viagem específica
            ).first()

            if not item_existente:
                novo_provento = ItemFolhaPagamento(
                    folha_pagamento_id=folha.id,
                    tipo='Provento',
                    descricao=f"Frete: {viagem.cliente} ({viagem.data_inicio.strftime('%d/%m')})",
                    valor=viagem.custo_motorista_variavel,
                    viagem_id=viagem.id # Linka o item à viagem
                )
                db.session.add(novo_provento)
        
        # 3. Recalcula os totais da folha após adicionar os proventos
        folha.recalcular_totais()
            
    try:
        db.session.commit()
        flash(f'{criadas} folha(s) criada(s) e {atualizadas} atualizada(s) com os fretes do mês.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao gerar/atualizar folhas de pagamento: {e}', 'error')
        
    return redirect(url_for('folha_pagamento_dashboard', mes=mes, ano=ano))


@app.route('/financeiro/folha_pagamento/<int:folha_id>', methods=['GET', 'POST'])
@login_required
@master_required
def detalhes_folha_pagamento(folha_id):
    folha = FolhaPagamento.query.options(
        db.joinedload(FolhaPagamento.motorista)
    ).filter_by(id=folha_id, empresa_id=current_user.empresa_id).first_or_404()

    if request.method == 'POST':
        if folha.status != 'Em Aberto':
            flash('Esta folha de pagamento está fechada e não pode ser editada.', 'error')
            return redirect(url_for('detalhes_folha_pagamento', folha_id=folha.id))

        # Deleta itens antigos para recriar
        folha.itens.delete()

        # Processa Proventos
        i = 1
        while f'provento_descricao_{i}' in request.form:
            descricao = request.form[f'provento_descricao_{i}']
            valor = request.form.get(f'provento_valor_{i}', type=float)
            if descricao and valor is not None:
                item = ItemFolhaPagamento(folha_pagamento_id=folha.id, tipo='Provento', descricao=descricao, valor=valor)
                db.session.add(item)
            i += 1
        
        # Processa Descontos
        i = 1
        while f'desconto_descricao_{i}' in request.form:
            descricao = request.form[f'desconto_descricao_{i}']
            valor = request.form.get(f'desconto_valor_{i}', type=float)
            if descricao and valor is not None:
                item = ItemFolhaPagamento(folha_pagamento_id=folha.id, tipo='Desconto', descricao=descricao, valor=valor)
                db.session.add(item)
            i += 1

        folha.salario_base_registro = request.form.get('salario_base', type=float, default=0.0)
        folha.observacoes = request.form.get('observacoes')
        folha.recalcular_totais()
        
        # Lógica de Ações (Salvar, Fechar, Pagar)
        action = request.form.get('action')
        if action == 'fechar':
            folha.status = 'Fechada'
            folha.data_fechamento = datetime.utcnow()
            flash_msg = 'Folha de pagamento fechada com sucesso!'
        elif action == 'pagar':
            folha.status = 'Paga'
            folha.data_pagamento = datetime.strptime(request.form.get('data_pagamento'), '%Y-%m-%d')
            folha.meio_pagamento = request.form.get('meio_pagamento')
            # Lógica de upload de comprovante aqui (se necessário)
            flash_msg = 'Folha de pagamento marcada como paga!'
        else: # Salvar
            flash_msg = 'Alterações salvas com sucesso!'
        
        try:
            db.session.commit()
            flash(flash_msg, 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao salvar: {e}', 'error')

        return redirect(url_for('detalhes_folha_pagamento', folha_id=folha.id))

    proventos = folha.itens.filter_by(tipo='Provento').all()
    descontos = folha.itens.filter_by(tipo='Desconto').all()
    
    return render_template('detalhes_folha_pagamento.html', 
                           folha=folha,
                           proventos=proventos,
                           descontos=descontos,
                           active_page='folha_pagamento')


@app.route('/financeiro/folha_pagamento/<int:folha_id>/holerite')
@login_required
@master_required
def gerar_holerite(folha_id):
    folha = FolhaPagamento.query.filter_by(id=folha_id, empresa_id=current_user.empresa_id).first_or_404()
    proventos = folha.itens.filter_by(tipo='Provento').all()
    descontos = folha.itens.filter_by(tipo='Desconto').all()
    
    return render_template('holerite.html', 
                           folha=folha,
                           proventos=proventos,
                           descontos=descontos)

class DespesaVeiculo(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    veiculo_id = db.Column(db.Integer, db.ForeignKey('veiculo.id'), nullable=False)
    empresa_id = db.Column(db.Integer, db.ForeignKey('empresa.id'), nullable=False)
    data = db.Column(db.Date, nullable=False, default=date.today)
    
    # Categoria: 'IPVA', 'Seguro', 'Multa', 'Salário Fixo', 'Pneus', 'Outros'
    categoria = db.Column(db.String(100), nullable=False, index=True)
    descricao = db.Column(db.Text, nullable=True)
    valor = db.Column(db.Float, nullable=False)
    
    # Armazenará URLs dos comprovantes, separadas por vírgula
    anexos = db.Column(db.String(1024), nullable=True) 

    veiculo = db.relationship('Veiculo', backref=db.backref('despesas_diversas', lazy=True, cascade="all, delete-orphan"))

    def to_dict(self):
        return {
            "id": self.id,
            "data": self.data.strftime('%d/%m/%Y'),
            "categoria": self.categoria,
            "descricao": self.descricao,
            "valor": self.valor,
            "anexos": self.anexos.split(',') if self.anexos else []
        }
    
class Localizacao(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    motorista_id = db.Column(db.Integer, db.ForeignKey('motorista.id'), nullable=False)
    viagem_id = db.Column(db.Integer, db.ForeignKey('viagem.id'), nullable=True)
    latitude = db.Column(db.Float, nullable=False)
    longitude = db.Column(db.Float, nullable=False)
    endereco = db.Column(db.String(200), nullable=True)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)

class Documento(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    
    # Informações do Documento
    tipo_documento = db.Column(db.String(100), nullable=False, index=True) # Ex: CNH, CRLV, ANTT, Seguro
    descricao = db.Column(db.String(255), nullable=True) # Ex: Apólice de Seguro XYZ
    numero_documento = db.Column(db.String(100), nullable=True)
    data_emissao = db.Column(db.Date, nullable=True)
    data_validade = db.Column(db.Date, nullable=False, index=True) # Essencial para os alertas
    url_anexo = db.Column(db.String(500), nullable=True) # Link para o arquivo no Cloudflare R2
    
    # Chaves estrangeiras para ligar o documento a outras partes do sistema
    empresa_id = db.Column(db.Integer, db.ForeignKey('empresa.id'), nullable=False)
    
    # Relacionamento: Um documento pertence ou a um motorista OU a um veículo
    motorista_id = db.Column(db.Integer, db.ForeignKey('motorista.id'), nullable=True)
    veiculo_id = db.Column(db.Integer, db.ForeignKey('veiculo.id'), nullable=True)

    # Relacionamentos para facilitar o acesso aos objetos
    motorista = db.relationship('Motorista', backref=db.backref('documentos', lazy=True, cascade="all, delete-orphan"))
    veiculo = db.relationship('Veiculo', backref=db.backref('documentos', lazy=True, cascade="all, delete-orphan"))

    def __repr__(self):
        return f'<Documento {self.id} - {self.tipo_documento}>'    




def parse_nfe_xml(xml_file):
    """
    Extrai dados de um arquivo XML de NF-e, incluindo informações
    detalhadas do cliente e da viagem.
    """
    try:
        xml_file.seek(0)
        tree = ET.parse(xml_file)
        root = tree.getroot()
        ns = {'nfe': 'http://www.portalfiscal.inf.br/nfe'}

        # --- Dados do Destinatário (Cliente) ---
        dest_element = root.find('.//nfe:dest', ns)
        if dest_element is None: return None
        
        ender_dest = dest_element.find('nfe:enderDest', ns)
        
        # Determina se é CNPJ ou CPF
        cpf_cnpj_node = dest_element.find('nfe:CNPJ', ns)
        pessoa_tipo = 'juridica'
        if cpf_cnpj_node is None:
            cpf_cnpj_node = dest_element.find('nfe:CPF', ns)
            pessoa_tipo = 'fisica'
        
        cpf_cnpj = cpf_cnpj_node.text if cpf_cnpj_node is not None else None

        cpf_cnpj_limpo = re.sub(r'\D', '', cpf_cnpj) if cpf_cnpj else ''
        # Coleta todos os dados do cliente
        cliente_info = {
                'pessoa_tipo': pessoa_tipo,
                'cpf_cnpj': cpf_cnpj_limpo if cpf_cnpj_limpo else None,
                'nome_razao_social': dest_element.find('nfe:xNome', ns).text,
                'inscricao_estadual': getattr(dest_element.find('nfe:IE', ns), 'text', None),
                'logradouro': ender_dest.find('nfe:xLgr', ns).text,
                'numero': ender_dest.find('nfe:nro', ns).text,
                'complemento': getattr(ender_dest.find('nfe:xCpl', ns), 'text', None),
                'bairro': ender_dest.find('nfe:xBairro', ns).text,
                'cidade': ender_dest.find('nfe:xMun', ns).text,
                'estado': ender_dest.find('nfe:UF', ns).text,
                'cep': re.sub(r'\D', '', ender_dest.find('nfe:CEP', ns).text),
                'email': f"{cpf_cnpj_limpo}@email.xml", # <-- Agora esta linha funciona
                'telefone': '00000000000'
        }

        # --- Dados do Emissor (Endereço de Saída) ---
        emit_element = root.find('.//nfe:emit', ns)
        ender_emit = emit_element.find('nfe:enderEmit', ns)
        saida_rua = ender_emit.find('nfe:xLgr', ns).text
        saida_num = ender_emit.find('nfe:nro', ns).text
        saida_bairro = ender_emit.find('nfe:xBairro', ns).text
        saida_cidade = ender_emit.find('nfe:xMun', ns).text
        saida_uf = ender_emit.find('nfe:UF', ns).text
        endereco_saida_completo = f"{saida_rua}, {saida_num} - {saida_bairro}, {saida_cidade} - {saida_uf}"

        # --- Dados da Viagem ---
        viagem_info = {
            "cliente": cliente_info['nome_razao_social'],
            "endereco_saida": endereco_saida_completo,
            "endereco_destino": f"{cliente_info['logradouro']}, {cliente_info['numero']} - {cliente_info['bairro']}, {cliente_info['cidade']} - {cliente_info['estado']}",
            "nome_arquivo": getattr(xml_file, 'filename', 'N/A')
        }

        return {
            "viagem_info": viagem_info,
            "cliente_info": cliente_info
        }

    except Exception as e:
        logger.error(f"Erro ao processar XML: {e}", exc_info=True)
        return None



@app.route('/veiculo/<int:veiculo_id>/lancar_despesa_diversa', methods=['POST'])
@login_required
def lancar_despesa_diversa(veiculo_id):
    """
    Recebe os dados do formulário de lançamento rápido do dashboard do veículo.
    """
    veiculo = Veiculo.query.filter_by(id=veiculo_id, empresa_id=current_user.empresa_id).first_or_404()
    
    try:
        # Cria uma nova instância do modelo DespesaVeiculo com os dados do formulário
        nova_despesa = DespesaVeiculo(
            veiculo_id=veiculo.id,
            empresa_id=current_user.empresa_id,
            data=datetime.strptime(request.form['data'], '%Y-%m-%d').date(),
            categoria=request.form['categoria'],
            descricao=request.form.get('descricao', ''),
            valor=float(request.form['valor'])
        )

        # Lógica para Upload de Anexos
        urls_anexos = []
        files = request.files.getlist('anexos')
        if files and any(f and f.filename for f in files):
            s3_client = boto3.client(
                's3',
                endpoint_url=app.config['CLOUDFLARE_R2_ENDPOINT'],
                aws_access_key_id=app.config['CLOUDFLARE_R2_ACCESS_KEY'],
                aws_secret_access_key=app.config['CLOUDFLARE_R2_SECRET_KEY'],
                region_name='auto'
            )
            bucket_name = app.config['CLOUDFLARE_R2_BUCKET']
            public_url_base = app.config['CLOUDFLARE_R2_PUBLIC_URL']

            for file in files:
                if file and file.filename: # Checagem extra
                    filename = secure_filename(file.filename)
                    s3_path = f"despesas_veiculo/{veiculo.id}/{uuid.uuid4()}-{filename}"
                    
                    s3_client.upload_fileobj(
                        file, bucket_name, s3_path,
                        ExtraArgs={'ContentType': file.content_type or 'application/octet-stream'}
                    )
                    urls_anexos.append(f"{public_url_base}/{s3_path}")
        
        if urls_anexos:
            nova_despesa.anexos = ",".join(urls_anexos)

        # Adiciona e salva a nova despesa no banco de dados
        db.session.add(nova_despesa)
        db.session.commit()
        flash('Despesa registrada com sucesso!', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao registrar despesa: {e}', 'error')
        logger.error(f"Erro ao adicionar despesa para veiculo {veiculo_id}: {e}", exc_info=True)

    # Redireciona de volta para o dashboard do veículo após o lançamento
    return redirect(url_for('veiculo_dashboard', veiculo_id=veiculo_id))


def validate_cpf_cnpj(cpf_cnpj, pessoa_tipo):
    if pessoa_tipo == 'fisica':
        return bool(re.match(r'^\d{11}$', cpf_cnpj))
    return bool(re.match(r'^\d{14}$', cpf_cnpj))

def validate_telefone(telefone):
    return bool(re.match(r'^\d{10,11}$', telefone))

def validate_cnh(cnh):
    return bool(re.match(r'^\d{11}$', cnh))

def validate_placa(placa):
    return bool(re.match(r'^[A-Z0-9]{7}$', placa.upper()))

def haversine_distance(lat1, lon1, lat2, lon2):
    R = 6371
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    delta_phi = math.radians(lat2 - lat1)
    delta_lambda = math.radians(lon2 - lon1)
    a = math.sin(delta_phi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(delta_lambda / 2) ** 2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return R * c

@app.route('/painel')
@login_required
def painel():
    """
    Esta rota agora carrega o dashboard principal da sua aplicação.
    Ela contém toda a lógica para buscar os dados das viagens e KPIs.
    """
    # Query base para buscar todas as viagens da empresa, já carregando dados relacionados
    viagens_query = Viagem.query.filter_by(empresa_id=current_user.empresa_id).options(
        db.joinedload(Viagem.veiculo),
        db.joinedload(Viagem.motorista_formal)
    ).order_by(Viagem.data_inicio.desc()).all()

    # --- KPIs Operacionais ---
    viagens_em_andamento_kpi = sum(1 for v in viagens_query if v.status == 'em_andamento')
    viagens_pendentes_kpi = sum(1 for v in viagens_query if v.status == 'pendente')
    veiculos_disponiveis_kpi = Veiculo.query.filter_by(empresa_id=current_user.empresa_id, status='Disponível').count()

    # --- KPIs Financeiros (para o mês atual) ---
    receita_mes = 0.0
    custo_mes = 0.0

    if current_user.role in ['Admin', 'Master']:
        hoje = date.today()
        viagens_do_mes = Viagem.query.filter(
            extract('year', Viagem.data_inicio) == hoje.year,
            extract('month', Viagem.data_inicio) == hoje.month,
            Viagem.empresa_id == current_user.empresa_id,
            Viagem.status == 'concluida'
        ).options(db.joinedload(Viagem.custo_viagem), db.joinedload(Viagem.abastecimentos)).all()

        for v in viagens_do_mes:
            receita_mes += v.valor_recebido or 0.0
            custo_despesas = 0
            if v.custo_viagem:
                custo_despesas = (v.custo_viagem.pedagios or 0) + (v.custo_viagem.alimentacao or 0) + (v.custo_viagem.hospedagem or 0) + (v.custo_viagem.outros or 0)
            custo_abastecimento = sum(a.custo_total for a in v.abastecimentos)
            custo_mes += custo_despesas + custo_abastecimento

    lucro_mes = receita_mes - custo_mes
    
    # --- Processa a lista de viagens para exibir na página principal ---
    viagens_para_template = []
    for viagem in viagens_query:
        motorista_nome = 'N/A'
        if viagem.motorista_formal:
            motorista_nome = viagem.motorista_formal.nome

        destinos_list = sorted(
            [{'endereco': d.endereco, 'ordem': d.ordem} for d in viagem.destinos],
            key=lambda d: d.get('ordem', 0)
        )

        viagens_para_template.append({
            'id': viagem.id,
            'cliente': viagem.cliente,
            'motorista_nome': motorista_nome,
            'endereco_saida': viagem.endereco_saida,
            'destinos': destinos_list,
            'status': viagem.status,
            'veiculo_placa': viagem.veiculo.placa if viagem.veiculo else 'N/A',
            'veiculo_modelo': viagem.veiculo.modelo if viagem.veiculo else 'N/A',
            'data_inicio': viagem.data_inicio,
            'data_fim': viagem.data_fim
        })

    # Renderiza o template do painel, que agora se chama 'painel.html'
    return render_template('painel.html',
                           viagens=viagens_para_template,
                           Maps_API_KEY=Maps_API_KEY,
                           viagens_em_andamento=viagens_em_andamento_kpi,
                           viagens_pendentes=viagens_pendentes_kpi,
                           veiculos_disponiveis=veiculos_disponiveis_kpi,
                           receita_mes=receita_mes,
                           custo_mes=custo_mes,
                           lucro_mes=lucro_mes)

@app.route('/cadastrar_cliente', methods=['GET', 'POST'])
@login_required
def cadastrar_cliente():
    if request.method == 'POST':
        try:
            cpf_cnpj = re.sub(r'\D', '', request.form.get('cpf_cnpj', ''))
            if Cliente.query.filter_by(cpf_cnpj=cpf_cnpj, empresa_id=current_user.empresa_id).first():
                flash('Erro: Este CPF/CNPJ já está cadastrado para um cliente em sua empresa.', 'error')
                return redirect(url_for('cadastrar_cliente'))

            # <<< INÍCIO DA LÓGICA DE UPLOAD >>>
            anexos_urls = []
            files = request.files.getlist('anexos') # O 'name' do input no HTML é 'anexos'

            if files and any(f and f.filename for f in files):
                s3_client = boto3.client(
                    's3',
                    endpoint_url=app.config['CLOUDFLARE_R2_ENDPOINT'],
                    aws_access_key_id=app.config['CLOUDFLARE_R2_ACCESS_KEY'],
                    aws_secret_access_key=app.config['CLOUDFLARE_R2_SECRET_KEY'],
                    region_name='auto'
                )
                bucket_name = app.config['CLOUDFLARE_R2_BUCKET']
                public_url_base = app.config['CLOUDFLARE_R2_PUBLIC_URL']

                for file in files:
                    if file and file.filename:
                        filename = secure_filename(file.filename)
                        s3_path = f"clientes/{cpf_cnpj}/{uuid.uuid4()}-{filename}"
                        
                        s3_client.upload_fileobj(
                            file, bucket_name, s3_path,
                            ExtraArgs={'ContentType': file.content_type or 'application/octet-stream'}
                        )
                        anexos_urls.append(f"{public_url_base}/{s3_path}")
            # <<< FIM DA LÓGICA DE UPLOAD >>>

            novo_cliente = Cliente(
                pessoa_tipo=request.form.get('pessoa_tipo'),
                nome_razao_social=request.form.get('nome_razao_social'),
                nome_fantasia=request.form.get('nome_fantasia'),
                cpf_cnpj=cpf_cnpj,
                inscricao_estadual=request.form.get('inscricao_estadual'),
                cep=re.sub(r'\D', '', request.form.get('cep', '')),
                logradouro=request.form.get('logradouro'),
                numero=request.form.get('numero'),
                complemento=request.form.get('complemento'),
                bairro=request.form.get('bairro'),
                cidade=request.form.get('cidade'),
                estado=request.form.get('estado'),
                email=request.form.get('email'),
                telefone=re.sub(r'\D', '', request.form.get('telefone', '')),
                contato_principal=request.form.get('contato_principal'),
                anexos=','.join(anexos_urls) if anexos_urls else None, # Salva as URLs no banco
                cadastrado_por_id=current_user.id,
                empresa_id=current_user.empresa_id
            )
            
            db.session.add(novo_cliente)
            db.session.commit()
            flash('Cliente cadastrado com sucesso!', 'success')
            return redirect(url_for('consultar_clientes'))

        except Exception as e:
            db.session.rollback()
            logger.error(f"Erro ao cadastrar cliente: {e}", exc_info=True)
            flash(f'Ocorreu um erro ao cadastrar o cliente: {e}', 'error')
    
    return render_template('cadastrar_cliente.html', active_page='cadastrar_cliente')

@app.route('/consultar_clientes')
@login_required
def consultar_clientes():
    search_query = request.args.get('search', '').strip()
    
    # Base query for clients of the current company
    base_query = Cliente.query.filter_by(empresa_id=current_user.empresa_id)
    
    # Apply search filter if present
    if search_query:
        search_filter = f"%{search_query}%"
        base_query = base_query.filter(
            or_(
                Cliente.nome_razao_social.ilike(search_filter),
                Cliente.cpf_cnpj.ilike(search_filter)
            )
        )
        
    clientes = base_query.order_by(Cliente.nome_razao_social.asc()).all()

    # Calculate total revenue for ALL clients of the company, respecting no search filter for the total.
    nomes_clientes = [c.nome_razao_social for c in Cliente.query.filter_by(empresa_id=current_user.empresa_id).all()]
    
    total_receita_clientes = db.session.query(func.sum(Viagem.valor_recebido)).filter(
        Viagem.empresa_id == current_user.empresa_id,
        Viagem.cliente.in_(nomes_clientes)
    ).scalar() or 0.0

    # --- CORREÇÃO APLICADA AQUI ---
    return render_template('consultar_clientes.html', 
                           clientes=clientes, 
                           total_receita=total_receita_clientes, # A variável está sendo passada para o template
                           search_query=search_query, 
                           active_page='consultar_clientes')



@app.route('/track/<uuid:token>')
def public_tracking_page(token):
    """Exibe a página pública de acompanhamento para o cliente."""
    # Converte o token para string para buscar no banco
    token_str = str(token)
    
    # Busca a viagem usando o token seguro
    viagem = Viagem.query.filter_by(public_tracking_token=token_str).first_or_404()
    
    # Prepara os dados do motorista para exibição
    motorista_nome = 'Não informado'
    if viagem.motorista_formal:
        motorista_nome = viagem.motorista_formal.nome
    elif viagem.motorista_cpf_cnpj:
        usuario = Usuario.query.filter_by(cpf_cnpj=viagem.motorista_cpf_cnpj).first()
        if usuario:
            motorista_nome = f"{usuario.nome} {usuario.sobrenome}"

    current_year = datetime.utcnow().year

    return render_template('public_tracking.html', 
                           viagem=viagem, 
                           motorista_nome=motorista_nome,
                           Maps_API_KEY=Maps_API_KEY)

from sqlalchemy import or_ 
@app.route('/api/clientes/search')
@login_required
def search_clientes_simplificado():
    search_term = request.args.get('term', '')
    if not search_term or len(search_term) < 2:
        return jsonify([])

    # Agora a busca só acontece nos clientes da empresa do usuário logado
    clientes = Cliente.query.filter(
        Cliente.empresa_id == current_user.empresa_id,  # <-- FILTRO DE SEGURANÇA ADICIONADO AQUI
        or_(
            Cliente.nome_razao_social.ilike(f'%{search_term}%'),
            Cliente.nome_fantasia.ilike(f'%{search_term}%')
        )
    ).limit(10).all()
    
    resultados = set()
    for cliente in clientes:
        resultados.add(cliente.nome_razao_social)
        if cliente.nome_fantasia:
            resultados.add(cliente.nome_fantasia)
    
    return jsonify(list(resultados))

@app.route('/viagem/<int:viagem_id>/iniciar', methods=['POST'])
@login_required
def iniciar_viagem_motorista(viagem_id):
    if current_user.role != 'Motorista':
        return jsonify({'success': False, 'message': 'Acesso negado.'}), 403

    viagem = Viagem.query.filter_by(id=viagem_id, empresa_id=current_user.empresa_id).first_or_404()
    if viagem.status != 'pendente':
        return jsonify({'success': False, 'message': 'Esta viagem não está mais pendente.'}), 400

    data = request.get_json()
    # --- CORREÇÃO AQUI ---
    odometro_str = data.get('odometer') # Alterado de 'odometro' para 'odometer'

    try:
        odometro_inicial = float(odometro_str)
        if odometro_inicial < 0:
            raise ValueError("Odômetro não pode ser negativo.")
    except (ValueError, TypeError):
        return jsonify({'success': False, 'message': 'Odômetro inicial inválido. Por favor, insira um número válido.'}), 400

    viagem.status = 'em_andamento'
    viagem.data_inicio = datetime.utcnow()
    viagem.odometro_inicial = odometro_inicial
    
    motorista_formal = Motorista.query.filter_by(cpf_cnpj=current_user.cpf_cnpj).first()
    if motorista_formal:
        viagem.motorista_id = motorista_formal.id
    viagem.motorista_cpf_cnpj = current_user.cpf_cnpj
    
    db.session.commit()
    
    # Emite evento para atualizar telas de admin em tempo real
    socketio.emit('status_viagem_atualizado', {
        'viagem_id': viagem.id, 
        'status': 'em_andamento'
    }, room='admins')

    return jsonify({'success': True, 'message': 'Viagem iniciada com sucesso!'})

@app.route('/viagem/<int:viagem_id>/despesas_form')
@login_required
def despesas_form_modal(viagem_id):
    """Renderiza o formulário de despesas para ser carregado no modal."""
    Viagem.query.filter_by(id=viagem_id, empresa_id=current_user.empresa_id).first_or_404()
    custo = CustoViagem.query.filter_by(viagem_id=viagem_id).first()
    # Renderiza o NOVO template que criamos
    return render_template('despesas_form_modal.html', viagem=viagem, custo=custo)

@app.route('/certificados')
@login_required
@admin_required
def configuracao_fiscal():
    """Página de configuração de certificados"""
    certificados = CertificadoDigital.query.filter_by(empresa_id=current_user.empresa_id)\
        .order_by(CertificadoDigital.principal.desc(), CertificadoDigital.id.asc()).all()
    
    hoje = date.today()
    return render_template('configuracao_fiscal.html', 
                         certificados=certificados, 
                         hoje=hoje)


@app.route('/certificados/upload', methods=['POST'])
@login_required
@admin_required
def upload_certificado():
    """Upload de novo certificado"""
    try:
        # Validar dados do formulário
        if 'certificado' not in request.files:
            flash('Nenhum arquivo selecionado.', 'error')
            return redirect(url_for('configuracao_fiscal'))
        
        file = request.files['certificado']
        senha = request.form.get('senha_certificado')
        validade_str = request.form.get('validade_certificado')
        
        if not file or file.filename == '':
            flash('Nenhum arquivo selecionado.', 'error')
            return redirect(url_for('configuracao_fiscal'))
        
        if not senha:
            flash('Senha do certificado é obrigatória.', 'error')
            return redirect(url_for('configuracao_fiscal'))
        
        if not validade_str:
            flash('Data de validade é obrigatória.', 'error')
            return redirect(url_for('configuracao_fiscal'))
        
        # Validar extensão do arquivo
        if not file.filename.lower().endswith('.pfx'):
            flash('Apenas arquivos .pfx são aceitos.', 'error')
            return redirect(url_for('configuracao_fiscal'))
        
        # Converter data de validade
        try:
            data_validade = datetime.strptime(validade_str, '%Y-%m-%d').date()
        except ValueError:
            flash('Data de validade inválida.', 'error')
            return redirect(url_for('configuracao_fiscal'))
        
        # Validar se data não está no passado
        if data_validade <= date.today():
            flash('Data de validade deve ser futura.', 'error')
            return redirect(url_for('configuracao_fiscal'))
        
        # Validar certificado e senha
        file_content = file.read()
        try:
            private_key, certificate, _ = load_key_and_certificates(
                file_content, senha.encode('utf-8')
            )
        except Exception as e:
            flash('Erro ao validar certificado. Verifique o arquivo e a senha.', 'error')
            return redirect(url_for('configuracao_fiscal'))
        
        # Extrair CNPJ do certificado para validação
        cnpj_certificado = None
        try:
            serial_number_attrs = certificate.subject.get_attributes_for_oid(NameOID.SERIAL_NUMBER)
            if serial_number_attrs:
                cnpj_certificado = serial_number_attrs[0].value.split(':')[0]
            
            if not cnpj_certificado:
                common_name_attrs = certificate.subject.get_attributes_for_oid(NameOID.COMMON_NAME)
                if common_name_attrs:
                    match = re.search(r':(\d{14})', common_name_attrs[0].value)
                    if match:
                        cnpj_certificado = match.group(1)
        except:
            pass
        
        if not cnpj_certificado:
            flash('Não foi possível extrair o CNPJ do certificado.', 'error')
            return redirect(url_for('configuracao_fiscal'))
        
        # Upload para R2
        s3_client = boto3.client(
            's3',
            endpoint_url=app.config['CLOUDFLARE_R2_ENDPOINT'],
            aws_access_key_id=app.config['CLOUDFLARE_R2_ACCESS_KEY'],
            aws_secret_access_key=app.config['CLOUDFLARE_R2_SECRET_KEY'],
            region_name='auto'
        )
        
        bucket_name = app.config['CLOUDFLARE_R2_BUCKET']
        file_key = f"certificados/{current_user.empresa_id}/{uuid.uuid4()}.pfx"
        
        s3_client.put_object(
            Bucket=bucket_name,
            Key=file_key,
            Body=file_content,
            ContentType='application/x-pkcs12'
        )
        
        # Verificar se é o primeiro certificado da empresa
        certificados_existentes = CertificadoDigital.query.filter_by(
            empresa_id=current_user.empresa_id
        ).count()
        
        # Se é o primeiro certificado, torna-se principal automaticamente
        is_principal = certificados_existentes == 0
        
        # Criar registro no banco
        novo_certificado = CertificadoDigital(
            empresa_id=current_user.empresa_id,
            nome_arquivo=file.filename,
            caminho_r2=file_key,
            data_validade=data_validade,
            principal=is_principal
        )
        novo_certificado.set_senha(senha, app.cipher_suite)
        
        db.session.add(novo_certificado)
        db.session.commit()
        
        if is_principal:
            flash('Certificado enviado e definido como principal com sucesso!', 'success')
        else:
            flash('Certificado enviado com sucesso! Use "Tornar Principal" para ativá-lo.', 'success')
        
        return redirect(url_for('configuracao_fiscal'))
        
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Erro no upload do certificado: {str(e)}")
        flash('Erro interno. Tente novamente.', 'error')
        return redirect(url_for('configuracao_fiscal'))


@app.route('/certificados/definir-principal/<int:certificado_id>', methods=['POST'])
@login_required
@admin_required
def definir_certificado_principal(certificado_id):
    """Define um certificado como principal"""
    try:
        # Verificar se o certificado pertence à empresa do usuário
        certificado = CertificadoDigital.query.filter_by(
            id=certificado_id,
            empresa_id=current_user.empresa_id
        ).first()
        
        if not certificado:
            flash('Certificado não encontrado.', 'error')
            return redirect(url_for('configuracao_fiscal'))
        
        # Verificar se não está vencido
        if certificado.data_validade <= date.today():
            flash('Não é possível tornar principal um certificado vencido.', 'error')
            return redirect(url_for('configuracao_fiscal'))
        
        # Usar método seguro para definir principal
        sucesso = CertificadoDigital.definir_como_principal(certificado_id, current_user.empresa_id)
        
        if sucesso:
            db.session.commit()
            flash(f'Certificado "{certificado.nome_arquivo}" definido como principal!', 'success')
        else:
            flash('Erro ao definir certificado como principal.', 'error')
        
        return redirect(url_for('configuracao_fiscal'))
        
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Erro ao definir certificado principal: {str(e)}")
        flash('Erro interno. Tente novamente.', 'error')
        return redirect(url_for('configuracao_fiscal'))


@app.route('/certificados/excluir/<int:certificado_id>', methods=['POST'])
@login_required
@admin_required
def excluir_certificado(certificado_id):
    """Exclui um certificado (apenas se não for principal)"""
    try:
        certificado = CertificadoDigital.query.filter_by(
            id=certificado_id,
            empresa_id=current_user.empresa_id
        ).first()
        
        if not certificado:
            return jsonify({'success': False, 'message': 'Certificado não encontrado.'})
        
        # CORREÇÃO: Não permitir excluir o certificado principal
        if certificado.principal:
            return jsonify({
                'success': False, 
                'message': 'Não é possível excluir o certificado principal. Defina outro como principal primeiro.'
            })
        
        # Remover arquivo do R2
        try:
            s3_client = boto3.client(
                's3',
                endpoint_url=app.config['CLOUDFLARE_R2_ENDPOINT'],
                aws_access_key_id=app.config['CLOUDFLARE_R2_ACCESS_KEY'],
                aws_secret_access_key=app.config['CLOUDFLARE_R2_SECRET_KEY'],
                region_name='auto'
            )
            bucket_name = app.config['CLOUDFLARE_R2_BUCKET']
            s3_client.delete_object(Bucket=bucket_name, Key=certificado.caminho_r2)
        except Exception as e:
            app.logger.warning(f"Erro ao remover arquivo do R2: {str(e)}")
        
        # Remover do banco
        db.session.delete(certificado)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Certificado excluído com sucesso!'})
        
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Erro ao excluir certificado: {str(e)}")
        return jsonify({'success': False, 'message': 'Erro interno. Tente novamente.'})
    
@app.route('/certificados/status')
@login_required
@admin_required
def status_certificados():
    """API para obter status dos certificados"""
    try:
        from sefaz_service import get_status_consulta_sefaz
        status = get_status_consulta_sefaz(current_user.empresa_id)
        return jsonify(status)
    except Exception as e:
        app.logger.error(f"Erro ao obter status certificados: {str(e)}")
        return jsonify({
            'pode_consultar': False,
            'motivo_bloqueio': 'Erro interno ao verificar status',
            'certificados_status': []
        })

@app.route('/registrar_abastecimento', methods=['POST'])
@login_required
def registrar_abastecimento():
    try:
        # --- CORREÇÃO APLICADA AQUI ---
        # A busca agora é feita pela coluna correta 'cpf' em vez da propriedade 'cpf_cnpj'
        motorista_formal = Motorista.query.filter_by(cpf=current_user.cpf_cnpj).first()
        
        if not motorista_formal:
             return jsonify({'success': False, 'message': 'Perfil de motorista formal não encontrado para o usuário atual.'}), 400

        viagem_ativa = Viagem.query.filter(
            or_(
                Viagem.motorista_cpf_cnpj == current_user.cpf_cnpj,
                Viagem.motorista_id == motorista_formal.id
            ),
            Viagem.status == 'em_andamento'
        ).first()

        if not viagem_ativa:
            return jsonify({'success': False, 'message': 'Nenhuma viagem ativa encontrada para associar o abastecimento.'}), 400

        litros = float(request.form.get('litros'))
        preco_por_litro = float(request.form.get('preco_por_litro'))
        odometro = float(request.form.get('odometro'))
        custo_total = litros * preco_por_litro

        novo_abastecimento = Abastecimento(
            veiculo_id=viagem_ativa.veiculo_id,
            motorista_id=motorista_formal.id,
            viagem_id=viagem_ativa.id,
            litros=litros,
            preco_por_litro=preco_por_litro,
            custo_total=custo_total,
            odometro=odometro
        )

        anexo_url = None
        anexo = request.files.get('anexo_comprovante')
        
        # Verifica se o arquivo 'anexo_comprovante' foi realmente enviado
        if anexo and anexo.filename:
            s3_client = boto3.client(
                's3',
                endpoint_url=app.config['CLOUDFLARE_R2_ENDPOINT'],
                aws_access_key_id=app.config['CLOUDFLARE_R2_ACCESS_KEY'],
                aws_secret_access_key=app.config['CLOUDFLARE_R2_SECRET_KEY'],
                region_name='auto'
            )
            bucket_name = app.config['CLOUDFLARE_R2_BUCKET']
            filename = secure_filename(anexo.filename)
            s3_path = f"abastecimentos/{viagem_ativa.id}/{uuid.uuid4()}-{filename}"
            
            s3_client.upload_fileobj(
                anexo,
                bucket_name,
                s3_path,
                ExtraArgs={
                    'ContentType': anexo.content_type or 'application/octet-stream',
                    'ContentDisposition': 'attachment'
                }
            )
            anexo_url = f"{app.config['CLOUDFLARE_R2_PUBLIC_URL']}/{s3_path}"
        
        novo_abastecimento.anexo_comprovante = anexo_url

        db.session.add(novo_abastecimento)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Abastecimento registrado com sucesso!'})

    except Exception as e:
        db.session.rollback()
        logger.error(f"Erro ao registrar abastecimento: {e}", exc_info=True)
        return jsonify({'success': False, 'message': f'Erro interno: {e}'}), 500



@app.route('/registrar/<token>', methods=['GET', 'POST'])
def registrar_com_token(token):
    convite = Convite.query.filter_by(token=token, usado=False).first()

    if not convite or convite.data_expiracao < datetime.utcnow():
        flash('O link de convite é inválido ou já expirou.', 'error')
        return redirect(url_for('login'))

    if request.method == 'POST':
        nome = request.form.get('nome')
        sobrenome = request.form.get('sobrenome')
        email = request.form.get('email')
        senha = request.form.get('senha')
        cpf_cnpj = re.sub(r'\D', '', request.form.get('cpf_cnpj', ''))

        if email != convite.email:
            flash('O e-mail não corresponde ao convite.', 'error')
            return redirect(request.url)

        if Usuario.query.filter_by(email=email).first() or Usuario.query.filter_by(cpf_cnpj=cpf_cnpj).first():
            flash('Este e-mail ou CPF/CNPJ já está cadastrado.', 'error')
            return redirect(request.url)
        
        # Cria o usuário
        usuario = Usuario(
            nome=nome,
            sobrenome=sobrenome,
            email=email,
            role=convite.role,
            is_admin=(convite.role in ['Admin', 'Master']),
            cpf_cnpj=cpf_cnpj,
            empresa_id=convite.empresa_id
        )
        usuario.set_password(senha)
        db.session.add(usuario)

        # Marca o convite como usado
        convite.usado = True

        # Flush para garantir que `usuario.id` seja gerado antes de atualizar Motorista
        db.session.flush()

        # --- INÍCIO DA CORREÇÃO ---
        # Esta lógica agora só é executada se o convite for para um 'Motorista'.
        # Para Admins e outros papéis, ela é corretamente ignorada.
        if convite.role == 'Motorista':
            motorista = Motorista.query.filter_by(
                cpf=cpf_cnpj,  # Corrigido de 'cpf_cnpj' para 'cpf'
                empresa_id=convite.empresa_id
            ).first()
            if motorista:
                motorista.usuario_id = usuario.id
        # --- FIM DA CORREÇÃO ---

        # Finalmente commit de tudo em bloco único
        db.session.commit()

        flash('Conta criada com sucesso! Faça login.', 'success')
        return redirect(url_for('login'))

    return render_template('registrar_token.html', email=convite.email, role=convite.role)

def calcular_media_km_veiculo(veiculo_id):
    """Calcula a média de KM rodados por dia para um veículo (versão corrigida para SQLite)."""
    hoje = date.today()
    data_limite = hoje - timedelta(days=90)
    
    # O cálculo (odometro_final - odometro_inicial) é feito diretamente no SQL.
    resultado = db.session.query(
        func.sum(Viagem.odometro_final - Viagem.odometro_inicial).label('total_km'),
        func.count(func.distinct(func.date(Viagem.data_inicio))).label('dias_com_viagem')
    ).filter(
        Viagem.veiculo_id == veiculo_id,
        Viagem.status == 'concluida',
        Viagem.odometro_final.isnot(None),
        Viagem.odometro_inicial.isnot(None),
        Viagem.data_inicio >= data_limite
    ).first()

    if resultado and resultado.total_km and resultado.dias_com_viagem > 0:
        return resultado.total_km / resultado.dias_com_viagem
    
    return 100 # Retorna um valor padrão de 100 km/dia se não houver dados

def get_coordinates(endereco):
    url = 'https://maps.googleapis.com/maps/api/geocode/json'
    params = {'address': endereco, 'key': Maps_API_KEY}
    try:
        logger.debug(f"Obtendo coordenadas para: {endereco}")
        response = requests.get(url, params=params, timeout=5)
        response.raise_for_status()
        data = response.json()
        if data['status'] == 'OK' and data['results']:
            location = data['results'][0]['geometry']['location']
            return location['lat'], location['lng']
        logger.warning(f"Endereço não encontrado: {endereco}")
        return None, None
    except requests.exceptions.RequestException as e:
        logger.error(f"Erro ao obter coordenadas: {str(e)}")
        return None, None

@app.route('/enviar_convite', methods=['POST'])
@login_required
@master_required
def enviar_convite():
    # 1. Verifica se o usuário está vinculado a uma empresa
    if not current_user.empresa_id:
        flash('Você precisa estar vinculado a uma empresa para enviar convites.', 'error')
        return redirect(url_for('configuracoes'))

    empresa_admin = current_user.empresa
    if not empresa_admin:
        flash('Empresa não encontrada para o usuário atual.', 'error')
        return redirect(url_for('configuracoes'))

    # 2. Verifica se a empresa tem licença válida e disponível
    if empresa_admin.licenca:
        usuarios_atuais = len(empresa_admin.usuarios)
        max_permitido = empresa_admin.licenca.max_usuarios
        if usuarios_atuais >= max_permitido:
            flash(f'Limite de usuários atingido ({max_permitido}) para o plano da sua empresa.', 'error')
            return redirect(url_for('configuracoes'))

    # 3. Coleta e valida dados do formulário
    email = request.form.get('email')
    role = request.form.get('role')

    if not email or not role:
        flash('E-mail e papel são obrigatórios.', 'error')
        return redirect(url_for('configuracoes'))

    # 4. Restringe os papéis que podem ser atribuídos
    papeis_permitidos = ['Motorista', 'Master', 'Admin']
    if role not in papeis_permitidos:
        flash('Papel inválido. Escolha entre Motorista, Master ou Admin.', 'error')
        return redirect(url_for('configuracoes'))

    # 5. Verifica se o usuário atual tem permissão para o tipo de convite
    if current_user.role == 'Master' and role != 'Motorista':
        flash('Usuários do tipo Master só podem convidar Motoristas.', 'error')
        return redirect(url_for('configuracoes'))

    # 6. Cria o convite com validade de 3 dias
    token = str(uuid.uuid4())
    data_expiracao = datetime.utcnow() + timedelta(days=3)

    convite = Convite(
        email=email,
        token=token,
        criado_por=current_user.id,
        data_expiracao=data_expiracao,
        role=role,
        empresa_id=current_user.empresa_id
    )

    try:
        db.session.add(convite)
        db.session.commit()

        # 7. Envia o e-mail
        link_convite = url_for('registrar_com_token', token=token, _external=True)
        msg = Message(
            subject=f'Convite para acessar o sistema como {role}',
            recipients=[email],
            body=f'Você foi convidado a se registrar no sistema como {role}.\nClique no link abaixo para se cadastrar:\n\n{link_convite}'
        )
        mail.send(msg)

        flash(f'Convite enviado com sucesso para {email} como {role}!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao enviar convite: {str(e)}', 'error')

    return redirect(url_for('configuracoes'))


def validar_endereco(endereco):
    url = 'https://maps.googleapis.com/maps/api/geocode/json'
    params = {'address': endereco, 'key': Maps_API_KEY}
    try:
        logger.debug(f"Validando endereço: {endereco}")
        response = requests.get(url, params=params, timeout=5)
        response.raise_for_status()
        data = response.json()
        return data['status'] == 'OK' and len(data['results']) > 0
    except requests.exceptions.RequestException as e:
        logger.error(f"Erro na validação de endereço: {str(e)}")
        return False




def calcular_rota_otimizada_ors(enderecos):
    if len(enderecos) < 2:
        return None, None, None, None, "São necessários ao menos um endereço de origem e um de destino."

    try:
        coordenadas = []
        for end in enderecos:
            lat, lon = get_coordinates(end)
            if lat is None or lon is None:
                return None, None, None, None, f"Não foi possível encontrar coordenadas para: {end}"
            coordenadas.append((lon, lat))

        client = openrouteservice.Client(key=OPENROUTESERVICE_API_KEY)
        routes = client.directions(
            coordinates=coordenadas,
            profile='driving-car',
            optimize_waypoints=True,
            geometry=True # Pede a geometria da rota
        )

        route_data = routes['routes'][0]
        geometria_rota = route_data.get('geometry') # Captura a geometria
        
        distancia_total_m, duracao_total_s = 0, 0
        if 'summary' in route_data:
            summary = route_data['summary']
            distancia_total_m = summary.get('distance', 0)
            duracao_total_s = summary.get('duration', 0)
        else:
            for segment in route_data.get('segments', []):
                distancia_total_m += segment.get('distance', 0)
                duracao_total_s += segment.get('duration', 0)

        enderecos_processados = []
        if 'waypoint_order' in route_data:
            waypoint_order = route_data['waypoint_order']
            enderecos_processados = [enderecos[0]] 
            enderecos_processados.extend([enderecos[i] for i in waypoint_order])
            if len(enderecos) > 1: enderecos_processados.append(enderecos[-1])
        else:
            enderecos_processados = enderecos

        distancia_km = distancia_total_m / 1000.0
        duracao_segundos = int(duracao_total_s)

        from collections import OrderedDict
        enderecos_otimizados = list(OrderedDict.fromkeys(enderecos_processados))

        return enderecos_otimizados, distancia_km, duracao_segundos, geometria_rota, None

    except Exception as e:
        logger.error(f"Erro inesperado no cálculo de rota ORS: {e}", exc_info=True)
        return None, None, None, None, f"Ocorreu um erro inesperado ao otimizar a rota: {e}"
    

    
    
@app.route('/api/ocr-process', methods=['POST'])
@login_required
def ocr_process():
    """Processa uma imagem enviada e retorna o texto extraído via OCR."""
    if 'image' not in request.files:
        return jsonify({'success': False, 'message': 'Nenhum arquivo de imagem enviado.'}), 400
    
    file = request.files['image']
    if file.filename == '':
        return jsonify({'success': False, 'message': 'Nenhum arquivo selecionado.'}), 400

    try:
        # Lê a imagem em memória
        image_bytes = file.read()
        image = Image.open(io.BytesIO(image_bytes))
        
        # Usa o Pytesseract para extrair o texto (configure o idioma se necessário)
        texto_extraido = pytesseract.image_to_string(image, lang='por') # 'por' para português
        
        # Limpa o texto, removendo quebras de linha excessivas
        texto_limpo = " ".join(texto_extraido.split()).strip()

        return jsonify({'success': True, 'text': texto_limpo})

    except Exception as e:
        logger.error(f"Erro no processamento OCR: {e}", exc_info=True)
        return jsonify({'success': False, 'message': f'Erro ao processar a imagem: {e}'}), 500
    
@app.route('/editar_viagem/<int:viagem_id>', methods=['GET'])
@login_required
def editar_viagem_page(viagem_id):
    viagem = Viagem.query.options(db.joinedload(Viagem.destinos)).filter_by(id=viagem_id, empresa_id=current_user.empresa_id).first_or_404()

    motoristas = Motorista.query.filter_by(empresa_id=current_user.empresa_id).order_by(Motorista.nome).all()

    # LINHA A SER CORRIGIDA:
    veiculos_disponiveis = Veiculo.query.filter_by(status='Disponível', empresa_id=current_user.empresa_id, is_administrativo=False).all()
    veiculo_atual = db.session.get(Veiculo, viagem.veiculo_id)
    if veiculo_atual and veiculo_atual not in veiculos_disponiveis:
        veiculos_disponiveis.insert(0, veiculo_atual)

    return render_template('editar_viagem.html', 
                        viagem=viagem,
                        motoristas=motoristas, 
                        veiculos=veiculos_disponiveis,
                        ORS_API_KEY=OPENROUTESERVICE_API_KEY)

@app.route('/api/viagem/editar/<int:viagem_id>', methods=['POST'])
@login_required
def editar_viagem_api(viagem_id):
    viagem = Viagem.query.filter_by(id=viagem_id, empresa_id=current_user.empresa_id).first_or_404()
    
    try:
        data = request.get_json()
        
        viagem.motorista_id = data.get('motorista_id')
        viagem.veiculo_id = data.get('veiculo_id')
        viagem.cliente = data.get('cliente')
        viagem.data_inicio = datetime.strptime(data.get('data_inicio'), '%Y-%m-%dT%H:%M')
        viagem.forma_pagamento = data.get('forma_pagamento')
        viagem.valor_recebido = float(data.get('valor_recebido') or 0)
        viagem.observacoes = data.get('observacoes')
        
        enderecos_destino = data.get('enderecos_destino', [])
        
        if not enderecos_destino:
            return jsonify({'success': False, 'message': 'É necessário pelo menos um endereço de destino.'}), 400

        todos_enderecos = [viagem.endereco_saida] + enderecos_destino
        
        # --- LINHA CORRIGIDA ---
        rota_otimizada, distancia_km, duracao_segundos, geometria, erro = calcular_rota_otimizada_ors(todos_enderecos)

        if erro:
            return jsonify({'success': False, 'message': f'Erro ao recalcular a rota: {erro}'}), 400
        
        viagem.distancia_km = distancia_km
        viagem.duracao_segundos = duracao_segundos
        viagem.endereco_destino = rota_otimizada[-1]
        viagem.route_geometry = geometria

        Destino.query.filter_by(viagem_id=viagem_id).delete()
        db.session.flush()

        for ordem, endereco in enumerate(rota_otimizada[1:], 1):
            destino = Destino(viagem_id=viagem_id, endereco=endereco, ordem=ordem)
            db.session.add(destino)

        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Viagem atualizada com sucesso!',
            'roteiro': rota_otimizada,
            'distancia': f"{distancia_km:.2f}",
            'duracao_minutos': duracao_segundos // 60
        })
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Erro na API ao editar viagem {viagem_id}: {e}", exc_info=True)
        return jsonify({'success': False, 'message': f'Ocorreu um erro inesperado ao salvar: {e}'}), 500


@app.route('/cobrancas')
@login_required
def consultar_cobrancas():
    # Query corrigida para buscar apenas cobranças da empresa do usuário logado
    cobrancas = Cobranca.query.filter(
        Cobranca.empresa_id == current_user.empresa_id
    ).options(
        db.joinedload(Cobranca.cliente)
    ).order_by(Cobranca.data_vencimento.desc()).all()
    
    # Atualiza o status para 'Vencida' se necessário (apenas nas cobranças da empresa)
    for cobranca in cobrancas:
        if cobranca.is_vencida:
            cobranca.status = 'Vencida'
    db.session.commit()

    # Cálculos de totais agora são feitos apenas com os dados da empresa correta
    total_pendente = sum(c.valor_total for c in cobrancas if c.status in ['Pendente', 'Vencida'])
    total_pago = sum(c.valor_total for c in cobrancas if c.status == 'Paga')

    return render_template('consultar_cobrancas.html', 
                           cobrancas=cobrancas,
                           total_pendente=total_pendente,
                           total_pago=total_pago,
                           active_page='cobrancas')

@app.route('/borracharia')
@login_required
def borracharia_dashboard():
    """Dashboard principal da borracharia"""
    # KPIs básicos
    kpis = {
        'em_estoque': Pneu.query.filter_by(empresa_id=current_user.empresa_id, status='Estoque').count(),
        'em_uso': Pneu.query.filter_by(empresa_id=current_user.empresa_id, status='Em Uso').count(),
        'recapando': Pneu.query.filter_by(empresa_id=current_user.empresa_id, status='Recapando').count(),
        'descartados': Pneu.query.filter_by(empresa_id=current_user.empresa_id, status='Descartado').count(),
    }
    
    # Alertas
    alertas = []
    config = ConfiguracaoBorracharia.query.filter_by(empresa_id=current_user.empresa_id).first()
    if config:
        # Pneus com problemas
        pneus_alertas = Pneu.query.filter_by(empresa_id=current_user.empresa_id).filter(
            Pneu.status.in_(['Estoque', 'Em Uso'])
        ).all()
        
        for pneu in pneus_alertas:
            if pneu.precisa_descarte:
                alertas.append({
                    'pneu': pneu,
                    'tipo': 'Descarte Necessário',
                    'mensagem': 'Pneu deve ser descartado por sulco ou excesso de recapagens',
                    'gravidade': 'vermelho'
                })
            elif pneu.precisa_recapagem:
                alertas.append({
                    'pneu': pneu,
                    'tipo': 'Recapagem Necessária',
                    'mensagem': f'Sulco atual: {pneu.sulco_atual_mm}mm',
                    'gravidade': 'amarelo'
                })
            elif pneu.precisa_alerta_dot:
                alertas.append({
                    'pneu': pneu,
                    'tipo': 'Idade Avançada',
                    'mensagem': f'Pneu com {pneu.idade_anos} anos de idade',
                    'gravidade': 'amarelo'
                })
    
    # Buscar veículos e pneus
    veiculos = Veiculo.query.filter_by(empresa_id=current_user.empresa_id).options(
        db.joinedload(Veiculo.pneus)
    ).all()
    
    pneus_em_estoque = Pneu.query.filter_by(
        empresa_id=current_user.empresa_id, 
        status='Estoque'
    ).order_by(Pneu.numero_fogo).all()
    
    pneus_recapando = Pneu.query.filter_by(
        empresa_id=current_user.empresa_id,
        status='Recapando'
    ).order_by(Pneu.numero_fogo).all()
    
    return render_template('borracharia_dashboard.html',
                           kpis=kpis,
                           alertas=alertas,
                           veiculos=veiculos,
                           pneus_em_estoque=pneus_em_estoque,
                           pneus_recapando=pneus_recapando,
                           active_page='borracharia')

@app.route('/borracharia/configuracoes', methods=['GET', 'POST'])
@login_required
@master_required
def borracharia_configuracoes():
    """Configurações da borracharia"""
    config = ConfiguracaoBorracharia.query.filter_by(empresa_id=current_user.empresa_id).first()
    
    if not config:
        config = ConfiguracaoBorracharia(empresa_id=current_user.empresa_id)
        db.session.add(config)
        db.session.commit()
    
    if request.method == 'POST':
        try:
            config.vida_util_dot_anos = int(request.form.get('vida_util_dot_anos', 7))
            config.alerta_dot_dias = int(request.form.get('alerta_dot_dias', 90))
            config.sulco_minimo_recapagem_mm = float(request.form.get('sulco_minimo_recapagem_mm', 3.0))
            config.sulco_minimo_descarte_mm = float(request.form.get('sulco_minimo_descarte_mm', 1.6))
            config.max_recapagens = int(request.form.get('max_recapagens', 2))
            config.km_alerta_recapagem = int(request.form.get('km_alerta_recapagem', 5000))
            
            db.session.commit()
            flash('Configurações atualizadas com sucesso!', 'success')
            return redirect(url_for('borracharia_dashboard'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao salvar configurações: {e}', 'error')
    
    return render_template('configuracao_borracharia.html', config=config)

@app.route('/borracharia/pneu/novo', methods=['POST'])
@login_required
def pneu_novo():
    """Cadastrar novo pneu"""
    try:
        # Verificar se já existe pneu com mesmo número
        numero_fogo = request.form.get('numero_fogo')
        if Pneu.query.filter_by(empresa_id=current_user.empresa_id, numero_fogo=numero_fogo).first():
            flash(f'Já existe um pneu com o número {numero_fogo}', 'error')
            return redirect(url_for('borracharia_dashboard'))
        
        novo_pneu = Pneu(
            empresa_id=current_user.empresa_id,
            numero_fogo=numero_fogo,
            marca=request.form.get('marca'),
            modelo=request.form.get('modelo'),
            dimensao=request.form.get('dimensao'),
            dot=request.form.get('dot'),
            data_compra=datetime.strptime(request.form.get('data_compra'), '%Y-%m-%d').date(),
            valor_compra=float(request.form.get('valor_compra')),
            fornecedor=request.form.get('fornecedor'),
            status='Estoque'
        )
        
        db.session.add(novo_pneu)
        db.session.commit()
        
        # Registrar movimentação
        movimentacao = MovimentacaoPneu(
            pneu_id=novo_pneu.id,
            tipo_movimentacao='ENTRADA_ESTOQUE',
            observacoes='Pneu cadastrado no sistema',
            usuario_id=current_user.id
        )
        db.session.add(movimentacao)
        db.session.commit()
        
        flash('Pneu cadastrado com sucesso!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao cadastrar pneu: {e}', 'error')
        logger.error(f"Erro ao cadastrar pneu: {e}", exc_info=True)
    
    return redirect(url_for('borracharia_dashboard'))

@app.route('/api/pneu/instalar', methods=['POST'])
@login_required
def api_instalar_pneu():
    """API para instalar pneu em veículo"""
    try:
        data = request.get_json()
        pneu_id = data.get('pneu_id')
        veiculo_id = data.get('veiculo_id')
        posicao = data.get('posicao')
        km_veiculo = float(data.get('km_veiculo', 0))
        
        # Buscar pneu e veículo
        pneu = Pneu.query.filter_by(id=pneu_id, empresa_id=current_user.empresa_id).first()
        veiculo = Veiculo.query.filter_by(id=veiculo_id, empresa_id=current_user.empresa_id).first()
        
        if not pneu or not veiculo:
            return jsonify({'success': False, 'message': 'Pneu ou veículo não encontrado'}), 404
        
        if pneu.status != 'Estoque':
            return jsonify({'success': False, 'message': 'Pneu não está disponível no estoque'}), 400
        
        # Verificar se já existe pneu na posição
        pneu_existente = Pneu.query.filter_by(
            veiculo_id=veiculo_id, 
            posicao=posicao, 
            status='Em Uso'
        ).first()
        
        if pneu_existente:
            return jsonify({'success': False, 'message': f'Já existe um pneu na posição {posicao}'}), 400
        
        # Instalar pneu
        pneu.status = 'Em Uso'
        pneu.veiculo_id = veiculo_id
        pneu.posicao = posicao
        pneu.km_instalacao = km_veiculo
        pneu.data_instalacao = date.today()
        
        # Atualizar KM do veículo se necessário
        if veiculo.km_atual is None or km_veiculo > veiculo.km_atual:
            veiculo.km_atual = km_veiculo
        
        # Registrar movimentação
        movimentacao = MovimentacaoPneu(
            pneu_id=pneu.id,
            tipo_movimentacao='INSTALACAO',
            veiculo_id=veiculo_id,
            posicao_destino=posicao,
            km_veiculo=km_veiculo,
            observacoes=f'Instalado na posição {posicao}',
            usuario_id=current_user.id
        )
        db.session.add(movimentacao)
        
        db.session.commit()
        return jsonify({'success': True, 'message': 'Pneu instalado com sucesso!'})
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Erro ao instalar pneu: {e}", exc_info=True)
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/pneu/remover', methods=['POST'])
@login_required
def api_remover_pneu():
    """API para remover pneu de veículo"""
    try:
        data = request.get_json()
        pneu_id = data.get('pneu_id')
        km_veiculo = float(data.get('km_veiculo', 0))
        destino = data.get('destino')  # estoque, recapagem, descarte
        motivo_descarte = data.get('motivo_descarte')
        
        # Buscar pneu
        pneu = Pneu.query.filter_by(id=pneu_id, empresa_id=current_user.empresa_id).first()
        if not pneu:
            return jsonify({'success': False, 'message': 'Pneu não encontrado'}), 404
        
        if pneu.status != 'Em Uso':
            return jsonify({'success': False, 'message': 'Pneu não está em uso'}), 400
        
        # Atualizar KM do veículo
        if pneu.veiculo and (pneu.veiculo.km_atual is None or km_veiculo > pneu.veiculo.km_atual):
            pneu.veiculo.km_atual = km_veiculo
        
        # Registrar dados de remoção
        posicao_origem = pneu.posicao
        pneu.km_remocao = km_veiculo
        pneu.data_remocao = date.today()
        
        # Definir novo status baseado no destino
        if destino == 'estoque':
            pneu.status = 'Estoque'
            pneu.veiculo_id = None
            pneu.posicao = None
            tipo_mov = 'REMOCAO'
            obs = 'Removido para estoque'
        elif destino == 'recapagem':
            pneu.status = 'Recapando'
            pneu.veiculo_id = None
            pneu.posicao = None
            tipo_mov = 'RECAPAGEM_ENVIO'
            obs = 'Enviado para recapagem'
        elif destino == 'descarte':
            pneu.status = 'Descartado'
            pneu.veiculo_id = None
            pneu.posicao = None
            pneu.motivo_descarte = motivo_descarte
            pneu.data_descarte = date.today()
            tipo_mov = 'DESCARTE'
            obs = f'Descartado: {motivo_descarte}'
        else:
            return jsonify({'success': False, 'message': 'Destino inválido'}), 400
        
        # Registrar movimentação
        movimentacao = MovimentacaoPneu(
            pneu_id=pneu.id,
            tipo_movimentacao=tipo_mov,
            veiculo_id=pneu.veiculo_id,
            posicao_origem=posicao_origem,
            km_veiculo=km_veiculo,
            observacoes=obs,
            usuario_id=current_user.id
        )
        db.session.add(movimentacao)
        
        db.session.commit()
        return jsonify({'success': True, 'message': 'Pneu removido com sucesso!'})
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Erro ao remover pneu: {e}", exc_info=True)
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/veiculo/<int:veiculo_id>/pneus')
@login_required
def api_veiculo_pneus(veiculo_id):
    """API para listar pneus instalados em um veículo"""
    veiculo = Veiculo.query.filter_by(id=veiculo_id, empresa_id=current_user.empresa_id).first()
    if not veiculo:
        return jsonify({'success': False, 'message': 'Veículo não encontrado'}), 404
    
    pneus = Pneu.query.filter_by(veiculo_id=veiculo_id, status='Em Uso').all()
    
    pneus_data = [{
        'id': p.id,
        'numero_fogo': p.numero_fogo,
        'posicao': p.posicao,
        'marca': p.marca,
        'modelo': p.modelo
    } for p in pneus]
    
    return jsonify({
        'success': True,
        'pneus': pneus_data,
        'km_atual': veiculo.km_atual
    })

@app.route('/api/pneu/registrar_medicao', methods=['POST'])
@login_required
def api_registrar_medicao():
    """API para registrar medição de sulco"""
    try:
        data = request.get_json()
        pneu_id = data.get('pneu_id')
        sulco_mm = float(data.get('sulco_mm'))
        observacoes = data.get('observacoes')
        
        # Buscar pneu
        pneu = Pneu.query.filter_by(id=pneu_id, empresa_id=current_user.empresa_id).first()
        if not pneu:
            return jsonify({'success': False, 'message': 'Pneu não encontrado'}), 404
        
        # Registrar medição
        medicao = MedicaoSulco(
            pneu_id=pneu_id,
            sulco_mm=sulco_mm,
            observacoes=observacoes,
            usuario_id=current_user.id
        )
        db.session.add(medicao)
        
        # Atualizar sulco atual do pneu
        pneu.sulco_atual_mm = sulco_mm
        pneu.data_ultima_medicao = date.today()
        
        db.session.commit()
        return jsonify({'success': True, 'message': 'Medição registrada com sucesso!'})
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Erro ao registrar medição: {e}", exc_info=True)
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/pneu/retorno_recapagem', methods=['POST'])
@login_required
def api_retorno_recapagem():
    """API para registrar retorno de recapagem"""
    try:
        data = request.get_json()
        pneu_id = data.get('pneu_id')
        sulco_mm = float(data.get('sulco_mm'))
        custo = float(data.get('custo', 0))
        fornecedor = data.get('fornecedor')
        observacoes = data.get('observacoes')
        
        # Buscar pneu
        pneu = Pneu.query.filter_by(id=pneu_id, empresa_id=current_user.empresa_id).first()
        if not pneu:
            return jsonify({'success': False, 'message': 'Pneu não encontrado'}), 404
        
        if pneu.status != 'Recapando':
            return jsonify({'success': False, 'message': 'Pneu não está em recapagem'}), 400
        
        # Atualizar dados do pneu
        pneu.status = 'Estoque'
        pneu.numero_recapagens += 1
        pneu.data_ultima_recapagem = date.today()
        pneu.custo_total_recapagens += custo
        pneu.sulco_atual_mm = sulco_mm
        pneu.data_ultima_medicao = date.today()
        
        # Registrar movimentação
        movimentacao = MovimentacaoPneu(
            pneu_id=pneu.id,
            tipo_movimentacao='RECAPAGEM_RETORNO',
            custo_recapagem=custo,
            fornecedor_recapagem=fornecedor,
            observacoes=observacoes,
            usuario_id=current_user.id
        )
        db.session.add(movimentacao)
        
        # Registrar medição de sulco
        medicao = MedicaoSulco(
            pneu_id=pneu_id,
            sulco_mm=sulco_mm,
            observacoes=f'Medição após recapagem #{pneu.numero_recapagens}',
            usuario_id=current_user.id
        )
        db.session.add(medicao)
        
        db.session.commit()
        return jsonify({'success': True, 'message': 'Retorno de recapagem registrado com sucesso!'})
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Erro ao registrar retorno de recapagem: {e}", exc_info=True)
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/borracharia/pneu/importar_lote', methods=['POST'])
@login_required
def pneu_importar_lote():
    """Importar pneus em lote via CSV"""
    try:
        if 'csv_file' not in request.files:
            flash('Nenhum arquivo selecionado', 'error')
            return redirect(url_for('borracharia_dashboard'))
        
        file = request.files['csv_file']
        if file.filename == '' or not file.filename.endswith('.csv'):
            flash('Arquivo inválido. Selecione um arquivo CSV.', 'error')
            return redirect(url_for('borracharia_dashboard'))
        
        import csv
        import io
        
        # Ler arquivo CSV
        stream = io.StringIO(file.stream.read().decode("UTF8"), newline=None)
        csv_reader = csv.DictReader(stream)
        
        pneus_criados = 0
        erros = []
        
        for linha_num, row in enumerate(csv_reader, start=2):
            try:
                # Verificar se já existe
                numero_fogo = row.get('numero_fogo', '').strip()
                if not numero_fogo:
                    erros.append(f'Linha {linha_num}: Número do fogo é obrigatório')
                    continue
                
                if Pneu.query.filter_by(empresa_id=current_user.empresa_id, numero_fogo=numero_fogo).first():
                    erros.append(f'Linha {linha_num}: Pneu {numero_fogo} já existe')
                    continue
                
                # Criar pneu
                pneu = Pneu(
                    empresa_id=current_user.empresa_id,
                    numero_fogo=numero_fogo,
                    marca=row.get('marca', '').strip(),
                    modelo=row.get('modelo', '').strip(),
                    dimensao=row.get('dimensao', '').strip(),
                    dot=row.get('dot', '').strip(),
                    data_compra=datetime.strptime(row.get('data_compra', ''), '%Y-%m-%d').date(),
                    valor_compra=float(row.get('valor_compra', 0)),
                    fornecedor=row.get('fornecedor', '').strip(),
                    status='Estoque'
                )
                db.session.add(pneu)
                db.session.flush()
                
                # Registrar movimentação
                movimentacao = MovimentacaoPneu(
                    pneu_id=pneu.id,
                    tipo_movimentacao='ENTRADA_ESTOQUE',
                    observacoes='Importado via CSV',
                    usuario_id=current_user.id
                )
                db.session.add(movimentacao)
                
                pneus_criados += 1
                
            except Exception as e:
                erros.append(f'Linha {linha_num}: {str(e)}')
        
        if pneus_criados > 0:
            db.session.commit()
            flash(f'{pneus_criados} pneu(s) importado(s) com sucesso!', 'success')
        
        if erros:
            flash(f'Erros encontrados: {"; ".join(erros[:5])}{"..." if len(erros) > 5 else ""}', 'warning')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao importar arquivo: {e}', 'error')
        logger.error(f"Erro ao importar CSV de pneus: {e}", exc_info=True)
    
    return redirect(url_for('borracharia_dashboard'))

@app.route('/borracharia/pneu/modelo_csv')
@login_required
def pneu_exportar_modelo_csv():
    """Gerar arquivo modelo CSV para importação"""
    import csv
    import io
    
    output = io.BytesIO()
    wrapper = io.TextIOWrapper(output, encoding='utf-8', newline='')
    
    # --- CORREÇÃO APLICADA AQUI ---
    # Adicionamos o delimiter=';' para que o Excel (no Brasil) abra corretamente.
    writer = csv.writer(wrapper, delimiter=';') 
    
    # Cabeçalhos
    writer.writerow([
        'numero_fogo', 'marca', 'modelo', 'dimensao', 'dot', 
        'data_compra', 'valor_compra', 'fornecedor'
    ])
    
    # Linha de exemplo
    writer.writerow([
        'P001', 'Michelin', 'X Line Energy D', '295/80 R22.5', 'DOT1234',
        '2024-01-15', '1200.50', 'Fornecedor Exemplo'
    ])
    
    wrapper.detach()
    output.seek(0)
    
    return send_file(
        output,
        mimetype='text/csv',
        as_attachment=True,
        download_name='modelo_pneus.csv'
    )

@app.route('/borracharia/pneu/<int:pneu_id>/historico')
@login_required
def pneu_historico(pneu_id):
    """Histórico detalhado de um pneu"""
    pneu = Pneu.query.filter_by(id=pneu_id, empresa_id=current_user.empresa_id).first_or_404()
    
    # Buscar movimentações e medições
    movimentacoes = MovimentacaoPneu.query.filter_by(pneu_id=pneu_id).options(
        db.joinedload(MovimentacaoPneu.veiculo),
        db.joinedload(MovimentacaoPneu.usuario)
    ).order_by(MovimentacaoPneu.data_movimentacao.desc()).all()
    
    medicoes = MedicaoSulco.query.filter_by(pneu_id=pneu_id).options(
        db.joinedload(MedicaoSulco.usuario)
    ).order_by(MedicaoSulco.data_medicao.desc()).all()
    
    return render_template('pneu_historico.html',
                           pneu=pneu,
                           movimentacoes=movimentacoes,
                           medicoes=medicoes)


@app.route('/api/viagem/<int:viagem_id>/map_data')
@login_required
def get_viagem_map_data(viagem_id):
    viagem = Viagem.query.filter_by(id=viagem_id, empresa_id=current_user.empresa_id).first_or_404()
    
    ultima_localizacao = Localizacao.query.filter_by(viagem_id=viagem_id).order_by(Localizacao.timestamp.desc()).first()
    
    localizacao_data = None
    if ultima_localizacao:
        localizacao_data = {
            'latitude': ultima_localizacao.latitude,
            'longitude': ultima_localizacao.longitude,
            'endereco': ultima_localizacao.endereco,
            'timestamp': ultima_localizacao.timestamp.strftime('%d/%m/%Y %H:%M')
        }

    # Lógica para buscar as coordenadas de cada destino
    destinos_com_coords = []
    for destino in sorted(viagem.destinos, key=lambda x: x.ordem):
        lat, lon = get_coordinates(destino.endereco)
        destinos_com_coords.append({
            'endereco': destino.endereco,
            'latitude': lat,
            'longitude': lon
        })

    return jsonify({
        'success': True,
        'route_geometry': viagem.route_geometry,
        'ultima_localizacao': localizacao_data,
        'destinos': destinos_com_coords,
        'origem': viagem.endereco_saida
    })


@app.route('/cobranca/gerar', methods=['GET', 'POST'])
@login_required
def gerar_cobranca():
    if request.method == 'POST':
        try:
            cliente_id = request.form.get('cliente_id')
            viagem_ids = request.form.getlist('viagem_ids')
            data_vencimento_str = request.form.get('data_vencimento')
            observacoes = request.form.get('observacoes')

            if not all([cliente_id, viagem_ids, data_vencimento_str]):
                flash('Cliente, data de vencimento e ao menos uma viagem são obrigatórios.', 'error')
                return redirect(url_for('gerar_cobranca'))

            viagens_selecionadas = Viagem.query.filter(Viagem.id.in_(viagem_ids)).all()
            valor_total = sum(v.valor_recebido or 0 for v in viagens_selecionadas)

            nova_cobranca = Cobranca(
                cliente_id=cliente_id,
                usuario_id=current_user.id,
                valor_total=valor_total,
                data_vencimento=datetime.strptime(data_vencimento_str, '%Y-%m-%d').date(),
                observacoes=observacoes,
                empresa_id=current_user.empresa_id
            )
            
            # --- LÓGICA CORRIGIDA E MAIS ROBUSTA ---
            # Em vez de apenas definir o ID, nós explicitamente adicionamos as viagens
            # à coleção da nova cobrança. O SQLAlchemy cuidará do resto.
            for viagem in viagens_selecionadas:
                nova_cobranca.viagens.append(viagem)
            
            db.session.add(nova_cobranca)
            db.session.commit()
            
            flash('Cobrança gerada com sucesso! Visualizando a Nota de Débito.', 'success')
            return redirect(url_for('visualizar_nota_debito', cobranca_id=nova_cobranca.id))

        except Exception as e:
            db.session.rollback()
            logger.error(f"Erro ao gerar cobrança: {e}", exc_info=True)
            flash(f'Ocorreu um erro ao gerar a cobrança: {e}', 'error')
            return redirect(url_for('gerar_cobranca'))
    
    clientes = Cliente.query.filter_by(empresa_id=current_user.empresa_id).order_by(Cliente.nome_razao_social).all()
    return render_template('gerar_cobranca.html', clientes=clientes, active_page='cobrancas')

@app.route('/api/cobranca/<int:cobranca_id>/marcar_paga', methods=['POST'])
@login_required
def api_marcar_paga(cobranca_id):
    Cobranca.query.filter_by(id=cobranca_id, empresa_id=current_user.empresa_id).first_or_404()
    try:
        data = request.get_json()
        meio_pagamento = data.get('meio_pagamento')

        if not meio_pagamento:
            return jsonify({'success': False, 'message': 'Meio de pagamento é obrigatório.'}), 400

        cobranca.status = 'Paga'
        cobranca.data_pagamento = datetime.utcnow()
        cobranca.meio_pagamento = meio_pagamento
        db.session.commit()

        return jsonify({'success': True, 'message': 'Cobrança marcada como paga.'})
    except Exception as e:
        db.session.rollback()
        logger.error(f"Erro ao marcar cobrança como paga: {e}", exc_info=True)
        return jsonify({'success': False, 'message': str(e)}), 500


def get_address_geoapify(lat, lon):
    try:
        url = f'https://api.geoapify.com/v1/geocode/reverse?lat={lat}&lon={lon}&apiKey={GEOAPIFY_API_KEY}'
        response = requests.get(url)
        if response.status_code == 200:
            data = response.json()
            if data['features']:
                return data['features'][0]['properties']['formatted']
    except Exception as e:
        logger.error(f"Erro na geocodificação Geoapify: {str(e)}")
    return "Endereço não encontrado"
@app.route('/fiscal/cte/escolher_viagem')
@login_required
@master_required
def escolher_viagem_para_cte():
    """Mostra a lista de viagens concluídas sem CT-e para iniciar a emissão."""
    viagens_sem_cte = Viagem.query.filter(
        Viagem.empresa_id == current_user.empresa_id,
        Viagem.status == 'concluida',
        ~Viagem.ctes.any()  # Filtra viagens que NÃO têm CT-e
    ).order_by(Viagem.data_fim.desc()).all()
    
    return render_template('cte/escolher_viagem_cte.html', 
                           viagens=viagens_sem_cte, 
                           active_page='cte')


# ROTA 3: Para a tela de sucesso após a emissão
@app.route('/fiscal/cte/sucesso/<int:cte_id>')
@login_required
@master_required
def sucesso_cte(cte_id):
    """Página exibida após a emissão bem-sucedida de um CT-e."""
    cte = CTeEmitido.query.filter_by(id=cte_id, empresa_id=current_user.empresa_id).first_or_404()
    return render_template('cte/sucesso_cte.html', cte=cte)



@app.route('/api/cliente/<int:cliente_id>/viagens_nao_cobradas')
@login_required
def api_viagens_nao_cobradas(cliente_id):
    cliente = db.session.get(Cliente, cliente_id)
    if not cliente or cliente.empresa_id != current_user.empresa_id:
        return jsonify({'error': 'Cliente não encontrado ou não pertence à sua empresa.'}), 404

    viagens = Viagem.query.filter(
        Viagem.cliente == cliente.nome_razao_social, 
        Viagem.empresa_id == current_user.empresa_id,
        Viagem.cobranca_id.is_(None),         
        Viagem.valor_recebido.isnot(None)   
    ).all()
    
    viagens_data = [{
        'id': v.id,
        'data_inicio': v.data_inicio.strftime('%d/%m/%Y'),
        'destino': v.endereco_destino,
        'valor': v.valor_recebido or 0.0,
        'valor_formatado': f"R$ {v.valor_recebido or 0:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    } for v in viagens]
    
    return jsonify(viagens_data)

@app.route('/')
def index():

    return render_template('index.html')

# Em app.py, substitua a rota 'cadastrar_motorista' por esta versão completa

@app.route('/cadastrar_motorista', methods=['GET', 'POST'])
@login_required
def cadastrar_motorista():
    if request.method == 'POST':
        try:
            # Validações iniciais (CPF, CNH)
            cpf = re.sub(r'\D', '', request.form.get('cpf'))
            cnh_numero = re.sub(r'\D', '', request.form.get('cnh_numero'))

            if Motorista.query.filter_by(cpf=cpf, empresa_id=current_user.empresa_id).first():
                flash('Um motorista com este CPF já foi cadastrado.', 'error')
                return render_template('cadastrar_motorista.html', motorista=request.form)
            
            if Motorista.query.filter_by(cnh_numero=cnh_numero, empresa_id=current_user.empresa_id).first():
                flash('Um motorista com esta CNH já foi cadastrado.', 'error')
                return render_template('cadastrar_motorista.html', motorista=request.form)

            # --- INÍCIO DA CORREÇÃO 1: Lógica para múltiplos anexos e download ---
            
            anexos_urls = []
            anexos_files = request.files.getlist("anexos")

            if anexos_files and any(f.filename for f in anexos_files):
                # Validação das credenciais do Cloudflare R2
                bucket_name = app.config.get('CLOUDFLARE_R2_BUCKET')
                if not bucket_name:
                    flash('Erro de configuração: O nome do bucket do Cloudflare R2 não foi definido nas variáveis de ambiente.', 'error')
                    return render_template('cadastrar_motorista.html', motorista=request.form)

                s3_client = boto3.client(
                    's3',
                    endpoint_url=app.config['CLOUDFLARE_R2_ENDPOINT'],
                    aws_access_key_id=app.config['CLOUDFLARE_R2_ACCESS_KEY'],
                    aws_secret_access_key=app.config['CLOUDFLARE_R2_SECRET_KEY'],
                    region_name='auto'
                )
                public_url_base = app.config['CLOUDFLARE_R2_PUBLIC_URL']
                
                for anexo_file in anexos_files:
                    if anexo_file.filename:
                        filename = secure_filename(anexo_file.filename)
                        s3_path = f"motoristas/{cpf}/anexos/{uuid.uuid4()}-{filename}"
                        
                        s3_client.upload_fileobj(
                            anexo_file,
                            bucket_name,
                            s3_path,
                            ExtraArgs={
                                'ContentType': anexo_file.content_type or 'application/octet-stream',
                                'ContentDisposition': 'attachment'
                            }
                        )
                        anexos_urls.append(f"{public_url_base}/{s3_path}")

            # --- FIM DA CORREÇÃO 1 ---

            def to_date(date_string):
                return datetime.strptime(date_string, '%Y-%m-%d').date() if date_string else None

            usuario_correspondente = Usuario.query.filter_by(cpf_cnpj=cpf, empresa_id=current_user.empresa_id).first()

            novo_motorista = Motorista(
                empresa_id=current_user.empresa_id,
                usuario_id=usuario_correspondente.id if usuario_correspondente else None,
                nome=request.form.get('nome'),
                telefone=re.sub(r'\D', '', request.form.get('telefone')),
                cpf=cpf,
                data_nascimento=to_date(request.form.get('data_nascimento')),
                nacionalidade=request.form.get('nacionalidade'),
                naturalidade=request.form.get('naturalidade'),
                estado_civil=request.form.get('estado_civil'),
                sexo=request.form.get('sexo'),
                pai=request.form.get('pai'),
                mae=request.form.get('mae'),
                data_admissao=to_date(request.form.get('data_admissao')),
                situacao=request.form.get('situacao'),
                data_desativacao=to_date(request.form.get('data_desativacao')),
                classificacao=request.form.get('classificacao'),
                cod_departamento=request.form.get('cod_departamento'),
                numero_ficha=request.form.get('numero_ficha'),
                foto=None,
                anexos=','.join(anexos_urls) if anexos_urls else None,
                cep=re.sub(r'\D', '', request.form.get('cep')),
                tipo_logradouro=request.form.get('tipo_logradouro'),
                logradouro=request.form.get('logradouro'),
                numero=request.form.get('numero'),
                complemento=request.form.get('complemento'),
                bairro=request.form.get('bairro'),
                cidade=request.form.get('cidade'),
                uf=request.form.get('uf'),
                email=request.form.get('email'),
                tipo_imovel=request.form.get('tipo_imovel'),
                tempo_residencia=request.form.get('tempo_residencia'),
                cnh_numero=cnh_numero,
                cnh_data_primeira=to_date(request.form.get('cnh_data_primeira')),
                cnh_data_vencimento=to_date(request.form.get('cnh_data_vencimento')),
                cnh_categoria=request.form.get('cnh_categoria'),
                cnh_cod_seguranca=request.form.get('cnh_cod_seguranca'),
                rg=request.form.get('rg'),
                rg_uf=request.form.get('rg_uf'),
                pis=request.form.get('pis'),
                inss=request.form.get('inss'),
                titulo_eleitor=request.form.get('titulo_eleitor'),
                ctps=request.form.get('ctps'),
                funcao=request.form.get('funcao'),
                mopp_numero=request.form.get('mopp_numero'),
                mopp_vencimento=to_date(request.form.get('mopp_vencimento')),
                contato_nome=request.form.get('contato_nome'),
                contato_tipo_ref=request.form.get('contato_tipo_ref'),
                contato_tipo_fone=request.form.get('contato_tipo_fone'),
                contato_telefone=re.sub(r'\D', '', request.form.get('contato_telefone')),
                contato_operadora=request.form.get('contato_operadora'),
                contato_obs=request.form.get('contato_obs')
            )
            
            db.session.add(novo_motorista)
            db.session.commit()
            flash('Motorista cadastrado com sucesso!', 'success')
            return redirect(url_for('consultar_motoristas'))

        except Exception as e:
            db.session.rollback()
            logger.error(f"Erro ao cadastrar motorista: {e}", exc_info=True)
            flash(f'Ocorreu um erro ao salvar o motorista: {e}', 'error')
            return render_template('cadastrar_motorista.html', motorista=request.form)

    return render_template('cadastrar_motorista.html', motorista={}, active_page='cadastrar_motorista')

# Substitua sua função 'dateformat' por esta versão
@app.template_filter('dateformat')
def dateformat(value):
    # --- INÍCIO DA CORREÇÃO 2: Filtro de data inteligente ---
    if isinstance(value, (date, datetime)):
        return value.strftime('%Y-%m-%d')
    # Se o valor já for uma string (como ao recarregar o form após um erro),
    # apenas o retorna sem tentar formatar.
    if isinstance(value, str):
        return value
    return ''
    # --- FIM DA CORREÇÃO 2 ---


@app.route('/fiscal/cte/<int:cte_id>/pdf')
@login_required
@master_required # <--- SEU DECORATOR ESTAVA FALTANDO AQUI
def gerar_dacte_pdf(cte_id):
    """Gera e baixa o PDF do DACTe usando xhtml2pdf (versão simplificada)."""
    cte = CTeEmitido.query.filter_by(id=cte_id, empresa_id=current_user.empresa_id).first_or_404()
    
    # Renderiza o template HTML (isso não muda)
    html_renderizado = render_template('cte/dacte_template.html', cte=cte)
    
    # --- INÍCIO DA LÓGICA SIMPLIFICADA com xhtml2pdf ---
    pdf_buffer = io.BytesIO()
    
    # Converte o HTML para PDF e salva no buffer de memória
    pisa_status = pisa.CreatePDF(
        io.StringIO(html_renderizado),  # Fonte do HTML
        dest=pdf_buffer)                # Destino (o buffer)

    # Se a conversão falhou, retorna um erro
    if pisa_status.err:
        logger.error(f"Erro ao gerar PDF com xhtml2pdf: {pisa_status.err}")
        return "<h1>Erro ao gerar PDF</h1><p>Ocorreu um problema com a biblioteca de conversão.</p>", 500

    pdf_buffer.seek(0)
    # --- FIM DA LÓGICA SIMPLIFICADA ---
    
    return send_file(
        pdf_buffer,
        as_attachment=True,
        download_name=f'DACTe_{cte.chave_acesso}.pdf',
        mimetype='application/pdf'
    )

@app.route('/editar_cliente/<int:cliente_id>', methods=['GET', 'POST'])
@login_required
def editar_cliente(cliente_id):
    cliente = Cliente.query.filter_by(id=cliente_id, empresa_id=current_user.empresa_id).first_or_404()

    if request.method == 'POST':
        try:
            novo_cpf_cnpj = re.sub(r'\D', '', request.form.get('cpf_cnpj', ''))

            cliente_existente = Cliente.query.filter(Cliente.cpf_cnpj == novo_cpf_cnpj, Cliente.id != cliente_id).first()
            if cliente_existente:
                flash('Erro: O CPF/CNPJ informado já pertence a outro cliente.', 'error')
                return redirect(url_for('editar_cliente', cliente_id=cliente_id))

            cliente.pessoa_tipo = request.form.get('pessoa_tipo')
            cliente.nome_razao_social = request.form.get('nome_razao_social')
            cliente.nome_fantasia = request.form.get('nome_fantasia') if cliente.pessoa_tipo == 'juridica' else None
            cliente.cpf_cnpj = novo_cpf_cnpj
            cliente.inscricao_estadual = request.form.get('inscricao_estadual') if cliente.pessoa_tipo == 'juridica' else None
            cliente.cep = re.sub(r'\D', '', request.form.get('cep', ''))
            cliente.logradouro = request.form.get('logradouro')
            cliente.numero = request.form.get('numero')
            cliente.complemento = request.form.get('complemento')
            cliente.bairro = request.form.get('bairro')
            cliente.cidade = request.form.get('cidade')
            cliente.estado = request.form.get('estado')
            cliente.email = request.form.get('email')
            cliente.telefone = re.sub(r'\D', '', request.form.get('telefone', ''))
            cliente.contato_principal = request.form.get('contato_principal')

            # Adicionar lógica de upload de novos arquivos
            novos_anexos_files = request.files.getlist('anexos')
            if novos_anexos_files and any(f.filename for f in novos_anexos_files):
                anexos_atuais = cliente.anexos.split(',') if cliente.anexos else []
                s3_client = boto3.client(
                    's3',
                    endpoint_url=app.config['CLOUDFLARE_R2_ENDPOINT'],
                    aws_access_key_id=app.config['CLOUDFLARE_R2_ACCESS_KEY'],
                    aws_secret_access_key=app.config['CLOUDFLARE_R2_SECRET_KEY'],
                    region_name='auto'
                )
                bucket_name = app.config['CLOUDFLARE_R2_BUCKET']
                public_url_base = app.config['CLOUDFLARE_R2_PUBLIC_URL']

                for file in novos_anexos_files:
                    if file.filename:
                        filename = secure_filename(file.filename)
                        s3_path = f"clientes/{cliente.cpf_cnpj}/{uuid.uuid4()}-{filename}"
                        s3_client.upload_fileobj(
                            file, bucket_name, s3_path,
                            ExtraArgs={'ContentType': file.content_type or 'application/octet-stream'}
                        )
                        anexos_atuais.append(f"{public_url_base}/{s3_path}")
                
                cliente.anexos = ','.join(filter(None, anexos_atuais))

            db.session.commit()
            flash('Cliente atualizado com sucesso!', 'success')
            return redirect(url_for('consultar_clientes'))

        except Exception as e:
            db.session.rollback()
            logger.error(f"Erro ao editar cliente: {e}", exc_info=True)
            flash(f'Ocorreu um erro inesperado ao salvar as alterações: {e}', 'error')
            return redirect(url_for('editar_cliente', cliente_id=cliente_id))

    return render_template('editar_cliente.html', cliente=cliente, active_page='consultar_clientes')

@app.route('/nota_debito/<int:cobranca_id>')
@login_required
def visualizar_nota_debito(cobranca_id):
    
    
    cobranca = Cobranca.query.filter_by(id=cobranca_id, empresa_id=current_user.empresa_id).first_or_404()
    empresa_emissora = db.session.get(Empresa, cobranca.usuario.empresa_id) if cobranca.usuario else None   
    valor_por_extenso = num2words(cobranca.valor_total, lang='pt_BR', to='currency')
    return render_template(
        'nota_debito.html',
        cobranca=cobranca,
        cliente=cobranca.cliente,
        empresa=empresa_emissora,
        valor_extenso=valor_por_extenso
    )





@app.route('/consultar_motoristas', methods=['GET'])
@login_required
def consultar_motoristas():
    search_query = request.args.get('search', '').strip()
    query = Motorista.query.filter_by(empresa_id=current_user.empresa_id)

    if search_query:
        search_filter = f"%{search_query}%"
        query = query.filter(
            or_(
                Motorista.nome.ilike(search_filter),
                Motorista.cpf.ilike(search_filter),
                Motorista.cnh_numero.ilike(search_filter)
            )
        )

    motoristas_list = query.order_by(Motorista.nome.asc()).all()

    for motorista in motoristas_list:
        viagem_ativa = Viagem.query.filter(
            Viagem.motorista_id == motorista.id,
            Viagem.status == 'em_andamento'
        ).first()
        motorista.status = 'Em Viagem' if viagem_ativa else 'Disponível'

        # Prevenção de erro se a data de vencimento for nula
        if motorista.cnh_data_vencimento is None:
            motorista.cnh_data_vencimento = date.max

    # --- INÍCIO DA CORREÇÃO ---
    # Prepara os dados de data para o template
    hoje = date.today()
    data_alerta_cnh = hoje + timedelta(days=30)
    data_alerta_mopp = hoje + timedelta(days=30) # <-- NOVA LINHA ADICIONADA

    contexto = {
        'motoristas': motoristas_list,
        'search_query': search_query,
        'active_page': 'consultar_motoristas',
        'hoje': hoje,
        'data_alerta_cnh': data_alerta_cnh,
        'data_alerta_mopp': data_alerta_mopp # <-- VARIÁVEL ADICIONADA AO CONTEXTO
    }
    # --- FIM DA CORREÇÃO ---
    
    return render_template('consultar_motoristas.html', **contexto)


@app.route('/api/cliente/excluir_anexo', methods=['POST'])
@login_required
def api_excluir_anexo_cliente():
    """API para excluir um anexo específico de um cliente."""
    data = request.get_json()
    cliente_id = data.get('cliente_id')
    anexo_url = data.get('anexo_url')

    if not cliente_id or not anexo_url:
        return jsonify({'success': False, 'message': 'Dados incompletos.'}), 400

    cliente = Cliente.query.filter_by(id=cliente_id, empresa_id=current_user.empresa_id).first_or_404()

    anexos_atuais = cliente.anexos.split(',') if cliente.anexos else []
    if anexo_url not in anexos_atuais:
        return jsonify({'success': False, 'message': 'Anexo não encontrado ou permissão negada.'}), 404

    try:
        # Lógica para excluir do Cloudflare R2
        s3_client = boto3.client(
            's3',
            endpoint_url=app.config['CLOUDFLARE_R2_ENDPOINT'],
            aws_access_key_id=app.config['CLOUDFLARE_R2_ACCESS_KEY'],
            aws_secret_access_key=app.config['CLOUDFLARE_R2_SECRET_KEY'],
            region_name='auto'
        )
        bucket_name = app.config['CLOUDFLARE_R2_BUCKET']
        public_url_base = app.config['CLOUDFLARE_R2_PUBLIC_URL']
        
        # Extrai a chave do arquivo a partir da URL completa
        key = anexo_url.replace(f"{public_url_base}/", "")
        s3_client.delete_object(Bucket=bucket_name, Key=key)

        # Remove do banco de dados
        anexos_atuais.remove(anexo_url)
        cliente.anexos = ','.join(anexos_atuais) or None
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Anexo excluído com sucesso!'})

    except Exception as e:
        db.session.rollback()
        logger.error(f"Erro ao excluir anexo do cliente {cliente_id}: {e}", exc_info=True)
        return jsonify({'success': False, 'message': f'Erro ao remover anexo: {e}'}), 500

@app.route('/relatorios/rentabilidade/exportar_excel')
@login_required
@master_required
def exportar_rentabilidade_excel():
    """ 
    Exporta o relatório de rentabilidade para Excel.
    VERSÃO CORRIGIDA para usar a MESMA LÓGICA da tela, garantindo consistência.
    """
    try:
        # 1. Obter os mesmos filtros da página do dashboard
        data_inicio_str = request.args.get('data_inicio')
        data_fim_str = request.args.get('data_fim')
        veiculo_id_filtro = request.args.get('veiculo_id', type=int)

        data_inicio = datetime.strptime(data_inicio_str, '%Y-%m-%d').date()
        data_fim = datetime.strptime(data_fim_str, '%Y-%m-%d').date()

        # ==================== LÓGICA DE BUSCA DE DADOS UNIFICADA ====================
        
        # Query de veículos
        veiculos_query = Veiculo.query.filter(Veiculo.empresa_id == current_user.empresa_id)
        if veiculo_id_filtro:
            veiculos_query = veiculos_query.filter(Veiculo.id == veiculo_id_filtro)
        veiculos_no_filtro = veiculos_query.all()
        veiculo_ids = [v.id for v in veiculos_no_filtro]

        # A. Receitas e Custos Diretos
        todas_receitas = []
        todos_custos = []
        
        viagens_periodo = Viagem.query.filter(
            Viagem.veiculo_id.in_(veiculo_ids),
            Viagem.data_inicio.between(data_inicio, data_fim)
        ).options(db.joinedload(Viagem.veiculo)).all()

        for viagem in viagens_periodo:
            todas_receitas.append(viagem)
            for abast in viagem.abastecimentos:
                todos_custos.append({'veiculo': viagem.veiculo, 'data': abast.data_abastecimento.date(), 'categoria': 'Combustível', 'descricao': f"{abast.litros:.2f}L @ R${abast.preco_por_litro:.2f}/L", 'valor': float(abast.custo_total or 0)})
            if viagem.custo_viagem:
                if viagem.custo_viagem.pedagios: todos_custos.append({'veiculo': viagem.veiculo, 'data': viagem.data_inicio.date(), 'categoria': 'Pedágio', 'descricao': 'Custo com pedágios', 'valor': float(viagem.custo_viagem.pedagios)})
                if viagem.custo_viagem.alimentacao: todos_custos.append({'veiculo': viagem.veiculo, 'data': viagem.data_inicio.date(), 'categoria': 'Alimentação', 'descricao': 'Custo com alimentação', 'valor': float(viagem.custo_viagem.alimentacao)})

        # B. Custos de Rateio
        rateios = RateioVeiculo.query.join(LancamentoFluxoCaixa).filter(RateioVeiculo.veiculo_id.in_(veiculo_ids), LancamentoFluxoCaixa.data_pagamento.between(data_inicio, data_fim)).options(db.joinedload(RateioVeiculo.veiculo)).all()
        for r in rateios:
            todos_custos.append({'veiculo': r.veiculo, 'data': r.lancamento.data_pagamento, 'categoria': r.lancamento.categoria, 'descricao': r.lancamento.descricao, 'valor': float(r.valor_rateado)})
        
        # C. Custos de Folha
        motoristas_ids = {v.motorista_id for v in viagens_periodo if v.motorista_id}
        if motoristas_ids:
            folhas_pagas = FolhaPagamento.query.filter(
                FolhaPagamento.motorista_id.in_(motoristas_ids),
                FolhaPagamento.data_pagamento.between(data_inicio, data_fim)
            ).all()
            for f in folhas_pagas:
                for item in f.itens.filter(ItemFolhaPagamento.tipo=='Provento', ItemFolhaPagamento.viagem_id.isnot(None)):
                    viagem_associada = db.session.get(Viagem, item.viagem_id)
                    if viagem_associada and viagem_associada.veiculo_id in veiculo_ids:
                        todos_custos.append({'veiculo': viagem_associada.veiculo, 'data': f.data_pagamento.date(), 'categoria': 'Pagamento Pessoal', 'descricao': f'{item.descricao} - {f.motorista.nome}', 'valor': float(item.valor)})

        # ==================== FIM DA LÓGICA UNIFICADA ====================

        # 3. Organizar dados por veículo
        dados_por_veiculo = defaultdict(lambda: {'receitas': [], 'custos': []})
        for receita in todas_receitas:
            dados_por_veiculo[receita.veiculo.placa]['receitas'].append(receita)
        for custo in todos_custos:
            dados_por_veiculo[custo['veiculo'].placa]['custos'].append(custo)

        # 4. Gerar o arquivo Excel (lógica de formatação mantida)
        output = io.BytesIO()
        workbook = Workbook()
        workbook.remove(workbook.active)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        currency_style = NamedStyle(name='currency_style', number_format='R$ #,##0.00')
        
        resumo_geral = []
        for placa, dados in sorted(dados_por_veiculo.items()):
            sheet = workbook.create_sheet(title=placa)
            sheet.append(["Receitas (Fretes)"])
            sheet.cell(row=sheet.max_row, column=1).font = Font(bold=True, size=14)
            sheet.append(["Data", "Cliente", "Destino", "Material", "Peso (TN)", "Valor (R$)"])
            for cell in sheet[sheet.max_row]: cell.font = header_font; cell.fill = header_fill
            
            total_receitas_veiculo = 0
            for r in dados['receitas']:
                sheet.append([r.data_inicio.date(), r.cliente, r.endereco_destino, r.material_transportado or 'N/A', r.peso_toneladas or 0, r.valor_recebido or 0])
                sheet.cell(row=sheet.max_row, column=6).style = currency_style
                total_receitas_veiculo += r.valor_recebido or 0
            
            sheet.append([])
            sheet.append(["Custos e Pagamentos"])
            sheet.cell(row=sheet.max_row, column=1).font = Font(bold=True, size=14)
            sheet.append(["Data Pag.", "Categoria", "Descrição", "Valor (R$)"])
            for cell in sheet[sheet.max_row]: cell.font = header_font; cell.fill = header_fill

            total_custos_veiculo = 0
            for c in sorted(dados['custos'], key=lambda x: x['data']):
                sheet.append([c['data'], c['categoria'], c['descricao'], c['valor']])
                sheet.cell(row=sheet.max_row, column=4).style = currency_style
                total_custos_veiculo += c['valor']
            
            sheet.append([])
            lucro_veiculo = total_receitas_veiculo - total_custos_veiculo
            sheet.append(["", "", "Total Receitas:", total_receitas_veiculo])
            sheet.append(["", "", "Total Custos:", total_custos_veiculo])
            sheet.append(["", "", "Lucro Líquido:", lucro_veiculo])
            for i in range(3):
                sheet.cell(row=sheet.max_row - i, column=3).font = Font(bold=True)
                sheet.cell(row=sheet.max_row - i, column=4).style = currency_style
            
            resumo_geral.append({'placa': placa, 'receita': total_receitas_veiculo, 'custo': total_custos_veiculo, 'lucro': lucro_veiculo})
            for col_idx, column_cells in enumerate(sheet.columns, 1):
                sheet.column_dimensions[get_column_letter(col_idx)].width = 20

        sheet_resumo = workbook.create_sheet(title="Resumo Geral", index=0)
        sheet_resumo.append(["Veículo", "Total Receitas (R$)", "Total Custos (R$)", "Lucro Líquido (R$)"])
        for cell in sheet_resumo[1]: cell.font = header_font; cell.fill = header_fill
        
        grand_total_receita = 0; grand_total_custo = 0
        for resumo in resumo_geral:
            sheet_resumo.append([resumo['placa'], resumo['receita'], resumo['custo'], resumo['lucro']])
            grand_total_receita += resumo['receita']; grand_total_custo += resumo['custo']
            for col in range(2, 5): sheet_resumo.cell(row=sheet_resumo.max_row, column=col).style = currency_style
        
        sheet_resumo.append([])
        sheet_resumo.append(["TOTAL GERAL", grand_total_receita, grand_total_custo, grand_total_receita - grand_total_custo])
        for cell in sheet_resumo[sheet_resumo.max_row]: cell.font = Font(bold=True, size=12)
        for col in range(2, 5): sheet_resumo.cell(row=sheet_resumo.max_row, column=col).style = currency_style
        for col_idx, column_cells in enumerate(sheet_resumo.columns, 1): sheet_resumo.column_dimensions[get_column_letter(col_idx)].width = 25

        workbook.save(output)
        output.seek(0)
        filename = f"Relatorio_Rentabilidade_{data_inicio_str}_a_{data_fim_str}.xlsx"
        return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    except Exception as e:
        logger.error(f"Erro ao exportar rentabilidade para Excel: {e}", exc_info=True)
        flash("Ocorreu um erro ao gerar o relatório Excel.", "error")
        return redirect(url_for('relatorio_rentabilidade_veiculo', **request.args))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        # Se o usuário já estiver logado, redireciona para o painel correto
        if current_user.role == 'Owner':
            return redirect(url_for('owner_dashboard'))
        elif current_user.role == 'Motorista':
            return redirect(url_for('motorista_dashboard'))
        else:
            # --- CORREÇÃO APLICADA AQUI ---
            # Para qualquer outro usuário logado (Admin, Master), o destino é o 'painel'
            return redirect(url_for('painel'))

    if request.method == 'POST':
        email = request.form.get('email')
        senha = request.form.get('senha')
        
        usuario = Usuario.query.filter_by(email=email).first()
        
        if not usuario or not usuario.check_password(senha):
            flash('Email ou senha incorretos. Por favor, tente novamente.', 'error')
            return redirect(url_for('login'))
            
        login_user(usuario)
        flash('Login realizado com sucesso!', 'success')
        
        # O redirecionamento após o login bem-sucedido já estava correto
        if usuario.role == 'Owner':
            return redirect(url_for('owner_dashboard'))
        elif usuario.role == 'Motorista':
            return redirect(url_for('motorista_dashboard'))
        else:
            return redirect(url_for('painel'))
            
    return render_template('login.html')

@app.route('/registrar', methods=['GET', 'POST'])
def registrar():
    if request.method == 'POST':
        nome = request.form.get('nome')
        sobrenome = request.form.get('sobrenome')
        email = request.form.get('email')
        senha = request.form.get('senha')
        
        if Usuario.query.filter_by(email=email).first():
            flash('Email já cadastrado', 'error')
            return redirect(url_for('registrar'))
            
        novo_usuario = Usuario(
            nome=nome,
            sobrenome=sobrenome,
            email=email
        )
        novo_usuario.set_password(senha)
        
        try:
            db.session.add(novo_usuario)
            db.session.commit()
            flash('Conta criada com sucesso! Faça login.', 'success')
            return redirect(url_for('login'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao criar conta: {str(e)}', 'error')
            
    return render_template('registrar.html')

@app.route('/promover_admin')
def promover_admin():
    user = Usuario.query.filter_by(email='adminadmin@admin.com').first()
    if user:
        user.is_admin = True
        db.session.commit()
        return "Admin atualizado!"
    return "Usuário não encontrado."


@app.template_filter('dateformat')
def dateformat(value):
    if value:
        return value.strftime('%Y-%m-%d')
    return ''

@app.route('/editar_motorista/<int:motorista_id>', methods=['GET', 'POST'])
@login_required
def editar_motorista(motorista_id):
    # Busca o motorista que pertence à empresa do usuário, ou retorna erro 404.
    motorista = Motorista.query.filter_by(id=motorista_id, empresa_id=current_user.empresa_id).first_or_404()

    if request.method == 'POST':
        try:
            # Coleta os dados do formulário
            cpf = re.sub(r'\D', '', request.form.get('cpf', ''))
            cnh_numero = re.sub(r'\D', '', request.form.get('cnh_numero', ''))

            # --- PRINCÍPIO DE VALIDAÇÃO PARA EDIÇÃO APLICADO ---

            # 1. Valida se o CPF informado já pertence a OUTRO motorista, ignorando o motorista atual.
            motorista_com_mesmo_cpf = Motorista.query.filter(
                Motorista.id != motorista_id,  # A CONDIÇÃO MAIS IMPORTANTE: Exclui o próprio registro da busca.
                Motorista.cpf == cpf,
                Motorista.empresa_id == current_user.empresa_id
            ).first()

            if motorista_com_mesmo_cpf:
                flash('Erro: Este CPF já está cadastrado para outro motorista.', 'error')
                return redirect(url_for('editar_motorista', motorista_id=motorista_id))
            
            # 2. Valida se a CNH informada já pertence a OUTRO motorista, aplicando o mesmo princípio.
            motorista_com_mesma_cnh = Motorista.query.filter(
                Motorista.id != motorista_id, # A CONDIÇÃO MAIS IMPORTANTE: Exclui o próprio registro da busca.
                Motorista.cnh_numero == cnh_numero,
                Motorista.empresa_id == current_user.empresa_id
            ).first()

            if motorista_com_mesma_cnh:
                flash('Erro: Esta CNH já está cadastrada para outro motorista.', 'error')
                return redirect(url_for('editar_motorista', motorista_id=motorista_id))
            
            # --- FIM DA APLICAÇÃO DO PRINCÍPIO ---

            def to_date(date_string):
                return datetime.strptime(date_string, '%Y-%m-%d').date() if date_string else None
            
            # Atualiza todos os campos do objeto motorista com os dados do formulário
            motorista.nome = request.form.get('nome')
            motorista.telefone = re.sub(r'\D', '', request.form.get('telefone', ''))
            motorista.cpf = cpf
            motorista.data_nascimento = to_date(request.form.get('data_nascimento'))
            motorista.nacionalidade = request.form.get('nacionalidade')
            motorista.naturalidade = request.form.get('naturalidade')
            motorista.estado_civil = request.form.get('estado_civil')
            motorista.sexo = request.form.get('sexo')
            motorista.pai = request.form.get('pai')
            motorista.mae = request.form.get('mae')
            motorista.data_admissao = to_date(request.form.get('data_admissao'))
            motorista.situacao = request.form.get('situacao')
            motorista.data_desativacao = to_date(request.form.get('data_desativacao'))
            motorista.classificacao = request.form.get('classificacao')
            motorista.cod_departamento = request.form.get('cod_departamento')
            motorista.numero_ficha = request.form.get('numero_ficha')
            motorista.cep = re.sub(r'\D', '', request.form.get('cep', ''))
            motorista.tipo_logradouro = request.form.get('tipo_logradouro')
            motorista.logradouro = request.form.get('logradouro')
            motorista.numero = request.form.get('numero')
            motorista.complemento = request.form.get('complemento')
            motorista.bairro = request.form.get('bairro')
            motorista.cidade = request.form.get('cidade')
            motorista.uf = request.form.get('uf')
            motorista.email = request.form.get('email')
            motorista.tipo_imovel = request.form.get('tipo_imovel')
            motorista.tempo_residencia = request.form.get('tempo_residencia')
            motorista.cnh_numero = cnh_numero
            motorista.cnh_data_primeira = to_date(request.form.get('cnh_data_primeira'))
            motorista.cnh_data_vencimento = to_date(request.form.get('cnh_data_vencimento'))
            motorista.cnh_categoria = request.form.get('cnh_categoria')
            motorista.cnh_cod_seguranca = request.form.get('cnh_cod_seguranca')
            motorista.rg = request.form.get('rg')
            motorista.rg_uf = request.form.get('rg_uf')
            motorista.pis = request.form.get('pis')
            motorista.inss = request.form.get('inss')
            motorista.titulo_eleitor = request.form.get('titulo_eleitor')
            motorista.ctps = request.form.get('ctps')
            motorista.funcao = request.form.get('funcao')

            # ▼▼▼ LINHA ADICIONADA AQUI ▼▼▼
            motorista.salario_base = request.form.get('salario_base', type=float, default=0.0)
            
            motorista.mopp_numero = request.form.get('mopp_numero')
            motorista.mopp_vencimento = to_date(request.form.get('mopp_vencimento'))
            motorista.contato_nome = request.form.get('contato_nome')
            motorista.contato_tipo_ref = request.form.get('contato_tipo_ref')
            motorista.contato_tipo_fone = request.form.get('contato_tipo_fone')
            motorista.contato_telefone = re.sub(r'\D', '', request.form.get('contato_telefone', ''))
            motorista.contato_operadora = request.form.get('contato_operadora')
            motorista.contato_obs = request.form.get('contato_obs')

            # Salva as alterações no banco de dados
            db.session.commit()
            flash('Dados do motorista atualizados com sucesso!', 'success')
            return redirect(url_for('editar_motorista', motorista_id=motorista_id))

        except Exception as e:
            db.session.rollback()
            logger.error(f"Erro ao editar motorista {motorista_id}: {e}", exc_info=True)
            flash(f'Ocorreu um erro ao salvar as alterações: {e}', 'error')
            return redirect(url_for('editar_motorista', motorista_id=motorista_id))

    # Método GET: apenas renderiza a página com os dados do motorista
    return render_template('editar_motorista.html', motorista=motorista)

@app.route('/relatorios/dre')
@login_required
def relatorio_financeiro_dre():
    """
    Rota para gerar um Demonstrativo de Resultado do Exercício (DRE) simplificado.
    Calcula receitas e custos fixos/variáveis dentro de um período.
    """
    try:
        # Pega as datas do filtro, com valores padrão para o mês atual
        hoje = date.today()
        primeiro_dia_mes_str = hoje.replace(day=1).strftime('%Y-%m-%d')
        data_inicio_str = request.args.get('data_inicio', primeiro_dia_mes_str)
        data_fim_str = request.args.get('data_fim', hoje.strftime('%Y-%m-%d'))
        
        data_inicio = datetime.strptime(data_inicio_str, '%Y-%m-%d')
        # Adiciona um dia ao fim para incluir o dia inteiro na consulta
        data_fim = datetime.strptime(data_fim_str, '%Y-%m-%d') + timedelta(days=1)

        empresa_id = current_user.empresa_id

        # 1. RECEITA BRUTA (Viagens concluídas no período)
        receita_bruta = db.session.query(func.sum(Viagem.valor_recebido)).filter(
            Viagem.empresa_id == empresa_id,
            Viagem.status == 'concluida',
            Viagem.data_fim >= data_inicio,
            Viagem.data_fim < data_fim
        ).scalar() or 0.0

        # 2. CUSTOS VARIÁVEIS (Diretamente ligados às viagens)
        custos_variaveis = defaultdict(float)
        viagens_no_periodo = Viagem.query.filter(
            Viagem.empresa_id == empresa_id,
            Viagem.data_inicio >= data_inicio,
            Viagem.data_inicio < data_fim
        ).options(db.joinedload(Viagem.custo_viagem), db.joinedload(Viagem.abastecimentos)).all()

        for viagem in viagens_no_periodo:
            if viagem.custo_viagem:
                custos_variaveis['Pedágios'] += viagem.custo_viagem.pedagios or 0
                custos_variaveis['Alimentação'] += viagem.custo_viagem.alimentacao or 0
                custos_variaveis['Hospedagem'] += viagem.custo_viagem.hospedagem or 0
                custos_variaveis['Outros (Viagem)'] += viagem.custo_viagem.outros or 0
            for abast in viagem.abastecimentos:
                custos_variaveis['Combustível'] += abast.custo_total or 0
        
        custos_variaveis['Total'] = sum(custos_variaveis.values())

        # 3. CUSTOS FIXOS (Despesas gerais dos veículos no período)
        custos_fixos = defaultdict(float)
        despesas_veiculares = DespesaVeiculo.query.filter(
            DespesaVeiculo.empresa_id == empresa_id,
            DespesaVeiculo.data >= data_inicio.date(),
            DespesaVeiculo.data < data_fim.date()
        ).all()
        for despesa in despesas_veiculares:
            custos_fixos[despesa.categoria] += despesa.valor
        
        custos_fixos_total = sum(custos_fixos.values())

        # 4. Montagem do DRE
        margem_contribuicao = receita_bruta - custos_variaveis['Total']
        lucro_operacional = margem_contribuicao - custos_fixos_total
        
        dre = {
            'receita_bruta': receita_bruta,
            'custos_variaveis': dict(custos_variaveis),
            'margem_contribuicao': margem_contribuicao,
            'custos_fixos': {'Total': custos_fixos_total, 'Detalhado': dict(custos_fixos)},
            'lucro_operacional': lucro_operacional
        }

        return render_template('relatorio_dre.html', dre=dre)
        
    except Exception as e:
        logger.error(f"Erro ao gerar relatório DRE: {e}", exc_info=True)
        flash("Ocorreu um erro ao gerar o relatório DRE.", "error")
        return redirect(url_for('relatorios'))
    

@app.cli.command("limpar-db-links")
def limpar_links_anexos_command():
    """
    Limpa TODAS as colunas de anexos e fotos no banco de dados.
    """
    if not click.confirm(
        click.style(
            "ATENÇÃO: Você está prestes a apagar todas as referências (links) a anexos "
            "e fotos no banco de dados. Isso complementa a limpeza dos arquivos. "
            "Deseja continuar?",
            fg='yellow', bold=True
        ),
        default=False
    ):
        click.echo("Operação cancelada.")
        return

    try:
        click.echo("Iniciando limpeza dos links no banco de dados...")
        
        # Mapeamento das tabelas e colunas a serem limpas
        modelos_e_colunas = {
            'Cliente': ['anexos'],
            'Motorista': ['anexos', 'foto'],
            'Veiculo': ['fotos_urls'],
            'CustoViagem': ['anexos'],
            'Abastecimento': ['anexo_comprovante'],
            'DespesaVeiculo': ['anexos'],
            'ManutencaoItem': [] # Manutenção não parece ter anexo direto no item, mas verificamos
        }
        
        total_afetado = 0
        
        # Itera sobre cada modelo e limpa as colunas especificadas
        for nome_modelo, colunas in modelos_e_colunas.items():
            modelo = globals()[nome_modelo] # Acessa a classe do modelo pelo nome
            
            for coluna in colunas:
                # O método update() é muito eficiente para operações em massa
                registros_afetados = db.session.query(modelo).filter(
                    getattr(modelo, coluna).isnot(None)
                ).update({coluna: None}, synchronize_session=False)
                
                if registros_afetados > 0:
                    click.echo(f"  -> Limpando coluna '{coluna}' da tabela '{nome_modelo}': {registros_afetados} registro(s) atualizado(s).")
                    total_afetado += registros_afetados

        db.session.commit()
        
        if total_afetado > 0:
            click.secho(f"\nOperação concluída! {total_afetado} referências de anexos foram removidas do banco de dados.", fg='green', bold=True)
        else:
            click.secho("\nOperação concluída! Nenhum link de anexo foi encontrado no banco de dados para ser limpo.", fg='green')

    except Exception as e:
        db.session.rollback()
        click.secho(f"\nOcorreu um erro: {e}", fg='red', bold=True)
        logger.error(f"Erro ao limpar links do banco de dados: {e}", exc_info=True)
    
@app.route('/adicionar_anexo/motorista/<int:motorista_id>', methods=['POST'])
@login_required
def adicionar_anexo_motorista(motorista_id):
    motorista = Motorista.query.filter_by(id=motorista_id, empresa_id=current_user.empresa_id).first_or_404()
    
    novos_anexos = request.files.getlist("anexos")

    if not novos_anexos or not any(f.filename for f in novos_anexos):
        flash('Nenhum arquivo selecionado para upload.', 'warning')
        return redirect(url_for('editar_motorista', motorista_id=motorista_id))

    try:
        anexos_atuais = motorista.anexos.split(',') if motorista.anexos else []
        
        # Lógica de Upload (reutilizada da sua função original)
        bucket_name = app.config.get('CLOUDFLARE_R2_BUCKET')
        s3_client = boto3.client(
            's3',
            endpoint_url=app.config['CLOUDFLARE_R2_ENDPOINT'],
            aws_access_key_id=app.config['CLOUDFLARE_R2_ACCESS_KEY'],
            aws_secret_access_key=app.config['CLOUDFLARE_R2_SECRET_KEY'],
            region_name='auto'
        )
        public_url_base = app.config['CLOUDFLARE_R2_PUBLIC_URL']
        
        for anexo in novos_anexos:
            if anexo.filename:
                filename = secure_filename(anexo.filename)
                s3_path = f"motoristas/{motorista.cpf}/anexos/{uuid.uuid4()}-{filename}"
                s3_client.upload_fileobj(
                    anexo, bucket_name, s3_path,
                    ExtraArgs={'ContentType': anexo.content_type or 'application/octet-stream', 'ContentDisposition': 'attachment'}
                )
                anexos_atuais.append(f"{public_url_base}/{s3_path}")
        
        motorista.anexos = ','.join(filter(None, anexos_atuais))
        db.session.commit()
        flash('Anexo(s) adicionado(s) com sucesso!', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao fazer upload do anexo: {e}', 'error')
        logger.error(f"Erro no upload de anexo para motorista {motorista_id}: {e}", exc_info=True)

    return redirect(url_for('editar_motorista', motorista_id=motorista_id))

@app.route('/relatorios/veiculos')
@login_required
def relatorio_desempenho_veiculos():
    """
    Rota para gerar um relatório detalhado do desempenho de cada veículo.
    """
    try:
        empresa_id = current_user.empresa_id
        veiculos = Veiculo.query.filter_by(empresa_id=empresa_id).all()
        stats_finais = []

        for veiculo in veiculos:
            viagens = Viagem.query.filter_by(veiculo_id=veiculo.id).all()
            abastecimentos = Abastecimento.query.filter_by(veiculo_id=veiculo.id).all()
            despesas_gerais = DespesaVeiculo.query.filter_by(veiculo_id=veiculo.id).all()
            manutencoes = Manutencao.query.filter_by(veiculo_id=veiculo.id).all()

            km_rodados = sum(v.distancia_percorrida for v in viagens)
            receita_total = sum(v.valor_recebido or 0 for v in viagens)
            
            custo_combustivel = sum(a.custo_total for a in abastecimentos)
            custo_despesas_viagens = sum(
                (v.custo_viagem.pedagios or 0) + (v.custo_viagem.alimentacao or 0) + 
                (v.custo_viagem.hospedagem or 0) + (v.custo_viagem.outros or 0)
                for v in viagens if v.custo_viagem
            )
            custo_despesas_fixas = sum(d.valor for d in despesas_gerais)
            custo_manutencoes = sum(m.custo_total or 0 for m in manutencoes)
            
            custo_total = custo_combustivel + custo_despesas_viagens + custo_despesas_fixas + custo_manutencoes
            lucro_total = receita_total - custo_total
            total_litros = sum(a.litros for a in abastecimentos)

            stats = {
                'info': veiculo,
                'km_rodados': km_rodados,
                'consumo_medio': (km_rodados / total_litros) if total_litros > 0 else 0,
                'receita_km': (receita_total / km_rodados) if km_rodados > 0 else 0,
                'custo_real_km': (custo_total / km_rodados) if km_rodados > 0 else 0,
                'lucro_km': (lucro_total / km_rodados) if km_rodados > 0 else 0,
                'custo_total': custo_total
            }
            stats_finais.append(stats)

        return render_template('relatorio_veiculos.html', veiculos_stats=stats_finais)

    except Exception as e:
        logger.error(f"Erro ao gerar relatório de veículos: {e}", exc_info=True)
        flash("Ocorreu um erro ao gerar o relatório de desempenho dos veículos.", "error")
        return redirect(url_for('relatorios'))

@app.route('/relatorios/contas_a_receber')
@login_required
def relatorio_contas_a_receber():
    """
    Rota para gerar um relatório de contas a receber, categorizado por data de vencimento.
    """
    try:
        empresa_id = current_user.empresa_id
        hoje = date.today()
        
        contas = Cobranca.query.filter(
            Cobranca.empresa_id == empresa_id,
            Cobranca.status.in_(['Pendente', 'Vencida'])
        ).order_by(Cobranca.data_vencimento.asc()).all()

        relatorio = {
            'vencidas': [], 'vence_hoje': [], 'vence_7_dias': [],
            'vence_30_dias': [], 'futuras': []
        }
        total_a_receber = 0.0

        for c in contas:
            total_a_receber += c.valor_total
            delta_dias = (c.data_vencimento - hoje).days
            
            if delta_dias < 0:
                relatorio['vencidas'].append(c)
            elif delta_dias == 0:
                relatorio['vence_hoje'].append(c)
            elif 1 <= delta_dias <= 7:
                relatorio['vence_7_dias'].append(c)
            elif 8 <= delta_dias <= 30:
                relatorio['vence_30_dias'].append(c)
            else:
                relatorio['futuras'].append(c)

        return render_template('relatorio_contas_receber.html', relatorio=relatorio, total_a_receber=total_a_receber)

    except Exception as e:
        logger.error(f"Erro ao gerar relatório de contas a receber: {e}", exc_info=True)
        flash("Ocorreu um erro ao gerar o relatório de contas a receber.", "error")
        return redirect(url_for('relatorios'))

@app.route('/owner/dashboard')
@login_required
@owner_required # Decorador personalizado para garantir que apenas o Owner acesse.
def owner_dashboard():
    # A rota original estava correta, listando todas as empresas para o Owner.
    empresas = Empresa.query.options(
        db.joinedload(Empresa.licenca)
    ).order_by(Empresa.razao_social).all()
    return render_template('owner_dashboard.html', empresas=empresas)

@app.route('/owner/create_client', methods=['POST'])
@login_required
@owner_required
def owner_create_client():
    """
    Rota para o Owner criar uma nova Empresa e enviar um convite para o primeiro Admin.
    """
    try:
        razao_social = request.form.get('razao_social')
        cnpj = re.sub(r'\D', '', request.form.get('cnpj', ''))
        admin_email = request.form.get('admin_email')
        admin_nome = request.form.get('admin_nome')

        if Empresa.query.filter_by(cnpj=cnpj).first():
            flash('Erro: Já existe uma empresa com este CNPJ.', 'error')
            return redirect(url_for('owner_dashboard'))
        
        if Usuario.query.filter_by(email=admin_email).first():
            flash('Erro: Este e-mail de administrador já está em uso.', 'error')
            return redirect(url_for('owner_dashboard'))

        nova_empresa = Empresa(
            razao_social=razao_social,
            cnpj=cnpj,
            endereco="A ser preenchido pelo admin",
            cidade="A ser preenchido",
            estado="XX",
            cep="00000000"
        )
        db.session.add(nova_empresa)
        db.session.commit()

        token = str(uuid.uuid4())
        data_expiracao = datetime.utcnow() + timedelta(days=7)
        convite = Convite(
            email=admin_email,
            token=token,
            criado_por=current_user.id,
            data_expiracao=data_expiracao,
            role='Admin',
            empresa_id=nova_empresa.id
        )
        db.session.add(convite)
        db.session.commit()

        link_convite = url_for('registrar_com_token', token=token, _external=True)
        msg = Message(
            subject=f'Bem-vindo ao TrackGo, {admin_nome}!',
            recipients=[admin_email],
            body=f'''Olá {admin_nome},\n\nSua empresa, {razao_social}, foi cadastrada em nossa plataforma TrackGo!\n\nPara começar a gerenciar sua equipe e operações, por favor, clique no link abaixo para criar sua senha e finalizar seu cadastro como Administrador:\n\n{link_convite}\n\nEste link é válido por 7 dias.\n\nAtenciosamente,\nEquipe TrackGo'''
        )
        mail.send(msg)

        flash(f'Empresa "{razao_social}" criada e convite enviado para {admin_email} com sucesso!', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'Ocorreu um erro inesperado: {str(e)}', 'error')
        logger.error(f"Erro ao criar cliente pelo owner: {e}", exc_info=True)
    
    return redirect(url_for('owner_dashboard'))

@app.route('/owner/empresa/<int:empresa_id>', methods=['GET', 'POST'])
@login_required
@owner_required
def owner_empresa_detalhes(empresa_id):
    empresa = Empresa.query.get_or_404(empresa_id)
    # Garante que a empresa tenha uma licença; cria uma se não tiver
    if not empresa.licenca:
        licenca = Licenca(empresa_id=empresa.id)
        db.session.add(licenca)
        db.session.commit()
        # Recarrega a empresa para obter a licença recém-criada
        empresa = Empresa.query.get_or_404(empresa_id)

    if request.method == 'POST':
        try:
            licenca = empresa.licenca
            licenca.plano = request.form.get('plano')
            licenca.max_usuarios = int(request.form.get('max_usuarios'))
            licenca.max_veiculos = int(request.form.get('max_veiculos'))
            data_expiracao_str = request.form.get('data_expiracao')
            licenca.data_expiracao = datetime.strptime(data_expiracao_str, '%Y-%m-%d').date() if data_expiracao_str else None
            licenca.ativo = 'ativo' in request.form

            db.session.commit()
            flash('Licença da empresa atualizada com sucesso!', 'success')
            return redirect(url_for('owner_dashboard'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar a licença: {e}', 'error')

    return render_template('owner_empresa_detalhes.html', empresa=empresa)


@app.route('/excluir_anexo/motorista/<int:motorista_id>', methods=['POST'])
@login_required
def excluir_anexo_motorista(motorista_id):
    motorista = Motorista.query.filter_by(id=motorista_id, empresa_id=current_user.empresa_id).first_or_404()
    anexo_url_para_excluir = request.form.get('anexo_url')
    
    if not anexo_url_para_excluir:
        flash('Nenhum anexo especificado para exclusão.', 'error')
        return redirect(url_for('editar_motorista', motorista_id=motorista_id))

    anexos_atuais = motorista.anexos.split(',') if motorista.anexos else []
    
    if anexo_url_para_excluir in anexos_atuais:
        try:
            # Lógica para excluir do Cloudflare R2
            bucket_name = app.config.get('CLOUDFLARE_R2_BUCKET')
            if not bucket_name:
                flash('Erro de configuração: O nome do bucket do Cloudflare R2 não foi definido.', 'error')
                return redirect(url_for('editar_motorista', motorista_id=motorista_id))

            s3_client = boto3.client(
                's3',
                endpoint_url=app.config['CLOUDFLARE_R2_ENDPOINT'],
                aws_access_key_id=app.config['CLOUDFLARE_R2_ACCESS_KEY'],
                aws_secret_access_key=app.config['CLOUDFLARE_R2_SECRET_KEY'],
                region_name='auto'
            )
            key = anexo_url_para_excluir.replace(app.config['CLOUDFLARE_R2_PUBLIC_URL'] + '/', '')
            s3_client.delete_object(Bucket=bucket_name, Key=key)
            
            anexos_atuais.remove(anexo_url_para_excluir)
            motorista.anexos = ','.join(filter(None, anexos_atuais)) or None
            db.session.commit()
            flash('Anexo removido com sucesso!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao remover anexo do Cloudflare: {e}', 'error')
    else:
        flash('Anexo não encontrado ou permissão negada.', 'error')
        
    return redirect(url_for('editar_motorista', motorista_id=motorista_id))

# E substitua sua função 'dateformat' por esta versão mais robusta:
@app.template_filter('dateformat')
def dateformat(value):
    if isinstance(value, (date, datetime)):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, str):
        # Tenta converter a string para data antes de formatar
        try:
            return datetime.strptime(value, '%Y-%m-%d').strftime('%Y-%m-%d')
        except ValueError:
            return value
    return ''
@app.route('/excluir_motorista/<int:motorista_id>')
@login_required
def excluir_motorista(motorista_id):
    motorista = Motorista.query.filter_by(id=motorista_id, empresa_id=current_user.empresa_id).first_or_404()
    if Viagem.query.filter_by(motorista_id=motorista_id).first():
        flash('Erro: Motorista possui viagens associadas.', 'error')
        return redirect(url_for('consultar_motoristas'))
    try:
        if motorista.anexos:
            s3_client = boto3.client(
                's3',
                endpoint_url=app.config['CLOUDFLARE_R2_ENDPOINT'],
                aws_access_key_id=app.config['CLOUDFLARE_R2_ACCESS_KEY'],
                aws_secret_access_key=app.config['CLOUDFLARE_R2_SECRET_KEY'],
                region_name='auto'
            )
            bucket_name = app.config['CLOUDFLARE_R2_BUCKET']
            for anexo in motorista.anexos.split(','):
                filename = anexo.replace(app.config['CLOUDFLARE_R2_PUBLIC_URL'] + '/', '')
                try:
                    s3_client.delete_object(Bucket=bucket_name, Key=filename)
                except Exception as e:
                    logger.error(f"Erro ao excluir anexo {filename}: {str(e)}")
        db.session.delete(motorista)
        db.session.commit()
        flash('Motorista excluído com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir motorista: {str(e)}', 'error')
    return redirect(url_for('consultar_motoristas'))

@app.route('/cadastrar_veiculo', methods=['GET', 'POST'])
@login_required
def cadastrar_veiculo():
    if request.method == 'POST':
        try:
            placa = request.form.get('placa', '').strip().upper()
            if not validate_placa(placa):
                flash('Placa inválida. Use o formato ABC1D23 ou ABC1234.', 'error')
                return render_template('cadastrar_veiculo.html', form_data=request.form)

            if Veiculo.query.filter_by(placa=placa, empresa_id=current_user.empresa_id).first():
                flash(f'Erro: Um veículo com a placa {placa} já foi cadastrado.', 'error')
                return render_template('cadastrar_veiculo.html', form_data=request.form)

            fotos_urls = []
            files = request.files.getlist('fotos[]')
            if files and any(f and f.filename for f in files):
                s3_client = boto3.client(
                    's3',
                    endpoint_url=app.config['CLOUDFLARE_R2_ENDPOINT'],
                    aws_access_key_id=app.config['CLOUDFLARE_R2_ACCESS_KEY'],
                    aws_secret_access_key=app.config['CLOUDFLARE_R2_SECRET_KEY'],
                    region_name='auto'
                )
                bucket_name = app.config['CLOUDFLARE_R2_BUCKET']
                public_url_base = app.config['CLOUDFLARE_R2_PUBLIC_URL']

                for file in files:
                    if file and file.filename:
                        filename = secure_filename(file.filename)
                        s3_path = f"veiculos/{placa}/fotos/{uuid.uuid4()}-{filename}"
                        
                        s3_client.upload_fileobj(
                            file, bucket_name, s3_path,
                            ExtraArgs={'ContentType': file.content_type or 'application/octet-stream'}
                        )
                        fotos_urls.append(f"{public_url_base}/{s3_path}")
            
            def to_date(date_string):
                return datetime.strptime(date_string, '%Y-%m-%d').date() if date_string else None
            
            def to_float(num_string):
                return float(num_string) if num_string else None

            def to_int(num_string):
                return int(num_string) if num_string else None

            novo_veiculo = Veiculo(
                empresa_id=current_user.empresa_id,
                
                placa=placa,
                categoria=request.form.get('categoria'),
                status=request.form.get('status'),
                modelo=request.form.get('modelo'),
                marca=request.form.get('marca'),
                ano_fabricacao=to_int(request.form.get('ano_fabricacao')),
                ano_modelo=to_int(request.form.get('ano_modelo')),
                cor=request.form.get('cor').strip() or None,
                combustivel=request.form.get('combustivel') or None,
                
                # --- CORREÇÃO APLICADA AQUI ---
                renavam=request.form.get('renavam').strip() or None,
                chassi=request.form.get('chassi').strip() or None,
                numero_motor=request.form.get('numero_motor').strip() or None,
                crlv_numero=request.form.get('crlv_numero').strip() or None,
                # ------------------------------

                crlv_vencimento=to_date(request.form.get('crlv_vencimento')),
                seguro_numero=request.form.get('seguro_numero').strip() or None,
                seguro_seguradora=request.form.get('seguro_seguradora').strip() or None,
                seguro_vencimento=to_date(request.form.get('seguro_vencimento')),
                
                capacidade_carga_kg=to_float(request.form.get('capacidade_carga')),
                peso_bruto_total_kg=to_float(request.form.get('peso_bruto')),
                eixos=to_int(request.form.get('eixos')),
                cilindrada=request.form.get('cilindrada').strip() or None,
                potencia_cv=to_int(request.form.get('potencia')),
                tanque_combustivel_litros=to_int(request.form.get('tanque_combustivel')),
                consumo_medio_km_l=to_float(request.form.get('consumo_medio')),
                
                valor_aquisicao=to_float(request.form.get('valor_aquisicao')),
                data_aquisicao=to_date(request.form.get('data_aquisicao')),
                km_atual=to_float(request.form.get('km_atual')),
                ultima_manutencao=to_date(request.form.get('ultima_manutencao')),
                km_ultima_manutencao=to_float(request.form.get('km_ultima_manutencao')),
                proxima_manutencao=to_date(request.form.get('proxima_manutencao')),
                motorista_padrao_id=to_int(request.form.get('motorista_padrao')),
                
                observacoes=request.form.get('observacoes').strip() or None,
                fotos_urls=','.join(fotos_urls) if fotos_urls else None
            )

            db.session.add(novo_veiculo)
            db.session.commit()
            flash('Veículo cadastrado com sucesso!', 'success')
            return redirect(url_for('consultar_veiculos'))

        except (ValueError, TypeError) as e:
            db.session.rollback()
            logger.error(f"Erro de tipo/valor ao cadastrar veículo: {e}", exc_info=True)
            flash('Erro de valor inválido. Verifique se os números e datas estão corretos.', 'error')
        except Exception as e:
            db.session.rollback()
            logger.error(f"Erro ao cadastrar veículo: {e}", exc_info=True)
            flash(f'Ocorreu um erro inesperado ao cadastrar o veículo: {e}', 'error')
        
        return render_template('cadastrar_veiculo.html', form_data=request.form, active_page='cadastrar_veiculo')

    return render_template('cadastrar_veiculo.html', active_page='cadastrar_veiculo')

@app.route('/consultar_veiculos', methods=['GET'])
@login_required
def consultar_veiculos():
    search_query = request.args.get('search', '').strip()
    query = Veiculo.query.filter_by(empresa_id=current_user.empresa_id, is_administrativo=False)
    
    if search_query:
        search_filter = f"%{search_query}%"
        query = query.filter(
            or_(
                Veiculo.placa.ilike(search_filter),
                Veiculo.modelo.ilike(search_filter),
                Veiculo.categoria.ilike(search_filter)
            )
        )
    

    veiculos_obj = query.order_by(Veiculo.placa.asc()).all()
    
    veiculos_json = [v.to_dict() for v in veiculos_obj]

    # Passo 3: Envia a lista de dicionários (veiculos_json) para o template
    return render_template('consultar_veiculos.html', veiculos=veiculos_json, search_query=search_query, active_page='consultar_veiculos')

@app.route('/editar_veiculo/<int:veiculo_id>', methods=['GET', 'POST'])
@login_required
def editar_veiculo(veiculo_id):
    veiculo = Veiculo.query.filter_by(id=veiculo_id, empresa_id=current_user.empresa_id).first_or_404()
    motoristas = Motorista.query.filter_by(empresa_id=current_user.empresa_id).order_by(Motorista.nome).all()

    if request.method == 'POST':
        try:
            # Lógica para upload de novas fotos
            novas_fotos_urls = []
            files = request.files.getlist('fotos[]') # O nome do input é 'fotos[]'
            if files and any(f and f.filename for f in files):
                s3_client = boto3.client(
                    's3',
                    endpoint_url=app.config['CLOUDFLARE_R2_ENDPOINT'],
                    aws_access_key_id=app.config['CLOUDFLARE_R2_ACCESS_KEY'],
                    aws_secret_access_key=app.config['CLOUDFLARE_R2_SECRET_KEY'],
                    region_name='auto'
                )
                bucket_name = app.config['CLOUDFLARE_R2_BUCKET']
                public_url_base = app.config['CLOUDFLARE_R2_PUBLIC_URL']

                for file in files:
                    if file and file.filename:
                        filename = secure_filename(file.filename)
                        # Usando a placa do veículo para organizar os arquivos
                        s3_path = f"veiculos/{veiculo.placa}/fotos/{uuid.uuid4()}-{filename}"
                        
                        s3_client.upload_fileobj(
                            file, bucket_name, s3_path,
                            ExtraArgs={'ContentType': file.content_type or 'application/octet-stream'}
                        )
                        novas_fotos_urls.append(f"{public_url_base}/{s3_path}")
            
            # Combina as fotos existentes com as novas
            fotos_atuais = veiculo.fotos_urls.split(',') if veiculo.fotos_urls else []
            todas_as_fotos = fotos_atuais + novas_fotos_urls
            if todas_as_fotos:
                veiculo.fotos_urls = ','.join(todas_as_fotos)

            # Funções auxiliares para conversão segura de tipos
            def to_int(val): return int(val) if val and val.strip() else None
            def to_float(val): return float(val) if val and val.strip() else None
            def to_date(val): return datetime.strptime(val, '%Y-%m-%d').date() if val and val.strip() else None

            # Atualização dos outros campos do formulário
            veiculo.categoria = request.form.get('categoria')
            veiculo.status = request.form.get('status')
            veiculo.modelo = request.form.get('modelo')
            veiculo.marca = request.form.get('marca')
            veiculo.ano_fabricacao = to_int(request.form.get('ano_fabricacao'))
            veiculo.ano_modelo = to_int(request.form.get('ano_modelo'))
            veiculo.cor = request.form.get('cor').strip() or None
            veiculo.combustivel = request.form.get('combustivel') or None
            veiculo.renavam = request.form.get('renavam').strip() or None
            veiculo.chassi = request.form.get('chassi').strip() or None
            veiculo.crlv_numero = request.form.get('crlv_numero').strip() or None
            veiculo.crlv_vencimento = to_date(request.form.get('crlv_vencimento'))
            veiculo.km_atual = to_float(request.form.get('km_atual'))
            veiculo.motorista_padrao_id = to_int(request.form.get('motorista_padrao_id'))
            veiculo.observacoes = request.form.get('observacoes').strip() or None

            db.session.commit()
            
            flash('Veículo atualizado com sucesso!', 'success')
            return redirect(url_for('consultar_veiculos'))

        except Exception as e:
            db.session.rollback()
            logger.error(f"Erro ao editar o veículo {veiculo_id}: {e}", exc_info=True)
            flash(f'Ocorreu um erro inesperado ao salvar: {str(e)}', 'error')
            return redirect(url_for('editar_veiculo', veiculo_id=veiculo_id))

    return render_template('editar_veiculo.html', 
                           veiculo=veiculo, 
                           motoristas=motoristas, 
                           active_page='consultar_veiculos')
                           
@app.route('/excluir_veiculo/<int:veiculo_id>')
@login_required
def excluir_veiculo(veiculo_id):
    veiculo = Veiculo.query.filter_by(id=veiculo_id, empresa_id=current_user.empresa_id).first_or_404()
    if Viagem.query.filter_by(veiculo_id=veiculo_id).first():
        flash('Erro: Veículo possui viagens associadas.', 'error')
        return redirect(url_for('consultar_veiculos'))
    try:
        db.session.delete(veiculo)
        db.session.commit()
        flash('Veículo excluído com sucesso!', 'success')
    except:
        db.session.rollback()
        flash('Erro ao excluir veículo.', 'error')
    return redirect(url_for('consultar_veiculos'))

@app.route('/api/viagem/processar_nfe', methods=['POST'])
@login_required
def processar_nfe_api():
    if 'nfe_files' not in request.files:
        return jsonify({'success': False, 'message': 'Nenhum arquivo enviado.'}), 400
    
    files = request.files.getlist('nfe_files')
    viagens_processadas = []
    clientes_criados = 0

    for file in files:
        if file and file.filename.endswith('.xml'):
            parsed_data = parse_nfe_xml(file)
            if not parsed_data:
                continue

            cliente_info = parsed_data['cliente_info']
            viagem_info = parsed_data['viagem_info']
            
            try:
                if cliente_info and cliente_info['cpf_cnpj']:
                    cliente_existente = Cliente.query.filter_by(
                        cpf_cnpj=cliente_info['cpf_cnpj'],
                        empresa_id=current_user.empresa_id
                    ).first()

                    if not cliente_existente:
                        novo_cliente = Cliente(
                            empresa_id=current_user.empresa_id,
                            cadastrado_por_id=current_user.id,
                            **cliente_info
                        )
                        db.session.add(novo_cliente)
                        clientes_criados += 1
                
                viagens_processadas.append(viagem_info)

            except Exception as e:
                db.session.rollback()
                logger.error(f"Erro ao criar cliente a partir do XML {viagem_info['nome_arquivo']}: {e}", exc_info=True)
                continue

    if clientes_criados > 0:
        db.session.commit()

    if not viagens_processadas:
        return jsonify({'success': False, 'message': 'Nenhum dado de viagem pôde ser extraído dos arquivos XML fornecidos.'}), 400

    mensagem_sucesso = f"{len(viagens_processadas)} NF-e(s) processada(s). "
    if clientes_criados > 0:
        mensagem_sucesso += f"E {clientes_criados} novo(s) cliente(s) foram cadastrados automaticamente."

    return jsonify({'success': True, 'viagens': viagens_processadas, 'message': mensagem_sucesso})

def buscar_lancamento_consolidado_por_id(item_id, empresa_id):
    """Busca e formata dados de um lançamento específico"""
    try:
        if item_id.startswith('manual_'):
            lancamento_id = int(item_id.replace('manual_', ''))
            lancamento = LancamentoFluxoCaixa.query.filter_by(
                id=lancamento_id, 
                empresa_id=empresa_id
            ).first()
            
            if not lancamento:
                return None
            
            # Verificar valores de parcela para evitar None
            parcela_numero = getattr(lancamento, 'parcela_numero', 1) or 1
            parcela_total = getattr(lancamento, 'parcela_total', 1) or 1
            
            # Verificar campos de data disponíveis
            data_criacao_field = getattr(lancamento, 'data_criacao', None) or \
                               getattr(lancamento, 'created_at', None) or \
                               getattr(lancamento, 'data_cadastro', None) or \
                               getattr(lancamento, 'data_lancamento', None)
            
            data_atualizacao_field = getattr(lancamento, 'data_atualizacao', None) or \
                                   getattr(lancamento, 'updated_at', None) or \
                                   getattr(lancamento, 'data_modificacao', None)
            
            # Formatar dados do lançamento manual
            dados_lancamento = {
                'id': f"manual_{lancamento.id}",
                'tipo_origem': 'MANUAL',
                'tipo_movimento': lancamento.tipo,
                'descricao': lancamento.descricao,
                'categoria': getattr(lancamento, 'categoria', None),
                'valor': float(lancamento.valor_total),
                'valor_formatado': f"{lancamento.valor_total:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
                'data_emissao': data_criacao_field.strftime('%d/%m/%Y') if data_criacao_field else None,
                'data_vencimento': lancamento.data_vencimento.strftime('%d/%m/%Y'),
                'data_pagamento': lancamento.data_pagamento.strftime('%d/%m/%Y') if lancamento.data_pagamento else None,
                'data_criacao': data_criacao_field.strftime('%d/%m/%Y %H:%M') if data_criacao_field else None,
                'data_atualizacao': data_atualizacao_field.strftime('%d/%m/%Y %H:%M') if data_atualizacao_field else None,
                'fornecedor_cliente': getattr(lancamento, 'fornecedor_cliente', None),
                'documento': getattr(lancamento, 'documento_numero', None),
                'status': getattr(lancamento, 'status_pagamento', 'PENDENTE'),
                'meio_pagamento': getattr(lancamento, 'meio_pagamento', None),
                'parcela': f"{parcela_numero}/{parcela_total}" if parcela_total > 1 else None,
                'observacoes': getattr(lancamento, 'observacoes', None),
                'tem_rateio': getattr(lancamento, 'tem_rateio', False)
            }
            
        elif item_id.startswith('nfe_'):
            lancamento_id = int(item_id.replace('nfe_', ''))
            lancamento = LancamentoNotaFiscal.query.filter_by(
                id=lancamento_id, 
                empresa_id=empresa_id
            ).first()
            
            if not lancamento:
                return None
            
            # Extrair número da NFe da chave de acesso (posições 25-33)
            numero_nfe = lancamento.chave_acesso[25:34] if len(lancamento.chave_acesso) >= 34 else lancamento.chave_acesso[-9:]
            
            # Verificar valores de parcela para evitar None
            parcela_numero = getattr(lancamento, 'parcela_numero', 1) or 1
            parcela_total = getattr(lancamento, 'parcela_total', 1) or 1
            
            # Formatar dados do lançamento NFe
            dados_lancamento = {
                'id': f"nfe_{lancamento.id}",
                'tipo_origem': 'NFE',
                'tipo_movimento': 'DESPESA',
                'descricao': f"NFe {numero_nfe} - {lancamento.emitente_nome}",
                'categoria': 'Fornecedores (NFe)',
                'valor': float(lancamento.valor_total),
                'valor_formatado': f"{lancamento.valor_total:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
                'data_emissao': lancamento.data_emissao.strftime('%d/%m/%Y') if lancamento.data_emissao else None,
                'data_vencimento': lancamento.data_vencimento.strftime('%d/%m/%Y'),
                'data_pagamento': lancamento.data_pagamento.strftime('%d/%m/%Y') if lancamento.data_pagamento else None,
                'data_criacao': lancamento.data_lancamento.strftime('%d/%m/%Y %H:%M') if lancamento.data_lancamento else None,
                'data_atualizacao': None,  # Campo não existe no modelo
                'fornecedor_cliente': lancamento.emitente_nome,
                'documento': lancamento.chave_acesso,
                'status': lancamento.status_pagamento,
                'meio_pagamento': None,  # Campo não existe no modelo
                'parcela': f"{parcela_numero}/{parcela_total}" if parcela_total > 1 else None,
                'observacoes': getattr(lancamento, 'observacoes', None),
                'tem_rateio': getattr(lancamento, 'tem_rateio', False)
            }
        else:
            return None
            
        return dados_lancamento
        
    except Exception as e:
        logger.error(f"Erro ao buscar lançamento consolidado: {e}", exc_info=True)
        return None
    
@app.route('/api/fluxo_caixa/detalhes/<string:item_id>')
@login_required
def api_detalhes_lancamento(item_id):
    """ API para buscar os detalhes de um lançamento específico do fluxo de caixa (versão corrigida). """
    try:
        lancamento_id = int(item_id.replace('manual_', ''))
        lancamento = db.session.get(LancamentoFluxoCaixa, lancamento_id)
        
        if not lancamento or lancamento.empresa_id != current_user.empresa_id:
            return jsonify({'success': False, 'message': 'Lançamento não encontrado.'}), 404

        nfe_original = None
        is_nfe = lancamento.documento_numero and lancamento.documento_numero.startswith('NFE-')
        if is_nfe:
            chave_acesso_base = lancamento.documento_numero.split('-')[1]
            nfe_original = NFeImportada.query.filter_by(chave_acesso=chave_acesso_base, empresa_id=current_user.empresa_id).first()

        dados_lancamento = {
            'id': f"manual_{lancamento.id}",
            'tipo_origem': 'NFE' if is_nfe else 'MANUAL',
            'tipo_movimento': lancamento.tipo,
            'descricao': lancamento.descricao,
            'categoria': lancamento.categoria,
            'valor': float(lancamento.valor_total),
            'valor_formatado': f"{lancamento.valor_total:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
            'data_emissao': nfe_original.data_emissao.strftime('%d/%m/%Y') if nfe_original else lancamento.data_lancamento.strftime('%d/%m/%Y'),
            'data_vencimento': lancamento.data_vencimento.strftime('%d/%m/%Y'),
            'data_pagamento': lancamento.data_pagamento.strftime('%d/%m/%Y') if lancamento.data_pagamento else '-',
            'fornecedor_cliente': lancamento.fornecedor_cliente,
            'documento': nfe_original.chave_acesso if nfe_original else lancamento.documento_numero,
            'status': lancamento.status_pagamento,
            'meio_pagamento': lancamento.meio_pagamento,
            'parcela': f"{lancamento.parcela_numero}/{lancamento.parcela_total}" if lancamento.parcela_total and lancamento.parcela_total > 1 else None,
            'observacoes': lancamento.observacoes,
            'tem_rateio': lancamento.tem_rateio
        }
        
        return jsonify({'success': True, 'dados': dados_lancamento})

    except Exception as e:
        logger.error(f"Erro ao buscar detalhes do lançamento {item_id}: {e}", exc_info=True)
        return jsonify({'success': False, 'message': f'Erro interno ao buscar detalhes: {str(e)}'}), 500

@app.route('/viagem/importar_nfe')
@login_required
def importar_nfe_page():
    """Renderiza a nova página de importação."""
    motoristas = Motorista.query.filter_by(empresa_id=current_user.empresa_id, situacao='NORMAL / LIBERADO').order_by(Motorista.nome).all()
    veiculos = Veiculo.query.filter_by(status='Disponível', empresa_id=current_user.empresa_id, is_administrativo=False).order_by(Veiculo.placa).all()
    return render_template('importar_nfe.html', motoristas=motoristas, veiculos=veiculos)

@app.route('/iniciar_viagem', methods=['GET'])
@login_required
def iniciar_viagem_page():
    """Apenas renderiza a página do formulário de iniciar viagem."""
    
    motoristas = Motorista.query.filter_by(empresa_id=current_user.empresa_id).order_by(Motorista.nome).all()
    veiculos = Veiculo.query.filter_by(status='Disponível', empresa_id=current_user.empresa_id, is_administrativo=False).order_by(Veiculo.placa).all()
    # A lógica de POST foi movida para uma rota de API separada.
    return render_template('iniciar_viagem.html', motoristas=motoristas, veiculos=veiculos, ORS_API_KEY=OPENROUTESERVICE_API_KEY)

@app.route('/api/motorista/<int:motorista_id>/details')
@login_required
def get_motorista_details_api(motorista_id):
    """
    Esta rota de API retorna as estatísticas e o histórico de viagens
    de um motorista em formato JSON para ser consumido pelo modal.
    """
    # Garante que o usuário só possa ver motoristas da sua própria empresa
    motorista = Motorista.query.filter_by(id=motorista_id, empresa_id=current_user.empresa_id).first_or_404()

    # Busca as viagens do motorista, carregando os dados relacionados para o cálculo de custos
    viagens = Viagem.query.filter(Viagem.motorista_id == motorista.id).options(
        db.joinedload(Viagem.custo_viagem),
        db.joinedload(Viagem.abastecimentos)
    ).order_by(Viagem.data_inicio.desc()).all()

    # Lógica de cálculo de estatísticas (reaproveitando e melhorando a da página de perfil)
    total_receita = 0
    total_custo_detalhado = 0
    
    for v in viagens:
        total_receita += v.valor_recebido or 0
        
        # Calcula o custo detalhado da viagem
        custo_despesas = 0
        if v.custo_viagem:
            custo_despesas = (v.custo_viagem.pedagios or 0) + (v.custo_viagem.alimentacao or 0) + (v.custo_viagem.hospedagem or 0) + (v.custo_viagem.outros or 0)
        
        custo_abastecimento = sum(a.custo_total for a in v.abastecimentos)
        total_custo_detalhado += custo_despesas + custo_abastecimento

    stats = {
        'total_viagens': len(viagens),
        'total_distancia': round(sum(v.distancia_km or 0 for v in viagens), 2),
        'total_receita': round(total_receita, 2),
        'total_custo': round(total_custo_detalhado, 2),
        'lucro_total': round(total_receita - total_custo_detalhado, 2)
    }

    # Formata os dados das viagens para o JSON
    viagens_data = []
    for v in viagens:
        viagens_data.append({
            'id': v.id,
            'cliente': v.cliente,
            'data_inicio': v.data_inicio.isoformat(),
            'endereco_saida': v.endereco_saida,
            'endereco_destino': v.endereco_destino,
            'status': v.status
        })

    # Retorna o JSON completo que o JavaScript espera
    return jsonify({
        'success': True,
        'stats': stats,
        'viagens': viagens_data
    })



@app.route('/api/viagem/criar', methods=['POST'])
@login_required
def criar_viagem_api():
    try:
        data = request.get_json()
        
        motorista_id = data.get('motorista_id')
        veiculo_id = data.get('veiculo_id')
        cliente = data.get('cliente')
        endereco_saida = data.get('endereco_saida')
        enderecos_destino = data.get('enderecos_destino', [])
        data_inicio_str = data.get('data_inicio')
        valor_recebido = float(data.get('valor_recebido') or 0) # Captura o valor recebido
        
        if not all([motorista_id, veiculo_id, cliente, endereco_saida, enderecos_destino, data_inicio_str]):
            return jsonify({'success': False, 'message': 'Todos os campos são obrigatórios.'}), 400

        motorista = db.session.get(Motorista, int(motorista_id))
        veiculo = db.session.get(Veiculo, int(veiculo_id))
        if not motorista or not veiculo:
            return jsonify({'success': False, 'message': 'Motorista ou Veículo não encontrado.'}), 404
        if veiculo.status != 'Disponível':
            return jsonify({'success': False, 'message': f'Veículo {veiculo.placa} não está disponível (Status: {veiculo.status}).'}), 409
            
        todos_enderecos = [endereco_saida] + enderecos_destino
        
        rota_otimizada, distancia_km, duracao_segundos, geometria, erro = calcular_rota_otimizada_ors(todos_enderecos)

        if erro:
            return jsonify({'success': False, 'message': erro}), 400

        # --- CÁLCULO DA ESTIMATIVA DE CUSTO (Lógica reutilizada da API de estimativa) ---
        consumo_real = calcular_consumo_medio_real(veiculo.id)
        consumo_a_ser_usado = consumo_real or veiculo.consumo_medio_km_l or 1.0
        preco_combustivel_para_calculo = obter_preco_medio_combustivel_recente(current_user.empresa_id)
        litros_estimados = distancia_km / consumo_a_ser_usado
        custo_combustivel = litros_estimados * preco_combustivel_para_calculo

        salario_base = motorista.salario_base or 0.0
        custo_hora_motorista = salario_base / 220
        duracao_horas = duracao_segundos / 3600
        custo_motorista = duracao_horas * custo_hora_motorista

        custo_fixo_km = calcular_custo_fixo_por_km(veiculo.id)
        custo_manutencao_km = calcular_custo_manutencao_por_km(veiculo.id)
        custo_desgaste_veiculo = distancia_km * (custo_fixo_km + custo_manutencao_km)
        
        custo_total_estimado = custo_combustivel + custo_motorista + custo_desgaste_veiculo
        lucro_estimado = valor_recebido - custo_total_estimado
        # --- FIM DO CÁLCULO ---

        nova_viagem = Viagem(
            motorista_id=motorista_id,
            motorista_cpf_cnpj=motorista.cpf_cnpj,
            veiculo_id=veiculo_id,
            cliente=cliente,
            valor_recebido=valor_recebido,
            forma_pagamento=data.get('forma_pagamento'),
            endereco_saida=endereco_saida,
            endereco_destino=rota_otimizada[-1],
            distancia_km=distancia_km,
            data_inicio=datetime.strptime(data_inicio_str, '%Y-%m-%dT%H:%M'),
            duracao_segundos=duracao_segundos,
            status='pendente',
            observacoes=data.get('observacoes'),
            route_geometry=geometria,
            empresa_id=current_user.empresa_id
        )
        veiculo.status = 'Em Rota'
        db.session.add(nova_viagem)
        db.session.flush()

        for ordem, endereco in enumerate(rota_otimizada[1:], 1):
            destino = Destino(viagem_id=nova_viagem.id, endereco=endereco, ordem=ordem)
            db.session.add(destino)

        db.session.commit()

        # Adiciona os novos dados ao JSON de retorno
        return jsonify({
            'success': True,
            'message': 'Viagem criada com sucesso!',
            'viagem_id': nova_viagem.id,
            'roteiro': rota_otimizada,
            'distancia': f"{distancia_km:.2f}",
            'duracao_minutos': duracao_segundos // 60,
           
            'custo_estimado': f"{custo_total_estimado:.2f}",
            'lucro_estimado': f"{lucro_estimado:.2f}"
        })

    except Exception as e:
        db.session.rollback()
        logger.error(f"Erro na API ao criar viagem: {e}", exc_info=True)
        return jsonify({'success': False, 'message': f'Ocorreu um erro interno: {e}'}), 500

@app.route('/excluir_viagem/<int:viagem_id>')
@login_required 
def excluir_viagem(viagem_id):
    viagem = Viagem.query.filter_by(id=viagem_id, empresa_id=current_user.empresa_id).first_or_404()

    try:
        if not viagem.data_fim and viagem.veiculo:
            # LINHA A SER CORRIGIDA:
            viagem.veiculo.status = 'Disponível' #
        db.session.delete(viagem)
        db.session.commit()
        flash('Viagem excluída com sucesso!', 'success')

    except Exception as e:
        db.session.rollback()
        logger.error(f"Erro ao excluir viagem: {str(e)}")
        flash(f'Erro ao excluir viagem: {str(e)}', 'error')

    return redirect(url_for('consultar_viagens'))



@app.route('/fiscal/importar', methods=['GET', 'POST'])
@login_required
@master_required
def importar_notas_fiscais():
    """Página e ação de importação de notas fiscais"""
    if request.method == 'GET':
        from sefaz_service import get_status_consulta_sefaz
        status = get_status_consulta_sefaz(current_user.empresa_id)
        
        notas_importadas = NFeImportada.query.filter_by(
            empresa_id=current_user.empresa_id
        ).order_by(NFeImportada.data_emissao.desc()).all()

        # Linha com a indentação corrigida
        unidades_negocio = UnidadeNegocio.query.filter_by(empresa_id=current_user.empresa_id).order_by(UnidadeNegocio.nome).all()
        
        return render_template('importar_notas_fiscais.html', 
                             notas=notas_importadas,
                             status_certificados=status,
                             unidades_negocio=unidades_negocio)
    
    elif request.method == 'POST':
        try:
            from sefaz_service import consultar_notas_sefaz
            
            # Verificar se pode consultar
            from sefaz_service import get_status_consulta_sefaz
            status = get_status_consulta_sefaz(current_user.empresa_id)
            
            if not status['pode_consultar']:
                flash(status['motivo_bloqueio'], 'error')
                return redirect(url_for('importar_notas_fiscais'))
            
            # Realizar consulta
            resultado = consultar_notas_sefaz(current_user.empresa_id)
            
            if resultado['success']:
                if resultado.get('notas_processadas', 0) > 0:
                    flash(resultado['message'], 'success')
                else:
                    flash(resultado['message'], 'info')
            else:
                flash(resultado['message'], 'error')
            
            return redirect(url_for('importar_notas_fiscais'))
            
        except Exception as e:
            app.logger.error(f"Erro na importação: {str(e)}")
            flash('Erro interno durante a consulta à SEFAZ.', 'error')
            return redirect(url_for('importar_notas_fiscais'))
        
@app.route('/financeiro/fluxo_caixa/editar/<int:lancamento_id>', methods=['GET', 'POST'])
@login_required
@master_required
def editar_lancamento_fluxo(lancamento_id):
    """Editar lançamento manual no fluxo de caixa"""
    lancamento = LancamentoFluxoCaixa.query.filter_by(
        id=lancamento_id, 
        empresa_id=current_user.empresa_id
    ).first_or_404()
    
    if request.method == 'POST':
        try:
            # Verificar se pode ser editado
            if lancamento.status_pagamento == 'PAGO':
                flash('Não é possível editar lançamentos já pagos.', 'error')
                return redirect(url_for('fluxo_caixa'))
            
            # Atualizar dados
            lancamento.tipo = request.form.get('tipo')
            lancamento.descricao = request.form.get('descricao')
            lancamento.categoria = request.form.get('categoria')
            lancamento.valor_total = float(request.form.get('valor_total', 0))
            lancamento.data_vencimento = datetime.strptime(request.form.get('data_vencimento'), '%Y-%m-%d').date()
            lancamento.fornecedor_cliente = request.form.get('fornecedor_cliente')
            lancamento.documento_numero = request.form.get('documento_numero')
            lancamento.observacoes = request.form.get('observacoes')
            
            db.session.commit()
            flash('Lançamento atualizado com sucesso!', 'success')
            return redirect(url_for('fluxo_caixa'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar lançamento: {e}', 'error')
            logger.error(f"Erro ao editar lançamento {lancamento_id}: {e}", exc_info=True)
    
    return render_template('editar_lancamento_fluxo.html', lancamento=lancamento)

# Em app.py, substitua a função /api/fluxo_caixa/excluir por esta versão corrigida

@app.route('/api/fluxo_caixa/excluir', methods=['POST'])
@login_required
@master_required
def api_excluir_lancamento_fluxo():
    """API para excluir lançamento, garantindo que a NFe original e os rateios sejam tratados corretamente."""
    data = request.get_json()
    item_id = data.get('item_id')
    
    if not item_id or not item_id.startswith('manual_'):
        return jsonify({'success': False, 'message': 'ID do item inválido ou não é um lançamento manual.'}), 400
    
    try:
        lancamento_id = int(item_id.replace('manual_', ''))
        lancamento_clicado = LancamentoFluxoCaixa.query.filter_by(
            id=lancamento_id, 
            empresa_id=current_user.empresa_id
        ).first_or_404()

        if lancamento_clicado.status_pagamento == 'PAGO':
            return jsonify({'success': False, 'message': 'Não é possível excluir lançamentos já pagos.'}), 400

        # --- INÍCIO DA LÓGICA CORRIGIDA ---
        
        # 1. Identifica se o lançamento veio de uma NF-e e encontra todas as suas parcelas
        is_nfe_related = lancamento_clicado.documento_numero and lancamento_clicado.documento_numero.startswith('NFE-')
        
        lancamentos_para_excluir = []
        if is_nfe_related:
            lancamentos_para_excluir = LancamentoFluxoCaixa.query.filter_by(
                documento_numero=lancamento_clicado.documento_numero,
                empresa_id=current_user.empresa_id
            ).all()
        else:
            lancamentos_para_excluir.append(lancamento_clicado)

        # 2. Coleta os IDs de todos os lançamentos que serão excluídos
        ids_para_excluir = [lanc.id for lanc in lancamentos_para_excluir]

        if ids_para_excluir:
            # 3. (A CORREÇÃO) Exclui PRIMEIRO os registros 'filhos' (rateios)
            RateioVeiculo.query.filter(
                RateioVeiculo.lancamento_id.in_(ids_para_excluir)
            ).delete(synchronize_session=False)

            # 4. Agora exclui os registros 'pai' (os próprios lançamentos)
            LancamentoFluxoCaixa.query.filter(
                LancamentoFluxoCaixa.id.in_(ids_para_excluir)
            ).delete(synchronize_session=False)

        # 5. Se era de uma NF-e, reabilita a nota original
        chave_acesso = None
        if is_nfe_related and lancamento_clicado.observacoes and 'Chave de acesso: ' in lancamento_clicado.observacoes:
            chave_acesso = lancamento_clicado.observacoes.split('Chave de acesso: ')[1]
            nfe_original = NFeImportada.query.filter_by(
                chave_acesso=chave_acesso,
                empresa_id=current_user.empresa_id
            ).first()
            if nfe_original:
                nfe_original.status = 'IMPORTADA'
        # --- FIM DA LÓGICA CORRIGIDA ---

        db.session.commit()
        
        if is_nfe_related:
            return jsonify({'success': True, 'message': 'Lançamento(s) e seu rateio foram excluídos! A nota fiscal foi reabilitada.'})
        else:
            return jsonify({'success': True, 'message': 'Lançamento manual e seu rateio foram excluídos com sucesso!'})

    except Exception as e:
        db.session.rollback()
        logger.error(f"Erro ao excluir lançamento {item_id}: {e}", exc_info=True)
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/fluxo_caixa/salvar_rateio', methods=['POST'])
@login_required
@master_required
def api_salvar_rateio():
    data = request.get_json()
    item_id_str = data.get('item_id', '')
    
    # Validações iniciais
    if not item_id_str:
        return jsonify({'success': False, 'message': 'ID do item é obrigatório.'}), 400
        
    lancamento_id_clicado = int(item_id_str.replace('manual_', '').replace('nfe_', ''))
    veiculos_data = data.get('veiculos', [])
    
    lancamento_clicado = db.session.get(LancamentoFluxoCaixa, lancamento_id_clicado)
    if not lancamento_clicado or lancamento_clicado.empresa_id != current_user.empresa_id:
        return jsonify({'success': False, 'message': 'Lançamento não encontrado.'}), 404

    # --- INÍCIO DA CORREÇÃO ---
    # Agora, a busca por parcelas funciona tanto para NFE quanto para lançamentos MANUAIS.
    lancamentos_do_grupo = []
    if lancamento_clicado.documento_numero and (lancamento_clicado.documento_numero.startswith('NFE-') or lancamento_clicado.documento_numero.startswith('MANUAL-')):
        # Busca todas as parcelas que compartilham o mesmo número de documento (grupo).
        lancamentos_do_grupo = LancamentoFluxoCaixa.query.filter_by(
            documento_numero=lancamento_clicado.documento_numero,
            empresa_id=current_user.empresa_id
        ).all()
    else:
        # Se não for um lançamento agrupado (parcela única), usa apenas ele.
        lancamentos_do_grupo.append(lancamento_clicado)
    # --- FIM DA CORREÇÃO ---
        
    try:
        # Itera sobre cada lançamento (parcela) do grupo
        for lancamento in lancamentos_do_grupo:
            # Impede a edição de rateio se qualquer parcela do grupo já estiver paga
            if lancamento.status_pagamento == 'PAGO':
                db.session.rollback()
                return jsonify({'success': False, 'message': 'Não é possível alterar o rateio de um lançamento com parcelas já pagas.'}), 400

            # Exclui o rateio antigo para esta parcela específica para recriá-lo
            RateioVeiculo.query.filter_by(lancamento_id=lancamento.id).delete()

            if veiculos_data: # Se há novos dados de rateio, cria-os
                for veiculo in veiculos_data:
                    percentual = float(veiculo['percentual'])
                    # O valor rateado é calculado com base no valor INDIVIDUAL da parcela
                    valor_rateado_parcela = lancamento.valor_total * (percentual / 100)
                    
                    novo_rateio = RateioVeiculo(
                        lancamento_id=lancamento.id,
                        veiculo_id=int(veiculo['veiculo_id']),
                        valor_rateado=valor_rateado_parcela,
                        percentual=percentual
                    )
                    db.session.add(novo_rateio)
                lancamento.tem_rateio = True
            else: # Se a lista de veículos está vazia, significa que o rateio foi removido
                lancamento.tem_rateio = False

        db.session.commit()
        return jsonify({'success': True, 'message': 'Rateio salvo com sucesso para todas as parcelas!'})

    except Exception as e:
        db.session.rollback()
        logger.error(f"Erro ao salvar rateio para o grupo do lançamento {lancamento_id_clicado}: {e}", exc_info=True)
        return jsonify({'success': False, 'message': f'Erro interno: {str(e)}'}), 500
    
@app.route('/api/fluxo_caixa/novo_lancamento_completo', methods=['POST'])
@login_required
@master_required
def api_novo_lancamento_completo():
    """
    API final para criar lançamentos, com suporte a rateio para múltiplos veículos.
    --- VERSÃO CORRIGIDA ---
    """
    data = request.get_json()
    try:
        valor_total_nota = float(data.get('valor_total', 0))
        tem_rateio = data.get('tem_rateio', False)
        rateios_data = data.get('veiculos', []) # Dados dos veículos para rateio
        parcelas_data = data.get('parcelas', [])
        
        # ▼▼▼ TRECHO ADICIONADO ▼▼▼
        data_emissao_str = data.get('data_emissao')
        data_emissao = datetime.strptime(data_emissao_str, '%Y-%m-%d').date() if data_emissao_str else date.today()
        # ▲▲▲ FIM DO TRECHO ▲▲▲

        if not parcelas_data:
            # Se não houver parcelas, cria uma parcela única com o valor total
            parcelas_data = [{
                'data_vencimento': data.get('data_vencimento'), # Supondo que venha no corpo se não for parcelado
                'valor': valor_total_nota
            }]

        if tem_rateio:
            soma_rateios = sum(float(r.get('valor', 0)) for r in rateios_data)
            if abs(soma_rateios - valor_total_nota) > 0.01:
                return jsonify({'success': False, 'message': 'A soma dos valores do rateio não confere com o valor total do lançamento.'}), 400

        # Cria um ID único para agrupar todas as parcelas deste lançamento
        grupo_id = f"MANUAL-{uuid.uuid4().hex[:8]}"

        # 1. Loop principal pelas PARCELAS
        for i, parcela_info in enumerate(parcelas_data, 1):
            valor_parcela = float(parcela_info['valor'])
            data_vencimento_parcela = datetime.strptime(parcela_info['data_vencimento'], '%Y-%m-%d').date()

            # 2. Cria UM lançamento para CADA parcela
            novo_lancamento = LancamentoFluxoCaixa(
                empresa_id=current_user.empresa_id,
                unidade_negocio_id=data.get('unidade_negocio_id'),
                tipo=data['tipo'],
                descricao=f"{data['descricao']} - Parcela {i}/{len(parcelas_data)}" if len(parcelas_data) > 1 else data['descricao'],
                categoria=data.get('categoria'),
                valor_total=valor_parcela,
                
                # ▼▼▼ LINHA ALTERADA ▼▼▼
                data_lancamento=data_emissao,
                
                data_vencimento=data_vencimento_parcela,
                fornecedor_cliente=data.get('fornecedor_cliente'),
                documento_numero=grupo_id,
                observacoes=data.get('observacoes'),
                parcela_numero=i,
                parcela_total=len(parcelas_data),
                tem_rateio=tem_rateio
            )
            db.session.add(novo_lancamento)
            db.session.flush() # ESSENCIAL: Gera o ID para o novo_lancamento

            # 3. Se tiver rateio, cria os registros na tabela RateioVeiculo
            if tem_rateio:
                for rateio_info in rateios_data:
                    percentual_rateio = (float(rateio_info['valor']) / valor_total_nota) * 100
                    valor_rateado_para_parcela = valor_parcela * (percentual_rateio / 100)
                    
                    # 4. Cria o registro de rateio VINCULADO ao lançamento da parcela
                    novo_rateio_db = RateioVeiculo(
                        lancamento_id=novo_lancamento.id, # <--- Ponto chave da correção
                        veiculo_id=int(rateio_info['veiculo_id']),
                        valor_rateado=valor_rateado_para_parcela,
                        percentual=percentual_rateio
                    )
                    db.session.add(novo_rateio_db)

        db.session.commit()
        return jsonify({'success': True, 'message': f'Lançamento criado com sucesso em {len(parcelas_data)} parcela(s)!'})

    except Exception as e:
        db.session.rollback()
        logger.error(f"Erro ao criar lançamento com rateio: {e}", exc_info=True)
        return jsonify({'success': False, 'message': f'Erro interno: {str(e)}'}), 500

@app.route('/api/fluxo_caixa/rateio/<string:item_id>')
@login_required
def api_buscar_rateio(item_id):
    """
    Busca os dados de rateio de um lançamento e de todas as suas parcelas irmãs.
    Esta é a API chamada quando o usuário clica no botão 'Rateio' no fluxo de caixa.
    """
    try:
        lancamento_id_real = int(item_id.replace('manual_', ''))
        lancamento_clicado = db.session.get(LancamentoFluxoCaixa, lancamento_id_real)

        if not lancamento_clicado or lancamento_clicado.empresa_id != current_user.empresa_id:
            return jsonify({'success': False, 'message': 'Lançamento não encontrado.'}), 404

        if lancamento_clicado.tipo != 'DESPESA':
            return jsonify({'success': False, 'message': 'Rateio só é aplicável a despesas.'}), 400

        # Encontra todos os lançamentos do mesmo grupo
        lancamentos_do_grupo = []
        if lancamento_clicado.documento_numero and (lancamento_clicado.documento_numero.startswith('NFE-') or lancamento_clicado.documento_numero.startswith('MANUAL-')):
            lancamentos_do_grupo = LancamentoFluxoCaixa.query.filter_by(
                documento_numero=lancamento_clicado.documento_numero,
                empresa_id=current_user.empresa_id
            ).all()
        else:
            lancamentos_do_grupo.append(lancamento_clicado)

        valor_total_grupo = sum(l.valor_total for l in lancamentos_do_grupo)

        primeiro_lancamento_com_rateio = next((l for l in lancamentos_do_grupo if l.tem_rateio), None)

        veiculos_data = []
        if primeiro_lancamento_com_rateio:
            rateios_db = RateioVeiculo.query.filter_by(lancamento_id=primeiro_lancamento_com_rateio.id).all()
            for r in rateios_db:
                percentual = float(r.percentual)
                valor_total_rateado = valor_total_grupo * (percentual / 100)
                
                veiculos_data.append({
                    'veiculo_id': r.veiculo_id,
                    'veiculo_nome': f"{r.veiculo.placa} - {r.veiculo.modelo}",
                    'valor': valor_total_rateado,
                    'percentual': percentual
                })
        
        dados_formatados = {
            'id': f"manual_{lancamento_clicado.id}",
            'descricao': lancamento_clicado.descricao.split(' - Parcela')[0],
            'valor_total': valor_total_grupo,
            'categoria': lancamento_clicado.categoria,
            'rateio_veiculos': veiculos_data
        }
        
        return jsonify({'success': True, 'dados': dados_formatados})

    except Exception as e:
        logger.error(f"Erro ao buscar dados de rateio para {item_id}: {e}", exc_info=True)
        return jsonify({'success': False, 'message': f'Erro interno: {str(e)}'}), 500


@app.route('/api/fluxo_caixa/novo_lancamento_rateio', methods=['POST'])
@login_required  
@master_required
def api_novo_lancamento_com_rateio():
    """API para criar novo lançamento (com ou sem rateio)"""
    try:
        data = request.get_json()
        
        # Dados básicos
        tipo_movimento = data.get('tipo_movimento', 'DESPESA')
        descricao = data.get('descricao')
        categoria = data.get('categoria')
        valor = float(data.get('valor', 0))
        data_vencimento = datetime.strptime(data.get('data_vencimento'), '%Y-%m-%d').date()
        
        # Dados opcionais
        data_emissao = data.get('data_emissao')
        if data_emissao:
            data_emissao = datetime.strptime(data_emissao, '%Y-%m-%d').date()
            
        fornecedor_cliente = data.get('fornecedor_cliente', '')
        documento = data.get('documento', '')
        observacoes = data.get('observacoes', '')
        
        # Status baseado em "marcar como pago"
        pago_agora = data.get('marcar_pago_agora', False)
        status = 'PAGO' if pago_agora else 'PENDENTE'
        
        # Dados de pagamento
        forma_pagamento = data.get('forma_pagamento') if pago_agora else None
        data_pagamento = None
        if pago_agora and data.get('data_pagamento'):
            data_pagamento = datetime.strptime(data.get('data_pagamento'), '%Y-%m-%d').date()
            
        # Verificar parcelas
        parcelado = data.get('parcelado', False)
        parcelas_data = data.get('parcelas', [])
        
        if parcelado and parcelas_data:
            # Criar múltiplas parcelas
            for i, parcela in enumerate(parcelas_data, 1):
                novo_lancamento = LancamentoFluxoCaixa(
                    empresa_id=current_user.empresa_id,
                    tipo=tipo_movimento,  # ou tipo_movimento dependendo do seu modelo
                    descricao=f"{descricao} - Parcela {i}/{len(parcelas_data)}",
                    categoria=categoria,
                    valor_total=float(parcela.get('valor', 0)),
                    data_emissao=data_emissao,
                    data_vencimento=datetime.strptime(parcela.get('data_vencimento'), '%Y-%m-%d').date(),
                    data_pagamento=data_pagamento,
                    fornecedor_cliente=fornecedor_cliente,
                    documento_numero=documento,
                    status=status,
                    meio_pagamento=forma_pagamento,
                    tipo_origem='MANUAL',
                    observacoes=observacoes,
                    parcela=f"{i}/{len(parcelas_data)}"
                )
                db.session.add(novo_lancamento)
        else:
            # Criar lançamento único
            novo_lancamento = LancamentoFluxoCaixa(
                empresa_id=current_user.empresa_id,
                tipo=tipo_movimento,
                descricao=descricao,
                categoria=categoria,
                valor_total=valor,
                data_emissao=data_emissao,
                data_vencimento=data_vencimento,
                data_pagamento=data_pagamento,
                fornecedor_cliente=fornecedor_cliente,
                documento_numero=documento,
                status=status,
                meio_pagamento=forma_pagamento,
                tipo_origem='MANUAL',
                observacoes=observacoes
            )
            db.session.add(novo_lancamento)
        
        db.session.flush()
        
        # Processar rateio se aplicável
        tem_rateio = data.get('tem_rateio', False)
        veiculos_rateio = data.get('veiculos', [])
        
        if tem_rateio and veiculos_rateio:
            # Aplicar rateio em todos os lançamentos criados
            lancamentos = [novo_lancamento] if not parcelado else db.session.query(LancamentoFluxoCaixa).filter(
                LancamentoFluxoCaixa.empresa_id == current_user.empresa_id,
                LancamentoFluxoCaixa.descricao.like(f"{descricao} - Parcela%")
            ).all()
            
            for lancamento in lancamentos:
                for veiculo in veiculos_rateio:
                    valor_rateado = lancamento.valor_total * (float(veiculo['percentual']) / 100)
                    
                    rateio = RateioVeiculo(
                        lancamento_id=lancamento.id,
                        veiculo_id=int(veiculo['veiculo_id']),
                        valor_rateado=valor_rateado,
                        percentual=float(veiculo['percentual'])
                    )
                    db.session.add(rateio)
                    
                lancamento.tem_rateio = True
        
        db.session.commit()
        return jsonify({'success': True, 'message': 'Lançamento criado com sucesso!'})
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Erro ao criar lançamento: {e}", exc_info=True)
        return jsonify({'success': False, 'message': f'Erro: {str(e)}'}), 500

@app.route('/api/veiculos/search')
@login_required
def api_buscar_veiculos():
    """API para buscar veículos por termo de pesquisa"""
    try:
        term = request.args.get('term', '').strip()
        if len(term) < 2:
            return jsonify([])
        
        veiculos = Veiculo.query.filter(
            Veiculo.empresa_id == current_user.empresa_id,
            # A linha "Veiculo.ativo == True," que causava o erro foi removida.
            # Também removemos o filtro is_administrativo daqui para que o caminhão fantasma apareça na busca.
            db.or_(
                Veiculo.placa.ilike(f'%{term}%'),
                Veiculo.modelo.ilike(f'%{term}%'),
                Veiculo.marca.ilike(f'%{term}%')
            )
        ).limit(10).all()
        
        resultado = []
        for veiculo in veiculos:
            resultado.append({
                'id': veiculo.id,
                'display': f"{veiculo.placa} - {veiculo.marca} {veiculo.modelo}"
            })
        
        return jsonify(resultado)
        
    except Exception as e:
        logger.error(f"Erro ao buscar veículos: {e}", exc_info=True)
        return jsonify([])

@app.route('/api/fluxo_caixa/exportar_relatorio_rateio')
@login_required
def api_exportar_relatorio_rateio():
    """API para exportar relatório de rateio em Excel"""
    try:
        # Aplicar filtros da tela atual
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        categoria_filtro = request.args.get('categoria')
        
        # Query base
        query = db.session.query(LancamentoFluxoCaixa, RateioVeiculo, Veiculo).join(
            RateioVeiculo, LancamentoFluxoCaixa.id == RateioVeiculo.lancamento_id
        ).join(
            Veiculo, RateioVeiculo.veiculo_id == Veiculo.id
        ).filter(
            LancamentoFluxoCaixa.empresa_id == current_user.empresa_id,
            LancamentoFluxoCaixa.tem_rateio == True
        )
        
        # Aplicar filtros
        if data_inicio:
            query = query.filter(LancamentoFluxoCaixa.data_vencimento >= datetime.strptime(data_inicio, '%Y-%m-%d').date())
        if data_fim:
            query = query.filter(LancamentoFluxoCaixa.data_vencimento <= datetime.strptime(data_fim, '%Y-%m-%d').date())
        if categoria_filtro:
            query = query.filter(LancamentoFluxoCaixa.categoria == categoria_filtro)
        
        dados = query.all()
        
        # Criar Excel
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet('Relatório de Rateio')
        
        # Estilos
        header_format = workbook.add_format({
            'bold': True, 'bg_color': '#2e7d32', 'font_color': 'white'
        })
        money_format = workbook.add_format({'num_format': 'R$ #,##0.00'})
        percent_format = workbook.add_format({'num_format': '0.0%'})
        
        # Cabeçalhos
        headers = [
            'Data Vencimento', 'Descrição', 'Categoria', 'Valor Total',
            'Veículo', 'Placa', 'Valor Rateado', 'Percentual', 'Status'
        ]
        
        for col, header in enumerate(headers):
            worksheet.write(0, col, header, header_format)
        
        # Dados
        for row, (lancamento, rateio, veiculo) in enumerate(dados, 1):
            worksheet.write(row, 0, lancamento.data_vencimento.strftime('%d/%m/%Y'))
            worksheet.write(row, 1, lancamento.descricao)
            worksheet.write(row, 2, lancamento.categoria)
            worksheet.write(row, 3, float(lancamento.valor_total), money_format)
            worksheet.write(row, 4, f"{veiculo.marca} {veiculo.modelo}")
            worksheet.write(row, 5, veiculo.placa)
            worksheet.write(row, 6, float(rateio.valor_rateado), money_format)
            worksheet.write(row, 7, float(rateio.percentual)/100, percent_format)
            worksheet.write(row, 8, lancamento.status_pagamento)
        
        workbook.close()
        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name=f'relatorio_rateio_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Erro ao exportar relatório de rateio: {e}", exc_info=True)
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/fluxo_caixa/buscar_dados/<string:item_id>')
@login_required
@master_required
def api_buscar_dados_lancamento(item_id):
    """Buscar dados de um lançamento para edição"""
    try:
        if item_id.startswith('manual_'):
            lancamento_id = int(item_id.replace('manual_', ''))
            lancamento = LancamentoFluxoCaixa.query.filter_by(
                id=lancamento_id, 
                empresa_id=current_user.empresa_id
            ).first_or_404()
            
            dados = {
                'id': lancamento.id,
                'tipo': lancamento.tipo,
                'descricao': lancamento.descricao,
                'categoria': lancamento.categoria,
                'valor_total': lancamento.valor_total,
                'data_vencimento': lancamento.data_vencimento.strftime('%Y-%m-%d'),
                'fornecedor_cliente': lancamento.fornecedor_cliente,
                'documento_numero': lancamento.documento_numero,
                'observacoes': lancamento.observacoes,
                'status': lancamento.status_pagamento,
                'pode_editar': lancamento.status_pagamento != 'PAGO',
                'tem_rateio': getattr(lancamento, 'tem_rateio', False)
            }
            
        elif item_id.startswith('nfe_'):
            # NFe não pode ser editada, apenas seus dados de pagamento
            return jsonify({'success': False, 'message': 'Lançamentos de NFe não podem ser editados. Use a tela de importação fiscal para reprocessar.'}), 400
            
        else:
            return jsonify({'success': False, 'message': 'Formato de ID inválido'}), 400
        
        return jsonify({'success': True, 'dados': dados})
        
    except Exception as e:
        logger.error(f"Erro ao buscar dados do lançamento: {e}", exc_info=True)
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/fluxo_caixa/estornar_pagamento', methods=['POST'])
@login_required
@master_required
def api_estornar_pagamento():
    data = request.get_json()
    item_id_str = data.get('item_id')
    
    if not item_id_str:
        return jsonify({'success': False, 'message': 'ID do item é obrigatório'}), 400
        
    # Lida com ambos os tipos de lançamento
    if item_id_str.startswith('manual_'):
        lancamento_id = int(item_id_str.replace('manual_', ''))
        lancamento = db.session.get(LancamentoFluxoCaixa, lancamento_id)
        if lancamento and lancamento.empresa_id == current_user.empresa_id:
            lancamento.status_pagamento = 'PENDENTE'
            lancamento.data_pagamento = None
            lancamento.meio_pagamento = None
    elif item_id_str.startswith('nfe_'):
        lancamento_id = int(item_id_str.replace('nfe_', ''))
        lancamento = db.session.get(LancamentoNotaFiscal, lancamento_id)
        if lancamento and lancamento.empresa_id == current_user.empresa_id:
            lancamento.status_pagamento = 'A Pagar'
            lancamento.data_pagamento = None
    else:
        return jsonify({'success': False, 'message': 'Tipo de lançamento inválido.'}), 400
        
    db.session.commit()
    return jsonify({'success': True, 'message': 'Pagamento estornado com sucesso!'})
    
@app.route('/api/fluxo_caixa/salvar_edicao', methods=['POST'])
@login_required
@master_required
def api_salvar_edicao_lancamento():
    """API para salvar edições de um lançamento manual a partir do modal (recebe JSON)."""
    data = request.get_json()
    lancamento_id = data.get('lancamento_id')
    
    if not lancamento_id:
        return jsonify({'success': False, 'message': 'ID do lançamento é obrigatório.'}), 400
    
    lancamento = db.session.get(LancamentoFluxoCaixa, lancamento_id)
    
    if not lancamento or lancamento.empresa_id != current_user.empresa_id:
        return jsonify({'success': False, 'message': 'Lançamento não encontrado.'}), 404
        
    if lancamento.status_pagamento == 'PAGO':
        return jsonify({'success': False, 'message': 'Não é possível editar lançamentos pagos.'}), 400

    try:
        # Se o lançamento já tem um rateio, impede a mudança para "Receita".
        if lancamento.tem_rateio and data['tipo'] == 'RECEITA':
            return jsonify({
                'success': False, 
                'message': 'Não é possível alterar o tipo para "Receita" em um lançamento que já possui um rateio de despesa.'
            }), 400

        lancamento.tipo = data['tipo']
        lancamento.descricao = data['descricao']
        lancamento.categoria = data.get('categoria')
        lancamento.valor_total = float(data['valor_total'])
        lancamento.data_vencimento = datetime.strptime(data['data_vencimento'], '%Y-%m-%d').date()
        lancamento.fornecedor_cliente = data.get('fornecedor_cliente')
        lancamento.documento_numero = data.get('documento_numero')
        lancamento.observacoes = data.get('observacoes')
        
        db.session.commit()
        return jsonify({'success': True, 'message': 'Lançamento atualizado com sucesso!'})
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Erro ao salvar edição via API: {e}", exc_info=True)
        return jsonify({'success': False, 'message': f'Erro interno: {str(e)}'}), 500

@app.route('/api/fluxo_caixa/duplicar', methods=['POST'])
@login_required
@master_required
def api_duplicar_lancamento():
    data = request.get_json()
    item_id = int(data.get('item_id').replace('manual_', ''))
    original = db.session.get(LancamentoFluxoCaixa, item_id)
    
    if not original or original.empresa_id != current_user.empresa_id:
        return jsonify({'success': False, 'message': 'Lançamento original não encontrado.'}), 404
        
    novo_lancamento = LancamentoFluxoCaixa(
        empresa_id=original.empresa_id,
        tipo=original.tipo,
        descricao=f"{original.descricao} (Cópia)",
        categoria=original.categoria,
        valor_total=original.valor_total,
        data_vencimento=datetime.strptime(data['nova_data_vencimento'], '%Y-%m-%d').date(),
        fornecedor_cliente=original.fornecedor_cliente,
        documento_numero=original.documento_numero,
        observacoes=f"Cópia do lançamento ID {original.id}",
        parcela_numero=1,
        parcela_total=1
    )
    db.session.add(novo_lancamento)
    db.session.commit()
    return jsonify({'success': True, 'message': 'Lançamento duplicado com sucesso!'})

@app.route('/api/fluxo_caixa/alterar_categoria_massa', methods=['POST'])
@login_required
@master_required
def api_alterar_categoria_em_massa():
    data = request.get_json()
    item_ids = [int(i.replace('manual_', '')) for i in data.get('item_ids', [])]
    nova_categoria = data.get('nova_categoria')
    
    atualizados = LancamentoFluxoCaixa.query.filter(
        LancamentoFluxoCaixa.id.in_(item_ids),
        LancamentoFluxoCaixa.empresa_id == current_user.empresa_id
    ).update({'categoria': nova_categoria}, synchronize_session=False)
    
    db.session.commit()
    return jsonify({'success': True, 'message': f'{atualizados} item(s) tiveram a categoria alterada.'})


@app.route('/api/fluxo_caixa/dashboard_resumo')
@login_required 
@master_required
def api_dashboard_resumo():
    """API para dados resumidos do dashboard financeiro"""
    try:
        hoje = date.today()
        inicio_mes = hoje.replace(day=1)
        fim_mes = (inicio_mes + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        
        # Totais do mês atual
        query_mes = LancamentoFluxoCaixa.query.filter(
            LancamentoFluxoCaixa.empresa_id == current_user.empresa_id,
            LancamentoFluxoCaixa.data_vencimento >= inicio_mes,
            LancamentoFluxoCaixa.data_vencimento <= fim_mes
        )
        
        receitas_mes = query_mes.filter(LancamentoFluxoCaixa.tipo == 'RECEITA').all()
        despesas_mes = query_mes.filter(LancamentoFluxoCaixa.tipo == 'DESPESA').all()
        
        total_receitas_mes = sum(l.valor_total for l in receitas_mes)
        total_despesas_mes = sum(l.valor_total for l in despesas_mes)
        
        # Pendências importantes
        vencidos_hoje = query_mes.filter(
            LancamentoFluxoCaixa.data_vencimento == hoje,
            LancamentoFluxoCaixa.status_pagamento.in_(['PENDENTE', 'A Pagar'])
        ).count()
        
        vencidos_total = query_mes.filter(
            LancamentoFluxoCaixa.data_vencimento < hoje,
            LancamentoFluxoCaixa.status_pagamento.in_(['PENDENTE', 'A Pagar'])
        ).count()
        
        # Próximos 7 dias
        proximos_7_dias = query_mes.filter(
            LancamentoFluxoCaixa.data_vencimento >= hoje,
            LancamentoFluxoCaixa.data_vencimento <= hoje + timedelta(days=7),
            LancamentoFluxoCaixa.status_pagamento.in_(['PENDENTE', 'A Pagar'])
        ).count()
        
        return jsonify({
            'success': True,
            'resumo': {
                'receitas_mes': total_receitas_mes,
                'despesas_mes': total_despesas_mes,
                'saldo_mes': total_receitas_mes - total_despesas_mes,
                'vencidos_hoje': vencidos_hoje,
                'vencidos_total': vencidos_total,
                'proximos_7_dias': proximos_7_dias
            }
        })
        
    except Exception as e:
        logger.error(f"Erro ao gerar resumo do dashboard: {e}", exc_info=True)
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/financeiro/fluxo_caixa/relatorio_analitico')
@login_required
@master_required
def relatorio_analitico_fluxo():
    """Relatório analítico detalhado do fluxo de caixa"""
    try:
        # Filtros
        data_inicio = request.args.get('data_inicio', '')
        data_fim = request.args.get('data_fim', '')
        
        if not data_inicio or not data_fim:
            hoje = date.today()
            data_inicio = hoje.replace(day=1).strftime('%Y-%m-%d')
            data_fim = hoje.strftime('%Y-%m-%d')
        
        data_inicio_obj = datetime.strptime(data_inicio, '%Y-%m-%d').date()
        data_fim_obj = datetime.strptime(data_fim, '%Y-%m-%d').date()
        
        # Buscar todos os lançamentos do período
        lancamentos_manuais = LancamentoFluxoCaixa.query.filter(
            LancamentoFluxoCaixa.empresa_id == current_user.empresa_id,
            LancamentoFluxoCaixa.data_vencimento >= data_inicio_obj,
            LancamentoFluxoCaixa.data_vencimento <= data_fim_obj
        ).all()
        
        lancamentos_nfe = LancamentoNotaFiscal.query.filter(
            LancamentoNotaFiscal.empresa_id == current_user.empresa_id,
            LancamentoNotaFiscal.data_vencimento >= data_inicio_obj,
            LancamentoNotaFiscal.data_vencimento <= data_fim_obj
        ).all()
        
        # Análises por categoria
        analise_categorias = {}
        analise_fornecedores = defaultdict(lambda: {'receitas': 0, 'despesas': 0, 'saldo': 0})
        analise_mensal = defaultdict(lambda: {'receitas': 0, 'despesas': 0, 'saldo': 0})
        
        # Processar lançamentos manuais
        for lanc in lancamentos_manuais:
            categoria = lanc.categoria or 'Sem categoria'
            if categoria not in analise_categorias:
                analise_categorias[categoria] = {'receitas': 0, 'despesas': 0, 'total': 0}
            
            if lanc.tipo == 'RECEITA':
                analise_categorias[categoria]['receitas'] += lanc.valor_total
                if lanc.fornecedor_cliente:
                    analise_fornecedores[lanc.fornecedor_cliente]['receitas'] += lanc.valor_total
            else:
                analise_categorias[categoria]['despesas'] += lanc.valor_total
                if lanc.fornecedor_cliente:
                    analise_fornecedores[lanc.fornecedor_cliente]['despesas'] += lanc.valor_total
            
            analise_categorias[categoria]['total'] = analise_categorias[categoria]['receitas'] - analise_categorias[categoria]['despesas']
            
            # Análise mensal
            mes_ano = lanc.data_vencimento.strftime('%Y-%m')
            if lanc.tipo == 'RECEITA':
                analise_mensal[mes_ano]['receitas'] += lanc.valor_total
            else:
                analise_mensal[mes_ano]['despesas'] += lanc.valor_total
            analise_mensal[mes_ano]['saldo'] = analise_mensal[mes_ano]['receitas'] - analise_mensal[mes_ano]['despesas']
        
        # Processar lançamentos NFe
        for lanc in lancamentos_nfe:
            categoria = 'Fornecedores (NFe)'
            if categoria not in analise_categorias:
                analise_categorias[categoria] = {'receitas': 0, 'despesas': 0, 'total': 0}
            
            analise_categorias[categoria]['despesas'] += lanc.valor_total
            analise_categorias[categoria]['total'] = analise_categorias[categoria]['receitas'] - analise_categorias[categoria]['despesas']
            
            if lanc.emitente_nome:
                analise_fornecedores[lanc.emitente_nome]['despesas'] += lanc.valor_total
            
            # Análise mensal
            mes_ano = lanc.data_vencimento.strftime('%Y-%m')
            analise_mensal[mes_ano]['despesas'] += lanc.valor_total
            analise_mensal[mes_ano]['saldo'] = analise_mensal[mes_ano]['receitas'] - analise_mensal[mes_ano]['despesas']
        
        # Calcular saldos dos fornecedores
        for fornecedor in analise_fornecedores:
            analise_fornecedores[fornecedor]['saldo'] = (
                analise_fornecedores[fornecedor]['receitas'] - 
                analise_fornecedores[fornecedor]['despesas']
            )
        
        return render_template('relatorio_analitico_fluxo.html',
                             analise_categorias=analise_categorias,
                             analise_fornecedores=dict(analise_fornecedores),
                             analise_mensal=dict(analise_mensal),
                             data_inicio=data_inicio,
                             data_fim=data_fim,
                             periodo_str=f"{data_inicio_obj.strftime('%d/%m/%Y')} a {data_fim_obj.strftime('%d/%m/%Y')}")
        
    except Exception as e:
        logger.error(f"Erro ao gerar relatório analítico: {e}", exc_info=True)
        flash('Erro ao gerar relatório analítico.', 'error')
        return redirect(url_for('fluxo_caixa'))
    
@app.route('/fiscal/cte/preparar/<int:viagem_id>')
@login_required
@master_required
def preparar_cte_de_viagem_nova(viagem_id):
    """Preparar CT-e a partir de uma viagem específica"""
    try:
        # Buscar dados da viagem
        viagem = Viagem.query.filter_by(
            id=viagem_id,
            empresa_id=current_user.empresa_id
        ).options(
            db.joinedload(Viagem.motorista_formal),
            db.joinedload(Viagem.veiculo)
        ).first_or_404()
        
        # Verificar se viagem já tem CT-e
        cte_existente = CTeEmitido.query.filter_by(viagem_id=viagem_id).first()
        if cte_existente:
            flash('Esta viagem já possui uma CT-e emitida!', 'warning')
            return redirect(url_for('visualizar_cte', cte_id=cte_existente.id))
        
        # Buscar empresa emitente
        empresa = db.session.get(Empresa, current_user.empresa_id)
        
        # Buscar cliente como destinatário
        cliente_destinatario = Cliente.query.filter_by(
            nome_razao_social=viagem.cliente,
            empresa_id=current_user.empresa_id
        ).first()
        
        # Preparar dados para o template
        dados = {
            'viagem_id': viagem_id,
            'viagem': viagem,
            'empresa': empresa,
            'remetente': empresa,  # Empresa é o remetente
            'destinatario': cliente_destinatario,
            'veiculo': viagem.veiculo,
            'motorista': viagem.motorista_formal
        }
        
        return render_template('nova_cte.html', dados=dados)
        
    except Exception as e:
        logger.error(f"Erro ao preparar CT-e da viagem {viagem_id}: {e}", exc_info=True)
        flash('Erro ao carregar dados da viagem', 'error')
        return redirect(url_for('nova_cte_page'))

@app.route('/api/fluxo_caixa/exportar_excel')
@login_required
@master_required
def api_exportar_fluxo_excel():
    """
    Exporta um relatório financeiro detalhado para Excel, com abas por unidade de negócio,
    resumo consolidado e formatação aprimorada para contabilidade.
    """
    try:
        # 1. Coleta e validação de todos os filtros da URL
        hoje = date.today()
        data_inicio_str = request.args.get('data_inicio', hoje.strftime('%Y-%m-%d'))
        data_fim_str = request.args.get('data_fim', (hoje + timedelta(days=30)).strftime('%Y-%m-%d'))
        data_emissao_inicio = request.args.get('data_emissao_inicio', '')
        data_emissao_fim = request.args.get('data_emissao_fim', '')
        categoria_filtro = request.args.get('categoria', '')
        status_filtro = request.args.get('status', '')
        meio_pagamento_filtro = request.args.get('meio_pagamento', '')
        tipo_filtro = request.args.get('tipo', '')
        unidade_negocio_filtro_id = request.args.get('unidade_negocio_id', type=int)

        data_inicio_obj = datetime.strptime(data_inicio_str, '%Y-%m-%d').date()
        data_fim_obj = datetime.strptime(data_fim_str, '%Y-%m-%d').date()

        # 2. Query otimizada para buscar todos os dados necessários com base nos filtros
        query = LancamentoFluxoCaixa.query.options(
            db.joinedload(LancamentoFluxoCaixa.unidade_negocio),
            db.joinedload(LancamentoFluxoCaixa.rateios).joinedload(RateioVeiculo.veiculo),
            db.joinedload(LancamentoFluxoCaixa.centro_custo)
        ).filter(
            LancamentoFluxoCaixa.empresa_id == current_user.empresa_id,
            LancamentoFluxoCaixa.data_vencimento.between(data_inicio_obj, data_fim_obj)
        )

        # Aplica os filtros da tela
        if data_emissao_inicio:
            query = query.filter(LancamentoFluxoCaixa.data_lancamento >= datetime.strptime(data_emissao_inicio, '%Y-%m-%d'))
        if data_emissao_fim:
            query = query.filter(LancamentoFluxoCaixa.data_lancamento <= datetime.strptime(data_emissao_fim, '%Y-%m-%d'))
        if categoria_filtro:
            query = query.filter(LancamentoFluxoCaixa.categoria.ilike(f'%{categoria_filtro}%'))
        if status_filtro:
            status_pagos = ['PAGO', 'Pago']
            if status_filtro == 'PAGO':
                query = query.filter(LancamentoFluxoCaixa.status_pagamento.in_(status_pagos))
            else:
                query = query.filter(LancamentoFluxoCaixa.status_pagamento.notin_(status_pagos))
        if tipo_filtro:
            query = query.filter(LancamentoFluxoCaixa.tipo == tipo_filtro)
        if meio_pagamento_filtro:
            query = query.filter(LancamentoFluxoCaixa.meio_pagamento == meio_pagamento_filtro)
        if unidade_negocio_filtro_id:
            query = query.filter(LancamentoFluxoCaixa.unidade_negocio_id == unidade_negocio_filtro_id)

        lancamentos = query.order_by(LancamentoFluxoCaixa.unidade_negocio_id, LancamentoFluxoCaixa.data_vencimento).all()

        # 3. Organiza os dados por Unidade de Negócio
        dados_por_unidade = defaultdict(list)
        for lanc in lancamentos:
            unidade_nome = lanc.unidade_negocio.nome if lanc.unidade_negocio else "Sem Unidade"
            dados_por_unidade[unidade_nome].append(lanc)

        # 4. Cria o arquivo Excel em memória
        output = io.BytesIO()
        workbook = Workbook()
        workbook.remove(workbook.active)

        # 5. Define estilos de formatação
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        currency_format = 'R$ #,##0.00'
        summary_font = Font(bold=True)

        # 6. Função auxiliar para criar cada aba da planilha
        def create_sheet(sheet_name, data):
            sheet = workbook.create_sheet(title=sheet_name)
            
            headers = [
                "Data Venc.", "Data Lanç.", "Data Pag.", "Unidade de Negócio", "Tipo", "Descrição",
                "Categoria", "Fornecedor/Cliente", "Nº Documento", "Chave Acesso NFe",
                "Valor", "Status", "Meio Pag.", "Parcela", "Rateado?", "Alocação do Rateio", "Obs."
            ]
            sheet.append(headers)
            
            for cell in sheet[1]:
                cell.font = header_font
                cell.fill = header_fill

            total_receitas = 0
            total_despesas = 0

            for lanc in data:
                is_nfe = lanc.documento_numero and lanc.documento_numero.startswith('NFE-')
                chave_acesso = lanc.observacoes.split('Chave de acesso: ')[1] if is_nfe and 'Chave de acesso: ' in lanc.observacoes else ''
                
                alocacao_rateio = ""
                if lanc.tem_rateio:
                    if lanc.rateios:
                        alocacao_rateio = ", ".join([r.veiculo.placa for r in lanc.rateios if r.veiculo])
                    elif lanc.centro_custo:
                        alocacao_rateio = lanc.centro_custo.nome

                row = [
                    lanc.data_vencimento,
                    lanc.data_lancamento.date() if lanc.data_lancamento else None,
                    lanc.data_pagamento,
                    lanc.unidade_negocio.nome if lanc.unidade_negocio else "N/A",
                    lanc.tipo,
                    lanc.descricao,
                    lanc.categoria,
                    lanc.fornecedor_cliente,
                    lanc.documento_numero,
                    chave_acesso,
                    lanc.valor_total,
                    lanc.status_pagamento,
                    lanc.meio_pagamento,
                    f"{lanc.parcela_numero}/{lanc.parcela_total}" if lanc.parcela_total and lanc.parcela_total > 1 else "1/1",
                    "Sim" if lanc.tem_rateio else "Não",
                    alocacao_rateio,
                    lanc.observacoes
                ]
                sheet.append(row)

                sheet.cell(row=sheet.max_row, column=11).number_format = currency_format

                if lanc.tipo == 'RECEITA':
                    total_receitas += lanc.valor_total
                else:
                    total_despesas += lanc.valor_total
            
            sheet.append([])
            summary_row_start = sheet.max_row + 1
            sheet.cell(row=summary_row_start, column=10, value="TOTAL RECEITAS:").font = summary_font
            sheet.cell(row=summary_row_start, column=11, value=total_receitas).number_format = currency_format
            sheet.cell(row=summary_row_start, column=11).font = summary_font

            sheet.cell(row=summary_row_start + 1, column=10, value="TOTAL DESPESAS:").font = summary_font
            sheet.cell(row=summary_row_start + 1, column=11, value=total_despesas).number_format = currency_format
            sheet.cell(row=summary_row_start + 1, column=11).font = summary_font
            
            sheet.cell(row=summary_row_start + 2, column=10, value="SALDO:").font = summary_font
            sheet.cell(row=summary_row_start + 2, column=11, value=(total_receitas - total_despesas)).number_format = currency_format
            sheet.cell(row=summary_row_start + 2, column=11).font = summary_font

            for col_idx, column_cells in enumerate(sheet.columns, 1):
                max_length = 0
                column = get_column_letter(col_idx)
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[column].width = min(adjusted_width, 50)

        # 7. Gera as planilhas: uma consolidada e uma para cada unidade de negócio
        if not unidade_negocio_filtro_id and len(dados_por_unidade) > 1:
            create_sheet("Resumo Geral Consolidado", lancamentos)

        for unidade, dados_unidade in sorted(dados_por_unidade.items()):
            # --- INÍCIO DA CORREÇÃO ---
            # Remove caracteres inválidos para nomes de planilhas do Excel
            sanitized_name = re.sub(r'[\\/*?:\[\]]', '', unidade)
            # Garante que o nome não exceda o limite de 31 caracteres do Excel
            safe_sheet_name = sanitized_name[:31]
            # --- FIM DA CORREÇÃO ---
            
            create_sheet(safe_sheet_name, dados_unidade)

        # 8. Salva o arquivo e envia como resposta
        workbook.save(output)
        output.seek(0)
        
        filename = f"Relatorio_Financeiro_{data_inicio_obj.strftime('%d%m%Y')}_a_{data_fim_obj.strftime('%d%m%Y')}.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except Exception as e:
        logger.error(f"Erro ao exportar relatório detalhado de fluxo de caixa: {e}", exc_info=True)
        flash('Ocorreu um erro inesperado ao gerar o relatório Excel.', 'error')
        return redirect(url_for('fluxo_caixa', **request.args))
    
@app.route('/enviar_contato', methods=['POST'])
def enviar_contato():
    """
    Processa os dados do formulário de contato e retorna uma resposta JSON para o modal.
    """
    try:
        nome = request.form.get('nome')
        telefone = request.form.get('telefone')
        email = request.form.get('email')
        empresa = request.form.get('empresa')
        frota = request.form.get('frota')

        if not all([nome, telefone, email, empresa]):
            return jsonify({'success': False, 'message': 'Por favor, preencha todos os campos obrigatórios.'}), 400

        corpo_email = f"""
        Novo Contato Recebido (Site TrackBras):

        - Nome: {nome}
        - Telefone: {telefone}
        - E-mail: {email}
        - Empresa: {empresa}
        - Tamanho da Frota: {frota}
        """

        msg = Message(
            subject=f"Novo Lead do Site: {empresa}",
            recipients=["seu-email-de-vendas@gmail.com"], # LEMBRE-SE de colocar seu e-mail aqui
            body=corpo_email
        )
        mail.send(msg)

        # Resposta de sucesso em JSON para ativar o modal
        return jsonify({'success': True, 'message': 'E-mail enviado com sucesso!'})

    except Exception as e:
        logger.error(f"Erro ao enviar e-mail de contato: {e}", exc_info=True)
        # Resposta de erro em JSON
        return jsonify({'success': False, 'message': 'Ocorreu um erro interno ao enviar sua solicitação.'}), 500
    
@app.route('/planos')
def planos_page():
    """Renderiza a página de planos."""
    return render_template('planos.html')

@app.route('/contato')
def contato_page():
    """Renderiza a página de contato."""
    return render_template('contato_ini.html')

@app.route('/api/fluxo_caixa/reprocessar_nfe/<string:chave_acesso>', methods=['POST'])
@login_required
@master_required
def api_reprocessar_nfe(chave_acesso):
    """API para reprocessar uma NFe (excluir lançamentos antigos e permitir nova configuração)"""
    try:
        # Buscar todos os lançamentos relacionados a esta NFe
        lancamentos_nfe = LancamentoNotaFiscal.query.filter(
            LancamentoNotaFiscal.chave_acesso == chave_acesso,
            LancamentoNotaFiscal.empresa_id == current_user.empresa_id
        ).all()
        
        if not lancamentos_nfe:
            return jsonify({'success': False, 'message': 'Nenhum lançamento encontrado para esta NFe'}), 404
        
        # Verificar se algum já foi pago
        pagos = [l for l in lancamentos_nfe if l.status_pagamento == 'Pago']
        if pagos:
            return jsonify({
                'success': False, 
                'message': f'Não é possível reprocessar: {len(pagos)} parcela(s) já foi(ram) paga(s)'
            }), 400
        
        # Excluir todos os lançamentos
        for lancamento in lancamentos_nfe:
            db.session.delete(lancamento)
        
        # Marcar a NFe original como não processada
        nfe_original = NFeImportada.query.filter_by(
            chave_acesso=chave_acesso,
            empresa_id=current_user.empresa_id
        ).first()
        
        if nfe_original:
            nfe_original.status = 'IMPORTADA'
        
        db.session.commit()
        
        return jsonify({
            'success': True, 
            'message': 'NFe marcada para reprocessamento. Acesse a tela de importação fiscal para configurar novamente.'
        })
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Erro ao reprocessar NFe {chave_acesso}: {e}", exc_info=True)
        return jsonify({'success': False, 'message': str(e)}), 500
    
@app.route('/api/fiscal/consultar_sefaz', methods=['POST'])
@login_required
def api_consultar_sefaz():
    # Esta rota chama a função do sefaz_service, que foi corrigido para não causar o loop.
    resultado = consultar_notas_sefaz(current_user.empresa_id)
    return jsonify(resultado)

@app.cli.command("associar-notas-antigas")
def associar_notas_antigas_command():
    """
    Busca por NFeImportada que não têm um certificado_id associado e as vincula
    ao certificado principal de sua respectiva empresa.
    """
    print("Iniciando associação de notas fiscais antigas...")
    
    try:
        # Encontra todas as empresas que possuem certificados
        empresas = Empresa.query.join(CertificadoDigital).distinct().all()
        
        total_atualizado = 0
        for empresa in empresas:
            print(f"Processando empresa: {empresa.razao_social} (ID: {empresa.id})")
            
            # Encontra o certificado principal desta empresa
            certificado_principal = CertificadoDigital.query.filter_by(empresa_id=empresa.id, principal=True).first()
            
            if not certificado_principal:
                print(f"  -> Aviso: Nenhuma certificado principal encontrado para esta empresa. Pulando.")
                continue
            
            print(f"  -> Certificado principal encontrado: ID {certificado_principal.id}")
            
            # Atualiza todas as notas "órfãs" (com certificado_id nulo) desta empresa
            # para que apontem para o certificado principal.
            notas_atualizadas = NFeImportada.query.filter_by(
                empresa_id=empresa.id,
                certificado_id=None
            ).update({'certificado_id': certificado_principal.id})
            
            if notas_atualizadas > 0:
                print(f"  -> {notas_atualizadas} nota(s) antiga(s) foram associadas com sucesso!")
                total_atualizado += notas_atualizadas
        
        db.session.commit()
        print(f"\nConcluído! Um total de {total_atualizado} notas foram atualizadas.")
        
    except Exception as e:
        db.session.rollback()
        print(f"\nOcorreu um erro: {e}")

# Adicione estas rotas ao seu app.py



@app.route('/fiscal/cte')
@login_required

def listar_ctes():
    """Lista todas as CT-e emitidas pela empresa"""
    ctes = CTeEmitido.query.filter_by(empresa_id=current_user.empresa_id)\
        .order_by(CTeEmitido.created_at.desc()).all()
    
    return render_template('cte/listar_ctes.html', 
                         ctes=ctes, 
                         active_page='cte')

@app.route('/fiscal/cte/nova')
@login_required

def nova_cte_page():
    """
    Página para emitir nova CT-e.
    Carrega dados de uma viagem se um 'viagem_id' for fornecido na URL,
    caso contrário, abre um formulário para emissão avulsa.
    """
    viagem_id = request.args.get('viagem_id', type=int)
    empresa = db.session.get(Empresa, current_user.empresa_id)
    dados = {}

    if viagem_id:
        # Se veio de uma viagem, carrega todos os dados para preencher o formulário
        viagem = Viagem.query.filter_by(
            id=viagem_id,
            empresa_id=current_user.empresa_id
        ).options(
            db.joinedload(Viagem.motorista_formal),
            db.joinedload(Viagem.veiculo)
        ).first_or_404()
        
        # Busca o cliente no cadastro para pegar todos os dados
        cliente_destinatario = Cliente.query.filter_by(
            nome_razao_social=viagem.cliente,
            empresa_id=current_user.empresa_id
        ).first()

        dados = {
            'viagem_id': viagem_id,
            'viagem': viagem,
            'remetente': empresa,
            'destinatario': cliente_destinatario,
            'veiculo': viagem.veiculo,
            'motorista': viagem.motorista_formal
        }
    else:
        # Se não veio de uma viagem, prepara dados vazios para uma CT-e avulsa
        dados = {
            'viagem_id': None,
            'viagem': None,
            'remetente': empresa,
            'destinatario': None,
            'veiculo': None,
            'motorista': None
        }
    
    # O template 'nova_cte.html' agora é usado para os dois cenários
    return render_template('cte/nova_cte.html', 
                           dados=dados, 
                           active_page='cte')


@app.route('/api/cte/emitir', methods=['POST'])
@login_required
def api_emitir_cte():
    """API para emitir nova CT-e"""
    try:
        data = request.get_json()
        
        # Validações básicas
        required_fields = ['destinatario', 'remetente', 'carga', 'valores', 'impostos']
        for field in required_fields:
            if field not in data:
                return jsonify({'success': False, 'message': f'Campo {field} é obrigatório'}), 400
        
        # Buscar dados da empresa emitente
        empresa = db.session.get(Empresa, current_user.empresa_id)
        if not empresa:
            return jsonify({'success': False, 'message': 'Empresa não encontrada'}), 404
        
        # --- INÍCIO DA CORREÇÃO ---
        # Trata o 'viagem_id' que vem do formulário.
        # Se for a string "None" ou vazio, converte para o nulo (None) do Python.
        # Se for um número, converte para inteiro.
        viagem_id_str = data.get('viagem_id')
        viagem_id_final = int(viagem_id_str) if viagem_id_str and viagem_id_str != 'None' else None
        # --- FIM DA CORREÇÃO ---
        
        # Preparar dados para emissão
        dados_cte = {
            'empresa': {
                'cnpj': empresa.cnpj,
                'razao_social': empresa.razao_social,
                'nome_fantasia': empresa.nome_fantasia,
                'endereco': empresa.endereco,
                'cidade': empresa.cidade,
                'estado': empresa.estado,
                'cep': empresa.cep,
                'inscricao_estadual': empresa.inscricao_estadual
            },
            'destinatario': data['destinatario'],
            'remetente': data['remetente'],
            'carga': data['carga'],
            'valores': data['valores'],
            'impostos': data['impostos'],
            'observacoes': data.get('observacoes', ''),
            'viagem_id': viagem_id_final, # <-- USA A VARIÁVEL CORRIGIDA AQUI
            'veiculo': data.get('veiculo', {}),
            'motorista': data.get('motorista', {}),
            'transmitir_automaticamente': data.get('transmitir_automaticamente', True)
        }
        
        # Determinar ambiente (produção ou homologação)
        ambiente = app.config.get('SEFAZ_AMBIENTE', 'PRODUCAO')
        
        # Emitir CT-e
        resultado = emitir_cte(current_user.empresa_id, dados_cte, ambiente)
        
        if resultado.get('sucesso'):
            return jsonify({
                'success': True,
                'message': 'CT-e emitida com sucesso!',
                'cte_id': resultado.get('cte_id'),
                'chave_acesso': resultado.get('chave_acesso'),
                'numero_cte': resultado.get('numero_cte'),
                'protocolo': resultado.get('protocolo')
            })
        else:
            return jsonify({
                'success': False,
                'message': f'Erro ao emitir CT-e: {resultado.get("erro")}'
            }), 400
            
    except Exception as e:
        logger.error(f"Erro na API de emissão de CT-e: {e}", exc_info=True)
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/fiscal/cte/<int:cte_id>')
@login_required

def visualizar_cte(cte_id):
    """Visualizar CT-e específica"""
    cte = CTeEmitido.query.filter_by(
        id=cte_id, 
        empresa_id=current_user.empresa_id
    ).first_or_404()
    
    return render_template('cte/visualizar_cte.html', 
                         cte=cte, 
                         active_page='cte')

@app.route('/api/cte/<int:cte_id>/consultar_status', methods=['POST'])
@login_required

def api_consultar_status_cte(cte_id):
    """API para consultar status da CT-e na SEFAZ"""
    try:
        cte = CTeEmitido.query.filter_by(
            id=cte_id,
            empresa_id=current_user.empresa_id
        ).first_or_404()
        
        if not cte.chave_acesso:
            return jsonify({'success': False, 'message': 'CT-e não possui chave de acesso'}), 400
        
        ambiente = app.config.get('SEFAZ_AMBIENTE', 'PRODUCAO')
        service = CTeService(current_user.empresa_id, ambiente)
        
        resultado = service.consultar_cte(cte.chave_acesso)
        
        if resultado.get('sucesso'):
            # Atualizar status da CT-e se necessário
            # Você pode implementar lógica para atualizar o status baseado na resposta
            
            return jsonify({
                'success': True,
                'status': resultado.get('status', 'CONSULTADO'),
                'protocolo': resultado.get('protocolo'),
                'message': 'Status consultado com sucesso'
            })
        else:
            return jsonify({
                'success': False,
                'message': f'Erro ao consultar status: {resultado.get("erro")}'
            }), 400
            
    except Exception as e:
        logger.error(f"Erro ao consultar status da CT-e {cte_id}: {e}", exc_info=True)
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/cte/<int:cte_id>/cancelar', methods=['POST'])
@login_required

def api_cancelar_cte(cte_id):
    """API para cancelar CT-e"""
    try:
        data = request.get_json()
        justificativa = data.get('justificativa', '')
        
        if len(justificativa) < 15:
            return jsonify({'success': False, 'message': 'Justificativa deve ter pelo menos 15 caracteres'}), 400
        
        cte = CTeEmitido.query.filter_by(
            id=cte_id,
            empresa_id=current_user.empresa_id
        ).first_or_404()
        
        if cte.status == 'CANCELADO':
            return jsonify({'success': False, 'message': 'CT-e já está cancelada'}), 400
        
        if cte.status != 'AUTORIZADO':
            return jsonify({'success': False, 'message': 'Apenas CT-e autorizadas podem ser canceladas'}), 400
        
        # Implementar cancelamento na SEFAZ aqui
        # Por enquanto, apenas atualizar o status localmente
        cte.status = 'CANCELADO'
        cte.observacoes = (cte.observacoes or '') + f"\n\nCANCELADA: {justificativa}"
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'CT-e cancelada com sucesso!'
        })
        
    except Exception as e:
        logger.error(f"Erro ao cancelar CT-e {cte_id}: {e}", exc_info=True)
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/cte/<int:cte_id>/xml')
@login_required

def api_download_xml_cte(cte_id):
    """API para download do XML da CT-e"""
    try:
        cte = CTeEmitido.query.filter_by(
            id=cte_id,
            empresa_id=current_user.empresa_id
        ).first_or_404()
        
        if not cte.xml_content:
            return jsonify({'success': False, 'message': 'XML não disponível'}), 404
        
        import io
        xml_buffer = io.BytesIO(cte.xml_content)
        
        return send_file(
            xml_buffer,
            mimetype='text/xml',
            as_attachment=True,
            download_name=f'CTE_{cte.chave_acesso}.xml'
        )
        
    except Exception as e:
        logger.error(f"Erro ao download XML da CT-e {cte_id}: {e}", exc_info=True)
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/viagem/<int:viagem_id>/dados_cte')
@login_required

def api_dados_viagem_para_cte(viagem_id):
    """API para buscar dados da viagem para preenchimento automático da CT-e"""
    try:
        viagem = Viagem.query.filter_by(
            id=viagem_id,
            empresa_id=current_user.empresa_id
        ).options(
            db.joinedload(Viagem.motorista_formal),
            db.joinedload(Viagem.veiculo)
        ).first_or_404()
        
        # Buscar cliente como destinatário
        cliente_destinatario = Cliente.query.filter_by(
            nome_razao_social=viagem.cliente,
            empresa_id=current_user.empresa_id
        ).first()
        
        dados = {
            'viagem': {
                'id': viagem.id,
                'cliente': viagem.cliente,
                'valor_recebido': viagem.valor_recebido or 0,
                'origem': viagem.endereco_saida,
                'destino': viagem.endereco_destino,
                'distancia_km': viagem.distancia_km or 0
            },
            'destinatario': {
                'nome': cliente_destinatario.nome_razao_social if cliente_destinatario else viagem.cliente,
                'cnpj_cpf': cliente_destinatario.cpf_cnpj if cliente_destinatario else '',
                'endereco': cliente_destinatario.logradouro if cliente_destinatario else '',
                'cidade': cliente_destinatario.cidade if cliente_destinatario else '',
                'uf': cliente_destinatario.estado if cliente_destinatario else '',
                'cep': cliente_destinatario.cep if cliente_destinatario else '',
                'ie': getattr(cliente_destinatario, 'inscricao_estadual', '') if cliente_destinatario else ''
            } if cliente_destinatario else None,
            'veiculo': {
                'placa': viagem.veiculo.placa if viagem.veiculo else '',
                'renavam': viagem.veiculo.renavam if viagem.veiculo else ''
            } if viagem.veiculo else None,
            'motorista': {
                'nome': viagem.motorista_formal.nome if viagem.motorista_formal else '',
                'cpf': viagem.motorista_formal.cpf if viagem.motorista_formal else ''
            } if viagem.motorista_formal else None
        }
        
        return jsonify({'success': True, 'dados': dados})
        
    except Exception as e:
        logger.error(f"Erro ao buscar dados da viagem {viagem_id} para CT-e: {e}", exc_info=True)
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/fiscal/cte/configuracoes', methods=['GET', 'POST'])
@login_required

def configuracoes_cte():
    """Configurações dos parâmetros de CT-e"""
    parametros = CTeParametros.query.filter_by(empresa_id=current_user.empresa_id).first()
    
    if not parametros:
        parametros = CTeParametros(empresa_id=current_user.empresa_id)
        db.session.add(parametros)
        db.session.commit()
    
    if request.method == 'POST':
        try:
            parametros.serie_padrao = request.form.get('serie_padrao', '1')
            parametros.aliquota_icms_padrao = float(request.form.get('aliquota_icms_padrao', 17.0))
            parametros.natureza_operacao = request.form.get('natureza_operacao', 'PRESTAÇÃO DE SERVIÇO DE TRANSPORTE')
            parametros.codigo_cfop = request.form.get('codigo_cfop', '5353')
            parametros.tipo_documento = request.form.get('tipo_documento', '0')
            parametros.tipo_servico = request.form.get('tipo_servico', '0')
            
            db.session.commit()
            flash('Configurações de CT-e atualizadas com sucesso!', 'success')
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar configurações: {e}', 'error')
            logger.error(f"Erro ao atualizar configurações CT-e: {e}", exc_info=True)
    
    return render_template('cte/configuracoes_cte.html', 
                         parametros=parametros,
                         active_page='cte')

@app.route('/api/cliente/<int:cliente_id>/dados_cte')
@login_required

def api_dados_cliente_para_cte(cliente_id):
    """API para buscar dados do cliente para CT-e"""
    try:
        cliente = Cliente.query.filter_by(
            id=cliente_id,
            empresa_id=current_user.empresa_id
        ).first_or_404()
        
        dados = {
            'nome': cliente.nome_razao_social,
            'cnpj_cpf': cliente.cpf_cnpj,
            'endereco': cliente.logradouro,
            'numero': cliente.numero,
            'bairro': cliente.bairro,
            'cidade': cliente.cidade,
            'uf': cliente.estado,
            'cep': cliente.cep,
            'ie': getattr(cliente, 'inscricao_estadual', '')
        }
        
        return jsonify({'success': True, 'dados': dados})
        
    except Exception as e:
        logger.error(f"Erro ao buscar dados do cliente {cliente_id} para CT-e: {e}", exc_info=True)
        return jsonify({'success': False, 'message': str(e)}), 500



@app.route('/api/fiscal/buscar_dados_nota/<string:chave_acesso>')
@login_required
def api_buscar_dados_nota(chave_acesso):
    """
    Esta API busca os dados de uma nota para exibi-los no modal de lançamento.
    Agora inclui informações do certificado de origem.
    """
    nota = NFeImportada.query.filter_by(
        chave_acesso=chave_acesso,
        empresa_id=current_user.empresa_id
    ).first()

    if not nota:
        return jsonify({'success': False, 'message': 'Nota não encontrada na base local. Realize a consulta à SEFAZ primeiro.'}), 404

    if nota.status == 'PROCESSADA':
        return jsonify({'success': False, 'message': 'Esta nota já foi processada anteriormente.'}), 409
    
    # Buscar informações do certificado - CORRIGIDO para usar Session.get()
    certificado = None
    if nota.certificado_id:
        certificado = db.session.get(CertificadoDigital, nota.certificado_id)
    
    # Prepara os dados da nota para serem retornados como JSON
    dados_da_nota = {
        "emitente_nome": nota.emitente_nome,
        "emitente_cnpj": nota.emitente_cnpj,
        "valor_total": nota.valor_total,
        "data_emissao": nota.data_emissao.isoformat(),
        "certificado_origem": {
            "id": certificado.id if certificado else None,
            "nome": certificado.nome_arquivo if certificado else "Certificado não identificado",
            "principal": certificado.principal if certificado else False
        }
    }
    
    return jsonify({'success': True, 'nota': dados_da_nota})

@app.route('/salvar_custo_viagem', methods=['POST'])
@login_required
def salvar_custo_viagem():
    viagem_id = request.form.get('viagem_id', type=int)
    if not viagem_id:
        return jsonify({'success': False, 'message': 'ID da viagem não foi fornecido.'}), 400

    viagem = Viagem.query.filter_by(id=viagem_id, empresa_id=current_user.empresa_id).first()
    if not viagem:
        return jsonify({'success': False, 'message': 'Viagem não encontrada ou acesso negado.'}), 404
    
    try:
        custo = CustoViagem.query.filter_by(viagem_id=viagem_id).first()
        if not custo:
            custo = CustoViagem(viagem_id=viagem_id)
            db.session.add(custo)

        custo.pedagios = float(request.form.get('pedagios') or 0)
        custo.alimentacao = float(request.form.get('alimentacao') or 0)
        custo.hospedagem = float(request.form.get('hospedagem') or 0)
        custo.outros = float(request.form.get('outros') or 0)
        custo.descricao_outros = request.form.get('descricao_outros', '').strip()
        
        # <<< INÍCIO DA CORREÇÃO >>>
        
        urls_anexos = custo.anexos.split(',') if custo.anexos else []
        if 'anexos_despesa' in request.files:
            # Pega as configurações corretas do app
            s3_client = boto3.client(
                's3',
                endpoint_url=app.config['CLOUDFLARE_R2_ENDPOINT'],
                aws_access_key_id=app.config['CLOUDFLARE_R2_ACCESS_KEY'],
                aws_secret_access_key=app.config['CLOUDFLARE_R2_SECRET_KEY'],
                region_name='auto'
            )
            bucket_name = app.config['CLOUDFLARE_R2_BUCKET']
            public_url_base = app.config['CLOUDFLARE_R2_PUBLIC_URL']
            
            for anexo in request.files.getlist('anexos_despesa'):
                if anexo and anexo.filename != '':
                    filename = secure_filename(anexo.filename)
                    # Corrigido o caminho para 'custos_viagem' e o separador para '-'
                    s3_path = f"custos_viagem/{viagem_id}/{uuid.uuid4()}-{filename}"
                    
                    s3_client.upload_fileobj(anexo, bucket_name, s3_path, ExtraArgs={'ContentType': anexo.content_type})
                    urls_anexos.append(f"{public_url_base}/{s3_path}")

        if urls_anexos:
            custo.anexos = ",".join(filter(None, urls_anexos))

        # <<< FIM DA CORREÇÃO >>>

        custo_total_geral = (custo.pedagios or 0) + (custo.alimentacao or 0) + (custo.hospedagem or 0) + (custo.outros or 0)
        viagem.custo = custo_total_geral

        db.session.commit()
        return jsonify({'success': True, 'message': 'Despesas salvas com sucesso!'})

    except Exception as e:
        db.session.rollback()
        logger.error(f"Erro ao salvar custo da viagem {viagem_id}: {str(e)}")
        return jsonify({'success': False, 'message': f'Erro interno do servidor: {str(e)}'}), 500
    
@app.route('/excluir_anexo_custo', methods=['POST'])
@login_required
def excluir_anexo_custo():
    data = request.get_json()
    viagem_id = data.get('viagem_id')
    anexo_url = data.get('anexo_url')

    if not viagem_id or not anexo_url:
        return jsonify({'success': False, 'message': 'Dados incompletos.'}), 400

    try:
        custo = CustoViagem.query.filter_by(viagem_id=viagem_id).first()
        if not custo or not custo.anexos:
            return jsonify({'success': False, 'message': 'Anexo não encontrado.'}), 404

        anexos_atuais = custo.anexos.split(',')
        if anexo_url not in anexos_atuais:
            return jsonify({'success': False, 'message': 'URL do anexo não corresponde.'}), 404

        # 1. Excluir do Cloudflare R2
        s3_client = boto3.client(
            's3',
            endpoint_url=app.config['CLOUDFLARE_R2_ENDPOINT'],
            aws_access_key_id=app.config['CLOUDFLARE_R2_ACCESS_KEY'],
            aws_secret_access_key=app.config['CLOUDFLARE_R2_SECRET_KEY'],
            region_name='auto'
        )
        bucket_name = app.config['CLOUDFLARE_R2_BUCKET']
        key = anexo_url.replace(app.config['CLOUDFLARE_R2_PUBLIC_URL'] + '/', '')
        s3_client.delete_object(Bucket=bucket_name, Key=key)

        # 2. Excluir do Banco de Dados
        anexos_atuais.remove(anexo_url)
        custo.anexos = ','.join(anexos_atuais) if anexos_atuais else None
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Anexo excluído com sucesso!'})

    except Exception as e:
        db.session.rollback()
        logger.error(f"Erro ao excluir anexo da viagem {viagem_id}: {str(e)}")
        return jsonify({'success': False, 'message': f'Erro interno do servidor: {str(e)}'}), 500

@app.route('/consultar_despesas/<int:viagem_id>', methods=['GET'])
@login_required
def consultar_despesas(viagem_id):
    try:
        viagem = Viagem.query.get_or_404(viagem_id)
        custo_viagem = CustoViagem.query.filter_by(viagem_id=viagem_id).first()
        custo_dict = {
            'combustivel': custo_viagem.combustivel if custo_viagem else 0.0,
            'pedagios': custo_viagem.pedagios if custo_viagem else 0.0,
            'alimentacao': custo_viagem.alimentacao if custo_viagem else 0.0,
            'hospedagem': custo_viagem.hospedagem if custo_viagem else 0.0,
            'outros': custo_viagem.outros if custo_viagem else 0.0,
            'descricao_outros': custo_viagem.descricao_outros if custo_viagem else 'Nenhuma',
            'anexos': custo_viagem.anexos.split(',') if custo_viagem and custo_viagem.anexos else []
        }
        return jsonify(custo_dict)
    except Exception as e:
        logger.error(f"Erro ao consultar despesas da viagem {viagem_id}: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/atualizar_status_viagem/<int:viagem_id>', methods=['POST'])
@login_required # Garante que apenas usuários logados possam alterar o status
def atualizar_status_viagem(viagem_id):
    try:
        data = request.get_json()
        novo_status = data.get('status')

        if novo_status not in ['pendente', 'em_andamento', 'concluida', 'cancelada']:
            return jsonify({'success': False, 'message': 'Status inválido'}), 400

        viagem = Viagem.query.filter_by(id=viagem_id, empresa_id=current_user.empresa_id).first_or_404()

        status_antigo = viagem.status
        viagem.status = novo_status

        if novo_status in ['concluida', 'cancelada']:
            if not viagem.data_fim:
                viagem.data_fim = datetime.utcnow()
            if viagem.veiculo:
                viagem.veiculo.status = 'Disponível' #

        elif novo_status == 'em_andamento' and status_antigo in ['concluida', 'cancelada']:
            viagem.data_fim = None
            if viagem.veiculo:
                outra_viagem_ativa = Viagem.query.filter(
                    Viagem.veiculo_id == viagem.veiculo_id,
                    Viagem.veiculo.has(status='Em Rota'), # Certifique-se de que o filtro aqui também está correto.
                    Viagem.id != viagem.id
                ).first()

                if outra_viagem_ativa:
                    flash(f'Erro: O veículo {viagem.veiculo.placa} já está em uso na viagem #{outra_viagem_ativa.id}.', 'error')
                    db.session.rollback()
                    return jsonify({'success': False, 'message': f'Veículo já está em outra viagem.'}), 409

                # LINHA A SER CORRIGIDA:
                viagem.veiculo.status = 'Em Rota' #

        db.session.commit()
        flash('Status da viagem atualizado com sucesso!', 'success')
        return jsonify({'success': True, 'message': 'Status atualizado com sucesso.'})

    except Exception as e:
        db.session.rollback()
        logger.error(f"Erro ao atualizar status da viagem {viagem_id}: {e}", exc_info=True)
        return jsonify({'success': False, 'message': str(e)}), 500
    
@app.route('/api/viagem/<int:viagem_id>/despesas_detalhes')
@login_required
def api_despesas_detalhes(viagem_id):
    """Retorna um HTML renderizado com os detalhes de todas as despesas da viagem."""
    viagem = Viagem.query.options(
        db.joinedload(Viagem.custo_viagem),
        db.joinedload(Viagem.abastecimentos)
    ).get_or_404(viagem_id)

    # Prepara uma lista combinada de despesas para o template
    despesas_detalhadas = []
    
    # Adiciona abastecimentos
    for abast in viagem.abastecimentos:
        despesas_detalhadas.append({
            'tipo': 'Abastecimento',
            'data': abast.data_abastecimento.strftime('%d/%m/%Y'),
            'descricao': f"{abast.litros:.2f}L @ R$ {abast.preco_por_litro:.2f}/L",
            'valor': abast.custo_total
        })

    # Adiciona outras despesas
    if viagem.custo_viagem:
        if viagem.custo_viagem.pedagios:
            despesas_detalhadas.append({'tipo': 'Pedágios', 'descricao': 'Total em pedágios', 'valor': viagem.custo_viagem.pedagios})
        if viagem.custo_viagem.alimentacao:
            despesas_detalhadas.append({'tipo': 'Alimentação', 'descricao': 'Total em alimentação', 'valor': viagem.custo_viagem.alimentacao})
        if viagem.custo_viagem.hospedagem:
            despesas_detalhadas.append({'tipo': 'Hospedagem', 'descricao': 'Total em hospedagem', 'valor': viagem.custo_viagem.hospedagem})
        if viagem.custo_viagem.outros:
            despesas_detalhadas.append({'tipo': 'Outros', 'descricao': viagem.custo_viagem.descricao_outros or 'Sem descrição', 'valor': viagem.custo_viagem.outros})

    # Renderiza um novo template com os detalhes
    return render_template('detalhes_despesas_modal.html', despesas=despesas_detalhadas, viagem_id=viagem_id)

@app.route('/consultar_viagens')
@login_required
def consultar_viagens():
    # Pega todos os argumentos do request de uma só vez
    args = request.args
    status_filter = args.get('status', '')
    search_query = args.get('search', '')
    data_inicio = args.get('data_inicio', '')
    data_fim = args.get('data_fim', '')
    motorista_id_filter = args.get('motorista_id', type=int)
    veiculo_id_filter = args.get('veiculo_id', type=int)

    # A query base já começa filtrando pela empresa do usuário
    query = Viagem.query.filter_by(empresa_id=current_user.empresa_id).options(
        db.joinedload(Viagem.motorista_formal),
        db.joinedload(Viagem.veiculo),
        db.joinedload(Viagem.custo_viagem),
        db.joinedload(Viagem.abastecimentos)
    )

    # Aplica os filtros
    if status_filter:
        query = query.filter(Viagem.status == status_filter)
    if motorista_id_filter:
        query = query.filter(Viagem.motorista_id == motorista_id_filter)
    if veiculo_id_filter:
        query = query.filter(Viagem.veiculo_id == veiculo_id_filter)

    if search_query:
        search_term = f'%{search_query}%'
        query = query.outerjoin(Motorista, Viagem.motorista_id == Motorista.id).filter(
            or_(
                Viagem.cliente.ilike(search_term),
                Motorista.nome.ilike(search_term),
                Viagem.veiculo.has(Veiculo.placa.ilike(search_term)),
                Viagem.endereco_saida.ilike(search_term),
                Viagem.endereco_destino.ilike(search_term)
            )
        )
        
    if data_inicio:
        try:
            query = query.filter(Viagem.data_inicio >= datetime.strptime(data_inicio, '%Y-%m-%d'))
        except ValueError:
            flash('Data de início inválida.', 'error')
            
    if data_fim:
        try:
            # Adiciona 1 dia e subtrai 1 segundo para incluir o dia inteiro
            data_fim_obj = datetime.strptime(data_fim, '%Y-%m-%d') + timedelta(days=1, seconds=-1)
            query = query.filter(Viagem.data_inicio <= data_fim_obj)
        except ValueError:
            flash('Data de fim inválida.', 'error')

    viagens_objetos = query.order_by(Viagem.data_inicio.desc()).all()

    # Processa os dados para o template
    for v in viagens_objetos:
        # ... (cálculo de custo total continua igual) ...
        custo_despesas = 0
        if v.custo_viagem:
            custo_despesas = (v.custo_viagem.pedagios or 0) + (v.custo_viagem.alimentacao or 0) + (v.custo_viagem.hospedagem or 0) + (v.custo_viagem.outros or 0)
        custo_abastecimento = sum(a.custo_total for a in v.abastecimentos)
        v.custo_total_calculado = custo_despesas + custo_abastecimento
        
        v.motorista_nome = v.motorista_formal.nome if v.motorista_formal else 'N/A'
        v.motorista_telefone = v.motorista_formal.telefone if v.motorista_formal else None

        # Busca o cliente no banco de dados para obter o telefone
        cliente_obj = Cliente.query.filter_by(nome_razao_social=v.cliente, empresa_id=current_user.empresa_id).first()
        v.cliente_telefone = cliente_obj.telefone if cliente_obj else None

    # Busca os dados para os dropdowns de filtro
    motoristas_filtro = Motorista.query.filter_by(empresa_id=current_user.empresa_id).order_by(Motorista.nome).all()
    veiculos_filtro = Veiculo.query.filter_by(empresa_id=current_user.empresa_id, is_administrativo=False).order_by(Veiculo.placa).all()
    
    return render_template(
        'consultar_viagens.html',
        active_page='consultar_viagens',
        viagens=viagens_objetos,
        motoristas=motoristas_filtro,
        veiculos=veiculos_filtro,
        request=request
    )

@app.route('/politica_privacidade')
def politica_privacidade():
    return render_template('politica_privacidade.html')

@app.route('/viagem/<int:viagem_id>/gerenciar_despesas')
@login_required
def gerenciar_despesas_viagem(viagem_id):
    # Garante que a viagem pertence à empresa do usuário
    viagem = Viagem.query.filter_by(id=viagem_id, empresa_id=current_user.empresa_id).first_or_404()
    
    # Busca os custos e abastecimentos associados
    custo = CustoViagem.query.filter_by(viagem_id=viagem_id).first()
    abastecimentos = Abastecimento.query.filter_by(viagem_id=viagem_id).order_by(Abastecimento.data_abastecimento.desc()).all()

    # Renderiza o novo template do modal
    return render_template('gerenciar_despesas_modal.html', 
                           viagem=viagem, 
                           custo=custo, 
                           abastecimentos=abastecimentos)

from flask import jsonify, request
from datetime import datetime

@app.route('/viagem/<int:viagem_id>/finalizar', methods=['POST'])
@login_required
def finalizar_viagem(viagem_id):
    if current_user.role != 'Motorista':
        return jsonify({'success': False, 'message': 'Acesso negado.'}), 403

    viagem = Viagem.query.filter_by(id=viagem_id, empresa_id=current_user.empresa_id, status='em_andamento').first_or_404()
    
    data = request.get_json()
    odometro_str = data.get('odometer')

    try:
        odometro_final = float(odometro_str)
        if viagem.odometro_inicial is not None and odometro_final < viagem.odometro_inicial:
            return jsonify({'success': False, 'message': 'Odômetro final não pode ser menor que o inicial.'}), 400
    except (ValueError, TypeError):
        return jsonify({'success': False, 'message': 'Odômetro final inválido. Por favor, insira um número válido.'}), 400

    viagem.status = 'concluida'
    viagem.data_fim = datetime.utcnow()
    viagem.odometro_final = odometro_final
    
    if viagem.veiculo:
        viagem.veiculo.status = 'Disponível'
        viagem.veiculo.km_rodados = odometro_final
        
    db.session.commit()
    
    socketio.emit('status_viagem_atualizado', {
        'viagem_id': viagem.id, 
        'status': 'concluida'
    }, room='admins')

    return jsonify({'success': True, 'message': 'Viagem finalizada com sucesso!'})
    

from collections import defaultdict

from collections import defaultdict

@app.route('/relatorios')
@login_required
def relatorios():
    try:
        data_inicio_str = request.args.get('data_inicio', '')
        data_fim_str = request.args.get('data_fim', '')
        motorista_id_filter = request.args.get('motorista_id', '')
        veiculo_id_filter = request.args.get('veiculo_id', '')
        
        query = Viagem.query.filter_by(empresa_id=current_user.empresa_id).options(
            db.joinedload(Viagem.custo_viagem),
            db.joinedload(Viagem.motorista_formal),
            db.joinedload(Viagem.veiculo),
            db.joinedload(Viagem.abastecimentos)
        )

        if data_inicio_str:
            query = query.filter(Viagem.data_inicio >= datetime.strptime(data_inicio_str, '%Y-%m-%d'))
        if data_fim_str:
            data_fim_obj = datetime.strptime(data_fim_str, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(Viagem.data_inicio < data_fim_obj)
        if motorista_id_filter:
            query = query.filter(Viagem.motorista_id == int(motorista_id_filter))
        if veiculo_id_filter:
            query = query.filter(Viagem.veiculo_id == int(veiculo_id_filter))

        viagens_filtradas = query.order_by(Viagem.data_inicio.desc()).all()

        total_receita = 0.0
        total_custo_outros = 0.0
        total_custo_combustivel = 0.0
        total_distancia = 0.0
        total_litros = 0.0
        
        dados_grafico_mensal = defaultdict(lambda: {'receita': 0.0, 'custo': 0.0})
        dados_grafico_categorias = defaultdict(float)
        clientes_stats = defaultdict(lambda: {'nome': '', 'viagens': 0, 'receita': 0.0, 'custo': 0.0, 'lucro': 0.0})
        motoristas_stats_dict = defaultdict(lambda: {'id': None, 'nome': 'N/A', 'viagens': 0, 'receita': 0.0}) # Renomeado para clareza
        veiculos_stats = defaultdict(lambda: {'id': None, 'modelo': 'N/A', 'placa': 'N/A', 'km': 0.0, 'custo': 0.0, 'litros': 0.0})

        for v in viagens_filtradas:
            receita_viagem = v.valor_recebido or 0.0
            custo_combustivel_viagem = sum(a.custo_total for a in v.abastecimentos)
            litros_viagem = sum(a.litros for a in v.abastecimentos)
            custo_outros_viagem = 0
            if v.custo_viagem:
                custo_outros_viagem += (v.custo_viagem.pedagios or 0) + (v.custo_viagem.alimentacao or 0) + (v.custo_viagem.hospedagem or 0) + (v.custo_viagem.outros or 0)
            
            custo_total_viagem = custo_combustivel_viagem + custo_outros_viagem
            total_receita += receita_viagem
            total_custo_combustivel += custo_combustivel_viagem
            total_custo_outros += custo_outros_viagem
            total_distancia += v.distancia_km or 0.0
            total_litros += litros_viagem

            if v.data_inicio:
                mes = v.data_inicio.strftime('%Y-%m')
                dados_grafico_mensal[mes]['receita'] += receita_viagem
                dados_grafico_mensal[mes]['custo'] += custo_total_viagem

            dados_grafico_categorias['Combustível'] += custo_combustivel_viagem
            if v.custo_viagem:
                dados_grafico_categorias['Pedágios'] += v.custo_viagem.pedagios or 0
                dados_grafico_categorias['Alimentação'] += v.custo_viagem.alimentacao or 0
                dados_grafico_categorias['Hospedagem'] += v.custo_viagem.hospedagem or 0
                dados_grafico_categorias['Outros'] += v.custo_viagem.outros or 0
            
            if v.cliente:
                clientes_stats[v.cliente]['nome'] = v.cliente
                clientes_stats[v.cliente]['viagens'] += 1
                clientes_stats[v.cliente]['receita'] += receita_viagem
                clientes_stats[v.cliente]['custo'] += custo_total_viagem
                clientes_stats[v.cliente]['lucro'] = clientes_stats[v.cliente]['receita'] - clientes_stats[v.cliente]['custo']

            if v.motorista_formal:
                motoristas_stats_dict[v.motorista_formal.id].update({'id': v.motorista_formal.id, 'nome': v.motorista_formal.nome})
                motoristas_stats_dict[v.motorista_formal.id]['viagens'] += 1
                motoristas_stats_dict[v.motorista_formal.id]['receita'] += receita_viagem

            if v.veiculo:
                veiculos_stats[v.veiculo.id].update({'id': v.veiculo.id, 'placa': v.veiculo.placa, 'modelo': v.veiculo.modelo})
                veiculos_stats[v.veiculo.id]['km'] += v.distancia_km or 0.0
                veiculos_stats[v.veiculo.id]['custo'] += custo_total_viagem
                veiculos_stats[v.veiculo.id]['litros'] += litros_viagem
        
        # --- LÓGICA DE ORDENAÇÃO ADICIONADA AQUI ---
        motoristas_ordenados = sorted(motoristas_stats_dict.values(), key=lambda m: m['receita'], reverse=True)

        total_custo = total_custo_combustivel + total_custo_outros
        consumo_medio_geral = (total_distancia / total_litros) if total_litros > 0 else 0

        motoristas_para_filtro = Motorista.query.filter_by(empresa_id=current_user.empresa_id).order_by(Motorista.nome).all()
        veiculos_para_filtro = Veiculo.query.filter_by(empresa_id=current_user.empresa_id).order_by(Veiculo.placa).all()

        return render_template(
            'relatorios.html',
            request=request,
            total_viagens=len(viagens_filtradas),
            total_receita=total_receita,
            total_custo=total_custo,
            consumo_medio_geral=consumo_medio_geral,
            motoristas_filtro=motoristas_para_filtro,
            veiculos_filtro=veiculos_para_filtro,
            dados_grafico_mensal=dict(sorted(dados_grafico_mensal.items())),
            dados_grafico_categorias=dict(dados_grafico_categorias),
            clientes_stats=list(clientes_stats.values()),
            motoristas_stats=motoristas_ordenados, # Enviando a LISTA JÁ ORDENADA
            veiculos_stats=list(veiculos_stats.values())
        )

    except Exception as e:
        logger.error(f"Erro ao gerar relatórios: {e}", exc_info=True)
        flash(f"Ocorreu um erro inesperado ao gerar os relatórios: {e}", "error")
        return redirect(url_for('index'))


@app.route('/api/viagem/<int:viagem_id>/despesas', methods=['GET'])
@login_required
def get_viagem_despesas(viagem_id):
    """Busca os custos de uma viagem para preencher o formulário de edição."""
    custo = CustoViagem.query.filter_by(viagem_id=viagem_id).first()
    if custo:
        return jsonify({
            'success': True,
            'pedagios': custo.pedagios,
            'alimentacao': custo.alimentacao,
            'hospedagem': custo.hospedagem,
            'outros': custo.outros,
            'descricao_outros': custo.descricao_outros
        })
    return jsonify({'success': False, 'message': 'Nenhuma despesa encontrada.'})

@app.route('/api/viagem/<int:viagem_id>/abastecimentos', methods=['GET'])
@login_required
def get_viagem_abastecimentos(viagem_id):
    """Busca o último abastecimento de uma viagem para preencher o formulário."""
    abastecimento = Abastecimento.query.filter_by(viagem_id=viagem_id).order_by(Abastecimento.data_abastecimento.desc()).first()
    if abastecimento:
        return jsonify({
            'success': True,
            'litros': abastecimento.litros,
            'preco_por_litro': abastecimento.preco_por_litro,
            'odometro': abastecimento.odometro
        })
    return jsonify({'success': False, 'message': 'Nenhum abastecimento encontrado.'})

@app.route('/oficina')
@login_required
def oficina():
    """ Rota principal do dashboard da oficina. """
    
    # <<< CORREÇÃO APLICADA AQUI >>>
    # Adicionado o filtro is_administrativo=False para não listar o veículo ADM01 na oficina
    todos_veiculos_obj = Veiculo.query.filter_by(
        empresa_id=current_user.empresa_id,
        is_administrativo=False
    ).order_by(Veiculo.modelo).all()
    
    alertas = []
    veiculos_data = []
    limite_km_alerta = 1000
    limite_dias_alerta = 30

    for veiculo in todos_veiculos_obj:
        planos_progresso = []
        progresso_maximo = -1
        next_maint_summary = {}

        for plano_assoc in veiculo.planos_associados:
            alerta_gerado = False
            mensagem = ""
            km_desde_ultima = 0
            progresso_km = 0
            
            if plano_assoc.intervalo_km and plano_assoc.intervalo_km > 0 and veiculo.km_rodados is not None:
                km_desde_ultima = (veiculo.km_rodados or 0) - (plano_assoc.km_ultima_manutencao or 0)
                if km_desde_ultima < 0: km_desde_ultima = 0
                
                progresso_km = (km_desde_ultima / plano_assoc.intervalo_km) * 100

                if km_desde_ultima >= plano_assoc.intervalo_km:
                    alerta_gerado = True
                    mensagem = "Vencido por KM"
                elif plano_assoc.intervalo_km - km_desde_ultima <= limite_km_alerta:
                    alerta_gerado = True
                    mensagem = "Próximo por KM"

            planos_progresso.append({
                "descricao": plano_assoc.plano.descricao,
                "progresso": int(progresso_km),
                "km_desde_ultima": km_desde_ultima,
                "intervalo_km": plano_assoc.intervalo_km,
            })

            if not alerta_gerado and plano_assoc.intervalo_meses and plano_assoc.data_ultima_manutencao:
                data_proxima = plano_assoc.data_ultima_manutencao + timedelta(days=plano_assoc.intervalo_meses * 30)
                dias_restantes = (data_proxima - date.today()).days
                if dias_restantes <= 0:
                    alerta_gerado = True
                    mensagem = "Vencido por tempo"
                elif dias_restantes <= limite_dias_alerta:
                    alerta_gerado = True
                    mensagem = "Próximo por tempo"

            if alerta_gerado:
                alertas.append({
                    "veiculo": veiculo,
                    "plano": plano_assoc.plano,
                    "km_ultima_manutencao": plano_assoc.km_ultima_manutencao,
                    "intervalo_km": plano_assoc.intervalo_km,
                    "mensagem": mensagem
                })

            if progresso_km > progresso_maximo:
                progresso_maximo = progresso_km
                next_maint_summary = {
                    "descricao": plano_assoc.plano.descricao,
                    "progresso": int(progresso_km),
                    "km_desde_ultima": km_desde_ultima,
                    "intervalo_km": plano_assoc.intervalo_km
                }

        manutencao_id_ativa = None
        if veiculo.status == 'Em Manutenção':
            manutencao_ativa = Manutencao.query.filter_by(veiculo_id=veiculo.id, status='Em Andamento').first()
            if manutencao_ativa:
                manutencao_id_ativa = manutencao_ativa.id
                
        veiculos_data.append({
            "id": veiculo.id,
            "modelo": veiculo.modelo,
            "placa": veiculo.placa,
            "km_rodados": veiculo.km_rodados or 0,
            "status": veiculo.status,
            "manutencao_id": manutencao_id_ativa,
            "proxima_manutencao": next_maint_summary if progresso_maximo > -1 else None,
            "planos_progresso": sorted(planos_progresso, key=lambda x: x['progresso'], reverse=True)
        })

    manutencoes_em_andamento = Manutencao.query.filter(
        Manutencao.veiculo.has(empresa_id=current_user.empresa_id),
        Manutencao.status.in_(['Em Andamento', 'Agendada'])
    ).all()
    
    hoje = date.today()
    primeiro_dia_mes = hoje.replace(day=1)
    q_kpis = db.session.query(
        func.count(Manutencao.id).label('concluidas_no_mes'),
        func.sum(Manutencao.custo_total).label('custo_mes_atual')
    ).join(Veiculo).filter(
        Veiculo.empresa_id == current_user.empresa_id,
        Manutencao.status == 'Concluída',
        Manutencao.data_saida >= primeiro_dia_mes
    ).first()
    kpis = {
        'concluidas_no_mes': q_kpis.concluidas_no_mes or 0,
        'custo_mes_atual': q_kpis.custo_mes_atual or 0.0,
    }

    alertas_de_estoque = Insumo.query.filter(
        Insumo.empresa_id == current_user.empresa_id,
        Insumo.ponto_ressuprimento.isnot(None),
        Insumo.quantidade_em_estoque <= Insumo.ponto_ressuprimento
    ).all()

    return render_template(
        'oficina.html',
        alertas=alertas,
        manutencoes_em_andamento=manutencoes_em_andamento,
        todos_veiculos=veiculos_data,
        kpis=kpis,
        alertas_de_estoque=alertas_de_estoque
    )

@app.route('/oficina/iniciar', methods=['POST'])
@login_required
def iniciar_manutencao_oficina():
    """ Inicia uma nova manutenção, com automação de kit de peças. """
    veiculo_id = request.form.get('veiculo_id', type=int)
    hodometro = request.form.get('hodometro', type=int)
    status = request.form.get('status')
    plano_id_str = request.form.get('plano_id')
    descricao = request.form.get('descricao')

    veiculo = Veiculo.query.get_or_404(veiculo_id)
    
    nova_manutencao = Manutencao(
        veiculo_id=veiculo.id,
        odometro=hodometro,
        status=status,
        empresa_id=current_user.empresa_id
    )

    if plano_id_str and plano_id_str.isdigit():
        plano_id = int(plano_id_str)
        nova_manutencao.tipo_manutencao = 'Preventiva'
        
        plano_assoc = VeiculoPlano.query.filter_by(veiculo_id=veiculo_id, plano_id=plano_id).first()
        if plano_assoc:
            nova_manutencao.veiculo_plano_veiculo_id = veiculo_id
            nova_manutencao.veiculo_plano_plano_id = plano_id
        
        kit_de_insumos = PlanoInsumo.query.filter_by(plano_id=plano_id).all()
        
        for item_do_kit in kit_de_insumos:
            item_manutencao = ManutencaoItem(
                manutencao=nova_manutencao, 
                data=date.today(),
                descricao=item_do_kit.insumo.descricao,
                quantidade=item_do_kit.quantidade,
                custo_unitario=item_do_kit.insumo.custo_unitario_medio or 0.0
            )
            db.session.add(item_manutencao)
            
    else:
        nova_manutencao.tipo_manutencao = 'Corretiva'
        nova_manutencao.descricao_problema = descricao

    if status == 'Em Andamento':
        veiculo.status = 'Em Manutenção'
    
    veiculo.km_atual = max(veiculo.km_atual or 0, hodometro)

    db.session.add(nova_manutencao)
    db.session.commit()

    flash(f'Manutenção registrada! Status do veículo {veiculo.placa} atualizado.', 'success')
    return redirect(url_for('oficina'))

@app.route('/oficina/finalizar', methods=['POST'])
@login_required
def finalizar_manutencao_oficina():
    """ Finaliza uma manutenção, com automação de baixa de estoque. """
    manutencao_id = request.form.get('manutencao_id', type=int)
    manutencao = Manutencao.query.get_or_404(manutencao_id)
    
    veiculo = manutencao.veiculo
    
    manutencao.status = 'Concluída'
    manutencao.data_saida = datetime.utcnow()
    manutencao.servicos_executados = request.form.get('servicos_executados')
    manutencao.custo_total = request.form.get('custo_total', type=float)
    
    veiculo.status = 'Disponível'
    veiculo.km_atual = max(veiculo.km_atual or 0, manutencao.odometro)

    if manutencao.tipo_manutencao == 'Preventiva' and manutencao.veiculo_plano_associado:
        assoc = manutencao.veiculo_plano_associado
        assoc.km_ultima_manutencao = manutencao.odometro
        assoc.data_ultima_manutencao = manutencao.data_saida.date()

    try:
        itens_utilizados = ManutencaoItem.query.filter_by(manutencao_id=manutencao.id).all()
        for item in itens_utilizados:
            insumo_correspondente = Insumo.query.filter_by(
                descricao=item.descricao,
                empresa_id=current_user.empresa_id
            ).first()
            
            if insumo_correspondente:
                insumo_correspondente.quantidade_em_estoque -= item.quantidade
        
        db.session.commit()
        flash(f'Manutenção finalizada e estoque atualizado! Veículo {veiculo.placa} agora está disponível.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Manutenção finalizada, mas ocorreu um erro ao atualizar o estoque: {e}', 'error')
        logger.error(f"Erro na baixa de estoque para manutenção {manutencao_id}: {e}", exc_info=True)

    return redirect(url_for('oficina'))

@app.route('/api/planos', methods=['GET'])
@login_required
def api_get_todos_planos():
    planos = PlanoDeManutencao.query.order_by(PlanoDeManutencao.descricao).all()
    return jsonify([{
        'id': plano.id,
        'descricao': plano.descricao,
        'intervalo_km_padrao': plano.intervalo_km_padrao,
        'intervalo_meses_padrao': plano.intervalo_meses_padrao
    } for plano in planos])

@app.route('/api/veiculo/<int:veiculo_id>/planos', methods=['GET', 'POST'])
@login_required
def api_gerenciar_planos_veiculo(veiculo_id):
    veiculo = Veiculo.query.get_or_404(veiculo_id)
    if veiculo.empresa_id != current_user.empresa_id:
        return jsonify({'message': 'Acesso negado'}), 403

    if request.method == 'GET':
        planos_atribuidos = [{
            'id': assoc.plano.id,
            'descricao': assoc.plano.descricao,
            'intervalo_km': assoc.intervalo_km,
            'intervalo_meses': assoc.intervalo_meses,
            'km_ultima_manutencao': assoc.km_ultima_manutencao,
            'data_ultima_manutencao': assoc.data_ultima_manutencao.strftime('%Y-%m-%d') if assoc.data_ultima_manutencao else None,
        } for assoc in veiculo.planos_associados]
        return jsonify(planos_atribuidos)

    if request.method == 'POST':
        try:
            data = request.json
            # Deleta as associações antigas para recriá-las com os novos dados
            VeiculoPlano.query.filter_by(veiculo_id=veiculo_id).delete()
            
            for plano_data in data:
                plano = PlanoDeManutencao.query.filter_by(descricao=plano_data['descricao']).first()
                if not plano: # Se o plano não existe, cria um novo
                    plano = PlanoDeManutencao(descricao=plano_data['descricao'])
                    db.session.add(plano)
                    db.session.flush() # Garante que o 'plano.id' esteja disponível

                data_ultima = None
                if data_str := plano_data.get('data_ultima_manutencao'):
                    data_ultima = datetime.strptime(data_str, '%Y-%m-%d').date()

                nova_assoc = VeiculoPlano(
                    veiculo_id=veiculo.id,
                    plano_id=plano.id,
                    intervalo_km=plano_data.get('intervalo_km'),
                    intervalo_meses=plano_data.get('intervalo_meses'),
                    km_ultima_manutencao=plano_data.get('km_ultima_manutencao'), # Salva o valor exato (número ou null)
                    data_ultima_manutencao=data_ultima
                )
                db.session.add(nova_assoc)
                if plano_data.get('atualizar_padrao'):
                    plano.intervalo_km_padrao = int(plano_data['intervalo_km']) if plano_data.get('intervalo_km') else None
                    plano.intervalo_meses_padrao = int(plano_data['intervalo_meses']) if plano_data.get('intervalo_meses') else None
            
            db.session.commit()
            return jsonify({'message': 'Planos atualizados com sucesso!'}), 200
        except Exception as e:
            db.session.rollback()
            logger.error(f"Erro ao salvar planos para veiculo {veiculo_id}: {e}", exc_info=True)
            return jsonify({'message': f'Erro interno do servidor: {e}'}), 500

@app.route('/api/manutencao/<int:manutencao_id>/itens', methods=['GET'])
@login_required
def get_manutencao_itens(manutencao_id):
    manutencao = Manutencao.query.join(Veiculo).filter(
        Manutencao.id == manutencao_id,
        Veiculo.empresa_id == current_user.empresa_id
    ).first_or_404()
    
    itens = [{
        'id': item.id,
        'data': item.data.strftime('%d/%m/%Y'),
        'descricao': item.descricao,
        'quantidade': item.quantidade,
        'custo_unitario': item.custo_unitario,
        'custo_total_item': item.custo_total_item
    } for item in manutencao.itens]
    
    return jsonify(itens)

@app.route('/api/manutencao/<int:manutencao_id>/adicionar_item', methods=['POST'])
@login_required
def add_manutencao_item(manutencao_id):
    manutencao = Manutencao.query.join(Veiculo).filter(
        Manutencao.id == manutencao_id,
        Veiculo.empresa_id == current_user.empresa_id
    ).first_or_404()
    
    data = request.json
    try:
        novo_item = ManutencaoItem(
            manutencao_id=manutencao.id,
            data=datetime.strptime(data['data'], '%Y-%m-%d').date(),
            descricao=data['descricao'],
            quantidade=float(data['quantidade']),
            custo_unitario=float(data['custo_unitario'])
        )
        db.session.add(novo_item)
        db.session.commit()

        custo_atualizado = db.session.query(func.sum(ManutencaoItem.quantidade * ManutencaoItem.custo_unitario)).filter_by(manutencao_id=manutencao.id).scalar()
        manutencao.custo_total = custo_atualizado or 0
        db.session.commit()

        return jsonify({'success': True, 'message': 'Item adicionado com sucesso!'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/veiculo/<int:veiculo_id>/historico')
@login_required
def api_historico_veiculo(veiculo_id):
    query = db.session.query(
        Manutencao,
        PlanoDeManutencao.descricao.label('plano_descricao')
    ).outerjoin(
        VeiculoPlano, and_(
            Manutencao.veiculo_plano_veiculo_id == VeiculoPlano.veiculo_id,
            Manutencao.veiculo_plano_plano_id == VeiculoPlano.plano_id
        )
    ).outerjoin(PlanoDeManutencao).filter(Manutencao.veiculo_id == veiculo_id).order_by(Manutencao.data_entrada.desc())

    historico = [{
        'data_saida': m.data_saida.isoformat() if m.data_saida else None,
        'data_entrada': m.data_entrada.isoformat() if m.data_entrada else None,
        'tipo_manutencao': m.tipo_manutencao,
        'plano_descricao': plano_desc,
        'odometro': m.odometro,
        'servicos_executados': m.servicos_executados,
        'descricao_problema': m.descricao_problema,
        'custo_total': m.custo_total,
        'status': m.status
    } for m, plano_desc in query.all()]
    return jsonify(historico)

@app.route('/api/manutencoes/historico', methods=['GET'])
@login_required
def get_historico_manutencoes():
    query = Manutencao.query.join(Veiculo).filter(
        Veiculo.empresa_id == current_user.empresa_id,
        Manutencao.status == 'Concluída'
    )
    if placa := request.args.get('placa'):
        query = query.filter(Veiculo.placa.ilike(f'%{placa}%'))
    if data_inicio_str := request.args.get('data_inicio'):
        query = query.filter(Manutencao.data_saida >= datetime.strptime(data_inicio_str, '%Y-%m-%d').date())
    if data_fim_str := request.args.get('data_fim'):
        query = query.filter(Manutencao.data_saida <= datetime.strptime(data_fim_str, '%Y-%m-%d').date())
    if tipo := request.args.get('tipo'):
        query = query.filter(Manutencao.tipo_manutencao == tipo)

    manutencoes = query.order_by(Manutencao.data_saida.desc()).all()
    resultado = [{
        'veiculo': f"{m.veiculo.modelo} ({m.veiculo.placa})",
        'data_saida': m.data_saida.strftime('%d/%m/%Y') if m.data_saida else 'N/A',
        'tipo': m.tipo_manutencao,
        'servicos': m.servicos_executados or 'N/A',
        'custo': m.custo_total or 0
    } for m in manutencoes]
    return jsonify(resultado)

@app.route('/oficina/insumos')
@login_required
def gerenciar_insumos_page():
    """ Renderiza a página para o CRUD de Insumos. """
    insumos = Insumo.query.filter_by(empresa_id=current_user.empresa_id).order_by(Insumo.descricao).all()
    return render_template('gerenciar_insumos.html', insumos=insumos)

@app.route('/oficina/api/insumos', methods=['POST'])
@login_required
def api_criar_insumo():
    """ API para criar um novo insumo (versão corrigida). """
    data = request.json
    try:
        custo = float(data.get('custo_unitario_medio')) if data.get('custo_unitario_medio') else 0.0
        qtd_estoque = float(data.get('quantidade_em_estoque')) if data.get('quantidade_em_estoque') else 0.0
        ponto_ressuprimento = float(data.get('ponto_ressuprimento')) if data.get('ponto_ressuprimento') else None

        novo_insumo = Insumo(
            empresa_id=current_user.empresa_id,
            descricao=data['descricao'],
            codigo_peca=data.get('codigo_peca'),
            custo_unitario_medio=custo,
            quantidade_em_estoque=qtd_estoque,
            ponto_ressuprimento=ponto_ressuprimento
        )
        db.session.add(novo_insumo)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Insumo criado com sucesso!'}), 201
    except IntegrityError:
        db.session.rollback()
        return jsonify({'success': False, 'message': 'Já existe um insumo com esta descrição.'}), 409
    except Exception as e:
        db.session.rollback()
        logger.error(f"Erro ao criar insumo: {e}", exc_info=True)
        return jsonify({'success': False, 'message': str(e)}), 500
    
@app.route('/oficina/api/insumos/exportar_excel')
@login_required
def exportar_insumos_excel():
    """Exporta a lista de insumos para um arquivo Excel."""
    try:
        search_query = request.args.get('search', '').strip()

        # Query base filtrando pela empresa do usuário
        query = Insumo.query.filter_by(empresa_id=current_user.empresa_id)

        # Aplica o filtro de busca se houver
        if search_query:
            search_filter = f"%{search_query}%"
            query = query.filter(
                or_(
                    Insumo.descricao.ilike(search_filter),
                    Insumo.codigo_peca.ilike(search_filter)
                )
            )
        
        insumos_para_exportar = query.order_by(Insumo.descricao).all()

        # Cria o arquivo Excel em memória
        output = io.BytesIO()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Catálogo de Insumos"

        # Define os cabeçalhos e aplica estilo
        headers = ["Descrição", "Código da Peça", "Estoque Atual", "Estoque Mínimo", "Custo Unitário Médio (R$)"]
        sheet.append(headers)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill

        # Preenche os dados
        for insumo in insumos_para_exportar:
            sheet.append([
                insumo.descricao,
                insumo.codigo_peca,
                insumo.quantidade_em_estoque,
                insumo.ponto_ressuprimento,
                insumo.custo_unitario_medio
            ])
            # Formata a coluna de custo como moeda
            sheet.cell(row=sheet.max_row, column=5).number_format = 'R$ #,##0.00'

        # Ajusta a largura das colunas
        for col_idx, column_cells in enumerate(sheet.columns, 1):
            sheet.column_dimensions[get_column_letter(col_idx)].width = 25

        workbook.save(output)
        output.seek(0)

        # Envia o arquivo para download
        filename = f"catalogo_insumos_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        logger.error(f"Erro ao exportar insumos para Excel: {e}", exc_info=True)
        flash("Ocorreu um erro ao gerar o relatório Excel.", "error")
        return redirect(url_for('gerenciar_insumos_page'))

@app.route('/oficina/api/insumos/<int:insumo_id>/ajustar_estoque', methods=['POST'])
@login_required
def api_ajustar_estoque_insumo(insumo_id):
    """ API para registrar entradas ou saídas manuais de um insumo no estoque. """
    insumo = Insumo.query.filter_by(id=insumo_id, empresa_id=current_user.empresa_id).first_or_404()
    data = request.json
    try:
        quantidade_ajuste = float(data.get('quantidade', 0))
        if quantidade_ajuste == 0:
            return jsonify({'success': False, 'message': 'A quantidade do ajuste não pode ser zero.'}), 400

        novo_estoque = insumo.quantidade_em_estoque + quantidade_ajuste
        
        # Impede que o estoque fique negativo
        if novo_estoque < 0:
            return jsonify({'success': False, 'message': f'Ajuste inválido. O estoque não pode ficar negativo. Estoque atual: {insumo.quantidade_em_estoque}'}), 400

        insumo.quantidade_em_estoque = novo_estoque
        db.session.commit()
        
        # Mensagem dinâmica baseada no tipo de ajuste
        if quantidade_ajuste > 0:
            message = f'Entrada de {quantidade_ajuste} unidade(s) registrada com sucesso!'
        else:
            message = f'Baixa de {abs(quantidade_ajuste)} unidade(s) registrada com sucesso!'

        return jsonify({
            'success': True, 
            'message': message,
            'novo_estoque': insumo.quantidade_em_estoque 
        }), 200
        
    except (ValueError, TypeError):
        db.session.rollback()
        return jsonify({'success': False, 'message': 'Quantidade inválida. Use apenas números.'}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500
    
@app.route('/oficina/planos_de_manutencao')
@login_required
def gerenciar_planos_page():
    """ Renderiza a página para gerenciar os Planos de Manutenção e seus kits. """
    planos = PlanoDeManutencao.query.order_by(PlanoDeManutencao.descricao).all()
    insumos = Insumo.query.filter_by(empresa_id=current_user.empresa_id).order_by(Insumo.descricao).all()
    return render_template('gerenciar_planos.html', planos=planos, insumos=insumos)

@app.route('/api/planos/<int:plano_id>/insumos', methods=['GET'])
@login_required
def api_get_kit_do_plano(plano_id):
    """ Retorna a lista de insumos (kit) de um plano específico. """
    kit = PlanoInsumo.query.filter_by(plano_id=plano_id).all()
    kit_data = [{
        'insumo_id': item.insumo_id,
        'insumo_descricao': item.insumo.descricao,
        'quantidade': item.quantidade
    } for item in kit]
    return jsonify(kit_data)

@app.route('/api/planos/<int:plano_id>/insumos', methods=['POST'])
@login_required
def api_salvar_kit_do_plano(plano_id):
    """ Salva o kit de insumos E os detalhes de um plano específico (versão corrigida). """
    plano = PlanoDeManutencao.query.get_or_404(plano_id)
    data = request.json
    
    try:
        plano.intervalo_km_padrao = int(data['intervalo_km']) if data.get('intervalo_km') else None
        plano.intervalo_meses_padrao = int(data['intervalo_meses']) if data.get('intervalo_meses') else None

        PlanoInsumo.query.filter_by(plano_id=plano_id).delete()
        
        for item_raw in data.get('kit', []):
            item_data = json.loads(item_raw) if isinstance(item_raw, str) else item_raw
            novo_item_kit = PlanoInsumo(
                plano_id=plano.id,
                insumo_id=int(item_data['insumo_id']),
                quantidade=float(item_data['quantidade'])
            )
            db.session.add(novo_item_kit)
            
        db.session.commit()
        return jsonify({'success': True, 'message': 'Plano e Kit de peças salvos com sucesso!'})
    except Exception as e:
        db.session.rollback()
        logger.error(f"Erro ao salvar kit para o plano {plano_id}: {e}", exc_info=True)
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/oficina/api/insumos/<int:insumo_id>', methods=['DELETE'])
@login_required
def api_deletar_insumo(insumo_id):
    """ API para deletar um insumo (versão corrigida). """
    insumo = Insumo.query.filter_by(id=insumo_id, empresa_id=current_user.empresa_id).first_or_404()
    try:
        db.session.delete(insumo)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Insumo deletado com sucesso!'})
    
    # ### CORREÇÃO APLICADA AQUI ###
    # Capturamos o erro específico de integridade do banco de dados.
    except IntegrityError:
        db.session.rollback()
        return jsonify({
            'success': False, 
            'message': 'Erro: Este insumo não pode ser excluído pois está em uso em um ou mais Kits de Planos.'
        }), 409 # HTTP 409 Conflict é um bom status para este caso.
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Erro ao deletar insumo {insumo_id}: {e}", exc_info=True)
        return jsonify({'success': False, 'message': str(e)}), 500
    
@app.route('/api/fiscal/exportar_documentos', methods=['POST'])
@login_required
@master_required # Garante que apenas usuários com permissão possam exportar
def api_exportar_documentos_fiscais():
    """
    Exporta um relatório em Excel e os arquivos XML correspondentes em um arquivo .zip.
    Recebe uma lista de chaves de acesso para identificar quais notas exportar.
    """
    try:
        data = request.get_json()
        chaves_acesso = data.get('chaves_acesso', [])

        if not chaves_acesso:
            return jsonify({'success': False, 'message': 'Nenhuma nota fiscal foi selecionada para exportação.'}), 400

        # 1. Buscar as notas fiscais no banco de dados
        notas_para_exportar = NFeImportada.query.filter(
            NFeImportada.empresa_id == current_user.empresa_id,
            NFeImportada.chave_acesso.in_(chaves_acesso)
        ).all()

        if not notas_para_exportar:
            return jsonify({'success': False, 'message': 'Nenhuma das notas selecionadas foi encontrada.'}), 404

        # 2. Preparar o arquivo .zip em memória
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            
            # 3. Criar a planilha Excel
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Relatório de Notas Fiscais"
            
            # Cabeçalhos do Excel
            headers = ["Chave de Acesso", "Emitente", "CNPJ", "Data Emissão", "Valor Total", "Status", "Nome Arquivo XML"]
            sheet.append(headers)
            for cell in sheet[1]:
                cell.font = Font(bold=True)
            
            # 4. Iterar sobre as notas, preenchendo o Excel e adicionando XML ao zip
            for nota in notas_para_exportar:
                xml_filename = f"NFe_{nota.chave_acesso}.xml"
                
                # Adicionar linha ao Excel
                sheet.append([
                    nota.chave_acesso,
                    nota.emitente_nome,
                    nota.emitente_cnpj,
                    nota.data_emissao.strftime('%d/%m/%Y'),
                    nota.valor_total,
                    nota.status,
                    xml_filename
                ])

                # Adicionar o arquivo XML ao zip, dentro de uma pasta
                if nota.xml_content:
                    zip_file.writestr(f"XMLs/{xml_filename}", nota.xml_content)

            # 5. Salvar o Excel em memória e adicioná-lo ao zip
            excel_buffer = io.BytesIO()
            workbook.save(excel_buffer)
            excel_buffer.seek(0)
            zip_file.writestr("Relatorio_Notas_Fiscais.xlsx", excel_buffer.read())

        zip_buffer.seek(0)
        
        # 6. Enviar o arquivo .zip como resposta
        filename = f"Exportacao_Contabil_{datetime.now().strftime('%Y-%m-%d')}.zip"
        return send_file(
            zip_buffer,
            mimetype='application/zip',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        logger.error(f"Erro ao exportar documentos fiscais: {e}", exc_info=True)
        return jsonify({'success': False, 'message': f'Ocorreu um erro interno: {str(e)}'}), 500


@app.route('/oficina/previsao')
@login_required
def previsao_custos_manutencao():
    """ Rota para a página de previsão de custos e orçamento de frota. """
    previsoes_detalhadas = []
    custo_mensal = defaultdict(float)
    hoje = date.today()
    
    veiculos = Veiculo.query.filter_by(empresa_id=current_user.empresa_id).all()

    for veiculo in veiculos:
        media_km_dia = calcular_media_km_veiculo(veiculo.id)
        
        for plano_assoc in veiculo.planos_associados:
            if not plano_assoc.intervalo_km or media_km_dia <= 0:
                continue

            km_desde_ultima = (veiculo.km_rodados or 0) - (plano_assoc.km_ultima_manutencao or 0)
            km_restantes = plano_assoc.intervalo_km - km_desde_ultima

            if km_restantes > 0:
                dias_para_manutencao = km_restantes / media_km_dia
                data_prevista = hoje + timedelta(days=dias_para_manutencao)
                
                custo_estimado_plano = db.session.query(
                    func.sum(PlanoInsumo.quantidade * Insumo.custo_unitario_medio)
                ).join(Insumo).filter(PlanoInsumo.plano_id == plano_assoc.plano_id).scalar() or 0.0

                if custo_estimado_plano > 0:
                    previsao = {
                        'veiculo_modelo': veiculo.modelo,
                        'veiculo_placa': veiculo.placa,
                        'plano_descricao': plano_assoc.plano.descricao,
                        'km_restantes': km_restantes,
                        'data_prevista': data_prevista,
                        'custo_estimado': custo_estimado_plano
                    }
                    previsoes_detalhadas.append(previsao)
                    
                    mes_ano = data_prevista.strftime('%Y-%m')
                    custo_mensal[mes_ano] += custo_estimado_plano

    previsoes_detalhadas.sort(key=lambda x: x['data_prevista'])
    
    orcamento_formatado = {
        'labels': sorted(custo_mensal.keys()),
        'data': [custo_mensal[key] for key in sorted(custo_mensal.keys())]
    }
    
    return render_template(
        'previsao_custos.html',
        previsoes=previsoes_detalhadas,
        orcamento=orcamento_formatado,
        today=hoje # Passa a data de hoje para o template
    )

@app.route('/exportar_relatorio')
@login_required
def exportar_relatorio():
    """
    Gera e exporta um relatório detalhado de viagens em formato Excel (XLSX),
    com formatação aprimorada e quebras de custo.
    """
    try:
        # 1. Pega os mesmos filtros da página 'consultar_viagens'
        args = request.args
        query = Viagem.query.filter_by(empresa_id=current_user.empresa_id)

        if data_inicio_str := args.get('data_inicio'):
            query = query.filter(Viagem.data_inicio >= datetime.strptime(data_inicio_str, '%Y-%m-%d'))
        if data_fim_str := args.get('data_fim'):
            data_fim_obj = datetime.strptime(data_fim_str, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(Viagem.data_inicio < data_fim_obj)
        if motorista_id := args.get('motorista_id', type=int):
            query = query.filter(Viagem.motorista_id == motorista_id)
        if status := args.get('status'):
            query = query.filter(Viagem.status == status)
        if search_query := args.get('search'):
             search_term = f'%{search_query}%'
             query = query.outerjoin(Motorista, Viagem.motorista_id == Motorista.id).filter(
                or_(
                    Viagem.cliente.ilike(search_term),
                    Motorista.nome.ilike(search_term),
                    Viagem.veiculo.has(Veiculo.placa.ilike(search_term))
                )
            )

        # 2. Carrega eficientemente todos os dados relacionados para evitar múltiplas consultas
        viagens = query.options(
            db.joinedload(Viagem.motorista_formal),
            db.joinedload(Viagem.veiculo),
            db.joinedload(Viagem.custo_viagem),
            db.joinedload(Viagem.abastecimentos),
            db.joinedload(Viagem.destinos)
        ).order_by(Viagem.data_inicio.desc()).all()

        output = io.BytesIO()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Relatório Detalhado de Viagens"

        # 3. Define os novos cabeçalhos detalhados
        headers = [
            "ID Viagem", "Status", "Data Início", "Hora Início", "Data Fim", "Hora Fim", "Duração (HH:MM:SS)",
            "Cliente", "Motorista", "Veículo (Placa)", "Veículo (Modelo)",
            "Endereço Saída", "Destinos Intermediários", "Endereço Destino Final",
            "Distância Estimada (km)", "Distância Real (km)", "Odômetro Inicial", "Odômetro Final",
            "Receita (R$)", "Custo Real Total (R$)", "Lucro Real (R$)",
            "Custo: Combustível (R$)", "Custo: Pedágios (R$)", "Custo: Alimentação (R$)", "Custo: Hospedagem (R$)", "Custo: Outros (R$)",
            "Forma Pagamento", "Observações"
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 4. Preenche o Excel com os dados detalhados
        for row_num, viagem in enumerate(viagens, 2):
            # Extração de custos detalhados
            custo_combustivel = sum(a.custo_total for a in viagem.abastecimentos)
            custo_pedagios = viagem.custo_viagem.pedagios if viagem.custo_viagem else 0.0
            custo_alimentacao = viagem.custo_viagem.alimentacao if viagem.custo_viagem else 0.0
            custo_hospedagem = viagem.custo_viagem.hospedagem if viagem.custo_viagem else 0.0
            custo_outros = viagem.custo_viagem.outros if viagem.custo_viagem else 0.0
            
            # Formatação da duração
            duracao_fmt = "N/A"
            if viagem.duracao_segundos:
                duracao_fmt = str(timedelta(seconds=viagem.duracao_segundos))

            # Concatenação dos destinos intermediários
            destinos_ordenados = sorted(viagem.destinos, key=lambda d: d.ordem)
            # Exclui o último destino, que já tem sua própria coluna
            destinos_intermediarios = ", ".join([d.endereco for d in destinos_ordenados[:-1]])

            # Preenchimento das células
            sheet.cell(row=row_num, column=1, value=viagem.id)
            sheet.cell(row=row_num, column=2, value=viagem.status.replace('_', ' ').title())
            sheet.cell(row=row_num, column=3, value=viagem.data_inicio.strftime('%d/%m/%Y'))
            sheet.cell(row=row_num, column=4, value=viagem.data_inicio.strftime('%H:%M:%S'))
            sheet.cell(row=row_num, column=5, value=viagem.data_fim.strftime('%d/%m/%Y') if viagem.data_fim else 'N/A')
            sheet.cell(row=row_num, column=6, value=viagem.data_fim.strftime('%H:%M:%S') if viagem.data_fim else 'N/A')
            sheet.cell(row=row_num, column=7, value=duracao_fmt)
            sheet.cell(row=row_num, column=8, value=viagem.cliente)
            sheet.cell(row=row_num, column=9, value=viagem.motorista_formal.nome if viagem.motorista_formal else 'N/A')
            sheet.cell(row=row_num, column=10, value=viagem.veiculo.placa if viagem.veiculo else 'N/A')
            sheet.cell(row=row_num, column=11, value=viagem.veiculo.modelo if viagem.veiculo else 'N/A')
            sheet.cell(row=row_num, column=12, value=viagem.endereco_saida)
            sheet.cell(row=row_num, column=13, value=destinos_intermediarios)
            sheet.cell(row=row_num, column=14, value=viagem.endereco_destino)
            sheet.cell(row=row_num, column=15, value=viagem.distancia_km)
            sheet.cell(row=row_num, column=16, value=viagem.distancia_percorrida)
            sheet.cell(row=row_num, column=17, value=viagem.odometro_inicial)
            sheet.cell(row=row_num, column=18, value=viagem.odometro_final)
            sheet.cell(row=row_num, column=19, value=viagem.valor_recebido or 0).number_format = 'R$ #,##0.00'
            sheet.cell(row=row_num, column=20, value=viagem.custo_real_completo).number_format = 'R$ #,##0.00'
            sheet.cell(row=row_num, column=21, value=viagem.lucro_real).number_format = 'R$ #,##0.00'
            sheet.cell(row=row_num, column=22, value=custo_combustivel or 0).number_format = 'R$ #,##0.00'
            sheet.cell(row=row_num, column=23, value=custo_pedagios or 0).number_format = 'R$ #,##0.00'
            sheet.cell(row=row_num, column=24, value=custo_alimentacao or 0).number_format = 'R$ #,##0.00'
            sheet.cell(row=row_num, column=25, value=custo_hospedagem or 0).number_format = 'R$ #,##0.00'
            sheet.cell(row=row_num, column=26, value=custo_outros or 0).number_format = 'R$ #,##0.00'
            sheet.cell(row=row_num, column=27, value=viagem.forma_pagamento or 'N/A')
            sheet.cell(row=row_num, column=28, value=viagem.observacoes)

        # 5. Auto-ajuste da largura das colunas para melhor visualização
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter # Pega a letra da coluna
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

        workbook.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name=f"relatorio_detalhado_viagens_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        logger.error(f"Erro ao exportar relatório detalhado: {str(e)}", exc_info=True)
        flash('Ocorreu um erro inesperado ao gerar o relatório em Excel.', 'error')
        # Redireciona de volta para a página de consulta, mantendo os filtros
        return redirect(url_for('consultar_viagens', **request.args))
@app.route('/get_active_trip')
@login_required
def get_active_trip():
    viagem = Viagem.query.filter_by(data_fim=None, status='em_andamento').first()
    if viagem:
        horario_chegada = (viagem.data_inicio + timedelta(seconds=viagem.duracao_segundos)).strftime('%d/%m/%Y %H:%M') if viagem.duracao_segundos else 'Não calculado'
        
        motorista_nome = 'N/A'
        if viagem.motorista_id:
            motorista_nome = viagem.motorista_formal.nome if viagem.motorista_formal else 'N/A'
        elif viagem.motorista_cpf_cnpj:
            usuario_com_cpf = Usuario.query.filter_by(cpf_cnpj=viagem.motorista_cpf_cnpj).first()
            if usuario_com_cpf:
                motorista_nome = f"{usuario_com_cpf.nome} {usuario_com_cpf.sobrenome}"
            else:
                motorista_formal_cpf = Motorista.query.filter_by(cpf_cnpj=viagem.motorista_cpf_cnpj).first()
                if motorista_formal_cpf:
                    motorista_nome = motorista_formal_cpf.nome

        trip_data = {
            'trip': {
                'motorista_nome': motorista_nome,
                'veiculo_placa': viagem.veiculo.placa if viagem.veiculo else 'N/A',
                'veiculo_modelo': viagem.veiculo.modelo if viagem.veiculo else 'N/A',
                'endereco_saida': viagem.endereco_saida,
                'endereco_destino': viagem.endereco_destino,
                'horario_chegada': horario_chegada
            }
        }
        return jsonify(trip_data)
    return jsonify({'trip': None})




@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Você saiu do sistema com segurança.', 'success')
    return redirect(url_for('login'))

@app.route('/configuracoes', methods=['GET', 'POST'])
@login_required
def configuracoes():
    if request.method == 'POST':
        nome = request.form.get('nome', '').strip()
        sobrenome = request.form.get('sobrenome', '').strip()
        idioma = request.form.get('idioma', '').strip()

        if not nome or not sobrenome:
            flash('Nome e sobrenome são obrigatórios.', 'error')
            return redirect(url_for('configuracoes'))
        
        if idioma not in ['Português', 'Inglês', 'Espanhol']:
            flash('Idioma inválido.', 'error')
            return redirect(url_for('configuracoes'))

        current_user.nome = nome
        current_user.sobrenome = sobrenome
        current_user.idioma = idioma

        try:
            db.session.commit()
            flash('Configurações pessoais atualizadas com sucesso!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar configurações: {str(e)}', 'error')

        return redirect(url_for('configuracoes'))

    usuarios = []
    empresa = None

    if current_user.empresa_id:
        empresa = db.session.get(Empresa, current_user.empresa_id)

    if current_user.is_admin and current_user.empresa_id:
        usuarios = Usuario.query.filter_by(empresa_id=current_user.empresa_id).all()
    elif current_user.is_admin:
        usuarios = [current_user]

    return render_template('configuracoes.html', usuarios=usuarios, empresa=empresa)


@app.route('/editar_usuario/<int:usuario_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def editar_usuario(usuario_id):
    usuario = Usuario.query.get_or_404(usuario_id)

    if request.method == 'POST':
        nome = request.form.get('nome', '').strip()
        sobrenome = request.form.get('sobrenome', '').strip()
        email = request.form.get('email', '').strip()
        role = request.form.get('role', '').strip()
        senha = request.form.get('senha', '').strip()
        cpf_cnpj = request.form.get('cpf_cnpj', '').strip() # Pega CPF/CNPJ do form

        if not nome or not sobrenome or not email or not role:
            flash('Todos os campos obrigatórios devem ser preenchidos.', 'error')
            return redirect(url_for('editar_usuario', usuario_id=usuario_id))

        if role not in ['Motorista', 'Master', 'Admin']:
            flash('Papel inválido.', 'error')
            return redirect(url_for('editar_usuario', usuario_id=usuario_id))

        if email != usuario.email and Usuario.query.filter_by(email=email).first():
            flash('E-mail já cadastrado.', 'error')
            return redirect(url_for('editar_usuario', usuario_id=usuario_id))
        
        if cpf_cnpj and cpf_cnpj != usuario.cpf_cnpj and Usuario.query.filter_by(cpf_cnpj=cpf_cnpj).first():
            flash('CPF/CNPJ já cadastrado para outro usuário.', 'error')
            return redirect(url_for('editar_usuario', usuario_id=usuario_id))

        usuario.nome = nome
        usuario.sobrenome = sobrenome
        usuario.email = email
        usuario.role = role
        usuario.is_admin = (role == 'Admin')
        usuario.cpf_cnpj = cpf_cnpj if cpf_cnpj else None # Atualiza CPF/CNPJ do usuário

        if senha:
            usuario.set_password(senha)

        try:
            db.session.commit()
            flash('Usuário atualizado com sucesso!', 'success')
            return redirect(url_for('configuracoes'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar usuário: {str(e)}', 'error')
            return redirect(url_for('editar_usuario', usuario_id=usuario_id))

    return render_template('editar_usuario.html', usuario=usuario)




@app.route('/gerenciar_empresa', methods=['GET', 'POST'])
@login_required
def gerenciar_empresa():
    
    if current_user.role not in ['Admin', 'Master']:
        flash('Acesso negado. Apenas administradores podem gerenciar a empresa.', 'error')
        return redirect(url_for('index'))

    empresa = db.session.get(Empresa, current_user.empresa_id) if current_user.empresa_id else None

    if request.method == 'POST':
        cnpj = re.sub(r'\D', '', request.form.get('cnpj', '')) 

        if not validate_cpf_cnpj(cnpj, 'juridica'):
            flash('CNPJ inválido. Deve conter 14 dígitos numéricos.', 'error')
            
            return render_template('gerenciar_empresa.html', empresa=request.form)

        
        empresa_existente = Empresa.query.filter(Empresa.cnpj == cnpj).first()
        if empresa_existente and (not empresa or empresa.id != empresa_existente.id):
            flash('Este CNPJ já está cadastrado em outra empresa.', 'error')
            return render_template('gerenciar_empresa.html', empresa=request.form)

        if empresa:
            # --- LÓGICA DE ATUALIZAÇÃO ---
            empresa.razao_social = request.form.get('razao_social').strip()
            empresa.nome_fantasia = request.form.get('nome_fantasia').strip()
            empresa.cnpj = cnpj
            empresa.inscricao_estadual = request.form.get('inscricao_estadual').strip()
            empresa.endereco = request.form.get('endereco').strip()
            empresa.cidade = request.form.get('cidade').strip()
            empresa.estado = request.form.get('estado').strip().upper()
            empresa.cep = re.sub(r'\D', '', request.form.get('cep', ''))
            empresa.telefone = re.sub(r'\D', '', request.form.get('telefone', ''))
            empresa.email_contato = request.form.get('email_contato').strip()
            flash('Dados da empresa atualizados com sucesso!', 'success')
        else:
            # --- LÓGICA DE CRIAÇÃO ---
            nova_empresa = Empresa(
                razao_social=request.form.get('razao_social').strip(),
                nome_fantasia=request.form.get('nome_fantasia').strip(),
                cnpj=cnpj,
                inscricao_estadual=request.form.get('inscricao_estadual').strip(),
                endereco=request.form.get('endereco').strip(),
                cidade=request.form.get('cidade').strip(),
                estado=request.form.get('estado').strip().upper(),
                cep=re.sub(r'\D', '', request.form.get('cep', '')),
                telefone=re.sub(r'\D', '', request.form.get('telefone', '')),
                email_contato=request.form.get('email_contato').strip()
            )
            db.session.add(nova_empresa)
            db.session.flush() 

            
            current_user.empresa_id = nova_empresa.id
            flash('Empresa cadastrada com sucesso!', 'success')
        
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            flash(f'Ocorreu um erro ao salvar os dados: {e}', 'error')
            return render_template('gerenciar_empresa.html', empresa=request.form)
            
        return redirect(url_for('configuracoes'))

    # --- LÓGICA GET ---
    # Mostra o formulário preenchido para edição ou vazio para criação
    return render_template('gerenciar_empresa.html', empresa=empresa)


@app.route('/excluir_usuario/<int:usuario_id>')
@login_required
@admin_required
def excluir_usuario(usuario_id):
    usuario = Usuario.query.get_or_404(usuario_id)
    if usuario.id == current_user.id:
        flash('Você não pode excluir sua própria conta.', 'error')
        return redirect(url_for('configuracoes'))

    # ▼▼▼ VERIFICAÇÃO ADICIONADA AQUI ▼▼▼
    if Cobranca.query.filter_by(usuario_id=usuario.id).first():
        flash('Erro: Não é possível excluir este usuário, pois ele possui cobranças financeiras registradas.', 'error')
        return redirect(url_for('configuracoes'))
    # ▲▲▲ FIM DA VERIFICAÇÃO ▲▲▲

    try:
        db.session.delete(usuario)
        db.session.commit()
        flash('Usuário excluído com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir usuário: {str(e)}', 'error')
    return redirect(url_for('configuracoes'))



from sqlalchemy import or_

@app.route('/motorista_dashboard')
@login_required
def motorista_dashboard():
    if current_user.role != 'Motorista':
        flash('Acesso negado. Esta página é restrita a motoristas.', 'error')
        return redirect(url_for('index'))

    motorista = Motorista.query.filter_by(
        cpf_cnpj=current_user.cpf_cnpj,
        empresa_id=current_user.empresa_id
    ).first()

    # CORREÇÃO AQUI: Buscamos a viagem primeiro, sem o 'joinedload'
    viagem_ativa = Viagem.query.filter(
        or_(
            Viagem.motorista_id == (motorista.id if motorista else None),
            Viagem.motorista_cpf_cnpj == current_user.cpf_cnpj
        ),
        Viagem.status == 'em_andamento'
    ).first()

    destinos_com_coords = []
    # Só processamos os destinos se a viagem realmente existir
    if viagem_ativa:
        for destino in sorted(viagem_ativa.destinos, key=lambda x: x.ordem):
            lat, lon = get_coordinates(destino.endereco)
            destinos_com_coords.append({
                'endereco': destino.endereco,
                'latitude': lat,
                'longitude': lon
            })

    # A busca pelo histórico permanece a mesma
    viagens_concluidas = Viagem.query.filter(
        or_(
            Viagem.motorista_id == (motorista.id if motorista else None),
            Viagem.motorista_cpf_cnpj == current_user.cpf_cnpj
        ),
        Viagem.status.in_(['concluida', 'cancelada'])
    ).order_by(Viagem.data_inicio.desc()).all()

    return render_template(
        'motorista_dashboard.html',
        viagem_ativa=viagem_ativa,
        viagens=viagens_concluidas,
        destinos_viagem_ativa=destinos_com_coords
    )


@app.route('/atualizar_localizacao', methods=['POST'])
@login_required
def atualizar_localizacao():
    data = request.get_json()
    lat = data.get('latitude')
    lon = data.get('longitude')
    viagem_id = data.get('viagem_id')

    if not lat or not lon:
        return jsonify({'success': False, 'message': 'Coordenadas inválidas.'})

    try:
        endereco = get_address_geoapify(lat, lon)

        # Buscar o motorista formal vinculado ao usuário logado pelo cpf_cnpj
        motorista_formal = Motorista.query.filter_by(cpf_cnpj=current_user.cpf_cnpj).first()
        motorista_id_para_localizacao = motorista_formal.id if motorista_formal else None

        if not motorista_id_para_localizacao:
            logger.warning(f"Usuário {current_user.email} tentou atualizar localização sem motorista formal vinculado por CPF/CNPJ.")
            return jsonify({'success': False, 'message': 'Motorista formal não encontrado para vincular localização.'})


        nova_localizacao = Localizacao(
            motorista_id=motorista_id_para_localizacao,
            viagem_id=viagem_id,
            latitude=lat,
            longitude=lon,
            endereco=endereco
        )
        db.session.add(nova_localizacao)
        db.session.commit()

        return jsonify({'success': True, 'endereco': endereco})
    except Exception as e:
        logger.error(f"Erro ao atualizar localização: {str(e)}", exc_info=True)
        return jsonify({'success': False, 'message': str(e)})


@app.route('/selecionar_viagem/<int:viagem_id>', methods=['POST'])
@login_required
def selecionar_viagem(viagem_id):
    if current_user.role != 'Motorista':
        return jsonify({'success': False, 'message': 'Acesso negado'})
    
    if not current_user.cpf_cnpj:
        return jsonify({'success': False, 'message': 'Seu perfil de usuário não possui CPF/CNPJ. Preencha-o nas configurações para iniciar viagens.'})

    Viagem.query.filter_by(id=viagem_id, empresa_id=current_user.empresa_id).first_or_404()

    if not viagem:
        return jsonify({'success': False, 'message': 'Viagem não encontrada'})

    if viagem.status != 'pendente': # 'Pendente' precisa ser 'pendente' conforme o default do modelo
        return jsonify({'success': False, 'message': 'Viagem já foi iniciada ou está em outro status'})

    viagem.motorista_cpf_cnpj = current_user.cpf_cnpj # Vincula pelo CPF/CNPJ do usuário
    
    # Opcional: Se o usuário logado tiver um motorista formal vinculado, use o ID desse motorista também
    motorista_formal = Motorista.query.filter_by(usuario_id=current_user.id, cpf_cnpj=current_user.cpf_cnpj).first()
    if motorista_formal:
        viagem.motorista_id = motorista_formal.id # Linka com o ID do motorista formal se ele existir

    viagem.status = 'em_andamento' # 'Ativa' precisa ser 'em_andamento' conforme o default do modelo
    viagem.data_inicio = datetime.utcnow()

    db.session.commit()

    return jsonify({'success': True})

@app.route('/viagens_pendentes', methods=['GET'])
@login_required
def viagens_pendentes():
    if current_user.role != 'Motorista':
        return jsonify({'success': False, 'message': 'Acesso restrito a motoristas.'}), 403

    print("\n--- INICIANDO DIAGNÓSTICO DE VIAGENS PENDENTES ---")
    try:
        print(f"[INFO] Buscando para o usuário: {current_user.email} (Empresa ID: {current_user.empresa_id}, CPF: {current_user.cpf_cnpj})")

        # PASSO 1: O sistema tenta encontrar o perfil "Motorista" que corresponde ao "Usuário" logado.
        motorista_formal = Motorista.query.filter_by(
            cpf=current_user.cpf_cnpj, 
            empresa_id=current_user.empresa_id
        ).first()

        if not motorista_formal:
            print("[ERRO CRÍTICO] Nenhum perfil de 'Motorista' encontrado com o CPF/CNPJ e Empresa do usuário logado. A busca não pode continuar.")
            return jsonify({'success': True, 'viagens': []})
        
        motorista_id = motorista_formal.id
        print(f"[INFO] Perfil de motorista formal encontrado: ID={motorista_id}, Nome='{motorista_formal.nome}'")

        # PASSO 2: O sistema agora busca no banco de dados por viagens que cumpram as 3 regras.
        print("[INFO] Executando a query final para encontrar viagens que sejam:")
        print(f"       1. Da empresa ID: {current_user.empresa_id}")
        print(f"       2. Com status: 'pendente'")
        print(f"       3. E para o motorista ID: {motorista_id} (ou CPF/CNPJ: {current_user.cpf_cnpj})")

        viagens_encontradas = Viagem.query.filter(
            Viagem.status == 'pendente',
            Viagem.empresa_id == current_user.empresa_id,
            or_(
                Viagem.motorista_id == motorista_id,
                Viagem.motorista_cpf_cnpj == current_user.cpf_cnpj
            )
        ).all()
        
        if viagens_encontradas:
            print(f"[SUCESSO] Foram encontradas {len(viagens_encontradas)} viagem(ns) pendentes para este motorista.")
        else:
            print("[FALHA] Nenhuma viagem encontrada que cumpra TODOS os três critérios acima.")

        print("--- FIM DO DIAGNÓSTICO ---\n")

        # A lógica para retornar o JSON continua a mesma
        viagens_data = []
        for v in viagens_encontradas:
            lista_de_destinos = [d.endereco for d in sorted(v.destinos, key=lambda x: x.ordem)]
            viagens_data.append({'id': v.id, 'cliente': v.cliente, 'endereco_saida': v.endereco_saida, 'destinos': lista_de_destinos})
        return jsonify({'success': True, 'viagens': viagens_data})
        
    except Exception as e:
        print(f"[ERRO GERAL] Ocorreu uma exceção inesperada: {str(e)}")
        logger.error(f"Erro ao obter viagens pendentes para o motorista {current_user.id}: {str(e)}", exc_info=True)
        return jsonify({'success': False, 'message': 'Erro interno ao processar a solicitação.'}), 500
    
def seed_database(force=False):
    try:
        if force:
            logger.info("Forçando recriação de todas as tabelas do banco de dados...")
            # O jeito correto e seguro de apagar e recriar as tabelas com SQLAlchemy
            db.drop_all()
            db.create_all()
            logger.info("Tabelas recriadas com sucesso.")

        with app.app_context():
            # Se o banco de dados não tiver usuários, ele será populado.
            if Usuario.query.count() == 0:
                logger.info("Iniciando semeação completa do banco de dados...")

                # 1. Criar Empresa de Exemplo
                empresa_exemplo = Empresa(
                    razao_social="TrackGo Logistica LTDA",
                    nome_fantasia="TrackGo",
                    cnpj="11222333000144",
                    inscricao_estadual="123456789",
                    endereco="Rua da Tecnologia, 123",
                    cidade="Curitiba",
                    estado="PR",
                    cep="80000100",
                    telefone="41999998888",
                    email_contato="contato@trackgo.com"
                )
                db.session.add(empresa_exemplo)
                db.session.commit()
                logger.info("Empresa de exemplo criada.")

                # 2. Criar Usuários
                admin = Usuario(
                    nome="João", sobrenome="Admin", email="admin@trackgo.com",
                    role="Admin", is_admin=True, telefone="11987654321",
                    cpf_cnpj="00000000000", empresa_id=empresa_exemplo.id
                )
                admin.set_password("admin123")

                master = Usuario(
                    nome="Maria", sobrenome="Master", email="master@trackgo.com",
                    role="Master", telefone="11987654322",
                    cpf_cnpj="11111111111", empresa_id=empresa_exemplo.id
                )
                master.set_password("master123")

                motorista1_user = Usuario(
                    nome="Carlos", sobrenome="Silva", email="carlos@trackgo.com",
                    role="Motorista", telefone="11987654323",
                    cpf_cnpj="12345678901", empresa_id=empresa_exemplo.id
                )
                motorista1_user.set_password("motorista123")

                db.session.add_all([admin, master, motorista1_user])
                db.session.commit()
                logger.info("Usuários criados.")

                # 3. Criar Clientes
                cliente_exemplo_1 = Cliente(
                    pessoa_tipo="juridica", nome_razao_social="Indústrias ACME S.A.",
                    nome_fantasia="ACME", cpf_cnpj="99888777000166", inscricao_estadual="ISENTO",
                    cep="80230010", logradouro="Avenida Sete de Setembro", numero="3000",
                    bairro="Centro", cidade="Curitiba", estado="PR", email="compras@acme.com",
                    telefone="4133221100", cadastrado_por_id=admin.id,
                    empresa_id=empresa_exemplo.id
                )
                db.session.add(cliente_exemplo_1)
                db.session.commit()
                logger.info("Clientes criados.")

                # 4. Criar Motoristas
                motorista1_db = Motorista(
                    nome="Carlos Silva",
                    data_nascimento=datetime(1985, 5, 15).date(),
                    # Campos de endereço atualizados:
                    logradouro="Rua das Flores",
                    numero="123",
                    cidade="São Paulo",
                    uf="SP",
                    # Campos de documento atualizados:
                    cpf="12345678901",
                    rg="123456789",
                    telefone="11987654323",
                    cnh_numero="98765432101",
                    cnh_data_vencimento=datetime(2026, 12, 31).date(),
                    cnh_categoria="AB", # Adicionado campo obrigatório
                    # IDs de relacionamento:
                    usuario_id=motorista1_user.id,
                    empresa_id=empresa_exemplo.id
                )
                db.session.add(motorista1_db)
                db.session.commit()
                logger.info("Motoristas formais criados.")

                # 5. Criar Veículos
                veiculo1 = Veiculo(
                    placa="ABC1234", categoria="Caminhão", modelo="Volvo FH", ano=2020,
                    empresa_id=empresa_exemplo.id
                )
                db.session.add(veiculo1)
                db.session.commit()
                logger.info("Veículos criados.")

                logger.info("Semeação do banco de dados concluída com sucesso!")
            else:
                logger.info("Banco de dados já contém dados. Semeação não foi executada.")

    except Exception as e:
        db.session.rollback()
        logger.error(f"Erro ao semear o banco de dados: {e}", exc_info=True)
        raise

def get_address_geoapify(lat, lon):
    try:
        url = f'https://api.geoapify.com/v1/geocode/reverse?lat={lat}&lon={lon}&apiKey={GEOAPIFY_API_KEY}'
        response = requests.get(url)
        response.raise_for_status() 
        data = response.json()
        if data['features']:
            return data['features'][0]['properties']['formatted']
    except requests.exceptions.RequestException as e:
        logger.error(f"Erro de rede/API na geocodificação Geoapify: {str(e)}", exc_info=True)
    except Exception as e:
        logger.error(f"Erro inesperado na geocodificação Geoapify: {str(e)}", exc_info=True)
    return "Endereço não encontrado"



@app.route('/ultima_localizacao/<int:viagem_id>', methods=['GET'])
def ultima_localizacao(viagem_id):
    """Retorna a última localização registrada para uma viagem, incluindo coordenadas."""
    localizacao = Localizacao.query.filter_by(viagem_id=viagem_id).order_by(Localizacao.timestamp.desc()).first()
    
    if localizacao:
        return jsonify({
            'success': True, 
            'endereco': localizacao.endereco,
            'latitude': localizacao.latitude,    # <-- LINHA ADICIONADA
            'longitude': localizacao.longitude   # <-- LINHA ADICIONADA
        })
    
    return jsonify({'success': False, 'message': 'Nenhuma localização encontrada para esta viagem.'})


@app.route('/motorista/<int:motorista_id>/perfil')
@login_required
def perfil_motorista(motorista_id):
    # CORREÇÃO: Garante que o admin só veja motoristas da sua empresa.
    motorista = Motorista.query.filter_by(id=motorista_id, empresa_id=current_user.empresa_id).first_or_404()

    viagens = Viagem.query.filter(Viagem.motorista_id == motorista.id)\
        .order_by(Viagem.data_inicio.desc()).all()

    # Lógica de cálculo de estatísticas.
    total_receita = sum(v.valor_recebido or 0 for v in viagens)
    total_custo = sum(v.custo or 0 for v in viagens)

    stats = {
        'total_viagens': len(viagens),
        'total_distancia': round(sum(v.distancia_km or 0 for v in viagens), 2),
        'total_receita': round(total_receita, 2),
        'total_custo': round(total_custo, 2),
        'lucro_total': round(total_receita - total_custo, 2)
    }

    return render_template('perfil_motorista.html', motorista=motorista, viagens=viagens, stats=stats)

@app.route('/romaneio/viagem/<int:viagem_id>', methods=['GET', 'POST'])
@login_required
def gerar_romaneio(viagem_id):
    viagem = Viagem.query.filter_by(id=viagem_id, empresa_id=current_user.empresa_id).first_or_404()
    romaneio = Romaneio.query.filter_by(viagem_id=viagem_id, empresa_id=current_user.empresa_id).first()

    if request.method == 'POST':
        try:
            data_emissao_str = request.form.get('data_emissao')
            observacoes = request.form.get('observacoes')
            data_emissao = datetime.strptime(data_emissao_str, '%Y-%m-%d').date() if data_emissao_str else datetime.utcnow().date()

            if romaneio:
                romaneio.data_emissao = data_emissao
                romaneio.observacoes = observacoes
                ItemRomaneio.query.filter_by(romaneio_id=romaneio.id).delete()
                flash_message = 'Romaneio atualizado com sucesso!'
            else:
                romaneio = Romaneio(
                    viagem_id=viagem.id,
                    data_emissao=data_emissao,
                    observacoes=observacoes,
                    empresa_id=current_user.empresa_id
                )
                db.session.add(romaneio)
                db.session.flush()
                flash_message = 'Romaneio salvo com sucesso!'

            item_counter = 1
            while f'produto_{item_counter}' in request.form:
                produto = request.form.get(f'produto_{item_counter}')
                if produto:
                    item = ItemRomaneio(
                        romaneio_id=romaneio.id,
                        produto_descricao=produto,
                        quantidade=int(request.form.get(f'qtd_{item_counter}', 1)),
                        embalagem=request.form.get(f'embalagem_{item_counter}'),
                        peso_kg=float(request.form.get(f'peso_{item_counter}', 0))
                    )
                    db.session.add(item)
                item_counter += 1
                
            db.session.commit()
            flash(flash_message, 'success')
            return redirect(url_for('gerar_romaneio', viagem_id=viagem_id))
        
        except Exception as e:
            db.session.rollback()
            logger.error(f"Erro ao salvar romaneio para viagem {viagem_id}: {e}", exc_info=True)
            flash(f"Ocorreu um erro inesperado ao salvar o romaneio: {e}", "error")
            return redirect(url_for('gerar_romaneio', viagem_id=viagem_id))

    if romaneio:
        return render_template('cadastro_romaneio.html', viagem=viagem, romaneio=romaneio)
    else:
        motorista_nome = 'N/A'
        if viagem.motorista_formal:
            motorista_nome = viagem.motorista_formal.nome
        elif viagem.motorista_cpf_cnpj:
            usuario = Usuario.query.filter_by(cpf_cnpj=viagem.motorista_cpf_cnpj).first()
            if usuario: motorista_nome = f"{usuario.nome} {usuario.sobrenome}"
        
        dados_novo_romaneio = {
            'dest_nome': viagem.cliente,
            'dest_endereco': viagem.endereco_destino,
            'transportadora': motorista_nome,
            'placa_veiculo': viagem.veiculo.placa
        }
        ultimo_id = db.session.query(db.func.max(Romaneio.id)).scalar() or 0
        
        return render_template(
            'cadastro_romaneio.html', 
            viagem=viagem, 
            romaneio=None,
            dados=dados_novo_romaneio,
            numero_romaneio=ultimo_id + 1
        )
@app.route('/consultar_romaneios')
@login_required
def consultar_romaneios():
    search_query = request.args.get('search', '').strip()
    query = Romaneio.query.join(Viagem)  # Join com Viagem para filtros adicionais
    
    if search_query:
        query = query.filter(
            or_(
                Viagem.cliente.ilike(f'%{search_query}%'),
                Viagem.motorista_formal.has(Motorista.nome.ilike(f'%{search_query}%')),
                Viagem.veiculo.has(Veiculo.placa.ilike(f'%{search_query}%'))
            )
        )
    
    romaneios = query.order_by(Romaneio.data_emissao.desc()).all()
    
    return render_template(
        'consultar_romaneios.html',
        romaneios=romaneios,
        search_query=search_query,
        active_page='consultar_romaneios'
    )

@socketio.on('join_trip_room')
def handle_join_trip_room(data):
    viagem_id = data.get('viagem_id')
    if viagem_id:
        # Sala para o admin/consultas
        join_room(f"trip_{viagem_id}")
        logger.info(f"Cliente {request.sid} entrou na sala de consulta trip_{viagem_id}")
        
        # Sala específica para o motorista da viagem
        join_room(f"driver_{viagem_id}")
        logger.info(f"Cliente {request.sid} entrou na sala do motorista driver_{viagem_id}")

@app.route('/api/cliente/<int:cliente_id>/details')
@login_required
def get_cliente_details_api(cliente_id):
    cliente = Cliente.query.filter_by(id=cliente_id, empresa_id=current_user.empresa_id).first_or_404()
    
    viagens = Viagem.query.filter_by(cliente=cliente.nome_razao_social, empresa_id=current_user.empresa_id).options(
        db.joinedload(Viagem.custo_viagem),
        db.joinedload(Viagem.abastecimentos)
    ).order_by(Viagem.data_inicio.desc()).all()

    total_receita = 0
    total_custo_detalhado = 0
    for v in viagens:
        total_receita += v.valor_recebido or 0
        custo_despesas = (v.custo_viagem.pedagios or 0) + (v.custo_viagem.alimentacao or 0) + (v.custo_viagem.hospedagem or 0) + (v.custo_viagem.outros or 0) if v.custo_viagem else 0
        custo_abastecimento = sum(a.custo_total for a in v.abastecimentos)
        total_custo_detalhado += custo_despesas + custo_abastecimento
    
    stats = {
        'total_viagens': len(viagens),
        'total_receita': round(total_receita, 2),
        'total_custo': round(total_custo_detalhado, 2),
        'lucro_total': round(total_receita - total_custo_detalhado, 2)
    }

    viagens_data = [{
        'id': v.id,
        'data_inicio': v.data_inicio.strftime('%d/%m/%Y'),
        'origem': v.endereco_saida,
        'destino': v.endereco_destino,
        'status': v.status,
        'receita': v.valor_recebido or 0
    } for v in viagens]
    
    cliente_data = {
        'id': cliente.id,
        'nome_razao_social': cliente.nome_razao_social
    }

    return jsonify({
        'success': True,
        'cliente': cliente_data,
        'stats': stats,
        'viagens': viagens_data,
        'anexos': cliente.anexos.split(',') if cliente.anexos else []
    })


@app.route('/veiculo/<int:veiculo_id>/dashboard')
@login_required
def veiculo_dashboard(veiculo_id):
    veiculo = Veiculo.query.filter_by(id=veiculo_id, empresa_id=current_user.empresa_id).first_or_404()
    # PASSE A DATA DE HOJE PARA O TEMPLATE
    return render_template('veiculo_dashboard.html', veiculo=veiculo, hoje=date.today())


@app.template_filter('get_usuario')
def get_usuario(cpf_cnpj):
    return Usuario.query.filter_by(cpf_cnpj=cpf_cnpj).first()

@app.route('/financeiro/folha_pagamento/exportar')
@login_required
@master_required
def exportar_folha_excel():
    mes = request.args.get('mes', type=int)
    ano = request.args.get('ano', type=int)

    if not mes or not ano:
        flash('Mês e Ano são obrigatórios para exportar.', 'error')
        return redirect(url_for('folha_pagamento_dashboard'))

    folhas = FolhaPagamento.query.filter(
        FolhaPagamento.empresa_id == current_user.empresa_id,
        FolhaPagamento.mes_referencia == mes,
        FolhaPagamento.ano_referencia == ano
    ).join(Motorista).order_by(Motorista.nome).all()

    if not folhas:
        flash('Nenhuma folha de pagamento encontrada para este período.', 'warning')
        return redirect(url_for('folha_pagamento_dashboard', mes=mes, ano=ano))

    output = io.BytesIO()
    workbook = Workbook()
    
    # --- Aba 1: Resumo ---
    sheet_resumo = workbook.active
    sheet_resumo.title = "Resumo Folha"
    headers_resumo = ["Motorista", "Salário Base (R$)", "Total Proventos (R$)", "Total Descontos (R$)", "Salário Líquido (R$)"]
    sheet_resumo.append(headers_resumo)
    for cell in sheet_resumo[1]:
        cell.font = Font(bold=True)

    for folha in folhas:
        row = [
            folha.motorista.nome,
            folha.salario_base_registro,
            folha.total_proventos,
            folha.total_descontos,
            folha.salario_liquido
        ]
        sheet_resumo.append(row)

    # --- Aba 2: Detalhado ---
    sheet_detalhado = workbook.create_sheet("Detalhado")
    headers_detalhado = ["ID Folha", "Motorista", "Tipo", "Descrição", "Valor (R$)"]
    sheet_detalhado.append(headers_detalhado)
    for cell in sheet_detalhado[1]:
        cell.font = Font(bold=True)

    for folha in folhas:
        # Adiciona o salário base como um provento
        sheet_detalhado.append([folha.id, folha.motorista.nome, "Provento", "Salário Base", folha.salario_base_registro])
        
        itens_ordenados = sorted(folha.itens.all(), key=lambda item: item.tipo)
        for item in itens_ordenados:
            row = [
                folha.id,
                folha.motorista.nome,
                item.tipo,
                item.descricao,
                item.valor
            ]
            sheet_detalhado.append(row)
    
    # Autoajuste de colunas
    for sheet in workbook.worksheets:
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

    workbook.save(output)
    output.seek(0)
    
    return send_file(
        output,
        as_attachment=True,
        download_name=f"folha_pagamento_{mes:02d}_{ano}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route('/romaneio/<int:romaneio_id>', methods=['GET'])
@login_required
def visualizar_romaneio(romaneio_id):
    # 1. Busca o romaneio garantindo que pertence à empresa do usuário
    romaneio = (
        Romaneio.query
        .join(Viagem)
        .filter(Romaneio.id == romaneio_id, Viagem.empresa_id == current_user.empresa_id)
        .first_or_404()
    )

    # 2. CORREÇÃO DA BUSCA PELA EMPRESA: Usa o método moderno e seguro
    empresa = db.session.get(Empresa, romaneio.viagem.empresa_id)

    # 3. CORREÇÃO DA BUSCA PELO CLIENTE: Busca o objeto Cliente completo para ter acesso ao CNPJ
    cliente_obj = Cliente.query.filter_by(
        nome_razao_social=romaneio.viagem.cliente, 
        empresa_id=current_user.empresa_id
    ).first()

    # 4. CORREÇÃO DA ORDENAÇÃO: Usa a função sorted() do Python para ordenar a lista de destinos
    destinos_ordenados = sorted(romaneio.viagem.destinos, key=lambda d: d.ordem)
    
    # Monta a lista de endereços para o template
    lista_enderecos = [d.endereco for d in destinos_ordenados]

    return render_template(
        'visualizar_romaneio.html',
        romaneio=romaneio,
        empresa=empresa,
        cliente=cliente_obj,  # Passa o objeto cliente completo para o template
        destinos=lista_enderecos, # Passa a lista de endereços já ordenada
        active_page='consultar_romaneios'
    )

import click

@app.cli.command("create-owner")
@click.argument("email")
@click.argument("password")
def create_owner_command(email, password):
    """Cria um novo usuário com o papel de Owner."""
    
    if Usuario.query.filter_by(email=email).first():
        print(f"Erro: O usuário com o e-mail '{email}' já existe.")
        return

    try:
        owner = Usuario(
            nome='Proprietário',
            sobrenome='do Sistema',
            email=email,
            role='Owner',
            is_admin=True # Um Owner também pode ser admin
        )
        owner.set_password(password)
        db.session.add(owner)
        db.session.commit()
        print(f"Usuário Owner '{email}' criado com sucesso!")
    except Exception as e:
        db.session.rollback()
        print(f"Erro ao criar o usuário Owner: {e}")

@socketio.on('leave_trip_room')
def handle_leave_trip_room(data):
    viagem_id = data.get('viagem_id')
    if viagem_id:
        leave_room(f"trip_{viagem_id}")
        logger.info(f"Cliente {request.sid} saiu da sala trip_{viagem_id}")




@socketio.on('atualizar_localizacao_socket')
def handle_atualizar_localizacao_socket(data):
    latitude = data.get('latitude')
    longitude = data.get('longitude')
    viagem_id = data.get('viagem_id')

    if not all([latitude, longitude, viagem_id]):
        return

    try:
        viagem = db.session.get(Viagem, int(viagem_id))
        if not viagem or not viagem.motorista_id:
            return

        current_time = datetime.utcnow()
        trip_last_geocode = last_geocode_time.get(viagem.id)
        
        should_geocode = False
        if trip_last_geocode is None or (current_time - trip_last_geocode).total_seconds() > GEOCODE_INTERVAL_SECONDS:
            should_geocode = True

        endereco = None
        if should_geocode:
            endereco = get_address_geoapify(latitude, longitude)
            last_geocode_time[viagem.id] = current_time # Atualiza o tempo do último geocode
        
        # Salva a localização no banco de dados. Se não geocodificou, o endereço fica nulo.
        nova_localizacao = Localizacao(
            motorista_id=viagem.motorista_id,
            viagem_id=viagem.id,
            latitude=latitude,
            longitude=longitude,
            endereco=endereco # Pode ser None se should_geocode for False
        )
        db.session.add(nova_localizacao)
        db.session.commit()

        # Prepara os dados para enviar via socket
        socket_data = {
            'viagem_id': viagem.id,
            'latitude': latitude,
            'longitude': longitude
        }
        # Só envia o endereço se ele foi realmente buscado nesta chamada
        if endereco:
            socket_data['endereco'] = endereco

        # Emite para a sala da viagem (para o admin em consultar_viagens)
        emit('localizacao_atualizada', socket_data, room=f"trip_{viagem.id}")
        
        # Emite também para a sala do motorista para atualizar o display de endereço
        # Isso garante que o "Iniciando rastreamento..." seja substituído pelo primeiro endereço
        if endereco:
             emit('localizacao_atualizada', {'endereco': endereco, 'viagem_id': viagem.id}, room=f"driver_{viagem.id}")


    except Exception as e:
        logger.error(f"Erro ao salvar localização via socket: {e}", exc_info=True)
        db.session.rollback()

@app.route('/central_documentos')
@login_required
def central_documentos():
    """ Exibe uma central com todos os documentos anexados no sistema. """
    search_query = request.args.get('search', '').strip().lower()
    
    documentos_consolidados = []
    
    # 1. Documentos dos Clientes
    clientes_com_anexos = Cliente.query.filter(
        Cliente.empresa_id == current_user.empresa_id,
        Cliente.anexos.isnot(None),
        Cliente.anexos != ''
    ).all()
    for cliente in clientes_com_anexos:
        urls = cliente.anexos.split(',')
        for url in urls:
            if url: # Garante que não adicione strings vazias
                documentos_consolidados.append({
                    'tipo': 'Cliente',
                    'nome_entidade': cliente.nome_razao_social,
                    'id_entidade': cliente.id,
                    'url': url,
                    'nome_arquivo': url.split('/')[-1]
                })

    # 2. Documentos dos Motoristas
    motoristas_com_anexos = Motorista.query.filter(
        Motorista.empresa_id == current_user.empresa_id,
        Motorista.anexos.isnot(None),
        Motorista.anexos != ''
    ).all()
    for motorista in motoristas_com_anexos:
        urls = motorista.anexos.split(',')
        for url in urls:
            if url:
                documentos_consolidados.append({
                    'tipo': 'Motorista',
                    'nome_entidade': motorista.nome,
                    'id_entidade': motorista.id,
                    'url': url,
                    'nome_arquivo': url.split('/')[-1]
                })

    # 3. Anexos de Custos de Viagem
    custos_com_anexos = CustoViagem.query.join(Viagem).filter(
        Viagem.empresa_id == current_user.empresa_id,
        CustoViagem.anexos.isnot(None),
        CustoViagem.anexos != ''
    ).options(db.joinedload(CustoViagem.viagem).joinedload(Viagem.veiculo)).all()
    for custo in custos_com_anexos:
        urls = custo.anexos.split(',')
        for url in urls:
            if url:
                documentos_consolidados.append({
                    'tipo': 'Viagem',
                    'nome_entidade': f"Viagem #{custo.viagem_id} ({custo.viagem.veiculo.placa if custo.viagem.veiculo else 'N/A'})",
                    'id_entidade': custo.viagem_id,
                    'url': url,
                    'nome_arquivo': url.split('/')[-1]
                })

    # 4. Anexos de Abastecimentos
    abastecimentos_com_anexos = Abastecimento.query.join(Viagem).filter(
        Viagem.empresa_id == current_user.empresa_id,
        Abastecimento.anexo_comprovante.isnot(None),
        Abastecimento.anexo_comprovante != ''
    ).options(db.joinedload(Abastecimento.viagem).joinedload(Viagem.veiculo)).all()
    for abast in abastecimentos_com_anexos:
        url = abast.anexo_comprovante
        if url:
            documentos_consolidados.append({
                'tipo': 'Viagem',
                'nome_entidade': f"Viagem #{abast.viagem_id} ({abast.viagem.veiculo.placa if abast.viagem.veiculo else 'N/A'})",
                'id_entidade': abast.viagem_id,
                'url': url,
                'nome_arquivo': f"ComprovanteAbast_{url.split('/')[-1]}"
            })

    # 5. Anexos de Despesas Diversas de Veículos
    despesas_com_anexos = DespesaVeiculo.query.join(Veiculo).filter(
        Veiculo.empresa_id == current_user.empresa_id,
        DespesaVeiculo.anexos.isnot(None),
        DespesaVeiculo.anexos != ''
    ).options(db.joinedload(DespesaVeiculo.veiculo)).all()
    for despesa in despesas_com_anexos:
        urls = despesa.anexos.split(',')
        for url in urls:
            if url:
                documentos_consolidados.append({
                    'tipo': 'Veículo',
                    'nome_entidade': f"Despesa: {despesa.veiculo.placa} ({despesa.categoria})",
                    'id_entidade': despesa.veiculo_id,
                    'url': url,
                    'nome_arquivo': url.split('/')[-1]
                })

    # --- BLOCO ADICIONADO ---
    # 6. Fotos de Veículos (Anexos do cadastro do veículo)
    veiculos_com_fotos = Veiculo.query.filter(
        Veiculo.empresa_id == current_user.empresa_id,
        Veiculo.fotos_urls.isnot(None),
        Veiculo.fotos_urls != ''
    ).all()
    for veiculo in veiculos_com_fotos:
        urls = veiculo.fotos_urls.split(',')
        for url in urls:
            if url:
                documentos_consolidados.append({
                    'tipo': 'Veículo',
                    'nome_entidade': f"Fotos: {veiculo.placa} ({veiculo.modelo})",
                    'id_entidade': veiculo.id,
                    'url': url,
                    'nome_arquivo': url.split('/')[-1]
                })
    # --- FIM DO BLOCO ADICIONADO ---

    # Filtrar resultados se houver uma busca
    documentos_finais = []
    if search_query:
        for doc in documentos_consolidados:
            if search_query in doc['tipo'].lower() or \
               search_query in doc['nome_entidade'].lower() or \
               search_query in doc['nome_arquivo'].lower():
                documentos_finais.append(doc)
    else:
        documentos_finais = documentos_consolidados

    return render_template('central_documentos.html', 
                           documentos=documentos_finais, 
                           search_query=search_query,
                           active_page='central_documentos')

@app.route('/relatorios/rentabilidade_veiculo')
@login_required
@master_required
def relatorio_rentabilidade_veiculo():
    """ 
    Rota para o novo Dashboard de Rentabilidade de Veículos.
    VERSÃO COM A CORREÇÃO FINAL PARA O ERRO 'eager loading'.
    """
    try:
        # 1. Obter filtros da URL (sem alterações)
        hoje = date.today()
        data_inicio_str = request.args.get('data_inicio', hoje.replace(day=1).strftime('%Y-%m-%d'))
        proximo_mes = hoje.replace(day=28) + timedelta(days=4)
        ultimo_dia_mes = proximo_mes - timedelta(days=proximo_mes.day)
        data_fim_str = request.args.get('data_fim', ultimo_dia_mes.strftime('%Y-%m-%d'))
        veiculo_id_filtro = request.args.get('veiculo_id', type=int)
        data_inicio = datetime.strptime(data_inicio_str, '%Y-%m-%d').date()
        data_fim = datetime.strptime(data_fim_str, '%Y-%m-%d').date()

        # 2. Query de veículos (sem alterações)
        veiculos_query = Veiculo.query.filter(Veiculo.empresa_id == current_user.empresa_id)
        if veiculo_id_filtro:
            veiculos_query = veiculos_query.filter(Veiculo.id == veiculo_id_filtro)
        veiculos_no_filtro = veiculos_query.all()
        veiculo_ids = [v.id for v in veiculos_no_filtro]

        # 3. Estruturas de dados (sem alterações)
        relatorio_custos_detalhado = []
        custos_por_categoria = defaultdict(float)
        receitas_total = 0.0
        km_total = 0.0
        viagens_com_receita = []
        evolucao_mensal = defaultdict(lambda: {'receitas': 0.0, 'custos': 0.0})

        # A. RECEITAS E CUSTOS DIRETOS DE VIAGEM
        viagens_periodo = Viagem.query.filter(
            Viagem.veiculo_id.in_(veiculo_ids),
            Viagem.data_inicio.between(data_inicio, data_fim)
        ).options(
            db.joinedload(Viagem.abastecimentos), 
            db.joinedload(Viagem.custo_viagem),
            db.joinedload(Viagem.veiculo),
            db.joinedload(Viagem.motorista_formal)
        ).order_by(Viagem.data_inicio.asc()).all()

        for viagem in viagens_periodo:
            valor_recebido = float(viagem.valor_recebido or 0.0)
            receitas_total += valor_recebido
            km_total += float(viagem.distancia_percorrida or 0.0)
            if valor_recebido > 0:
                viagens_com_receita.append(viagem)
            
            mes_ano_receita = viagem.data_inicio.strftime('%Y-%m')
            evolucao_mensal[mes_ano_receita]['receitas'] += valor_recebido
            
            for abast in viagem.abastecimentos:
                relatorio_custos_detalhado.append({"data": abast.data_abastecimento.date(), "veiculo_obj": viagem.veiculo, "categoria": "Combustível", "fornecedor_motorista": viagem.motorista_formal.nome if viagem.motorista_formal else "N/A", "documento": f"Viagem #{viagem.id}", "descricao": f"{abast.litros:.2f}L @ R${abast.preco_por_litro:.2f}/L", "valor": float(abast.custo_total or 0), "viagem_obj": viagem})
            if viagem.custo_viagem:
                if viagem.custo_viagem.pedagios: relatorio_custos_detalhado.append({"data": viagem.data_inicio.date(), "veiculo_obj": viagem.veiculo, "categoria": "Pedágio", "fornecedor_motorista": "N/A", "documento": f"Viagem #{viagem.id}", "descricao": "Custo com pedágios", "valor": float(viagem.custo_viagem.pedagios), "viagem_obj": viagem})
                if viagem.custo_viagem.alimentacao: relatorio_custos_detalhado.append({"data": viagem.data_inicio.date(), "veiculo_obj": viagem.veiculo, "categoria": "Alimentação", "fornecedor_motorista": "N/A", "documento": f"Viagem #{viagem.id}", "descricao": "Custo com alimentação", "valor": float(viagem.custo_viagem.alimentacao), "viagem_obj": viagem})

        # B. CUSTOS DE RATEIO 
        rateios_periodo = RateioVeiculo.query.join(LancamentoFluxoCaixa).filter(
            RateioVeiculo.veiculo_id.in_(veiculo_ids), 
            LancamentoFluxoCaixa.data_pagamento.between(data_inicio, data_fim)
        ).options(db.joinedload(RateioVeiculo.veiculo)).all()
        
        for rateio in rateios_periodo:
            relatorio_custos_detalhado.append({
                "data": rateio.lancamento.data_pagamento, 
                "veiculo_obj": rateio.veiculo, 
                "categoria": rateio.lancamento.categoria or "Outros", 
                "fornecedor_motorista": rateio.lancamento.fornecedor_cliente or "N/A", 
                "documento": rateio.lancamento.documento_numero or "N/A", 
                "descricao": rateio.lancamento.descricao, 
                "valor": float(rateio.valor_rateado), 
                "viagem_obj": None
            })

        # C. PAGAMENTOS DA FOLHA
        motoristas_ids = {v.motorista_id for v in viagens_periodo if v.motorista_id}
        if motoristas_ids:
            # ==================== INÍCIO DA CORREÇÃO FINAL ====================
            folhas_pagas = FolhaPagamento.query.filter(
                FolhaPagamento.motorista_id.in_(motoristas_ids), 
                FolhaPagamento.data_pagamento.between(data_inicio, data_fim)
            ).options(db.joinedload(FolhaPagamento.motorista)).all() # Removido o joinedload para 'itens'
            # ===================== FIM DA CORREÇÃO FINAL ======================

            for folha in folhas_pagas:
                for item in folha.itens.filter(ItemFolhaPagamento.tipo=='Provento', ItemFolhaPagamento.viagem_id.isnot(None)):
                    viagem_associada = db.session.get(Viagem, item.viagem_id)
                    if viagem_associada and viagem_associada.veiculo_id in veiculo_ids:
                        relatorio_custos_detalhado.append({"data": folha.data_pagamento.date(), "veiculo_obj": viagem_associada.veiculo, "categoria": "Pagamento Pessoal", "fornecedor_motorista": folha.motorista.nome, "documento": f"Folha {folha.mes_referencia}/{folha.ano_referencia}", "descricao": item.descricao, "valor": float(item.valor), "viagem_obj": viagem_associada})

        # 4. Processamento final e cálculo de KPIs
        for custo in relatorio_custos_detalhado:
            custos_por_categoria[custo['categoria']] += custo['valor']
            mes_ano_custo = custo['data'].strftime('%Y-%m')
            evolucao_mensal[mes_ano_custo]['custos'] += custo['valor']

        custo_total = sum(item['valor'] for item in relatorio_custos_detalhado)
        lucro_liquido = receitas_total - custo_total
        pagamento_motoristas_total = custos_por_categoria.get('Pagamento Pessoal', 0.0)
        
        kpis = {
            "receita_total": receitas_total, "custo_total": custo_total, "pagamento_motoristas": pagamento_motoristas_total, 
            "lucro_liquido": lucro_liquido, "custo_km": (custo_total / km_total) if km_total > 0 else 0.0, 
            "receita_km": (receitas_total / km_total) if km_total > 0 else 0.0, 
            "margem_lucro": (lucro_liquido / receitas_total * 100) if receitas_total > 0 else 0.0
        }

        # 5. Preparar dados para o template
        meses_ordenados = sorted(evolucao_mensal.keys())
        grafico_evolucao = {"labels": [datetime.strptime(mes, '%Y-%m').strftime('%b/%y') for mes in meses_ordenados], "receitas": [evolucao_mensal[mes]['receitas'] for mes in meses_ordenados], "custos": [evolucao_mensal[mes]['custos'] for mes in meses_ordenados]}
        
        for custo in relatorio_custos_detalhado:
             custo['veiculo'] = f"{custo['veiculo_obj'].placa}"
        relatorio_custos_detalhado.sort(key=lambda x: x['data'], reverse=True)

        veiculos_para_filtro = Veiculo.query.filter_by(empresa_id=current_user.empresa_id, is_administrativo=False).order_by(Veiculo.placa).all()
        
        return render_template('relatorios/rentabilidade_veiculo.html', kpis=kpis, receitas=viagens_com_receita, custos=relatorio_custos_detalhado, grafico_custos={"labels": list(custos_por_categoria.keys()), "data": list(custos_por_categoria.values())}, grafico_evolucao=grafico_evolucao, veiculos_filtro=veiculos_para_filtro, filtros={"data_inicio": data_inicio_str, "data_fim": data_fim_str, "veiculo_id": veiculo_id_filtro})
    
    except Exception as e:
        logger.error(f"Erro ao gerar relatório de rentabilidade: {e}", exc_info=True)
        flash(f"Ocorreu um erro inesperado ao gerar o relatório: {e}", "error")
        return redirect(url_for('painel'))

@app.route('/api/viagem/<int:viagem_id>/detalhes_rentabilidade')
@login_required
def api_detalhes_viagem_rentabilidade(viagem_id):
    """API para buscar todos os detalhes de uma viagem para o modal de rentabilidade."""
    try:
        viagem = Viagem.query.filter_by(
            id=viagem_id, empresa_id=current_user.empresa_id
        ).options(
            db.joinedload(Viagem.motorista_formal),
            db.joinedload(Viagem.veiculo),
            db.joinedload(Viagem.custo_viagem),
            db.joinedload(Viagem.abastecimentos),
            db.joinedload(Viagem.destinos)
        ).first_or_404()

        custo_combustivel = sum(a.custo_total for a in viagem.abastecimentos)
        custo_pedagios = viagem.custo_viagem.pedagios if viagem.custo_viagem else 0.0
        custo_alimentacao = viagem.custo_viagem.alimentacao if viagem.custo_viagem else 0.0
        custo_outros = viagem.custo_viagem.outros if viagem.custo_viagem else 0.0
        custo_total_viagem = custo_combustivel + custo_pedagios + custo_alimentacao + custo_outros
        pagamento_motorista = viagem.custo_motorista_variavel or 0.0

        material = "N/A"
        if hasattr(viagem, 'material_transportado') and viagem.material_transportado:
            material = viagem.material_transportado
        elif viagem.observacoes:
            try:
                if 'Material:' in viagem.observacoes.split('|')[0]:
                    material = viagem.observacoes.split('|')[0].replace('Material:', '').strip()
                else:
                    material = viagem.observacoes
            except:
                material = viagem.observacoes

        # <<< A LÓGICA QUE FALTAVA FOI ADICIONADA AQUI >>>
        # Usa a distância percorrida (real) se for maior que zero, senão, usa a distância estimada da rota.
        distancia_a_exibir = viagem.distancia_percorrida if viagem.distancia_percorrida and viagem.distancia_percorrida > 0 else viagem.distancia_km
        
        dados_viagem = {
            'id': viagem.id,
            'cliente': viagem.cliente,
            'data_inicio': viagem.data_inicio.strftime('%d/%m/%Y'),
            'origem': viagem.endereco_saida,
            'destino_final': viagem.endereco_destino,
            'destinos_intermediarios': [d.endereco for d in sorted(viagem.destinos, key=lambda x: x.ordem)[:-1]],
            'distancia_real': f"{distancia_a_exibir or 0.0:.1f} km", # Agora a variável existe
            'peso_total': f"{viagem.peso_toneladas or 0.0:.3f} TN",
            'material': material,
            'motorista': viagem.motorista_formal.nome if viagem.motorista_formal else "N/A",
            'veiculo': f"{viagem.veiculo.placa} - {viagem.veiculo.modelo}" if viagem.veiculo else "N/A",
            'financeiro': {
                'receita_bruta': f"R$ {viagem.valor_recebido or 0.0:,.2f}",
                'pagamento_motorista': f"R$ {pagamento_motorista:,.2f}",
                'custo_combustivel': f"R$ {custo_combustivel:,.2f}",
                'custo_pedagios': f"R$ {custo_pedagios:,.2f}",
                'custo_alimentacao': f"R$ {custo_alimentacao:,.2f}",
                'custo_outros': f"R$ {custo_outros:,.2f}",
                'custo_operacional_total': f"R$ {custo_total_viagem:,.2f}",
                'lucro_bruto_operacional': f"R$ {(viagem.valor_recebido or 0.0) - custo_total_viagem - pagamento_motorista:,.2f}"
            }
        }
        for key, value in dados_viagem['financeiro'].items():
            dados_viagem['financeiro'][key] = value.replace(",", "X").replace(".", ",").replace("X", ".")
            
        return jsonify({'success': True, 'viagem': dados_viagem})
    except Exception as e:
        logger.error(f"Erro ao buscar detalhes da viagem {viagem_id}: {e}", exc_info=True)
        return jsonify({'success': False, 'message': str(e)}), 500
    
@app.route('/api/veiculo/<int:veiculo_id>/despesas_consolidadas')
@login_required
def get_veiculo_despesas_consolidadas(veiculo_id):
    Veiculo.query.filter_by(id=veiculo_id, empresa_id=current_user.empresa_id).first_or_404()
    
    extrato_final = []

    # 1. Buscar Manutenções (Lógica atualizada)
    manutencoes = Manutencao.query.filter_by(veiculo_id=veiculo_id).all()
    for despesa in manutencoes:
        extrato_final.append({
            "data": despesa.data_entrada.isoformat(),
            "categoria": f"Manutenção ({despesa.tipo_manutencao})",
            "descricao": despesa.servicos_executados or despesa.descricao_problema or 'N/A',
            "valor": despesa.custo_total,
            "anexos": []
        })

    # 2. Buscar Abastecimentos (Lógica existente)
    abastecimentos = Abastecimento.query.filter_by(veiculo_id=veiculo_id).all()
    for despesa in abastecimentos:
        extrato_final.append({
            "data": despesa.data_abastecimento.isoformat(),
            "categoria": "Combustível",
            "descricao": f"{despesa.litros:.2f}L @ R$ {despesa.preco_por_litro:.2f}/L (Odôm: {despesa.odometro}km)",
            "valor": despesa.custo_total,
            "anexos": [despesa.anexo_comprovante] if despesa.anexo_comprovante else []
        })

    # 3. Buscar Despesas de Viagem (Lógica existente)
    viagens_do_veiculo = Viagem.query.filter_by(veiculo_id=veiculo_id).options(db.joinedload(Viagem.custo_viagem)).all()
    for viagem in viagens_do_veiculo:
        if viagem.custo_viagem:
            custo = viagem.custo_viagem
            anexos_custo = custo.anexos.split(',') if custo.anexos else []
            if custo.pedagios and custo.pedagios > 0:
                extrato_final.append({"data": viagem.data_inicio.isoformat(), "categoria": "Pedágio", "descricao": f"Ref. Viagem #{viagem.id}", "valor": custo.pedagios, "anexos": anexos_custo})
            if custo.alimentacao and custo.alimentacao > 0:
                extrato_final.append({"data": viagem.data_inicio.isoformat(), "categoria": "Alimentação", "descricao": f"Ref. Viagem #{viagem.id}", "valor": custo.alimentacao, "anexos": anexos_custo})
            if custo.hospedagem and custo.hospedagem > 0:
                extrato_final.append({"data": viagem.data_inicio.isoformat(), "categoria": "Hospedagem", "descricao": f"Ref. Viagem #{viagem.id}", "valor": custo.hospedagem, "anexos": anexos_custo})
            if custo.outros and custo.outros > 0:
                extrato_final.append({"data": viagem.data_inicio.isoformat(), "categoria": "Outras Desp. (Viagem)", "descricao": custo.descricao_outros, "valor": custo.outros, "anexos": anexos_custo})

    # 4. Buscar Despesas Diversas (Lógica existente)
    despesas_diversas = DespesaVeiculo.query.filter_by(veiculo_id=veiculo_id).all()
    for despesa in despesas_diversas:
         extrato_final.append({
            "data": despesa.data.isoformat(),
            "categoria": despesa.categoria,
            "descricao": despesa.descricao,
            "valor": despesa.valor,
            "anexos": despesa.anexos.split(',') if despesa.anexos else []
        })

    extrato_ordenado = sorted(extrato_final, key=lambda d: d['data'], reverse=True)
    return jsonify(success=True, extrato=extrato_ordenado)

@app.route('/api/veiculo/<int:veiculo_id>/details')
@login_required
def get_veiculo_details_api(veiculo_id):
    veiculo = Veiculo.query.filter_by(id=veiculo_id, empresa_id=current_user.empresa_id).first_or_404()
    
    viagens = Viagem.query.filter_by(veiculo_id=veiculo.id).options(
        db.joinedload(Viagem.custo_viagem),
        db.joinedload(Viagem.abastecimentos)
    ).all()
    
    manutencoes = Manutencao.query.filter_by(veiculo_id=veiculo.id).order_by(Manutencao.data_entrada.desc()).all()
    abastecimentos = Abastecimento.query.filter_by(veiculo_id=veiculo.id).order_by(Abastecimento.data_abastecimento.desc()).all()

    total_km = veiculo.km_rodados or 0.0
    
    
    total_custo_viagens = 0
    for v in viagens:
        custo_despesas = 0
        if v.custo_viagem:
            custo_despesas = (v.custo_viagem.pedagios or 0) + (v.custo_viagem.alimentacao or 0) + (v.custo_viagem.hospedagem or 0) + (v.custo_viagem.outros or 0)
        custo_abastecimento_viagem = sum(a.custo_total for a in v.abastecimentos)
        total_custo_viagens += custo_despesas + custo_abastecimento_viagem

    total_custo_manutencao = sum(m.custo_total or 0 for m in manutencoes)
    
    custo_geral = total_custo_viagens + total_custo_manutencao
    custo_por_km = (custo_geral / total_km) if total_km > 0 else 0
    
    kpis = {
        "total_viagens": len(viagens),
        "total_km": round(total_km, 2),
        "custo_geral": round(custo_geral, 2),
        "custo_por_km": round(custo_por_km, 2)
    }

    manutencoes_data = [{
        "id": m.id,
        "data": m.data_entrada.strftime('%d/%m/%Y'),
        "odometro": m.odometro,
        "tipo": m.tipo_manutencao,
        "descricao": m.servicos_executados or m.descricao_problema or 'N/A',
        "custo": m.custo_total
    } for m in manutencoes]

    return jsonify({
        "success": True,
        "kpis": kpis,
        "manutencoes": manutencoes_data,
        "abastecimentos": [a.to_dict() for a in abastecimentos]
    })

@app.route('/veiculo/<int:veiculo_id>/lancar_receita', methods=['POST'])
@login_required
def lancar_receita_veiculo(veiculo_id):
    veiculo = Veiculo.query.filter_by(id=veiculo_id, empresa_id=current_user.empresa_id).first_or_404()
    
    try:
        # Cada frete cria um registro de viagem, mesmo que simplificado
        nova_viagem_receita = Viagem(
            veiculo_id=veiculo.id,
            empresa_id=current_user.empresa_id,
            # Usamos o campo 'cliente' para a Fazenda/Empresa do frete
            cliente=request.form.get('cliente'), 
            valor_recebido=float(request.form.get('valor_frete')),
            data_inicio=datetime.strptime(request.form.get('data'), '%Y-%m-%d'),
            # Usamos 'observacoes' para guardar detalhes como material, peso, etc.
            observacoes=f"Material: {request.form.get('material', '')} | Peso: {request.form.get('peso', '')} TN",
            # Preenchemos campos obrigatórios com valores padrão
            endereco_saida='N/A',
            endereco_destino='N/A',
            status='concluida' # Marcamos como concluída, pois é apenas um registro de faturamento
        )
        
        db.session.add(nova_viagem_receita)
        db.session.commit()
        flash('Receita (frete) lançada com sucesso!', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao lançar receita: {e}', 'error')
        logger.error(f"Erro ao lançar receita para o veículo {veiculo_id}: {e}", exc_info=True)

    return redirect(url_for('veiculo_dashboard', veiculo_id=veiculo_id))

@app.route('/manifest.json')
def manifest():
    return send_from_directory('templates', 'manifest.json')

@app.route('/reset-password-tool/<secret_key>', methods=['GET', 'POST'])
def secret_password_reset(secret_key):
    # CHAVE DE SEGURANÇA: Mude isso para qualquer coisa que só você saiba
    # e não compartilhe com ninguém.
    SUPER_SECRET_KEY = "trocar_para_uma_chave_muito_secreta_12345"

    if secret_key != SUPER_SECRET_KEY:
        return "Acesso Negado.", 403

    if request.method == 'POST':
        email = request.form.get('email')
        new_password = request.form.get('new_password')
        
        user = Usuario.query.filter_by(email=email).first()
        
        if not user:
            flash(f"Usuário com e-mail '{email}' não encontrado.", 'error')
            return redirect(url_for('secret_password_reset', secret_key=secret_key))

        try:
            user.set_password(new_password)
            db.session.commit()
            flash(f"Senha para '{email}' foi redefinida com sucesso!", 'success')
        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao redefinir a senha: {e}", 'error')

        return redirect(url_for('secret_password_reset', secret_key=secret_key))

    return render_template('reset_password_page.html')

@app.route('/financeiro/fluxo_caixa')
@login_required
@master_required
def fluxo_caixa():
    """Página principal do fluxo de caixa, com filtros avançados e consolidação de dados."""
    hoje = date.today()
    
    # Coleta de todos os filtros da URL, incluindo o novo filtro de unidade de negócio
    data_inicio = request.args.get('data_inicio', hoje.strftime('%Y-%m-%d'))
    data_fim = request.args.get('data_fim', (hoje + timedelta(days=30)).strftime('%Y-%m-%d'))
    data_emissao_inicio = request.args.get('data_emissao_inicio', '')
    data_emissao_fim = request.args.get('data_emissao_fim', '')
    categoria_filtro = request.args.get('categoria', '')
    status_filtro = request.args.get('status', '')
    meio_pagamento_filtro = request.args.get('meio_pagamento', '')
    tipo_filtro = request.args.get('tipo', '')
    # --- NOVO FILTRO ADICIONADO AQUI ---
    unidade_negocio_filtro_id = request.args.get('unidade_negocio_id', type=int)
    
    try:
        data_inicio_obj = datetime.strptime(data_inicio, '%Y-%m-%d').date()
        data_fim_obj = datetime.strptime(data_fim, '%Y-%m-%d').date()
    except ValueError:
        data_inicio_obj = hoje
        data_fim_obj = hoje + timedelta(days=30)

    # A query base já filtra pela empresa e pelo período de vencimento
    query = LancamentoFluxoCaixa.query.filter(
        LancamentoFluxoCaixa.empresa_id == current_user.empresa_id,
        LancamentoFluxoCaixa.data_vencimento.between(data_inicio_obj, data_fim_obj)
    )

    # Aplicação centralizada de todos os filtros
    if data_emissao_inicio:
        query = query.filter(LancamentoFluxoCaixa.data_lancamento >= datetime.strptime(data_emissao_inicio, '%Y-%m-%d'))
    if data_emissao_fim:
        query = query.filter(LancamentoFluxoCaixa.data_lancamento <= datetime.strptime(data_emissao_fim, '%Y-%m-%d'))
    if categoria_filtro:
        query = query.filter(LancamentoFluxoCaixa.categoria.ilike(f'%{categoria_filtro}%'))
    if status_filtro:
        status_pagos = ['PAGO', 'Pago']
        if status_filtro == 'PAGO':
            query = query.filter(LancamentoFluxoCaixa.status_pagamento.in_(status_pagos))
        else: # PENDENTE, A Pagar, VENCIDO
            query = query.filter(LancamentoFluxoCaixa.status_pagamento.notin_(status_pagos))
    if tipo_filtro:
        query = query.filter(LancamentoFluxoCaixa.tipo == tipo_filtro)
    if meio_pagamento_filtro:
        query = query.filter(LancamentoFluxoCaixa.meio_pagamento == meio_pagamento_filtro)
    
    # --- APLICAÇÃO DO NOVO FILTRO NA QUERY ---
    if unidade_negocio_filtro_id:
        query = query.filter(LancamentoFluxoCaixa.unidade_negocio_id == unidade_negocio_filtro_id)

    lancamentos_filtrados = query.all()

    # Consolida os lançamentos (sua função existente)
    fluxo_consolidado = consolidar_fluxo_caixa_unificado(lancamentos_filtrados)
    
    # Calcula totais e saldos (sua função existente)
    totais = calcular_totais_fluxo(fluxo_consolidado)
    
    # Busca categorias para o dropdown de filtro (lógica existente)
    categorias = sorted(set([
        cat[0] for cat in db.session.query(LancamentoFluxoCaixa.categoria).filter(
            LancamentoFluxoCaixa.empresa_id == current_user.empresa_id,
            LancamentoFluxoCaixa.categoria.isnot(None),
            LancamentoFluxoCaixa.categoria != ''
        ).distinct().all()
    ] + ['Fornecedores (NFe)']))

    # --- BUSCA A LISTA DE UNIDADES PARA ENVIAR AO TEMPLATE ---
    unidades_negocio = UnidadeNegocio.query.filter_by(empresa_id=current_user.empresa_id).order_by(UnidadeNegocio.nome).all()
    
    return render_template('fluxo_caixa.html',
                           fluxo_consolidado=fluxo_consolidado,
                           totais=totais,
                           categorias=categorias,
                           data_inicio=data_inicio,
                           data_fim=data_fim,
                           data_emissao_inicio=data_emissao_inicio,
                           data_emissao_fim=data_emissao_fim,
                           categoria_filtro=categoria_filtro,
                           status_filtro=status_filtro,
                           meio_pagamento_filtro=meio_pagamento_filtro,
                           tipo_filtro=tipo_filtro,
                           hoje=hoje,
                           # --- NOVAS VARIÁVEIS ENVIADAS PARA O TEMPLATE ---
                           unidades_negocio=unidades_negocio,
                           unidade_negocio_filtro_id=unidade_negocio_filtro_id)

def consolidar_fluxo_caixa_unificado(lancamentos):
    """Consolida lançamentos de uma única lista em uma estrutura unificada."""
    fluxo = []

    def formatar_brl(valor):
        if valor is None: return "0,00"
        return f"{valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

    for lancamento in lancamentos:
        # Determina a origem (NFE ou MANUAL) com base no documento
        is_nfe = lancamento.documento_numero and lancamento.documento_numero.startswith('NFE-')
        
        fluxo.append({
            'id': f"manual_{lancamento.id}", # O ID é sempre do LancamentoFluxoCaixa
            'tipo_origem': 'NFE' if is_nfe else 'MANUAL',
            'tipo_movimento': lancamento.tipo,
            'data_vencimento': lancamento.data_vencimento,
            'data_emissao': lancamento.data_lancamento,
            'data_pagamento': lancamento.data_pagamento,
            'descricao': lancamento.descricao,
            'categoria': lancamento.categoria or 'Sem categoria',
            'fornecedor_cliente': lancamento.fornecedor_cliente,
            'documento': lancamento.documento_numero,
            'valor': lancamento.valor_total,
            'valor_formatado': formatar_brl(lancamento.valor_total),
            'status': lancamento.status_pagamento,
            'meio_pagamento': getattr(lancamento, 'meio_pagamento', None),
            'parcela': f"{lancamento.parcela_numero}/{lancamento.parcela_total}" if lancamento.parcela_total and lancamento.parcela_total > 1 else None,
            'observacoes': lancamento.observacoes,
            'is_vencido': (lancamento.data_vencimento < date.today()) and lancamento.status_pagamento in ['PENDENTE', 'A Pagar'],
            'tem_rateio': lancamento.tem_rateio,
            'objeto': lancamento
        })

    fluxo.sort(key=lambda x: x.get('data_vencimento') or date.max)
    return fluxo

def calcular_totais_fluxo(fluxo_consolidado):
    """Calcula totais, saldos e estatísticas do fluxo"""
    hoje = date.today()
    
    total_receitas = 0
    total_despesas = 0
    total_receitas_pagas = 0
    total_despesas_pagas = 0
    total_vencido_receitas = 0
    total_vencido_despesas = 0
    
    pendente_hoje = 0
    vencido_total = 0
    
    for item in fluxo_consolidado:
        valor = item['valor']
        
        if item['tipo_movimento'] == 'RECEITA':
            total_receitas += valor
            if item['status'] in ['PAGO', 'Paga']:
                total_receitas_pagas += valor
            elif item['is_vencido']:
                total_vencido_receitas += valor
        else:  # DESPESA
            total_despesas += valor
            if item['status'] in ['PAGO', 'Paga']:
                total_despesas_pagas += valor
            elif item['is_vencido']:
                total_vencido_despesas += valor
        
        if item['data_vencimento'] == hoje and item['status'] in ['PENDENTE', 'A Pagar']:
            if item['tipo_movimento'] == 'RECEITA':
                pendente_hoje += valor
            else:
                pendente_hoje -= valor
        
        if item['is_vencido']:
            if item['tipo_movimento'] == 'RECEITA':
                vencido_total -= valor # Vencido a receber é positivo para o saldo
            else:
                vencido_total += valor # Vencido a pagar é negativo para o saldo
    
    saldo_previsto = total_receitas - total_despesas
    saldo_realizado = total_receitas_pagas - total_despesas_pagas
    saldo_pendente = (total_receitas - total_receitas_pagas) - (total_despesas - total_despesas_pagas)
    
    # Função auxiliar para formatar
    def formatar_brl(valor):
        if valor is None:
            return "0,00"
        return f"{valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

    return {
        'total_receitas': total_receitas,
        'total_despesas': total_despesas,
        'total_receitas_pagas': total_receitas_pagas,
        'total_despesas_pagas': total_despesas_pagas,
        'saldo_previsto': saldo_previsto,
        'saldo_realizado': saldo_realizado,
        'saldo_pendente': saldo_pendente,
        'total_vencido_receitas': total_vencido_receitas,
        'total_vencido_despesas': total_vencido_despesas,
        'pendente_hoje': pendente_hoje,
        'vencido_total': vencido_total,
        # --- CHAVES FORMATADAS ADICIONADAS AQUI ---
        'total_receitas_formatado': formatar_brl(total_receitas),
        'total_despesas_formatado': formatar_brl(total_despesas),
        'total_receitas_pagas_formatado': formatar_brl(total_receitas_pagas),
        'total_despesas_pagas_formatado': formatar_brl(total_despesas_pagas),
        'saldo_previsto_formatado': formatar_brl(saldo_previsto),
        'saldo_realizado_formatado': formatar_brl(saldo_realizado),
        'vencido_total_formatado': formatar_brl(abs(vencido_total)) # Mostra o valor absoluto dos vencidos
    }


@app.route('/financeiro/fluxo_caixa/novo_lancamento', methods=['GET', 'POST'])
@login_required
@master_required
def novo_lancamento_fluxo():
    """Criar novo lançamento manual no fluxo de caixa"""
    if request.method == 'POST':
        try:
            # Dados básicos
            tipo = request.form.get('tipo')
            descricao = request.form.get('descricao')
            categoria = request.form.get('categoria')
            valor_total = float(request.form.get('valor_total', 0))
            data_vencimento = datetime.strptime(request.form.get('data_vencimento'), '%Y-%m-%d').date()
            
            # Dados opcionais
            fornecedor_cliente = request.form.get('fornecedor_cliente')
            documento_numero = request.form.get('documento_numero')
            observacoes = request.form.get('observacoes')
            
            # Controle de parcelas
            parcela_total = int(request.form.get('parcela_total', 1))
            
            if parcela_total == 1:
                # Lançamento único
                novo_lancamento = LancamentoFluxoCaixa(
                    empresa_id=current_user.empresa_id,
                    tipo=tipo,
                    descricao=descricao,
                    categoria=categoria,
                    valor_total=valor_total,
                    data_vencimento=data_vencimento,
                    fornecedor_cliente=fornecedor_cliente,
                    documento_numero=documento_numero,
                    observacoes=observacoes,
                    parcela_numero=1,
                    parcela_total=1
                )
                db.session.add(novo_lancamento)
                
            else:
                # Criar lançamento pai
                lancamento_pai = LancamentoFluxoCaixa(
                    empresa_id=current_user.empresa_id,
                    tipo=tipo,
                    descricao=f"{descricao} (Parcelado)",
                    categoria=categoria,
                    valor_total=valor_total,
                    data_vencimento=data_vencimento,
                    fornecedor_cliente=fornecedor_cliente,
                    documento_numero=documento_numero,
                    observacoes=observacoes,
                    parcela_numero=0,  # Lançamento pai não conta como parcela
                    parcela_total=parcela_total,
                    status_pagamento='PARCELADO'  # Status especial para o pai
                )
                db.session.add(lancamento_pai)
                db.session.flush()  # Para obter o ID
                
                # Criar parcelas
                valor_parcela = valor_total / parcela_total
                for i in range(1, parcela_total + 1):
                    data_parcela = data_vencimento + timedelta(days=(i-1)*30)  # Mensais por padrão
                    
                    parcela = LancamentoFluxoCaixa(
                        empresa_id=current_user.empresa_id,
                        tipo=tipo,
                        descricao=f"{descricao} - Parcela {i}/{parcela_total}",
                        categoria=categoria,
                        valor_total=valor_parcela,
                        data_vencimento=data_parcela,
                        fornecedor_cliente=fornecedor_cliente,
                        documento_numero=f"{documento_numero}-{i:02d}" if documento_numero else None,
                        observacoes=observacoes,
                        parcela_numero=i,
                        parcela_total=parcela_total,
                        lancamento_pai_id=lancamento_pai.id
                    )
                    db.session.add(parcela)
            
            db.session.commit()
            flash(f'Lançamento criado com sucesso! {parcela_total} parcela(s) gerada(s).', 'success')
            return redirect(url_for('fluxo_caixa'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao criar lançamento: {e}', 'error')
            logger.error(f"Erro ao criar lançamento manual: {e}", exc_info=True)
            
    return render_template('novo_lancamento_fluxo.html')

# Em app.py

@app.route('/api/fluxo_caixa/marcar_pago', methods=['POST'])
@login_required
def api_marcar_pago_fluxo():
    """API para marcar lançamento como pago"""
    data = request.get_json()
    
    # --- INÍCIO DA CORREÇÃO ---
    # O nome do campo enviado pelo JavaScript é 'itemId' (camelCase).
    item_id = data.get('itemId') 
    # --- FIM DA CORREÇÃO ---
    
    data_pagamento_str = data.get('data_pagamento')
    meio_pagamento = data.get('meio_pagamento')

    if not item_id:
        return jsonify({'success': False, 'message': 'ID do item é obrigatório'}), 400
    
    if not meio_pagamento:
        return jsonify({'success': False, 'message': 'Meio de pagamento é obrigatório.'}), 400

    try:
        data_pagamento = datetime.strptime(data_pagamento_str, '%Y-%m-%d').date() if data_pagamento_str else date.today()
        
        if item_id.startswith('manual_'):
            lancamento_id = int(item_id.replace('manual_', ''))
            lancamento = LancamentoFluxoCaixa.query.filter_by(
                id=lancamento_id, 
                empresa_id=current_user.empresa_id
            ).first_or_404()
            
            lancamento.status_pagamento = 'PAGO'
            lancamento.data_pagamento = data_pagamento
            lancamento.meio_pagamento = meio_pagamento

        elif item_id.startswith('nfe_'):
            lancamento_id = int(item_id.replace('nfe_', ''))
            lancamento = LancamentoNotaFiscal.query.filter_by(
                id=lancamento_id, 
                empresa_id=current_user.empresa_id
            ).first_or_404()
            
            lancamento.status_pagamento = 'Pago'
            lancamento.data_pagamento = data_pagamento
            
        else:
            return jsonify({'success': False, 'message': 'Formato de ID inválido'}), 400
        
        db.session.commit()
        return jsonify({'success': True, 'message': 'Lançamento marcado como pago!'})
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Erro ao marcar como pago: {e}", exc_info=True)
        return jsonify({'success': False, 'message': str(e)}), 500

# Rota para servir o sw.js a partir da raiz do projeto
@app.route('/sw.js')
def service_worker():
    response = make_response(send_from_directory('.', 'sw.js'))
    response.headers['Content-Type'] = 'application/javascript'
    return response



@app.route('/fix-missing-tokens/f9a7b3c8d2e1f4a0b9c8d7e6f5a4b3c2')
def fix_old_trip_tokens():
    """
    Esta rota é uma correção única para gerar tokens para viagens antigas.
    DEVE SER REMOVIDA APÓS O PRIMEIRO USO EM PRODUÇÃO.
    """
    try:
        viagens_sem_token = Viagem.query.filter(Viagem.public_tracking_token.is_(None)).all()
        
        if not viagens_sem_token:
            return "<h1>Tudo certo!</h1><p>Nenhuma viagem precisava de correção.</p>", 200

        count = len(viagens_sem_token)
        for viagem in viagens_sem_token:
            viagem.public_tracking_token = str(uuid.uuid4())
        
        db.session.commit()
        
        return f"<h1>Correção Concluída!</h1><p>{count} viagem(ns) foram atualizadas com sucesso.</p>", 200

    except Exception as e:
        db.session.rollback()
        logger.error(f"Erro ao executar a correção de tokens: {e}")
        return f"<h1>Erro ao executar a correção.</h1><p>Detalhes: {e}</p>", 500
    

def parse_nfe_xml_content(xml_content_str):
    """
    Extrai dados essenciais de uma string de conteúdo XML de NF-e.
    Retorna um dicionário com os dados ou None se houver erro.
    """
    try:
        if isinstance(xml_content_str, bytes):
            xml_content_str = xml_content_str.decode('utf-8')
            
        root = ET.fromstring(xml_content_str)
        ns = {'nfe': 'http://www.portalfiscal.inf.br/nfe'}

        infNFe = root.find('.//nfe:infNFe', ns)
        if infNFe is None: return None
        chave_acesso = infNFe.get('Id').replace('NFe', '')
        
        emit = root.find('.//nfe:emit', ns)
        if emit is None: return None # Adiciona uma verificação de segurança

        # --- INÍCIO DA CORREÇÃO ---
        # Procura por CNPJ, se não achar, procura por CPF.
        emit_cnpj_node = emit.find('nfe:CNPJ', ns)
        emit_cpf_node = emit.find('nfe:CPF', ns)
        
        if emit_cnpj_node is not None:
            emit_cnpj = emit_cnpj_node.text
        elif emit_cpf_node is not None:
            emit_cnpj = emit_cpf_node.text
        else:
            # Se não encontrar nem CNPJ nem CPF, não é uma nota válida para o sistema.
            logger.error(f"XML com chave {chave_acesso} não possui CNPJ ou CPF do emitente.")
            return None
        
        # Acesso seguro ao nome do emitente
        emit_nome_node = emit.find('nfe:xNome', ns)
        emit_nome = emit_nome_node.text if emit_nome_node is not None else "Emitente Desconhecido"
        # --- FIM DA CORREÇÃO ---

        ide = root.find('.//nfe:ide', ns)
        data_emissao_str = getattr(ide.find('nfe:dhEmi', ns), 'text', None) or getattr(ide.find('nfe:dEmi', ns), 'text', None)

        if 'T' in data_emissao_str:
            data_emissao_str = data_emissao_str.split('T')[0]
            
        data_emissao = datetime.fromisoformat(data_emissao_str)

        total = root.find('.//nfe:ICMSTot', ns)
        valor_total = float(total.find('nfe:vNF', ns).text)

        return {
            "chave_acesso": chave_acesso,
            "emitente_cnpj": emit_cnpj,
            "emitente_nome": emit_nome,
            "data_emissao": data_emissao,
            "valor_total": valor_total,
            "xml_content": xml_content_str
        }

    except Exception as e:
        logger.error(f"Erro ao parsear XML: {e}", exc_info=True)
        return None


@app.route('/api/fiscal/importar_xmls', methods=['POST'])
@login_required
@master_required
def api_importar_xmls():
    """
    Recebe múltiplos arquivos XML, processa e salva no banco de dados
    um por um para garantir que erros em um arquivo não afetem os outros.
    """
    if 'xml_files' not in request.files:
        return jsonify({'success': False, 'message': 'Nenhum arquivo enviado.'}), 400

    files = request.files.getlist('xml_files')
    
    novas_notas_dict = []
    notas_ja_existentes = 0
    erros_parse = 0

    for file in files:
        if file and file.filename.lower().endswith('.xml'):
            try:
                xml_content = file.read() 
                dados_nota = parse_nfe_xml_content(xml_content)

                if not dados_nota:
                    erros_parse += 1
                    continue

                chave = dados_nota['chave_acesso']
                
                # A verificação de existência continua sendo uma boa prática
                if NFeImportada.query.filter_by(chave_acesso=chave, empresa_id=current_user.empresa_id).first():
                    notas_ja_existentes += 1
                    continue

                # Cria o objeto para a nova nota
                nova_nfe = NFeImportada(
                    chave_acesso=chave,
                    empresa_id=current_user.empresa_id,
                    nsu='000000000000000',
                    cnpj_consultado=current_user.empresa.cnpj,
                    emitente_cnpj=dados_nota['emitente_cnpj'],
                    emitente_nome=dados_nota['emitente_nome'],
                    data_emissao=dados_nota['data_emissao'],
                    valor_total=dados_nota['valor_total'],
                    xml_content=dados_nota['xml_content'],
                    status='BAIXADA',
                    certificado_id=None
                )
                db.session.add(nova_nfe)
                
                # --- MUDANÇA PRINCIPAL ---
                # Salva (commit) esta nota imediatamente.
                db.session.commit()
                
                # Se o commit foi bem-sucedido, adiciona à lista de resposta.
                novas_notas_dict.append(nova_nfe.to_dict())

            except IntegrityError:
                # Se o commit falhar por chave duplicada (UniqueViolation é um tipo de IntegrityError),
                # desfaz a transação e conta como "já existente".
                db.session.rollback()
                notas_ja_existentes += 1
                logger.warning(f"Chave duplicada encontrada ao tentar salvar: {file.filename}")

            except Exception as e:
                # Para qualquer outro erro, desfaz a transação e conta como erro de parse.
                db.session.rollback()
                logger.error(f"Erro genérico ao processar o arquivo {file.filename}: {e}", exc_info=True)
                erros_parse += 1

    # Monta a mensagem final para o usuário
    msg_sucesso = f"{len(novas_notas_dict)} nova(s) nota(s) importada(s) com sucesso."
    if notas_ja_existentes > 0:
        msg_sucesso += f" {notas_ja_existentes} nota(s) já existia(m) e foram ignorada(s)."
    if erros_parse > 0:
        msg_sucesso += f" {erros_parse} arquivo(s) não puderam ser lidos."
    
    return jsonify({
        'success': True, 
        'message': msg_sucesso,
        'novas_notas': novas_notas_dict
    })
    
@app.after_request
def add_header(response):
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response

from datetime import datetime

@app.cli.command("consultar-sefaz")
def consultar_sefaz_command():
    """
    Comando de terminal para executar a consulta de NFe na SEFAZ para todas as empresas ativas.
    Projetado para ser chamado por um Cron Job.
    """
    logger.info("=====================================================")
    logger.info("INICIANDO TAREFA AGENDADA: Consulta de NFe na SEFAZ")
    logger.info(f"Horário de início: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info("=====================================================")
    
    # É crucial rodar dentro do contexto da aplicação para ter acesso ao banco
    with app.app_context():
        # Busca todas as empresas que têm pelo menos um certificado digital válido
        empresas_para_consultar = db.session.query(Empresa).join(
            CertificadoDigital, Empresa.id == CertificadoDigital.empresa_id
        ).filter(
            CertificadoDigital.data_validade >= datetime.utcnow().date()
        ).distinct().all()

        if not empresas_para_consultar:
            logger.info("Nenhuma empresa com certificados válidos encontrada. Encerrando tarefa.")
            return

        logger.info(f"Encontradas {len(empresas_para_consultar)} empresa(s) para consultar.")
        
        for empresa in empresas_para_consultar:
            logger.info(f"--- Consultando para a empresa: {empresa.razao_social} (ID: {empresa.id}) ---")
            try:
                # A função principal do nosso outro arquivo é chamada aqui
                from sefaz_service import consultar_notas_sefaz
                resultado = consultar_notas_sefaz(empresa.id)
                
                if resultado.get('success'):
                    logger.info(f"Resultado para empresa {empresa.id}: {resultado.get('message', 'Sucesso.')}")
                else:
                    logger.error(f"Falha na consulta para empresa {empresa.id}: {resultado.get('message', 'Erro desconhecido.')}")
            
            except Exception as e:
                logger.error(f"ERRO CRÍTICO ao processar empresa {empresa.id}: {e}", exc_info=True)
                db.session.rollback()
    
    logger.info("=====================================================")
    logger.info("FIM DA TAREFA AGENDADA.")
    logger.info("=====================================================")


if __name__ == '__main__':
    # Inicia o servidor com suporte a SocketIO e debug
    print("Iniciando o servidor com SocketIO...")
    socketio.run(app, debug=True)
